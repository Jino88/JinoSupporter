# -*- coding: utf-8 -*-
"""
엑셀 DRM 우회 사본 생성기 — Excel 로 읽기만 하고 openpyxl 로 새 xlsx 작성.

핵심: Excel 의 SaveAs/SaveCopyAs/Export 등 *저장 경로* 를 일체 호출하지
않는다. 원본 .xlsx 를 Excel COM 으로 *읽기* 만 한 뒤, 셀 값/서식/이미지를
Python 메모리로 옮겨 openpyxl 이 직접 ZIP+XML 을 써내는 흐름.

DRM/DLP 가 Excel 의 저장 API 만 후킹하는 형태라면 이 흐름엔 후킹 지점이
없다 — 결과 *_clean.xlsx 는 그림이 더미바이트로 바꿔치기되지 않는다.

3단계 폴백 전략:
  ⓪ Excel 읽기 + openpyxl 쓰기  (선호)
  ① openpyxl 단독 load+save     (Excel 없이 — 단순 보호 파일에 효과)
  ② zipfile raw 재포장          (최후 폴백 — 그림 한 장도 안 빠짐)

요구사항:
  - Windows + Microsoft Excel 설치
  - pip install pywin32 openpyxl Pillow
"""

import datetime as _dt
import io
import math as _math
import os
import re
import shutil
import subprocess
import sys
import threading
import time
import tkinter as tk
import traceback
import zipfile
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

# ===========================================================================
# 크래시 로그 — GUI 가 갑자기 종료될 때 흔적 남기는 파일 로거.
# ===========================================================================
# 스크립트 옆에 _crash.log 로 append. Tk 위젯에만 의지하면 GUI 가 죽으며
# 같이 사라져 디버깅 단서가 없어진다. 모든 unhandled exception (UI 스레드,
# 워커 스레드, Tk 콜백) 을 잡아 이 파일에 timestamp 와 함께 추적 기록.
_CRASH_LOG_PATH = None


def _crash_log_path() -> Path:
    global _CRASH_LOG_PATH
    if _CRASH_LOG_PATH is None:
        try:
            base = Path(__file__).parent
        except NameError:
            base = Path.cwd()
        _CRASH_LOG_PATH = base / "excel_drm_clean_crash.log"
    return _CRASH_LOG_PATH


def _crash_log(msg: str):
    """크래시 로그 파일에 append. 어떤 스레드에서든 안전."""
    try:
        ts = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(_crash_log_path(), "a", encoding="utf-8") as f:
            f.write(f"[{ts}] {msg}\n")
            f.flush()
    except Exception:
        pass
    # 가능하면 stderr 에도 — VSCode debug console 등에서 볼 수 있게
    try:
        sys.stderr.write(msg + "\n")
        sys.stderr.flush()
    except Exception:
        pass


def _install_crash_hooks():
    """전역 unhandled exception 훅 설치. UI 스레드 + 워커 스레드 + Tk 콜백 커버."""
    def _main_thread_hook(exc_type, exc_value, exc_tb):
        tb_str = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
        _crash_log(f"!!! MAIN THREAD UNHANDLED EXCEPTION:\n{tb_str}")
        # 기본 동작 (stderr 출력) 도 유지
        sys.__excepthook__(exc_type, exc_value, exc_tb)

    sys.excepthook = _main_thread_hook

    # threading 의 default excepthook (Python 3.8+) — 모든 스레드 unhandled 잡음
    if hasattr(threading, "excepthook"):
        def _thread_hook(args):
            tb_str = "".join(
                traceback.format_exception(args.exc_type, args.exc_value, args.exc_traceback)
            )
            _crash_log(
                f"!!! THREAD '{args.thread.name}' UNHANDLED EXCEPTION:\n{tb_str}"
            )
        threading.excepthook = _thread_hook

try:
    from PIL import Image, ImageGrab
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

try:
    import pythoncom
    import pywintypes
    import win32clipboard
    import win32com.client
    import win32process
    HAS_PYWIN32 = True
except ImportError:
    HAS_PYWIN32 = False

try:
    import openpyxl  # type: ignore
    from openpyxl import Workbook as _OpxWorkbook
    from openpyxl.utils import get_column_letter as _opx_col_letter
    from openpyxl.drawing.image import Image as _OpxImage
    from openpyxl.drawing.spreadsheet_drawing import (
        TwoCellAnchor as _OpxTwoCellAnchor,
        AnchorMarker as _OpxAnchorMarker,
    )
    from openpyxl.styles import (
        Font as _OpxFont,
        PatternFill as _OpxFill,
        Alignment as _OpxAlign,
        Border as _OpxBorder,
        Side as _OpxSide,
    )
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# tkinterdnd2 — Tk 자체에는 OS 드래그앤드롭이 없어 외부 바인딩 필요.
# 미설치여도 앱은 정상 동작 (드롭만 비활성).
try:
    from tkinterdnd2 import TkinterDnD, DND_FILES  # type: ignore
    HAS_TKDND = True
except ImportError:
    HAS_TKDND = False


# ===========================================================================
# 상수
# ===========================================================================
# 1 포인트 = 1/72 인치 = 12700 EMU (English Metric Units, OOXML 좌표 단위)
_EMU_PER_POINT = 12700

# 클립보드는 시스템 단일 자원이라 직렬화 필요.
CLIPBOARD_LOCK = threading.Lock()

# Excel COM 상수.
_XL_NONE = -4142
_XL_LEFT = -4131
_XL_CENTER = -4108
_XL_RIGHT = -4152
_XL_TOP = -4160
_XL_BOTTOM = -4107
_XL_BORDER_TOP = 8
_XL_BORDER_LEFT = 7
_XL_BORDER_BOTTOM = 9
_XL_BORDER_RIGHT = 10
_XL_LINESTYLE_TO_OPX = {
    1: "thin",         # xlContinuous
    -4115: "dashed",   # xlDash
    -4118: "dotted",   # xlDot
    -4119: "double",   # xlDouble
    13: "dashDotDot",  # xlSlantDashDot
    4: "dashDot",      # xlDashDot
}
_XL_WEIGHT_TO_OPX = {
    1: "thin",
    2: "thin",
    -4138: "medium",
    4: "thick",
}

_INVALID_NAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')

# UsedRange 가 시트 전체(XFD열·1048576행)로 잡혀 있을 때 — 사용하지 않는
# 영역에 서식만 발려 있는 케이스 — per-cell COM 읽기로 시간이 폭발한다.
# 처리 영역을 이 한도로 클램프해 1000×1000 안쪽만 전사한다.
MAX_RANGE_ROWS = 1000
MAX_RANGE_COLS = 1000


# ===========================================================================
# 파일/문자열 헬퍼
# ===========================================================================
def safe_name(name: str) -> str:
    """Windows 에서 안전한 파일/폴더 이름. 불법 문자 → '_', 후행/선행 공백·점 제거."""
    cleaned = _INVALID_NAME_CHARS.sub("_", name)
    cleaned = cleaned.strip(" .")
    return cleaned or "untitled"


# ===========================================================================
# 클립보드 헬퍼
# ===========================================================================
def _open_clipboard_retry(timeout_sec: float = 1.5):
    """클립보드 매니저/백신이 잡고 있을 때를 대비해 짧은 재시도."""
    deadline = time.time() + timeout_sec
    last_exc = None
    while time.time() < deadline:
        try:
            win32clipboard.OpenClipboard()
            return
        except Exception as exc:
            last_exc = exc
            time.sleep(0.05)
    if last_exc:
        raise last_exc


def empty_clipboard_safely():
    """
    클립보드 내용을 비우는 best-effort 시도.

    Shape.CopyPicture / Range.CopyPicture 직전에 호출해 두면, 그 호출이
    내부적으로 실패했을 때 ImageGrab.grabclipboard() 가 *이전 잔여 비트맵*
    (예: 사용자가 직전에 Outlook 사인 이미지를 복사해 둔 것) 을 우리 도형의
    렌더 결과로 오인하는 사고를 막을 수 있다.
    """
    try:
        _open_clipboard_retry(timeout_sec=0.5)
    except Exception:
        return
    try:
        win32clipboard.EmptyClipboard()
    except Exception:
        pass
    finally:
        try:
            win32clipboard.CloseClipboard()
        except Exception:
            pass


# ===========================================================================
# 도형 → PNG 렌더링
# ===========================================================================
def capture_range_as_png(rng, log=None):
    """Range 전체를 단일 PNG 비트맵으로 캡처."""
    if not HAS_PIL:
        return None
    for appearance in (2, 1):  # xlPrinter 우선
        try:
            empty_clipboard_safely()
            rng.CopyPicture(Appearance=appearance, Format=2)
            try:
                pythoncom.OleFlushClipboard()
            except Exception:
                pass
            time.sleep(0.2)
            img = ImageGrab.grabclipboard()
            if img is None or isinstance(img, list):
                continue
            if img.mode == "RGBA" or img.info.get("transparency") is not None:
                img = img.convert("RGBA")
            else:
                img = img.convert("RGB")
            buf = io.BytesIO()
            img.save(buf, format="PNG", optimize=False, compress_level=6)
            return buf.getvalue()
        except Exception:
            pass
    return None


def render_shape_as_png(shape, log=None):
    """
    Excel Shape 을 비트맵으로 렌더링해서 PNG 바이트로 반환.

    여러 클립보드 메서드를 차례로 시도하고 가장 큰(=실제 콘텐츠 가능성 큰)
    결과를 채택. OLE-embedded 그림은 Shape.CopyPicture 가 placeholder 만
    돌려주는 경우가 있는데, Range.CopyPicture 로 도형의 앵커 셀 영역을
    캡처하면 Excel 화면에 그려진 그대로 받을 수 있다.

    시도 순서:
      1. Range.CopyPicture(xlScreen)
      2. Range.CopyPicture(xlPrinter)
      3. Shape.CopyPicture(xlPrinter)
      4. Shape.CopyPicture(xlScreen)
      5. Shape.Copy

    가장 면적이 큰 유효 결과 채택. 30x30 미만은 placeholder 로 간주해 거름.
    """
    if not HAS_PIL:
        return None

    def make_range_copypicture(appearance):
        def fn():
            ws = shape.Parent
            tl = shape.TopLeftCell
            br = shape.BottomRightCell
            rng = ws.Range(tl, br)
            rng.CopyPicture(Appearance=appearance, Format=2)
        return fn

    attempts = [
        ("Range.CopyPicture xlScreen", make_range_copypicture(1)),
        ("Range.CopyPicture xlPrinter", make_range_copypicture(2)),
        ("Shape.CopyPicture xlPrinter",
            lambda: shape.CopyPicture(Appearance=2, Format=2)),
        ("Shape.CopyPicture xlScreen",
            lambda: shape.CopyPicture(Appearance=1, Format=2)),
        ("Shape.Copy", lambda: shape.Copy()),
    ]

    candidates = []  # (area, png_size, png_bytes, method_name, w, h)

    for name, action in attempts:
        try:
            empty_clipboard_safely()
            action()
            try:
                pythoncom.OleFlushClipboard()
            except Exception:
                pass
            time.sleep(0.10)
            img = ImageGrab.grabclipboard()
            if img is None or isinstance(img, list):
                continue
            w, h = img.size
            if w < 30 or h < 30:
                continue
            if img.mode == "RGBA" or img.info.get("transparency") is not None:
                img = img.convert("RGBA")
            else:
                img = img.convert("RGB")
            buf = io.BytesIO()
            img.save(buf, format="PNG", optimize=False, compress_level=6)
            png = buf.getvalue()
            candidates.append((w * h, len(png), png, name, w, h))
        except Exception:
            pass

    if not candidates:
        return None

    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    chosen = candidates[0]
    if log:
        log(
            f"    선택: {chosen[3]} ({chosen[4]}x{chosen[5]}, "
            f"{chosen[1]:,} B / 후보 {len(candidates)}개)"
        )
    return chosen[2]


# ===========================================================================
# openpyxl 서식 전사 헬퍼
# ===========================================================================
def _excel_bgr_to_argb(color_int):
    """Excel COM 의 BGR 정수 색상 → openpyxl ARGB hex (FFRRGGBB)."""
    try:
        c = int(color_int)
    except Exception:
        return None
    if c < 0:
        return None
    b = (c >> 16) & 0xFF
    g = (c >> 8) & 0xFF
    r = c & 0xFF
    return f"FF{r:02X}{g:02X}{b:02X}"


def _ha_map(v):
    return {_XL_LEFT: "left", _XL_CENTER: "center", _XL_RIGHT: "right"}.get(int(v))


def _va_map(v):
    return {_XL_TOP: "top", _XL_CENTER: "center", _XL_BOTTOM: "bottom"}.get(int(v))


# 도형이 많은 파일은 PNG 렌더링이 도형당 수백 ms 이상 걸려서 전체 변환이
# 분 단위로 늘어난다. drm_clean_resave 진입 직후 사전 스캔으로 합산 도형 수가
# 임계치를 넘는 파일은 변환 대상에서 빼고 Heavy/ 폴더로 원본만 이동.
HEAVY_SHAPE_THRESHOLD = 100


def _open_workbook_robust(excel, path: str):
    """Excel.Workbooks.Open 의 late-binding 이슈 회피.

    pywin32 가 가끔 `.Open()` 결과를 타입 미지정 dispatch 로 돌려줘서
    `.Sheets` 같은 속성 접근이 `AttributeError: Open.Sheets` 로 깨진다.
    Open 직후 Workbooks 인덱스로 다시 잡으면 typed 객체로 받혀 우회됨.
    """
    excel.Workbooks.Open(path, ReadOnly=True, UpdateLinks=0)
    return excel.Workbooks(int(excel.Workbooks.Count))


# COM Range.NumberFormat 는 로케일 언어 접두사 ([ENG], [KOR], [JPN]…) 를 끼워서
# 돌려주는데, Excel 의 OOXML 파서는 이 형태를 거부해서 styles.xml 복구 메시지를
#낸다. LCID 매핑은 일부만 알려져 있어 안전하게 알파벳 3~4자 언어 접두사를
# 통째로 제거한다 (dd/mmm 같은 영문 약자는 로케일 없이도 동작).
_LANG_PREFIX_RE = re.compile(r"\[[A-Za-z]{2,4}\]")


def _sanitize_numfmt(fmt):
    """numFmt 코드에서 Excel 거부 요소 제거. None / "General" / 비ASCII 는 None 반환."""
    if not fmt:
        return None
    s = str(fmt).strip()
    if not s or s == "General":
        return None
    s = _LANG_PREFIX_RE.sub("", s).strip()
    if not s or s == "General":
        return None
    if not all(ord(ch) < 128 for ch in s):
        return None  # 한국어 "G/표준" 등 로케일화 General
    return s


def _copy_cell_format(src_cell, dest_cell):
    """COM Cell → openpyxl Cell 서식 전사 (best-effort, 모든 속성 try/except)."""
    # 글꼴
    try:
        f = src_cell.Font
        kwargs = {}
        try:
            kwargs["name"] = str(f.Name)
        except Exception:
            pass
        try:
            kwargs["size"] = float(f.Size)
        except Exception:
            pass
        try:
            if bool(f.Bold):
                kwargs["bold"] = True
        except Exception:
            pass
        try:
            if bool(f.Italic):
                kwargs["italic"] = True
        except Exception:
            pass
        try:
            argb = _excel_bgr_to_argb(f.Color)
            if argb and argb != "FF000000":
                kwargs["color"] = argb
        except Exception:
            pass
        if kwargs:
            dest_cell.font = _OpxFont(**kwargs)
    except Exception:
        pass

    # 배경
    try:
        interior = src_cell.Interior
        try:
            pat = int(interior.Pattern)
        except Exception:
            pat = _XL_NONE
        if pat != _XL_NONE:
            argb = _excel_bgr_to_argb(interior.Color)
            if argb and argb != "FFFFFFFF":
                dest_cell.fill = _OpxFill(fill_type="solid", fgColor=argb)
    except Exception:
        pass

    # 정렬
    try:
        h = _ha_map(src_cell.HorizontalAlignment)
        v = _va_map(src_cell.VerticalAlignment)
        try:
            wrap = bool(src_cell.WrapText)
        except Exception:
            wrap = False
        if h or v or wrap:
            dest_cell.alignment = _OpxAlign(
                horizontal=h, vertical=v, wrap_text=wrap or None
            )
    except Exception:
        pass

    # 숫자 서식 — _sanitize_numfmt 가 로케일 General / 언어 접두사 처리.
    try:
        fmt = _sanitize_numfmt(src_cell.NumberFormat)
        if fmt:
            dest_cell.number_format = fmt
    except Exception:
        pass

    # 테두리
    try:
        sides = {}
        for excel_idx, opx_name in (
            (_XL_BORDER_TOP, "top"),
            (_XL_BORDER_RIGHT, "right"),
            (_XL_BORDER_BOTTOM, "bottom"),
            (_XL_BORDER_LEFT, "left"),
        ):
            try:
                b = src_cell.Borders(excel_idx)
                ls = int(b.LineStyle)
                if ls == _XL_NONE:
                    continue
                style = None
                try:
                    style = _XL_WEIGHT_TO_OPX.get(int(b.Weight))
                except Exception:
                    pass
                if not style:
                    style = _XL_LINESTYLE_TO_OPX.get(ls, "thin")
                argb = _excel_bgr_to_argb(b.Color) or "FF000000"
                sides[opx_name] = _OpxSide(style=style, color=argb)
            except Exception:
                pass
        if sides:
            dest_cell.border = _OpxBorder(**sides)
    except Exception:
        pass


def _bulk_subdivide(rng, get_value, apply_uniform, _depth=0, _calls=None):
    """이분 분할로 균질 영역 탐색.

    get_value(rng) -> 균질값 또는 None
        Excel COM 의 속성 (Range.Font.Bold, Range.Interior.Color 등) 은
        다중 셀 Range 에서 모든 셀이 같으면 그 값을, 혼합이면 None 을 반환.
        이 시그널을 받아 None 이면 절반으로 잘라 재귀.
    apply_uniform(first_row, first_col, last_row, last_col, value)
        균질로 판명난 영역에 한 번에 값 적용. 셀별 누적 dict 에 기록.

    복잡도: O(서식블록수 × log(셀수)). 시트가 행 단위로 균질하면 (헤더 굵음
    + 데이터 평범 + 합계 색칠) 1000×1000 = 1M셀 시트에서 ~수백 COM 호출.
    """
    if _calls is not None:
        _calls[0] += 1
    try:
        first_row = int(rng.Row)
        first_col = int(rng.Column)
        n_rows = int(rng.Rows.Count)
        n_cols = int(rng.Columns.Count)
    except Exception:
        return
    if n_rows == 0 or n_cols == 0:
        return

    try:
        val = get_value(rng)
    except Exception:
        val = None  # 속성 접근 자체가 실패하면 혼합 취급

    if val is not None:
        try:
            apply_uniform(
                first_row, first_col,
                first_row + n_rows - 1, first_col + n_cols - 1, val,
            )
        except Exception:
            pass
        return

    # 더 잘게 못 쪼갬 — 1×1 인데 val 이 None 이면 그냥 포기 (드문 케이스)
    if n_rows <= 1 and n_cols <= 1:
        return

    # 큰 축 방향으로 절반 자르기 — 행 우선 (실제 서식은 행 단위 균질이 흔함)
    if n_rows >= n_cols and n_rows > 1:
        mid = n_rows // 2
        try:
            top = rng.Resize(mid, n_cols)
            bot = rng.Offset(mid, 0).Resize(n_rows - mid, n_cols)
        except Exception:
            return
        _bulk_subdivide(top, get_value, apply_uniform, _depth + 1, _calls)
        _bulk_subdivide(bot, get_value, apply_uniform, _depth + 1, _calls)
    elif n_cols > 1:
        mid = n_cols // 2
        try:
            left = rng.Resize(n_rows, mid)
            right = rng.Offset(0, mid).Resize(n_rows, n_cols - mid)
        except Exception:
            return
        _bulk_subdivide(left, get_value, apply_uniform, _depth + 1, _calls)
        _bulk_subdivide(right, get_value, apply_uniform, _depth + 1, _calls)


def _apply_formats_subdivision(ws_src, used, ws_new, content_cells, log, sheet_name):
    """이분 분할로 시트 전체 서식 + 병합 일괄 복사.

    각 속성 (Bold/Italic/색/정렬/...) 마다 별도 트리 순회. 셀당 COM 호출 0번.
    셀별 누적 dict 에 기록한 뒤 마지막에 openpyxl Font/Fill/Alignment 로 변환.

    수집 항목: Bold, Italic, Font name/size/color, 배경색, 가로/세로 정렬,
    WrapText, NumberFormat, 병합.
    테두리는 Range.Borders 의미가 셀별이 아니라 영역 가장자리라 이 트릭이
    안 통함 — 균등 분할 모드에선 스킵. 필요하면 정밀 모드 사용.
    """
    cell_styles = {}  # (r,c) -> {bold, italic, font_name, font_size, font_color,
                      #          fill, halign, valign, wrap, numfmt}
    # 속성별 어디까지 채워졌는지 진단 — 각 패스에서 stamp 가 호출된 영역 수.
    # 0이 많이 나오면 Excel COM 이 mixed 속성을 None 으로 안 돌려주는 환경.
    region_counts = {}

    def stamp(r1, c1, r2, c2, key, val):
        # content_cells 에 있는 셀만 — 값 없는 셀은 시각적 의미 거의 없음
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                if (r, c) in content_cells:
                    cs = cell_styles.get((r, c))
                    if cs is None:
                        cs = {}
                        cell_styles[(r, c)] = cs
                    cs[key] = val
        region_counts[key] = region_counts.get(key, 0) + 1

    calls = [0]
    t0 = time.perf_counter()
    log(f"    [{sheet_name}] 균등 분할 서식 추출 시작…")

    # ---- 굵게 ----
    def _get_bold(rng):
        v = rng.Font.Bold
        return None if v is None else bool(v)
    def _apply_bold(r1, c1, r2, c2, v):
        if v:  # False 는 기본값 — 적용 안 함
            stamp(r1, c1, r2, c2, "bold", True)
    _bulk_subdivide(used, _get_bold, _apply_bold, _calls=calls)

    # ---- 기울임 ----
    def _get_italic(rng):
        v = rng.Font.Italic
        return None if v is None else bool(v)
    def _apply_italic(r1, c1, r2, c2, v):
        if v:
            stamp(r1, c1, r2, c2, "italic", True)
    _bulk_subdivide(used, _get_italic, _apply_italic, _calls=calls)

    # ---- 글꼴 이름 ----
    def _get_font_name(rng):
        v = rng.Font.Name
        return str(v) if v is not None else None
    def _apply_font_name(r1, c1, r2, c2, v):
        if v:
            stamp(r1, c1, r2, c2, "font_name", v)
    _bulk_subdivide(used, _get_font_name, _apply_font_name, _calls=calls)

    # ---- 글꼴 크기 ----
    def _get_font_size(rng):
        v = rng.Font.Size
        try:
            return float(v) if v is not None else None
        except Exception:
            return None
    def _apply_font_size(r1, c1, r2, c2, v):
        stamp(r1, c1, r2, c2, "font_size", v)
    _bulk_subdivide(used, _get_font_size, _apply_font_size, _calls=calls)

    # ---- 글꼴색 ----
    def _get_font_color(rng):
        try:
            v = rng.Font.Color
            return None if v is None else int(v)
        except Exception:
            return None
    def _apply_font_color(r1, c1, r2, c2, v):
        argb = _excel_bgr_to_argb(v)
        if argb and argb != "FF000000":  # 검정 기본값 스킵
            stamp(r1, c1, r2, c2, "font_color", argb)
    _bulk_subdivide(used, _get_font_color, _apply_font_color, _calls=calls)

    # ---- 배경색 — Pattern 우선 체크, xlNone 영역은 색칠 안 함 ----
    def _get_fill_color(rng):
        try:
            pat = rng.Interior.Pattern
            if pat is None:
                return None  # 패턴 자체가 혼합 — 더 쪼개야
            if int(pat) == _XL_NONE:
                return -1   # 명시적 "칠 없음" sentinel — apply 에서 스킵
            v = rng.Interior.Color
            return None if v is None else int(v)
        except Exception:
            return None
    def _apply_fill_color(r1, c1, r2, c2, v):
        if v == -1:
            return  # 패턴 없음 — 칠하지 않음
        argb = _excel_bgr_to_argb(v)
        if argb and argb != "FFFFFFFF":  # 흰색 스킵
            stamp(r1, c1, r2, c2, "fill", argb)
    _bulk_subdivide(used, _get_fill_color, _apply_fill_color, _calls=calls)

    # ---- 가로 정렬 ----
    def _get_halign(rng):
        try:
            v = rng.HorizontalAlignment
            if v is None:
                return None
            mapped = _ha_map(v)
            return mapped if mapped else "__none__"  # General → sentinel
        except Exception:
            return None
    def _apply_halign(r1, c1, r2, c2, v):
        if v and v != "__none__":
            stamp(r1, c1, r2, c2, "halign", v)
    _bulk_subdivide(used, _get_halign, _apply_halign, _calls=calls)

    # ---- 세로 정렬 ----
    def _get_valign(rng):
        try:
            v = rng.VerticalAlignment
            if v is None:
                return None
            mapped = _va_map(v)
            return mapped if mapped else "__none__"
        except Exception:
            return None
    def _apply_valign(r1, c1, r2, c2, v):
        if v and v != "__none__":
            stamp(r1, c1, r2, c2, "valign", v)
    _bulk_subdivide(used, _get_valign, _apply_valign, _calls=calls)

    # ---- WrapText ----
    def _get_wrap(rng):
        try:
            v = rng.WrapText
            return None if v is None else bool(v)
        except Exception:
            return None
    def _apply_wrap(r1, c1, r2, c2, v):
        if v:
            stamp(r1, c1, r2, c2, "wrap", True)
    _bulk_subdivide(used, _get_wrap, _apply_wrap, _calls=calls)

    # ---- NumberFormat — _sanitize_numfmt 가 로케일 General/언어 접두사 처리 ----
    def _get_numfmt(rng):
        try:
            return _sanitize_numfmt(rng.NumberFormat) or "__skip__"
        except Exception:
            return None
    def _apply_numfmt(r1, c1, r2, c2, v):
        if v and v != "__skip__":
            stamp(r1, c1, r2, c2, "numfmt", v)
    _bulk_subdivide(used, _get_numfmt, _apply_numfmt, _calls=calls)

    fmt_elapsed = time.perf_counter() - t0

    # ---- 병합 ---- (val=True 영역의 MergeArea 를 진짜 병합 경계로 사용)
    t_merge = time.perf_counter()
    seen_merges = set()
    def _get_merge(rng):
        try:
            v = rng.MergeCells
            return None if v is None else bool(v)
        except Exception:
            return None
    def _apply_merge(r1, c1, r2, c2, v):
        if not v:
            return
        # val=True 인 영역의 TopLeft 셀이 속한 병합 경계가 진짜 병합 좌표
        try:
            area = ws_src.Cells(r1, c1).MergeArea
            mr1 = int(area.Row)
            mc1 = int(area.Column)
            mr2 = mr1 + int(area.Rows.Count) - 1
            mc2 = mc1 + int(area.Columns.Count) - 1
            key = (mr1, mc1, mr2, mc2)
            if key in seen_merges:
                return
            seen_merges.add(key)
            if mr1 == mr2 and mc1 == mc2:
                return  # 1×1 은 병합 아님
            ws_new.merge_cells(
                start_row=mr1, start_column=mc1,
                end_row=mr2, end_column=mc2,
            )
        except Exception:
            pass
    _bulk_subdivide(used, _get_merge, _apply_merge, _calls=calls)
    merge_elapsed = time.perf_counter() - t_merge

    # ---- dict → openpyxl 객체 변환 (셀별 적용) ----
    t_write = time.perf_counter()
    styled = 0
    for (r, c), cs in cell_styles.items():
        try:
            dest = ws_new.cell(row=r, column=c)

            font_kw = {}
            if "font_name" in cs:
                font_kw["name"] = cs["font_name"]
            if "font_size" in cs:
                font_kw["size"] = cs["font_size"]
            if cs.get("bold"):
                font_kw["bold"] = True
            if cs.get("italic"):
                font_kw["italic"] = True
            if "font_color" in cs:
                font_kw["color"] = cs["font_color"]
            if font_kw:
                dest.font = _OpxFont(**font_kw)

            if "fill" in cs:
                dest.fill = _OpxFill(fill_type="solid", fgColor=cs["fill"])

            if cs.get("halign") or cs.get("valign") or cs.get("wrap"):
                dest.alignment = _OpxAlign(
                    horizontal=cs.get("halign"),
                    vertical=cs.get("valign"),
                    wrap_text=cs.get("wrap") or None,
                )

            if "numfmt" in cs:
                dest.number_format = cs["numfmt"]

            styled += 1
        except Exception:
            pass
    write_elapsed = time.perf_counter() - t_write

    # 속성별 채워진 영역 수 — 0 인 속성은 mixed 가 None 으로 안 돌아오는 환경
    # 가능성. 모든 속성이 0 이면 _apply_formats_subdivision 자체가 무용지물.
    if region_counts:
        counts_str = ", ".join(
            f"{k}={v}" for k, v in sorted(region_counts.items())
        )
    else:
        counts_str = "(채워진 영역 없음 — Excel COM 이 mixed 속성을 None 으로 안 돌려줌)"
    log(
        f"    [{sheet_name}] 균등 분할 완료: 서식 {styled:,}셀 / "
        f"병합 {len(seen_merges)}개 / COM {calls[0]:,}회 "
        f"(서식 {fmt_elapsed:.2f}s + 병합 {merge_elapsed:.2f}s + "
        f"쓰기 {write_elapsed:.2f}s)"
    )
    log(f"    [{sheet_name}] 속성별 영역: {counts_str}")


# ===========================================================================
# 클립보드 "XML Spreadsheet" (SpreadsheetML 2003) 경유 시트 전사
# ===========================================================================
# Excel 이 Ctrl+C 시 클립보드에 깔아주는 20+ 포맷 중, "XML Spreadsheet" 는
# 공개 스펙 (MS-XLSP) 의 구조화된 XML — 셀 값/수식/스타일/병합/컬럼폭/행높이가
# 모두 한 덩어리. CF_HTML 처럼 mso-* 로 더럽혀지지 않음. CF_DIB 처럼 비트맵도
# 아님. 시트당 클립보드 1회 호출로 데이터 + 서식 + 병합 다 처리.
#
# 한계:
#   - SpreadsheetML 2003 은 65,536 행 제한 (보통 실무 시트는 안쪽)
#   - 수식은 R1C1 표기법 — A1 변환 미구현 → XMLSS 모드는 계산값만 보존
#   - 조건부 서식/차트 미포함
#   - 도형은 별도 (Shape.CopyPicture 그대로 사용)
# ===========================================================================
_XMLSS_NS = "urn:schemas-microsoft-com:office:spreadsheet"


class _PhaseSkip(Exception):
    """clipboard XMLSS 가 이미 처리한 phase 를 통째로 건너뛰기 위한 sentinel.

    기존 phase 의 `try: ... except Exception: log(...)` 구조에 한 줄만 더해서
    스킵할 수 있게 만들기 위한 장치. 각 phase try 본문 첫 줄에서
    `if clipboard_done: raise _PhaseSkip()` 으로 던지고, 동일 try 의
    `except _PhaseSkip: pass` 가 (Exception 보다 먼저) 잡는다.
    """
    pass


def _xmlss_q(tag: str) -> str:
    """ElementTree 의 Clark notation 으로 SS 네임스페이스 태그 만들기."""
    return f"{{{_XMLSS_NS}}}{tag}"


def _xmlss_parse_data(data_elem):
    """<Data ss:Type="..."> 의 값을 Python 객체로 변환."""
    type_attr = data_elem.get(_xmlss_q("Type"))
    text = data_elem.text or ""
    if type_attr == "Number":
        try:
            v = float(text)
            return int(v) if v.is_integer() else v
        except (ValueError, TypeError):
            return text
    if type_attr == "Boolean":
        return text == "1" or text.upper() == "TRUE"
    if type_attr == "DateTime":
        try:
            return _dt.datetime.fromisoformat(text.rstrip("Z"))
        except Exception:
            return text
    if type_attr == "Error":
        return text  # "#REF!", "#NAME?", "#DIV/0!" 등 — 문자열로 보존
    return text or None  # String / 미지정


def _parse_xmlss_style(style_elem):
    """<Style ss:ID="..."> 엘리먼트를 openpyxl 객체 dict 로.

    Returns: {"font": Font, "fill": PatternFill, "border": Border,
              "alignment": Alignment, "number_format": str} — 있는 키만.
    """
    result = {}

    # Font
    font = style_elem.find(_xmlss_q("Font"))
    if font is not None:
        kw = {}
        name = font.get(_xmlss_q("FontName"))
        if name:
            kw["name"] = name
        size = font.get(_xmlss_q("Size"))
        if size:
            try:
                s = float(size)
                # Excel 허용 폰트 크기 (1 ~ 409). 범위 밖이면 Excel 거부.
                if 1 <= s <= 409:
                    kw["size"] = s
            except (ValueError, TypeError):
                pass
        if font.get(_xmlss_q("Bold")) == "1":
            kw["bold"] = True
        if font.get(_xmlss_q("Italic")) == "1":
            kw["italic"] = True
        if font.get(_xmlss_q("StrikeThrough")) == "1":
            kw["strike"] = True
        color = font.get(_xmlss_q("Color"))
        if color and len(color) == 7 and color.startswith("#"):
            argb = f"FF{color[1:].upper()}"
            if argb != "FF000000":  # 기본 검정 스킵
                kw["color"] = argb
        underline = font.get(_xmlss_q("Underline"))
        if underline:
            kw["underline"] = {
                "Single": "single",
                "Double": "double",
                "SingleAccounting": "singleAccounting",
                "DoubleAccounting": "doubleAccounting",
            }.get(underline, "single")
        if kw:
            try:
                result["font"] = _OpxFont(**kw)
            except Exception:
                pass

    # Interior (Fill)
    interior = style_elem.find(_xmlss_q("Interior"))
    if interior is not None:
        pattern = interior.get(_xmlss_q("Pattern"))
        color = interior.get(_xmlss_q("Color"))
        if (
            pattern and pattern not in ("None", "")
            and color and color.startswith("#") and len(color) == 7
        ):
            argb = f"FF{color[1:].upper()}"
            if argb != "FFFFFFFF":  # 흰색 스킵 (시각적 의미 없음)
                try:
                    result["fill"] = _OpxFill(fill_type="solid", fgColor=argb)
                except Exception:
                    pass

    # Borders
    borders = style_elem.find(_xmlss_q("Borders"))
    if borders is not None:
        ls_map = {
            "Continuous": "thin",
            "Dash": "dashed",
            "Dot": "dotted",
            "DashDot": "dashDot",
            "DashDotDot": "dashDotDot",
            "Double": "double",
            "SlantDashDot": "slantDashDot",
        }
        pos_map = {
            "Left": "left", "Right": "right",
            "Top": "top", "Bottom": "bottom",
        }
        sides = {}
        for b in borders.findall(_xmlss_q("Border")):
            pos = b.get(_xmlss_q("Position"))
            opx_pos = pos_map.get(pos)
            if not opx_pos:
                continue  # Diagonal 은 일단 스킵
            line_style = b.get(_xmlss_q("LineStyle"))
            if not line_style or line_style == "None":
                continue
            opx_style = ls_map.get(line_style, "thin")
            # Weight 가 더 강하면 medium/thick 으로 격상
            weight = b.get(_xmlss_q("Weight"))
            if weight:
                try:
                    w = int(weight)
                    if w == 0:
                        opx_style = "hair"
                    elif w == 2:
                        opx_style = "medium"
                    elif w == 3:
                        opx_style = "thick"
                except (ValueError, TypeError):
                    pass
            color = b.get(_xmlss_q("Color"))
            opx_color = "FF000000"
            if color and color.startswith("#") and len(color) == 7:
                opx_color = f"FF{color[1:].upper()}"
            try:
                sides[opx_pos] = _OpxSide(style=opx_style, color=opx_color)
            except Exception:
                pass
        if sides:
            try:
                result["border"] = _OpxBorder(**sides)
            except Exception:
                pass

    # Alignment — Excel 시스템 기본 (horizontal=General, vertical=Bottom) 은
    # 명시 안 하는 게 정답. XMLSS Default 가 흔히 Vertical="Bottom" 만 들어 있는데
    # 그걸 그대로 cellXf 에 넣으면 styles.xml 에 의미 없는 중복 xf 가 쌓이고
    # Excel 이 "복구 필요" 트리거. Bottom-only 는 스킵.
    align = style_elem.find(_xmlss_q("Alignment"))
    if align is not None:
        h = align.get(_xmlss_q("Horizontal"))
        v = align.get(_xmlss_q("Vertical"))
        wrap = align.get(_xmlss_q("WrapText"))
        h_map = {
            "Left": "left", "Center": "center", "Right": "right",
            "Justify": "justify", "CenterAcrossSelection": "centerContinuous",
            "Fill": "fill", "Distributed": "distributed",
        }
        v_map = {
            "Top": "top", "Center": "center",
            "Justify": "justify", "Distributed": "distributed",
            # "Bottom" 의도적 제외 — Excel 기본값이라 명시하면 styles.xml 오염
        }
        h_val = h_map.get(h) if h else None
        v_val = v_map.get(v) if v else None
        wrap_val = (wrap == "1")
        if h_val or v_val or wrap_val:
            try:
                result["alignment"] = _OpxAlign(
                    horizontal=h_val, vertical=v_val,
                    wrap_text=wrap_val or None,
                )
            except Exception:
                pass

    # NumberFormat —
    #   1) 비ASCII (한국어 "G/표준" 등) 는 스킵 — 로케일화된 General 함정
    #   2) XMLSS 의 *이름* 포맷 (실제 코드 아닌 라벨) 은 실제 Excel 코드로 변환.
    #      "Standard" 같은 이름을 그대로 numFmt 로 박으면 Excel 이 못 읽음 →
    #      styles.xml 복구 메시지.
    _NAMED_NUMFMT = {
        "General Number": "General",
        "General Date": "m/d/yyyy h:mm",
        "Long Date": "[$-F800]dddd, mmmm dd, yyyy",
        "Medium Date": "mmm-dd-yy",
        "Short Date": "m/d/yyyy",
        "Long Time": "h:mm:ss AM/PM",
        "Medium Time": "h:mm AM/PM",
        "Short Time": "h:mm",
        "Currency": '"$"#,##0.00_);[Red]("$"#,##0.00)',
        "Euro Currency": "[$€-2] #,##0.00",
        "Fixed": "0.00",
        "Standard": "#,##0.00",
        "Percent": "0.00%",
        "Scientific": "0.00E+00",
        "Yes/No": '"Yes";"Yes";"No"',
        "True/False": '"True";"True";"False"',
        "On/Off": '"On";"On";"Off"',
    }
    nf = style_elem.find(_xmlss_q("NumberFormat"))
    if nf is not None:
        fmt = nf.get(_xmlss_q("Format"))
        if fmt and fmt != "General":
            if fmt in _NAMED_NUMFMT:
                result["number_format"] = _NAMED_NUMFMT[fmt]
            else:
                cleaned = _sanitize_numfmt(fmt)
                if cleaned:
                    result["number_format"] = cleaned

    return result


def _copy_sheet_via_xmlss(ws_src, ws_new, log, sheet_name):
    """클립보드 'XML Spreadsheet' 포맷 → openpyxl 매핑.

    UsedRange.Copy() 한 번으로 값/스타일/병합/폭/높이가 모두 단일 XML 로
    클립보드에 떨어짐. 그걸 파싱해서 ws_new 에 직접 매핑.

    Returns: (content_cells set, used_range_com_object)
        content_cells — 도형 phase 의 빈 셀 스킵 판정용
        used — 도형 위치 절대좌표 계산용 (그대로 image phase 에 흘림)
    """
    used = ws_src.UsedRange
    base_row = int(used.Row)
    base_col = int(used.Column)
    n_rows = int(used.Rows.Count)
    n_cols = int(used.Columns.Count)

    # 너무 큰 UsedRange 는 클립보드 크기 폭주 — 기존 코드와 동일 한도로 클램프
    if n_rows > MAX_RANGE_ROWS or n_cols > MAX_RANGE_COLS:
        new_rows = min(n_rows, MAX_RANGE_ROWS)
        new_cols = min(n_cols, MAX_RANGE_COLS)
        log(
            f"    [{sheet_name}] ⚠ UsedRange {n_rows}x{n_cols} 너무 큼 → "
            f"{new_rows}x{new_cols} 로 제한"
        )
        used = ws_src.Range(
            ws_src.Cells(base_row, base_col),
            ws_src.Cells(base_row + new_rows - 1, base_col + new_cols - 1),
        )

    log(f"    [{sheet_name}] 클립보드 XMLSS — Range.Copy 호출")
    t0 = time.perf_counter()

    # 클립보드 비우고 Copy. OleFlushClipboard 로 Excel 이 모든 포맷을 떨굴
    # 시간을 충분히 줌.
    empty_clipboard_safely()
    try:
        used.Copy()
    except Exception as exc:
        raise RuntimeError(f"Range.Copy 실패: {exc}") from exc
    try:
        pythoncom.OleFlushClipboard()
    except Exception:
        pass
    time.sleep(0.2)

    # 클립보드 → XMLSS 바이트 추출. UI 의 클립보드 폴러와 경합 가능 → 재시도.
    cf_xmlss = win32clipboard.RegisterClipboardFormat("XML Spreadsheet")
    xml_data = None
    last_exc = None
    for _ in range(15):
        try:
            win32clipboard.OpenClipboard()
            try:
                if not win32clipboard.IsClipboardFormatAvailable(cf_xmlss):
                    raise RuntimeError("XML Spreadsheet 포맷이 클립보드에 없음")
                xml_data = win32clipboard.GetClipboardData(cf_xmlss)
            finally:
                try:
                    win32clipboard.CloseClipboard()
                except Exception:
                    pass
            break
        except Exception as exc:
            last_exc = exc
            time.sleep(0.05)

    # Copy 모드 해제 — Excel 의 marching ants 정리 (안 해두면 다음 시트 Copy 깨짐)
    try:
        ws_src.Application.CutCopyMode = False
    except Exception:
        pass

    if xml_data is None:
        raise RuntimeError(f"클립보드 XMLSS 읽기 실패: {last_exc}")

    # bytes → str
    if isinstance(xml_data, (bytes, bytearray)):
        b = bytes(xml_data).rstrip(b"\x00")
        if b.startswith(b"\xef\xbb\xbf"):
            b = b[3:]
        xml_str = b.decode("utf-8", errors="replace")
    else:
        xml_str = str(xml_data).rstrip("\x00")

    # <?xml ?> 와 <?mso-application ?> 처리 지시자를 건너뛰고 <Workbook> 부터 잘라냄
    wb_idx = xml_str.find("<Workbook")
    if wb_idx < 0:
        raise RuntimeError("XMLSS 구조 인식 실패 — <Workbook> 못 찾음")
    xml_str = xml_str[wb_idx:]
    end_idx = xml_str.rfind("</Workbook>")
    if end_idx >= 0:
        xml_str = xml_str[: end_idx + len("</Workbook>")]

    log(
        f"    [{sheet_name}] XMLSS {len(xml_str):,}자 — 파싱 시작 "
        f"(클립보드 추출 {time.perf_counter() - t0:.2f}s)"
    )

    import xml.etree.ElementTree as ET
    try:
        root = ET.fromstring(xml_str)
    except ET.ParseError as exc:
        raise RuntimeError(f"XMLSS 파싱 실패: {exc}") from exc

    t_parse = time.perf_counter()

    # ---- 스타일 테이블 ----
    styles = {}
    for style_elem in root.iterfind(f".//{_xmlss_q('Style')}"):
        sid = style_elem.get(_xmlss_q("ID"))
        if sid:
            try:
                styles[sid] = _parse_xmlss_style(style_elem)
            except Exception:
                styles[sid] = {}

    # ---- Worksheet / Table ----
    ws_elem = root.find(_xmlss_q("Worksheet"))
    if ws_elem is None:
        raise RuntimeError("XMLSS <Worksheet> 엘리먼트 없음")
    table = ws_elem.find(_xmlss_q("Table"))
    if table is None:
        raise RuntimeError("XMLSS <Table> 엘리먼트 없음")

    content_cells = set()
    seen_merges = set()

    # 컬럼 폭 — XMLSS Width 는 포인트. openpyxl 은 글자수 기준 (1글자 ≈ 7pt)
    col_rel = 1
    for col_elem in table.iterfind(_xmlss_q("Column")):
        idx_attr = col_elem.get(_xmlss_q("Index"))
        if idx_attr:
            try:
                col_rel = int(idx_attr)
            except (ValueError, TypeError):
                pass
        span_attr = col_elem.get(_xmlss_q("Span"))
        span_count = 1
        if span_attr:
            try:
                span_count = int(span_attr) + 1
            except (ValueError, TypeError):
                pass
        width = col_elem.get(_xmlss_q("Width"))
        hidden = col_elem.get(_xmlss_q("Hidden"))
        if width:
            try:
                # XMLSS Width 는 포인트, openpyxl 은 character width.
                # 정확한 변환식은 폰트 의존이지만 1 char ≈ 5.25pt 가 Calibri 11 의
                # 보편 근사. openpyxl 허용 범위 (0.5, 255) 로 클램프.
                w_pt = float(width)
                w_chars = max(0.5, min(255.0, w_pt / 5.25))
                for offset in range(span_count):
                    abs_col = base_col + (col_rel - 1) + offset
                    if abs_col < 1 or abs_col > 16384:
                        continue
                    try:
                        cd = ws_new.column_dimensions[_opx_col_letter(abs_col)]
                        cd.width = w_chars
                        if hidden == "1":
                            cd.hidden = True
                    except Exception:
                        pass
            except (ValueError, TypeError):
                pass
        col_rel += span_count

    # ---- Row / Cell 본 처리 ----
    row_rel = 1
    rows_written = 0
    for row_elem in table.iterfind(_xmlss_q("Row")):
        idx_attr = row_elem.get(_xmlss_q("Index"))
        if idx_attr:
            try:
                row_rel = int(idx_attr)
            except (ValueError, TypeError):
                pass

        abs_row = base_row + (row_rel - 1)

        # 행 높이 — openpyxl 허용 범위 (0, 409) 로 클램프
        height = row_elem.get(_xmlss_q("Height"))
        if height:
            try:
                h = float(height)
                if 0 < h <= 409:
                    if 1 <= abs_row <= 1048576:
                        ws_new.row_dimensions[abs_row].height = h
            except (ValueError, TypeError):
                pass

        # 행에 StyleID 가 있으면 그 행의 모든 셀 기본 스타일로 — 셀 ss:StyleID 가
        # 있으면 그게 우선
        row_style_id = row_elem.get(_xmlss_q("StyleID"))

        col_rel = 1
        for cell_elem in row_elem.iterfind(_xmlss_q("Cell")):
            cell_idx_attr = cell_elem.get(_xmlss_q("Index"))
            if cell_idx_attr:
                try:
                    col_rel = int(cell_idx_attr)
                except (ValueError, TypeError):
                    pass

            abs_col = base_col + (col_rel - 1)
            style_id = cell_elem.get(_xmlss_q("StyleID")) or row_style_id

            ma_attr = cell_elem.get(_xmlss_q("MergeAcross"))
            md_attr = cell_elem.get(_xmlss_q("MergeDown"))
            merge_across = 0
            merge_down = 0
            try:
                if ma_attr:
                    merge_across = int(ma_attr)
                if md_attr:
                    merge_down = int(md_attr)
            except (ValueError, TypeError):
                pass

            # 값 — Formula 는 R1C1 표기라 일단 무시, Data 의 계산값만 사용
            data_elem = cell_elem.find(_xmlss_q("Data"))
            value = None
            if data_elem is not None:
                try:
                    value = _xmlss_parse_data(data_elem)
                except Exception:
                    value = None

            # 셀 좌표 안전성 확인 — XMLSS 의 ss:Index 가 비정상이거나
            # MergeAcross 가 폭발해 16384(XFD) 또는 1048576 을 넘으면
            # openpyxl 이 받아주더라도 Excel 이 "복구 필요" 처리.
            if abs_row < 1 or abs_col < 1 or abs_row > 1048576 or abs_col > 16384:
                col_rel += merge_across + 1
                continue

            # 셀 쓰기
            try:
                dest = ws_new.cell(row=abs_row, column=abs_col)
                if value is not None:
                    # 값이 문자열인데 '=' 로 시작하면 openpyxl 이 자동으로 수식
                    # 으로 해석해 <f> 태그로 써냄 → Excel 이 평가 실패 시 복구
                    # 다이얼로그. data_type='s' 로 강제 인라인 문자열 처리.
                    if isinstance(value, str) and value.startswith("="):
                        dest.value = value
                        try:
                            dest.data_type = "s"
                        except Exception:
                            pass
                    else:
                        dest.value = value
                    content_cells.add((abs_row, abs_col))

                # 스타일 적용 — Default 는 workbook 의 기본 cellStyleXfs/cellStyles
                # 와 정합성 깨질 수 있어 스킵. openpyxl 의 native 기본 스타일 사용.
                # 값 없는 셀에도 style 적용 (테두리/배경만 있는 셀 등 보존).
                if style_id and style_id != "Default" and style_id in styles:
                    sd = styles[style_id]
                    if "font" in sd:
                        dest.font = sd["font"]
                    if "fill" in sd:
                        dest.fill = sd["fill"]
                    if "border" in sd:
                        dest.border = sd["border"]
                    if "alignment" in sd:
                        dest.alignment = sd["alignment"]
                    if "number_format" in sd:
                        dest.number_format = sd["number_format"]
            except Exception:
                pass

            # 병합 — 경계 검사 추가. last_row/last_col 이 시트 상한 (1048576/16384)
            # 을 넘으면 Excel 에서 "복구 필요" 트리거.
            if merge_across > 0 or merge_down > 0:
                last_row = min(abs_row + merge_down, 1048576)
                last_col = min(abs_col + merge_across, 16384)
                if last_row >= abs_row and last_col >= abs_col and (
                    last_row > abs_row or last_col > abs_col
                ):
                    mkey = (abs_row, abs_col, last_row, last_col)
                    if mkey not in seen_merges:
                        seen_merges.add(mkey)
                        try:
                            ws_new.merge_cells(
                                start_row=abs_row, start_column=abs_col,
                                end_row=last_row, end_column=last_col,
                            )
                        except Exception:
                            pass

            col_rel += merge_across + 1

        row_rel += 1
        rows_written += 1

    log(
        f"    [{sheet_name}] XMLSS 완료: 값 {len(content_cells):,}셀 / "
        f"병합 {len(seen_merges)}개 / 스타일 {len(styles)}개 / "
        f"행 {rows_written}개 처리 (파싱+쓰기 {time.perf_counter() - t_parse:.2f}s)"
    )

    return content_cells, used


def _clean_value_for_openpyxl(val):
    """
    Excel COM 에서 받은 값을 openpyxl 이 받을 수 있는 형태로 정리.

    - timezone-aware datetime/time : tzinfo 제거 (openpyxl 이 거부함)
    - float NaN/Inf                : None (openpyxl 이 거부함)
    - bytes                        : str(decode)
    """
    if val is None:
        return None
    if isinstance(val, _dt.datetime) and val.tzinfo is not None:
        return val.replace(tzinfo=None)
    if isinstance(val, _dt.time) and val.tzinfo is not None:
        return val.replace(tzinfo=None)
    if isinstance(val, float):
        if _math.isnan(val) or _math.isinf(val):
            return None
    if isinstance(val, bytes):
        try:
            return val.decode("utf-8", errors="replace")
        except Exception:
            return None
    return val


# ===========================================================================
# 핵심 — Excel 읽기 → openpyxl 쓰기
# ===========================================================================
def _excel_read_to_openpyxl(
    excel, src_path: Path, dest_path: Path, log, format_mode: str = "fast",
    on_file_progress=None,
):
    """
    원본 xlsx 를 Excel 로 열어 메모리로 데이터를 읽은 뒤 openpyxl 로 새
    xlsx 를 작성. Excel 의 SaveAs/SaveCopyAs 등 저장 메서드는 일체 호출
    안 함 — Open / 읽기만.

    DRM 정책이 Excel 의 *저장 경로* 만 후킹하는 경우 (= 가장 흔한 DLP 패턴)
    이 흐름엔 후킹 지점이 없다. Excel 은 읽기만 하고, 새 파일 작성은
    Python(openpyxl) 이 ZIP+XML 을 직접 써내므로 후킹이 끼어들 자리가 없다.

    format_mode ∈ {"fast", "balanced", "precise", "clipboard"}:
      - fast      : 셀 서식 복사 안 함 (병합/이미지/폭/높이만). 가장 빠름
      - balanced  : 이분 분할로 Bold/색/정렬 등 빠르게 복사 (테두리 미포함)
      - precise   : 셀별 COM 으로 모든 서식 (테두리 포함) — 가장 느림
      - clipboard : Range.Copy → 클립보드 XML Spreadsheet 파싱 → openpyxl 직접 매핑
                    값/스타일/병합/폭/높이까지 한 시트당 클립보드 1회 호출로 처리.
                    테두리 포함 모든 시각 서식 보존, 매우 빠름. 실패 시 balanced 폴백.

    추출/보존 항목:
      - 셀 값 / 수식
      - 컬럼 폭 / 행 높이
      - 셀 서식 (모드에 따라 일부/전체)
      - 병합 셀
      - 도형 → PNG 렌더링 → openpyxl 이미지로 절대 좌표(EMU) 기반 앵커 삽입
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl 미설치 — pip install openpyxl")

    log(f"  Excel 로 원본 열기: {src_path.name}")
    t0 = time.perf_counter()
    wb_src = _open_workbook_robust(excel, os.path.abspath(str(src_path)))
    log(f"  원본 열림 ({time.perf_counter() - t0:.2f}s)")

    wb_new = _OpxWorkbook()
    if wb_new.active is not None:
        try:
            wb_new.remove(wb_new.active)
        except Exception:
            pass

    # CF_HTML 에서 비트맵으로 렌더링되는 후보 도형 타입.
    # msoChart=3, msoFreeform=5, msoGroup=6, msoEmbeddedOLEObject=7,
    # msoLinkedOLEObject=10, msoLinkedPicture=11, msoOLEControlObject=12,
    # msoPicture=13, msoTextBox=17, msoSmartArt=24, msoAutoShape=1
    image_likely_types = {1, 3, 5, 6, 7, 10, 11, 12, 13, 17, 24}
    shape_type_names = {
        1: "AutoShape", 3: "Chart", 5: "Freeform", 6: "Group",
        7: "EmbeddedOLE", 10: "LinkedOLE", 11: "LinkedPicture",
        12: "OLEControl", 13: "Picture", 17: "TextBox", 24: "SmartArt",
    }

    # 파일 전체 도형 합산 (진행률 분모) — 1회 사전 카운트.
    file_shapes_total = 0
    file_shapes_done = 0
    try:
        for _sh_idx in range(1, int(wb_src.Sheets.Count) + 1):
            try:
                file_shapes_total += int(wb_src.Sheets(_sh_idx).Shapes.Count)
            except Exception:
                pass
    except Exception:
        pass

    def _tick(label):
        """파일별 진행률 콜백 호출 (도형 1개 처리 단위)."""
        if on_file_progress is None:
            return
        try:
            on_file_progress(
                file_shapes_done, max(1, file_shapes_total), label,
            )
        except Exception:
            pass

    try:
        sheet_count = int(wb_src.Sheets.Count)
        log(f"  시트 {sheet_count}개 발견 (총 도형 {file_shapes_total}개)")
        _tick(f"{src_path.name} 시작")
        for sh_idx in range(1, sheet_count + 1):
            ws_src = wb_src.Sheets(sh_idx)
            sheet_name = (str(ws_src.Name)[:31] or f"Sheet{sh_idx}").strip()
            log(f"  ── 시트 {sh_idx}/{sheet_count}: {sheet_name}")
            t_sheet = time.perf_counter()
            try:
                ws_new = wb_new.create_sheet(title=sheet_name)
            except Exception:
                ws_new = wb_new.create_sheet(title=f"Sheet{sh_idx}")

            # 값 있는 셀 좌표 모음 — 서식 phase 에서 빈 셀 스킵용.
            # per-cell COM 호출 폭증 (셀당 ~22개 프로퍼티 read) 으로 인한
            # 5셀/s 수준 정체를 방지한다. 값 없는 셀의 서식은 보통 시각적
            # 의미가 없어 누락해도 결과물 체감 차이가 크지 않다.
            content_cells = set()
            used = None  # 도형 phase 가 참조 — XMLSS 또는 값 phase 에서 채움

            # ---- 클립보드 XMLSS (format_mode == "clipboard") ----
            # 성공 시 값/스타일/병합/폭/높이까지 한 번에 처리되므로 아래
            # COM 기반 phase 들은 모두 스킵. 실패 시 _force_to_balanced 로
            # 폴백 모드 결정 후 일반 경로로 진행.
            clipboard_done = False
            effective_format_mode = format_mode
            if format_mode == "clipboard":
                try:
                    content_cells, used = _copy_sheet_via_xmlss(
                        ws_src, ws_new, log, sheet_name
                    )
                    clipboard_done = True
                except Exception as exc:
                    log(
                        f"    [{sheet_name}] 클립보드 XMLSS 실패: {exc} — "
                        f"균등 분할로 폴백"
                    )
                    effective_format_mode = "balanced"
                    # XMLSS 가 부분 쓰기 후 실패했으면 ws_new 에 일관성 깨진
                    # 상태가 남음. 시트 통째로 갈아엎고 COM 경로로 새로 채움.
                    try:
                        wb_new.remove(ws_new)
                    except Exception:
                        pass
                    try:
                        ws_new = wb_new.create_sheet(title=sheet_name)
                    except Exception:
                        ws_new = wb_new.create_sheet(title=f"Sheet{sh_idx}")
                    content_cells = set()
                    used = None

            # ---- 값 + 수식 (한 번에 2D 튜플로 받기 — COM 호출 횟수 최소화) ----
            try:
                if clipboard_done:
                    raise _PhaseSkip()
                t_phase = time.perf_counter()
                used = ws_src.UsedRange
                base_row = int(used.Row)
                base_col = int(used.Column)
                n_rows = int(used.Rows.Count)
                n_cols = int(used.Columns.Count)
                log(
                    f"    [{sheet_name}] UsedRange: {n_rows}행 x {n_cols}열 "
                    f"(시작 R{base_row}C{base_col}, 총 {n_rows*n_cols:,}셀)"
                )

                # 끝열·끝행까지 서식만 발려 있는 시트는 UsedRange 가 시트
                # 전체로 잡혀 .Value 호출 자체가 메모리 폭발이 된다 — 벌크
                # 읽기 전에 범위를 새로 잘라 used 를 교체해 둔다.
                if n_rows > MAX_RANGE_ROWS or n_cols > MAX_RANGE_COLS:
                    new_rows = min(n_rows, MAX_RANGE_ROWS)
                    new_cols = min(n_cols, MAX_RANGE_COLS)
                    log(
                        f"    [{sheet_name}] ⚠ UsedRange {n_rows}x{n_cols} 너무 큼 "
                        f"→ {new_rows}x{new_cols} 로 제한 (1000×1000)"
                    )
                    n_rows = new_rows
                    n_cols = new_cols
                    last_row = base_row + n_rows - 1
                    last_col = base_col + n_cols - 1
                    used = ws_src.Range(
                        ws_src.Cells(base_row, base_col),
                        ws_src.Cells(last_row, last_col),
                    )

                log(f"    [{sheet_name}] 값/수식 벌크 읽기…")
                t_read = time.perf_counter()
                values = used.Value
                formulas = used.Formula
                log(
                    f"    [{sheet_name}] 값/수식 읽기 완료 "
                    f"({time.perf_counter() - t_read:.2f}s)"
                )

                t_write = time.perf_counter()
                values_written = 0
                if n_rows == 1 and n_cols == 1:
                    val = (
                        formulas
                        if isinstance(formulas, str) and formulas.startswith("=")
                        else values
                    )
                    val = _clean_value_for_openpyxl(val)
                    if val is not None:
                        try:
                            ws_new.cell(
                                row=base_row, column=base_col, value=val
                            )
                            content_cells.add((base_row, base_col))
                            values_written = 1
                        except Exception:
                            pass
                else:
                    if not isinstance(values, tuple):
                        values = ((values,),)
                    if not isinstance(formulas, tuple):
                        formulas = values
                    for r_idx in range(min(n_rows, len(values))):
                        v_row = values[r_idx]
                        f_row = formulas[r_idx] if r_idx < len(formulas) else v_row
                        if not isinstance(v_row, tuple):
                            v_row = (v_row,)
                        if not isinstance(f_row, tuple):
                            f_row = (f_row,)
                        for c_idx in range(min(n_cols, len(v_row))):
                            f_cell = f_row[c_idx] if c_idx < len(f_row) else None
                            v_cell = v_row[c_idx]
                            cell_val = (
                                f_cell
                                if isinstance(f_cell, str)
                                and f_cell.startswith("=")
                                else v_cell
                            )
                            cell_val = _clean_value_for_openpyxl(cell_val)
                            if cell_val is None:
                                continue
                            try:
                                ws_new.cell(
                                    row=base_row + r_idx,
                                    column=base_col + c_idx,
                                    value=cell_val,
                                )
                                content_cells.add(
                                    (base_row + r_idx, base_col + c_idx)
                                )
                                values_written += 1
                            except Exception:
                                pass
                log(
                    f"    [{sheet_name}] 값 {values_written:,}셀 기록 "
                    f"(쓰기 {time.perf_counter() - t_write:.2f}s, "
                    f"단계 누적 {time.perf_counter() - t_phase:.2f}s)"
                )
            except _PhaseSkip:
                pass
            except Exception as exc:
                log(f"    [{sheet_name}] 값 읽기 부분 실패: {exc}")

            # ---- 컬럼 폭 ----
            try:
                if clipboard_done:
                    raise _PhaseSkip()
                t_cw = time.perf_counter()
                base_col = int(used.Column)
                n_cols = int(used.Columns.Count)
                cols_done = 0
                for offset in range(n_cols):
                    col_idx = base_col + offset
                    try:
                        width = float(ws_src.Columns(col_idx).ColumnWidth)
                        if width > 0:
                            ws_new.column_dimensions[
                                _opx_col_letter(col_idx)
                            ].width = width
                            cols_done += 1
                    except Exception:
                        pass
                log(
                    f"    [{sheet_name}] 컬럼 폭 {cols_done}/{n_cols}개 적용 "
                    f"({time.perf_counter() - t_cw:.2f}s)"
                )
            except _PhaseSkip:
                pass
            except Exception as exc:
                log(f"    [{sheet_name}] 컬럼 폭 실패: {exc}")

            # ---- 행 높이 ----
            try:
                if clipboard_done:
                    raise _PhaseSkip()
                t_rh = time.perf_counter()
                base_row = int(used.Row)
                n_rows = int(used.Rows.Count)
                rows_done = 0
                for offset in range(n_rows):
                    row_idx = base_row + offset
                    try:
                        height = float(ws_src.Rows(row_idx).RowHeight)
                        if height > 0:
                            ws_new.row_dimensions[row_idx].height = height
                            rows_done += 1
                    except Exception:
                        pass
                log(
                    f"    [{sheet_name}] 행 높이 {rows_done}/{n_rows}개 적용 "
                    f"({time.perf_counter() - t_rh:.2f}s)"
                )
            except _PhaseSkip:
                pass
            except Exception as exc:
                log(f"    [{sheet_name}] 행 높이 실패: {exc}")

            # ---- 셀 서식 + 병합 — 모드별 디스패치 ----
            # clipboard_done 이면 XMLSS 가 이미 다 처리 — phase 통째 스킵.
            # 폴백 상태에서는 effective_format_mode 가 "balanced" 로 격하됨.
            if clipboard_done:
                pass
            elif effective_format_mode == "balanced":
                # 이분 분할: 시트당 ~수백 COM 호출, 셀별 루프 없음
                try:
                    _apply_formats_subdivision(
                        ws_src, used, ws_new, content_cells, log, sheet_name
                    )
                except Exception as exc:
                    log(f"    [{sheet_name}] 균등 분할 부분 실패: {exc}")
            else:
                # fast/precise — 셀별 루프 (precise 만 _copy_cell_format 실행)
                precise = (effective_format_mode == "precise")
                try:
                    t_fmt = time.perf_counter()
                    base_row = int(used.Row)
                    base_col = int(used.Column)
                    n_rows = int(used.Rows.Count)
                    n_cols = int(used.Columns.Count)
                    total_in_range = n_rows * n_cols
                    target_cells = sorted(content_cells)
                    target_total = len(target_cells)
                    if precise:
                        log(
                            f"    [{sheet_name}] 셀 서식+병합 전사 시작 — 정밀 모드 "
                            f"(값 있는 {target_total:,}셀, 빈 셀 "
                            f"{total_in_range - target_total:,}개 스킵)"
                        )
                    else:
                        log(
                            f"    [{sheet_name}] 셀 병합 스캔 시작 — 빠른 모드 "
                            f"(값 있는 {target_total:,}셀, 셀 서식 복사 OFF)"
                        )
                    seen_merge_ranges = set()
                    cells_done = 0
                    last_log = time.perf_counter()
                    for r, c in target_cells:
                        try:
                            src_cell = ws_src.Cells(r, c)
                            if precise:
                                dest_cell = ws_new.cell(row=r, column=c)
                                _copy_cell_format(src_cell, dest_cell)
                            cells_done += 1
                            # 진행률 — 500셀마다 시간 체크 후 1.5s 간격으로 로그
                            if cells_done % 500 == 0:
                                now = time.perf_counter()
                                if now - last_log >= 1.5:
                                    pct = 100 * cells_done / max(target_total, 1)
                                    rate = cells_done / max(now - t_fmt, 0.001)
                                    log(
                                        f"    [{sheet_name}] 서식 진행 "
                                        f"{cells_done:,}/{target_total:,} "
                                        f"({pct:.0f}%, {rate:.0f}셀/s)"
                                    )
                                    last_log = now
                            # 병합 — 좌표 직접 사용 (pywin32 의 Address 메서드
                            # 호출이 환경에 따라 실패하기 때문에 문자열 파싱
                            # 대신 Row/Column/Count 속성으로 키 만든다).
                            try:
                                if bool(src_cell.MergeCells):
                                    area = src_cell.MergeArea
                                    first_row = int(area.Row)
                                    first_col = int(area.Column)
                                    nrows = int(area.Rows.Count)
                                    ncols = int(area.Columns.Count)
                                    if nrows == 1 and ncols == 1:
                                        pass
                                    else:
                                        last_row = first_row + nrows - 1
                                        last_col = first_col + ncols - 1
                                        merge_key = (
                                            first_row, first_col,
                                            last_row, last_col,
                                        )
                                        if merge_key not in seen_merge_ranges:
                                            seen_merge_ranges.add(merge_key)
                                            try:
                                                ws_new.merge_cells(
                                                    start_row=first_row,
                                                    start_column=first_col,
                                                    end_row=last_row,
                                                    end_column=last_col,
                                                )
                                            except Exception:
                                                pass
                            except Exception:
                                pass
                        except Exception:
                            pass
                    fmt_elapsed = time.perf_counter() - t_fmt
                    phase_name = "서식+병합" if precise else "병합 스캔"
                    if seen_merge_ranges:
                        log(
                            f"    [{sheet_name}] {phase_name} 완료: "
                            f"{cells_done:,}셀 / 병합 {len(seen_merge_ranges)}개 "
                            f"({fmt_elapsed:.2f}s)"
                        )
                    else:
                        log(
                            f"    [{sheet_name}] {phase_name} 완료: "
                            f"{cells_done:,}셀 ({fmt_elapsed:.2f}s)"
                        )
                except Exception as exc:
                    log(f"    [{sheet_name}] 서식/병합 부분 실패: {exc}")

            # ---- 이미지 (도형 → PNG → openpyxl 삽입, 절대 좌표 기반) ----
            try:
                shape_count = int(ws_src.Shapes.Count)
            except Exception:
                shape_count = 0

            if shape_count == 0:
                log(f"    [{sheet_name}] 도형 없음")
            elif not HAS_PIL:
                log(
                    f"    [{sheet_name}] 도형 {shape_count}개 — Pillow 미설치라 스킵"
                )

            if shape_count > 0 and HAS_PIL:
                t_img = time.perf_counter()
                log(f"    [{sheet_name}] 도형 {shape_count}개 처리 시작…")
                inserted = 0
                skipped_type = 0
                skipped_render = 0
                for s_idx in range(1, shape_count + 1):
                    try:
                        shape = ws_src.Shapes.Item(s_idx)
                        try:
                            s_type = int(shape.Type)
                        except Exception:
                            s_type = 0
                        type_label = shape_type_names.get(s_type, f"type{s_type}")
                        if s_type not in image_likely_types:
                            log(
                                f"      도형 {s_idx}/{shape_count}: {type_label} "
                                f"(이미지 변환 대상 아님 — 스킵)"
                            )
                            skipped_type += 1
                            file_shapes_done += 1
                            _tick(f"{sheet_name} 도형 {s_idx}/{shape_count} (타입 스킵)")
                            continue
                        try:
                            sw = float(shape.Width)
                            sh = float(shape.Height)
                        except Exception:
                            sw = sh = 0
                        log(
                            f"      도형 {s_idx}/{shape_count}: {type_label} "
                            f"({sw:.0f}×{sh:.0f}pt) → 렌더링…"
                        )
                        t_render = time.perf_counter()
                        png = render_shape_as_png(shape, log=None)
                        if not png:
                            log(
                                f"        ! 렌더링 실패 — 스킵 "
                                f"({time.perf_counter() - t_render:.2f}s)"
                            )
                            skipped_render += 1
                            file_shapes_done += 1
                            _tick(f"{sheet_name} 도형 {s_idx}/{shape_count} (렌더 실패)")
                            continue
                        log(
                            f"        렌더링 OK ({len(png):,} B, "
                            f"{time.perf_counter() - t_render:.2f}s)"
                        )
                        opx_img = _OpxImage(io.BytesIO(png))

                        # 절대 위치/크기 → TwoCellAnchor 로 정확히 박기.
                        # 단순 셀 anchor("A1") 만 쓰면 같은 셀에 앵커된
                        # 도형들이 화면상 겹친다 — 원본은 셀 내부 오프셋으로
                        # 분리되어 있는데 그 정보가 사라지기 때문.
                        anchored = False
                        try:
                            shape_left = float(shape.Left)
                            shape_top = float(shape.Top)
                            shape_w = float(shape.Width)
                            shape_h = float(shape.Height)
                            tl = shape.TopLeftCell
                            br = shape.BottomRightCell
                            tl_row = int(tl.Row)
                            tl_col = int(tl.Column)
                            br_row = int(br.Row)
                            br_col = int(br.Column)
                            tl_cell = ws_src.Cells(tl_row, tl_col)
                            br_cell = ws_src.Cells(br_row, br_col)
                            tl_cell_left = float(tl_cell.Left)
                            tl_cell_top = float(tl_cell.Top)
                            br_cell_left = float(br_cell.Left)
                            br_cell_top = float(br_cell.Top)

                            from_off_x = (shape_left - tl_cell_left) * _EMU_PER_POINT
                            from_off_y = (shape_top - tl_cell_top) * _EMU_PER_POINT
                            to_off_x = (
                                shape_left + shape_w - br_cell_left
                            ) * _EMU_PER_POINT
                            to_off_y = (
                                shape_top + shape_h - br_cell_top
                            ) * _EMU_PER_POINT

                            from_marker = _OpxAnchorMarker(
                                col=tl_col - 1, colOff=int(from_off_x),
                                row=tl_row - 1, rowOff=int(from_off_y),
                            )
                            to_marker = _OpxAnchorMarker(
                                col=br_col - 1, colOff=int(to_off_x),
                                row=br_row - 1, rowOff=int(to_off_y),
                            )
                            opx_img.anchor = _OpxTwoCellAnchor(
                                editAs="oneCell",
                                _from=from_marker,
                                to=to_marker,
                            )
                            ws_new.add_image(opx_img)
                            anchored = True
                        except Exception:
                            pass

                        # 폴백 — 셀 단위 단순 앵커
                        if not anchored:
                            try:
                                tl = shape.TopLeftCell
                                anchor_col = int(tl.Column)
                                anchor_row = int(tl.Row)
                            except Exception:
                                anchor_col, anchor_row = 1, 1
                            try:
                                opx_img.width = max(
                                    1, int(float(shape.Width) * 96 / 72)
                                )
                                opx_img.height = max(
                                    1, int(float(shape.Height) * 96 / 72)
                                )
                            except Exception:
                                pass
                            anchor_addr = (
                                f"{_opx_col_letter(anchor_col)}{anchor_row}"
                            )
                            ws_new.add_image(opx_img, anchor_addr)
                        inserted += 1
                        file_shapes_done += 1
                        _tick(f"{sheet_name} 도형 {s_idx}/{shape_count} (삽입)")
                    except Exception as exc:
                        log(f"      ! 도형 #{s_idx} 처리 실패: {exc}")
                        file_shapes_done += 1
                        _tick(f"{sheet_name} 도형 {s_idx}/{shape_count} (실패)")
                log(
                    f"    [{sheet_name}] 이미지 단계 완료: 삽입 {inserted}개 / "
                    f"타입스킵 {skipped_type}개 / 렌더실패 {skipped_render}개 "
                    f"({time.perf_counter() - t_img:.2f}s)"
                )

            log(
                f"  ── 시트 {sh_idx}/{sheet_count} '{sheet_name}' 완료 "
                f"({time.perf_counter() - t_sheet:.2f}s)"
            )
    finally:
        try:
            wb_src.Close(SaveChanges=False)
            log("  원본 닫음")
        except Exception:
            pass

    # 새 워크북에 시트가 하나도 없으면 openpyxl 이 save 시 에러 — 빈 시트 추가
    if not wb_new.sheetnames:
        wb_new.create_sheet(title="Sheet1")

    log(f"  openpyxl 저장 시작 → {dest_path.name}")
    t_save = time.perf_counter()
    wb_new.save(str(dest_path))
    log(f"  저장 완료 ({time.perf_counter() - t_save:.2f}s)")

    # ---- 출력본 자가 검증 ----
    # 우리가 쓴 xlsx 를 openpyxl 이 *다시 읽어서* 무사히 통과하는지 확인.
    # openpyxl 이 못 읽는 파일은 Excel 도 거의 항상 거부 ("파일을 열 수 없음").
    # 통과해도 Excel 이 거부할 수 있지만 (스키마 차이) — 이건 큰 1차 거름망.
    try:
        _vwb = openpyxl.load_workbook(str(dest_path), read_only=True)
        v_sheet_count = len(_vwb.sheetnames)
        try:
            _vwb.close()
        except Exception:
            pass
        log(f"  ✓ 출력본 자가 검증 통과 (시트 {v_sheet_count}개)")
    except Exception as exc:
        log(f"  ⚠ 출력본 openpyxl 재읽기 실패 — Excel 도 못 열 가능성 큼: {exc}")

    # ---- styles.xml 직접 진단 ----
    # Excel 이 "복구 필요" 라고 하는 styles.xml 의 어떤 엔트리가 문제인지
    # 좁히기 위한 일회성 덤프. fonts/fills/borders/cellXfs/numFmts/cellStyles
    # 개수와 cellXfs 인덱스 참조 정합성을 점검.
    try:
        import xml.etree.ElementTree as _ET
        with zipfile.ZipFile(str(dest_path)) as _zf:
            with _zf.open("xl/styles.xml") as _sf:
                _styles_root = _ET.fromstring(_sf.read())
        _NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
        def _count(parent, child):
            el = parent.find(f"{_NS}{child}")
            if el is None:
                return 0
            tag = "numFmt" if child == "numFmts" else child[:-1]
            return len(el.findall(f"{_NS}{tag}"))
        n_fonts = _count(_styles_root, "fonts")
        n_fills = _count(_styles_root, "fills")
        n_borders = _count(_styles_root, "borders")
        n_cellxfs = _count(_styles_root, "cellXfs")
        n_numfmts = _count(_styles_root, "numFmts")
        n_cellstyles = _count(_styles_root, "cellStyles")
        n_cellstylexfs = _count(_styles_root, "cellStyleXfs")
        log(
            f"  styles.xml: fonts={n_fonts} fills={n_fills} borders={n_borders} "
            f"cellXfs={n_cellxfs} numFmts={n_numfmts} "
            f"cellStyles={n_cellstyles} cellStyleXfs={n_cellstylexfs}"
        )
        # cellXfs 인덱스 참조 검증 — 범위 벗어난 ID 가 있으면 styles.xml 복구 트리거
        _cellxfs_el = _styles_root.find(f"{_NS}cellXfs")
        bad_refs = []
        if _cellxfs_el is not None:
            for _idx, _xf in enumerate(_cellxfs_el.findall(f"{_NS}xf")):
                for _attr, _max in (
                    ("fontId", n_fonts), ("fillId", n_fills),
                    ("borderId", n_borders),
                ):
                    _v = _xf.get(_attr)
                    if _v is None:
                        continue
                    try:
                        if int(_v) >= _max:
                            bad_refs.append(f"xf#{_idx}.{_attr}={_v} (max {_max})")
                    except ValueError:
                        bad_refs.append(f"xf#{_idx}.{_attr}={_v} (not int)")
        if bad_refs:
            log(f"  ⚠ styles.xml 잘못된 인덱스 참조 {len(bad_refs)}개:")
            for ref in bad_refs[:5]:
                log(f"    {ref}")
        # 모든 numFmt 코드 덤프 — Excel 이 거부할 만한 패턴을 직접 비교용으로
        _numfmts_el = _styles_root.find(f"{_NS}numFmts")
        all_fmts = []
        if _numfmts_el is not None:
            for _nf in _numfmts_el.findall(f"{_NS}numFmt"):
                _code = _nf.get("formatCode", "")
                _id = _nf.get("numFmtId", "?")
                all_fmts.append((_id, _code))
        if all_fmts:
            log(f"  styles.xml 의 numFmts ({len(all_fmts)}개):")
            for _id, _code in all_fmts:
                log(f"    [{_id}] {_code!r}")

        # cellXfs 의 applyXxx 플래그와 실제 ID 일관성 — 가장 흔한 복구 원인 중 하나
        _xf_anomalies = []
        if _cellxfs_el is not None:
            for _idx, _xf in enumerate(_cellxfs_el.findall(f"{_NS}xf")):
                for _id_attr, _apply_attr in (
                    ("fontId", "applyFont"),
                    ("fillId", "applyFill"),
                    ("borderId", "applyBorder"),
                    ("numFmtId", "applyNumberFormat"),
                ):
                    _id_v = _xf.get(_id_attr)
                    _apply_v = _xf.get(_apply_attr)
                    # apply=1 인데 id 가 0 (default) 이거나 그 반대인 경우 의심
                    if _apply_v == "1" and _id_v in (None, "0"):
                        _xf_anomalies.append(
                            f"xf#{_idx}: {_apply_attr}=1 인데 {_id_attr}={_id_v}"
                        )
        if _xf_anomalies:
            log(f"  ⚠ cellXfs 의심 (apply* vs id 불일치) {len(_xf_anomalies)}개:")
            for a in _xf_anomalies[:5]:
                log(f"    {a}")
    except Exception as exc:
        log(f"  styles.xml 진단 실패 (무해): {exc}")


def _rezip_xlsx(src: Path, dest: Path):
    """
    원본 xlsx 를 ZIP 단위 raw 재포장. Excel·openpyxl 둘 다 못 쓸 때 최후 폴백.
    바이트 시퀀스만 다른 동일 내용 — 그림 한 장도 안 빠짐.
    """
    with zipfile.ZipFile(str(src), "r") as zin:
        with zipfile.ZipFile(
            str(dest), "w", zipfile.ZIP_DEFLATED, compresslevel=6
        ) as zout:
            for entry in zin.infolist():
                data = zin.read(entry.filename)
                new_info = zipfile.ZipInfo(
                    filename=entry.filename,
                    date_time=(2020, 1, 1, 0, 0, 0),
                )
                new_info.compress_type = zipfile.ZIP_DEFLATED
                zout.writestr(new_info, data)


def drm_clean_resave(
    src_paths,
    output_root: str,
    log,
    on_progress,
    on_done=None,
    format_mode: str = "fast",
    should_cancel=None,
    should_force_stop=None,
    excel_pid_holder=None,
    on_file_progress=None,
):
    """
    DRM 우회 사본 — Excel 의 SaveAs 를 거치지 않고 메모리에서 새 xlsx 작성.

    출력: <output_root>/<원본이름>_clean.xlsx  (출력 폴더에 직접 저장)

    format_mode ∈ {"fast", "balanced", "precise", "clipboard"}:
      - fast      : 값/병합/이미지/폭/높이만 (서식 미복사 — 가장 빠름)
      - balanced  : 이분 분할로 Bold/색/정렬 빠르게 복사 (테두리 제외)
      - precise   : 셀별 COM 으로 모든 서식 (테두리 포함, 가장 느림)
      - clipboard : Excel 클립보드 XMLSS → openpyxl. 모든 서식 + 매우 빠름.

    on_done(src, status, dest) — 파일 단위 콜백. status ∈ {"ok","skip","fail"}.
      GUI 가 완료 폴더 이동/Queue 제거 등 후처리에 사용.

    should_cancel — () -> bool 콜러블. True 반환 시 *다음 파일* 직전에 루프
      탈출 (진행 중인 파일은 끝까지 완료). 호출 측의 threading.Event.is_set
      을 그대로 넘기면 됨.

    Returns: (성공목록 [(원본, 사본)], 실패목록 [(원본, 오류문자열)])
    """
    succeeded = []
    failed = []
    total = len(src_paths)

    if not HAS_OPENPYXL:
        log(
            "openpyxl 미설치 — Excel 읽기 전략 비활성화. "
            "pip install openpyxl 권장 (지금은 zipfile 재포장만 동작)."
        )

    # 출력 폴더에 직접 *_clean.xlsx 생성 (하위 폴더 추가 없음).
    clean_dir = Path(output_root)
    try:
        clean_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        log(f"출력 폴더 생성 실패 ({clean_dir}): {exc}")
        return [], [(s, f"폴더 생성 실패: {exc}") for s in src_paths]
    log(f"출력 → {clean_dir}")

    pythoncom.CoInitialize()
    excel = None
    try:
        if HAS_PYWIN32 and HAS_OPENPYXL:
            log("Excel 인스턴스 시작 중…")
            t_excel = time.perf_counter()
            try:
                excel = win32com.client.DispatchEx("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                excel.AskToUpdateLinks = False
                log(f"Excel 시작 완료 ({time.perf_counter() - t_excel:.2f}s)")
                # 강제 종료용 PID 캡처. Visible=False 라도 Hwnd 는 존재 (숨은
                # 메인 윈도우). UI 의 강제 종료 버튼이 이 PID 를 taskkill 함.
                if excel_pid_holder is not None:
                    try:
                        hwnd = int(excel.Hwnd)
                        if hwnd:
                            _, pid = win32process.GetWindowThreadProcessId(hwnd)
                            excel_pid_holder[0] = int(pid)
                            log(f"  Excel PID = {pid}")
                    except Exception as exc:
                        log(f"  Excel PID 캡처 실패 (강제 종료 비활성): {exc}")
            except Exception as exc:
                log(f"Excel 시작 실패 — Excel-경유 전략 건너뜀: {exc}")
                excel = None

        def _force_stop():
            if should_force_stop is None:
                return False
            try:
                return bool(should_force_stop())
            except Exception:
                return False

        def _start_excel():
            """Excel 인스턴스 새로 띄움 + PID 등록. 실패하면 None."""
            log("  Excel 인스턴스 (재)시작 중…")
            try:
                ex = win32com.client.DispatchEx("Excel.Application")
                ex.Visible = False
                ex.DisplayAlerts = False
                ex.AskToUpdateLinks = False
                if excel_pid_holder is not None:
                    try:
                        hwnd = int(ex.Hwnd)
                        if hwnd:
                            _, pid = win32process.GetWindowThreadProcessId(hwnd)
                            excel_pid_holder[0] = int(pid)
                            log(f"  Excel PID = {pid}")
                    except Exception:
                        pass
                return ex
            except Exception as exc:
                log(f"  ✗ Excel 시작 실패: {exc}")
                return None

        def _is_rpc_dead(exc):
            """COM 에러가 'Excel 죽음' (RPC_S_SERVER_UNAVAILABLE = -2147023174) 인가?"""
            # pywintypes.com_error 의 args[0] 이 HRESULT
            try:
                code = getattr(exc, "hresult", None)
                if code is None and hasattr(exc, "args") and exc.args:
                    code = exc.args[0]
                if code in (-2147023174, -2147417848):
                    # -2147023174 RPC_S_SERVER_UNAVAILABLE
                    # -2147417848 The object invoked has disconnected from its clients
                    return True
            except Exception:
                pass
            return False

        for i, src in enumerate(src_paths, start=1):
            # 다음 파일 진입 직전에만 (soft) 중단 체크 — 진행 중 파일은 완료
            if should_cancel is not None:
                try:
                    if should_cancel():
                        log(f"⏹ 사용자 중단 요청 — 남은 {total - i + 1}개 건너뜀")
                        break
                except Exception:
                    pass
            src_p = Path(src)
            dest = clean_dir / f"{safe_name(src_p.stem)}_clean.xlsx"
            t_file = time.perf_counter()
            log(f"[{i}/{total}] ▶ {src_p.name}")
            # 파일별 진행률 초기화 — 실제 도형 수는 _excel_read_to_openpyxl 가
            # 파일 열고 직접 세서 갱신함.
            if on_file_progress:
                try:
                    on_file_progress(0, 1, f"준비: {src_p.name}")
                except Exception:
                    pass

            # 출력본이 이미 있으면 스킵 (재실행 비용 절감 + 덮어쓰기 방지)
            if dest.exists():
                log(f"  ↷ 스킵 — 이미 존재: {dest.name}")
                succeeded.append((src, dest))
                if on_done:
                    try:
                        on_done(src, "skip", dest)
                    except Exception as exc:
                        log(f"  ! on_done 콜백 오류: {exc}")
                on_progress(i, total, src)
                continue

            ok = False
            err_log = []

            # ⓪ Excel 읽기 + openpyxl 쓰기
            if excel is not None and HAS_OPENPYXL and not _force_stop():
                log("  [전략 ⓪] Excel 읽기 + openpyxl 쓰기 시도…")
                try:
                    _excel_read_to_openpyxl(
                        excel, src_p, dest, log,
                        format_mode=format_mode,
                        on_file_progress=on_file_progress,
                    )
                    log(f"  ✓ Excel 읽기 + openpyxl 쓰기 → {dest.name}")
                    ok = True
                except Exception as exc:
                    log(f"  ✗ 전략 ⓪ 실패: {exc}")
                    err_log.append(f"Excel→openpyxl 실패: {exc}")
                    # Excel 프로세스가 죽었으면 즉시 재시작 — 안 그러면 남은
                    # 모든 파일이 같은 RPC 에러로 줄줄이 실패함. 사용자가 강제
                    # 종료 눌렀을 때는 재시작 안 함 (의도된 중단).
                    if _is_rpc_dead(exc) and not _force_stop():
                        log("  ⚠ Excel 프로세스 사망 (RPC) — 인스턴스 재시작")
                        try:
                            excel = _start_excel()
                        except Exception as exc2:
                            log(f"  ✗ Excel 재시작 실패: {exc2}")
                            excel = None
                        if excel is not None:
                            # 재시작 직후 같은 파일 한 번 더 시도
                            log("  [전략 ⓪/재시도] Excel 재시작 후 재시도…")
                            try:
                                _excel_read_to_openpyxl(
                                    excel, src_p, dest, log,
                                    format_mode=format_mode,
                                    on_file_progress=on_file_progress,
                                )
                                log(f"  ✓ 재시도 성공 → {dest.name}")
                                ok = True
                                err_log.clear()
                            except Exception as exc3:
                                log(f"  ✗ 재시도도 실패: {exc3}")
                                err_log.append(f"재시도: {exc3}")
                                if _is_rpc_dead(exc3):
                                    # 재시도도 RPC 로 죽었으면 이 파일이 Excel 을
                                    # 죽이는 trigger — 다음 파일들 위해 또 살림
                                    log(
                                        "  ⚠ 재시도도 RPC — 이 파일이 Excel 을 "
                                        "죽이는 듯. 다음 파일 위해 한 번 더 재시작."
                                    )
                                    try:
                                        excel = _start_excel()
                                    except Exception:
                                        excel = None

            # ① openpyxl 단독 — 강제 종료 요청 시 폴백 스킵
            if not ok and HAS_OPENPYXL and not _force_stop():
                log("  [전략 ①] openpyxl 단독 load+save 시도…")
                try:
                    wb = openpyxl.load_workbook(
                        str(src_p), keep_vba=False, data_only=False
                    )
                    wb.save(str(dest))
                    log(f"  ✓ openpyxl 단독 → {dest.name}")
                    ok = True
                except Exception as exc:
                    log(f"  ✗ 전략 ① 실패: {exc}")
                    err_log.append(f"openpyxl 단독 실패: {exc}")

            # ② zipfile 재포장
            if not ok and not _force_stop():
                log("  [전략 ②] zipfile raw 재포장 시도…")
                try:
                    _rezip_xlsx(src_p, dest)
                    log(f"  ✓ zipfile 재포장 → {dest.name}")
                    ok = True
                except Exception as exc:
                    log(f"  ✗ 전략 ② 실패: {exc}")
                    err_log.append(f"zipfile 재포장 실패: {exc}")

            elapsed = time.perf_counter() - t_file
            if ok:
                log(f"[{i}/{total}] ✓ {src_p.name} 완료 ({elapsed:.2f}s)")
                succeeded.append((src, dest))
                if on_done:
                    try:
                        on_done(src, "ok", dest)
                    except Exception as exc:
                        log(f"  ! on_done 콜백 오류: {exc}")
            else:
                log(f"[{i}/{total}] ✗ {src_p.name} 실패 ({elapsed:.2f}s)")
                failed.append((src, " / ".join(err_log) or "unknown"))
                if on_done:
                    try:
                        on_done(src, "fail", None)
                    except Exception as exc:
                        log(f"  ! on_done 콜백 오류: {exc}")

            on_progress(i, total, src)
            # 이 파일 끝나자마자 강제 종료 체크 — soft cancel 보다 한 발 빠름
            if _force_stop():
                log(
                    f"⛔ 강제 종료 — 남은 "
                    f"{total - i}개 건너뜀 (Excel 도 이미 죽었음)"
                )
                break
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass

    return succeeded, failed


# ===========================================================================
# GUI
# ===========================================================================
class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("엑셀 DRM 우회 사본 생성기")
        self.root.geometry("1040x640")
        self.root.minsize(900, 540)

        self.src_paths = []
        # 출력 폴더 — 내부에 Clean/ 와 Org/ 가 자동 생성됨.
        # Clean/ 에는 *_clean.xlsx 결과, Org/ 에는 처리 끝난 원본을 이동.
        self.output_dir = tk.StringVar()
        self.open_output = tk.BooleanVar(value=False)
        # 서식 복사 모드.
        # fast      : 서식 미복사 (가장 빠름, 그러나 평문처럼 보임)
        # balanced  : 이분 분할로 Bold/색/정렬 빠르게 복사 (테두리 제외)
        # precise   : 셀별 COM, 테두리 포함 모든 서식 (가장 느림)
        # clipboard : Excel 의 'XML Spreadsheet' 클립보드 포맷 파싱 — 시트당
        #             클립보드 1회 호출로 값+모든 서식+병합+폭/높이 일괄 처리.
        #             테두리 포함 시각 서식 전부 보존, 매우 빠름. 권장 기본값.
        self.format_mode = tk.StringVar(value="clipboard")
        self.progress_text = tk.StringVar(value="")
        # 중단 요청 신호 — 워커 시작 시 clear(), 사용자 버튼 클릭 시 set().
        # 워커는 *다음 파일* 직전에 체크 → 진행 중 파일은 끝까지 완료.
        self.cancel_event = threading.Event()
        # 강제 종료 신호 — set 되면 Excel 프로세스를 즉시 kill (taskkill /F /T).
        # 진행 중 COM 호출이 com_error 로 깨지면서 현재 파일은 fail 처리, 그 뒤
        # 루프가 폴백 전략을 모두 건너뛰고 즉시 break.
        self.force_event = threading.Event()
        # Excel 프로세스 PID — 워커가 Excel 시작 직후 채워넣음 (0 = 미시작).
        # 리스트 형태로 두는 건 워커 스레드와 UI 스레드 사이에서 mutable 공유
        # 컨테이너로 쓰기 위함. tk Var 와 달리 락 없이 안전하게 정수 1개를 주고받음.
        self._excel_pid_holder = [0]
        # 워커가 클립보드/COM 을 잡고 있는 동안 UI 폴러를 일시정지하기 위한 플래그.
        # 폴러와 워커가 동시에 win32clipboard 호출하면 pywin32 가 가끔 segfault.
        self._worker_busy = False

        self._build_ui()

    def _build_ui(self):
        container = ttk.Frame(self.root, padding=12)
        container.pack(fill="both", expand=True)

        # 좌: 컨트롤 + 로그 / 우: Queue + 클립보드 모니터 (세로 스택)
        left = ttk.Frame(container)
        left.pack(side="left", fill="both", expand=True)

        right_col = ttk.Frame(container)
        right_col.pack(side="right", fill="y", padx=(8, 0))

        right = ttk.LabelFrame(right_col, text="대기열 Queue", padding=6)
        right.pack(fill="both", expand=True)

        # ---- 좌측 ----
        guide = ttk.LabelFrame(left, text="동작", padding=8)
        guide.pack(fill="x", pady=(0, 6))
        ttk.Label(
            guide,
            justify="left",
            text=(
                "원본 .xlsx 를 Excel 로 *읽기만* 하고, 셀/서식/이미지를 메모리에\n"
                "옮긴 뒤 openpyxl 로 새 .xlsx 를 직접 써냅니다 (Excel.SaveAs 미사용).\n"
                "DRM/DLP 가 Excel 저장 경로만 후킹하는 환경에서 우회 효과가 있습니다.\n"
                "출력: <출력폴더>/<원본이름>_clean.xlsx (출력 폴더에 바로 저장)\n"
                "이미 출력본이 있으면 스킵. 원본 이동 폴더가 지정되어 있으면\n"
                "처리 끝난 원본을 그쪽으로 이동 (비우면 이동 안 함).\n"
                "‘중단’ 클릭 시 현재 진행 중인 파일까지 완료 후 정지."
            ),
        ).pack(anchor="w")

        # 출력 폴더 — 내부에 Clean/ (결과) 와 Org/ (원본 이동) 가 자동 생성됨.
        out_frame = ttk.Frame(left)
        out_frame.pack(fill="x", pady=4)
        ttk.Label(out_frame, text="출력 폴더", width=12).pack(side="left")
        ttk.Entry(out_frame, textvariable=self.output_dir).pack(
            side="left", fill="x", expand=True, padx=6
        )
        ttk.Button(out_frame, text="찾아보기…", command=self._browse_output).pack(
            side="left"
        )

        # 옵션
        ttk.Checkbutton(
            left,
            text="완료 후 출력 폴더 열기",
            variable=self.open_output,
        ).pack(anchor="w", pady=2)

        # 서식 복사 모드 — 라디오 4개를 한 줄 LabelFrame 으로 묶음
        mode_frame = ttk.LabelFrame(left, text="서식 복사 모드", padding=6)
        mode_frame.pack(fill="x", pady=4)
        ttk.Radiobutton(
            mode_frame,
            text="클립보드 XML (테두리 포함 모든 서식, 매우 빠름 — 권장)",
            variable=self.format_mode, value="clipboard",
        ).pack(anchor="w")
        ttk.Radiobutton(
            mode_frame,
            text="균등 분할 (Bold/색/정렬 빠르게, 테두리 제외)",
            variable=self.format_mode, value="balanced",
        ).pack(anchor="w")
        ttk.Radiobutton(
            mode_frame,
            text="정밀 (셀별 — 테두리 포함, 가장 느림)",
            variable=self.format_mode, value="precise",
        ).pack(anchor="w")
        ttk.Radiobutton(
            mode_frame,
            text="빠름 (서식 미복사 — 가장 빠름, 평문처럼 보임)",
            variable=self.format_mode, value="fast",
        ).pack(anchor="w")

        # 실행 / 중단 버튼 — 한 줄에 나란히
        btn_row = ttk.Frame(left)
        btn_row.pack(fill="x", pady=6)
        drm_label = (
            "DRM 우회 — Excel 읽기 + openpyxl 쓰기"
            if HAS_OPENPYXL else
            "DRM 우회 — zipfile 재포장 (openpyxl 설치 권장: pip install openpyxl)"
        )
        self.drm_btn = ttk.Button(btn_row, text=drm_label, command=self._on_drm_clean)
        self.drm_btn.pack(side="left", fill="x", expand=True)
        # 스캔만 — 도형 개수만 확인해서 임계치 초과 파일을 큐에서 빼는 옵션
        self.scan_btn = ttk.Button(
            btn_row, text=f"스캔만 (>{HEAVY_SHAPE_THRESHOLD} 도형)",
            command=self._on_scan_only, width=20,
        )
        self.scan_btn.pack(side="left", padx=(6, 0))
        # 중단 버튼은 idle 상태에서 비활성, 작업 중에만 활성
        self.cancel_btn = ttk.Button(
            btn_row, text="중단", command=self._on_cancel, state="disabled", width=10,
        )
        self.cancel_btn.pack(side="left", padx=(6, 0))
        # 강제 종료 — Excel.exe 직접 kill. 중단보다 더 즉시성 있는 정지.
        self.force_btn = ttk.Button(
            btn_row, text="강제 종료", command=self._on_force_stop,
            state="disabled", width=10,
        )
        self.force_btn.pack(side="left", padx=(6, 0))

        # 진행률 — 파일별 (상단) / 전체 (하단) 두 줄.
        # 라벨에 anchor="w" + width 를 주면 텍스트가 길어도 폭이 고정되어
        # 윈도우 크기가 따라 변하지 않음.
        self.file_progress_text = tk.StringVar(value="대기")
        ttk.Label(
            left, textvariable=self.file_progress_text,
            anchor="w", width=80,
        ).pack(anchor="w", fill="x")
        self.file_progress = ttk.Progressbar(left, mode="determinate")
        self.file_progress.pack(fill="x", pady=(0, 4))

        ttk.Label(
            left, textvariable=self.progress_text,
            anchor="w", width=80,
        ).pack(anchor="w", fill="x")
        self.progress = ttk.Progressbar(left, mode="determinate")
        self.progress.pack(fill="x", pady=2)

        # 로그
        ttk.Label(left, text="로그").pack(anchor="w", pady=(8, 2))
        log_frame = ttk.Frame(left)
        log_frame.pack(fill="both", expand=True)
        self.log = tk.Text(log_frame, height=14, wrap="word")
        log_scroll = ttk.Scrollbar(
            log_frame, orient="vertical", command=self.log.yview
        )
        self.log.configure(yscrollcommand=log_scroll.set)
        self.log.pack(side="left", fill="both", expand=True)
        log_scroll.pack(side="right", fill="y")

        # ---- 우측 Queue ----
        queue_top = ttk.Frame(right)
        queue_top.pack(fill="both", expand=True)
        self.files_list = tk.Listbox(
            queue_top, height=22, width=40, selectmode="extended", activestyle="dotbox"
        )
        self.files_list.pack(side="left", fill="both", expand=True)
        list_scroll = ttk.Scrollbar(
            queue_top, orient="vertical", command=self.files_list.yview
        )
        self.files_list.configure(yscrollcommand=list_scroll.set)
        list_scroll.pack(side="right", fill="y")

        self.queue_count = tk.StringVar(value="대기 0개")
        ttk.Label(right, textvariable=self.queue_count).pack(anchor="w", pady=(4, 0))

        queue_btns = ttk.Frame(right)
        queue_btns.pack(fill="x", pady=(4, 0))
        ttk.Button(queue_btns, text="추가…", command=self._add_files).pack(
            side="left", padx=2
        )
        ttk.Button(queue_btns, text="선택 삭제", command=self._remove_selected).pack(
            side="left", padx=2
        )
        ttk.Button(queue_btns, text="모두 삭제", command=self._clear_files).pack(
            side="left", padx=2
        )

        # 드래그앤드롭 등록 — Listbox + LabelFrame + 그 안내 라벨까지
        # 모두 드롭 타겟으로 잡아 Queue 영역 어디에 떨궈도 받도록.
        if HAS_TKDND:
            for w in (self.files_list, right, queue_top):
                try:
                    w.drop_target_register(DND_FILES)
                    w.dnd_bind("<<Drop>>", self._on_drop)
                except Exception as exc:
                    self._log(f"[경고] DnD 등록 실패: {exc}")
            self.dnd_hint = ttk.Label(
                right, text="↧ 파일을 여기에 드롭하세요", foreground="#666"
            )
            self.dnd_hint.pack(anchor="w", pady=(4, 0))

        # ---- 클립보드 모니터 ---- (Queue 아래)
        # 워커가 Shape.CopyPicture / Range.CopyPicture 로 클립보드에 떨군 직후,
        # ImageGrab.grabclipboard 가 무엇을 받는지 가시화. Excel 이 내려놓는
        # 다중 포맷 (CF_HTML, CF_DIB, Biff12, Native ...) 도 같이 보임.
        # 폴링 방식 — 500ms 마다 GetClipboardSequenceNumber 만 비교 (싸다).
        # 변화 감지 시에만 OpenClipboard → EnumClipboardFormats.
        clip_frame = ttk.LabelFrame(right_col, text="클립보드 모니터", padding=6)
        clip_frame.pack(fill="both", expand=True, pady=(8, 0))

        self.clip_status = tk.StringVar(value="대기 중…")
        ttk.Label(clip_frame, textvariable=self.clip_status, foreground="#666").pack(
            anchor="w"
        )

        clip_tree_frame = ttk.Frame(clip_frame)
        clip_tree_frame.pack(fill="both", expand=True, pady=(4, 0))
        self.clip_tree = ttk.Treeview(
            clip_tree_frame,
            columns=("size",), show="tree headings", height=8,
        )
        self.clip_tree.heading("#0", text="포맷")
        self.clip_tree.heading("size", text="크기")
        self.clip_tree.column("#0", width=200, anchor="w")
        self.clip_tree.column("size", width=80, anchor="e")
        self.clip_tree.pack(side="left", fill="both", expand=True)
        clip_scroll = ttk.Scrollbar(
            clip_tree_frame, orient="vertical", command=self.clip_tree.yview
        )
        self.clip_tree.configure(yscrollcommand=clip_scroll.set)
        clip_scroll.pack(side="right", fill="y")

        ttk.Label(
            clip_frame, text="미리보기 (텍스트 포맷):", foreground="#666"
        ).pack(anchor="w", pady=(6, 0))
        self.clip_preview = tk.Text(
            clip_frame, height=5, width=40, wrap="word",
            state="disabled", font=("Consolas", 9),
        )
        self.clip_preview.pack(fill="both", expand=False)

        # 폴링 상태 — 마지막으로 본 sequence number. 0 = 미초기화.
        self._clip_last_seq = -1
        # 시작 — 500ms 후 첫 폴
        if HAS_PYWIN32:
            self.root.after(500, self._poll_clipboard)
        else:
            self.clip_status.set("pywin32 미설치 — 클립보드 모니터 비활성")

        # 환경 경고
        warnings = []
        if not HAS_PYWIN32:
            warnings.append("pywin32 미설치 — pip install pywin32")
        if not HAS_OPENPYXL:
            warnings.append("openpyxl 미설치 — pip install openpyxl")
        if not HAS_PIL:
            warnings.append("Pillow 미설치 — pip install Pillow")
        if not HAS_TKDND:
            warnings.append(
                "tkinterdnd2 미설치 — 드래그앤드롭 비활성. pip install tkinterdnd2"
            )
        for w in warnings:
            self._log(f"[경고] {w}")

    # ---- 파일 목록 (Queue) ----
    def _refresh_queue_count(self):
        try:
            self.queue_count.set(f"대기 {len(self.src_paths)}개")
        except Exception:
            pass

    def _add_files(self):
        paths = filedialog.askopenfilenames(
            title="엑셀 파일 선택 (다중 선택 가능)",
            filetypes=[
                ("엑셀 파일", "*.xlsx;*.xlsm;*.xlsb;*.xls"),
                ("모든 파일", "*.*"),
            ],
        )
        added = 0
        for p in paths:
            if p not in self.src_paths:
                self.src_paths.append(p)
                # Queue 는 파일명만 표시 (긴 경로로 가로 폭 폭주 방지)
                self.files_list.insert("end", Path(p).name)
                added += 1
        if added:
            self._log(f"파일 {added}개 추가 (총 {len(self.src_paths)}개)")
            if not self.output_dir.get().strip() and self.src_paths:
                self.output_dir.set(str(Path(self.src_paths[0]).parent))
        self._refresh_queue_count()

    def _remove_selected(self):
        for idx in reversed(list(self.files_list.curselection())):
            self.files_list.delete(idx)
            del self.src_paths[idx]
        self._refresh_queue_count()

    def _clear_files(self):
        self.files_list.delete(0, "end")
        self.src_paths.clear()
        self._refresh_queue_count()

    def _remove_from_queue(self, path: str):
        """완료/스킵된 항목을 Queue 와 src_paths 에서 동기 제거 (UI 스레드에서 호출)."""
        try:
            idx = self.src_paths.index(path)
        except ValueError:
            return
        del self.src_paths[idx]
        try:
            self.files_list.delete(idx)
        except Exception:
            pass
        self._refresh_queue_count()

    @staticmethod
    def _parse_dnd_paths(data: str) -> list:
        """tkdnd 의 경로 문자열 파싱.

        포맷: 공백 구분, 공백 포함 경로는 `{...}` 로 감싸짐.
        예) `{C:/My Files/a.xlsx} D:/b.xlsx {E:/c d.xlsm}`
        """
        paths = []
        i, n = 0, len(data)
        while i < n:
            while i < n and data[i] == ' ':
                i += 1
            if i >= n:
                break
            if data[i] == '{':
                j = data.find('}', i + 1)
                if j == -1:
                    paths.append(data[i + 1:])
                    break
                paths.append(data[i + 1:j])
                i = j + 1
            else:
                j = i
                while j < n and data[j] != ' ':
                    j += 1
                paths.append(data[i:j])
                i = j
        return paths

    def _on_drop(self, event):
        """tkinterdnd2 <<Drop>> 핸들러. 폴더면 그 안의 엑셀 파일 1단계만 수집."""
        valid_exts = {'.xlsx', '.xlsm', '.xlsb', '.xls'}
        try:
            raw = event.data
        except Exception:
            return
        candidates = []
        for p in self._parse_dnd_paths(raw):
            p = p.strip().strip('"')
            if not p:
                continue
            pth = Path(p)
            if pth.is_dir():
                # 폴더 드롭 — 1단계 엑셀 파일만 (재귀 X — 의도치 않은 폭증 방지)
                for ext in valid_exts:
                    candidates.extend(str(x) for x in pth.glob(f"*{ext}"))
            elif pth.is_file():
                candidates.append(str(pth))

        added = 0
        skipped_dup = 0
        skipped_ext = 0
        for c in candidates:
            if Path(c).suffix.lower() not in valid_exts:
                skipped_ext += 1
                continue
            if c in self.src_paths:
                skipped_dup += 1
                continue
            self.src_paths.append(c)
            self.files_list.insert("end", Path(c).name)
            added += 1

        if added or skipped_dup or skipped_ext:
            parts = [f"드롭 — {added}개 추가"]
            if skipped_dup:
                parts.append(f"중복 {skipped_dup}개 스킵")
            if skipped_ext:
                parts.append(f"비엑셀 {skipped_ext}개 스킵")
            parts.append(f"총 {len(self.src_paths)}개")
            self._log(" / ".join(parts))
            if added and not self.output_dir.get().strip() and self.src_paths:
                self.output_dir.set(str(Path(self.src_paths[0]).parent))
        self._refresh_queue_count()

    def _browse_output(self):
        path = filedialog.askdirectory(title="출력 폴더 선택")
        if path:
            self.output_dir.set(path)

    # ---- 로그 / busy / progress ----
    def _log(self, msg: str):
        def append():
            self.log.insert("end", msg + "\n")
            self.log.see("end")
        self.root.after(0, append)

    def _show_msgbox(self, kind: str, title: str, msg: str):
        """messagebox 를 UI 스레드에서 호출하도록 root.after 로 위임.

        Tkinter 의 messagebox 는 thread-safe 가 아니라 워커 스레드에서 직접
        부르면 Windows 에서 가끔 segfault → 프로세스 통째 종료. 모든 워커발
        다이얼로그는 이 메서드 경유.
        """
        def show():
            try:
                getattr(messagebox, kind)(title, msg)
            except Exception:
                pass
        self.root.after(0, show)

    def _set_busy(self, busy: bool):
        self.drm_btn.config(state="disabled" if busy else "normal")
        self.scan_btn.config(state="disabled" if busy else "normal")
        # 중단/강제 종료 버튼은 작업 중에만 활성. idle 시점에 텍스트 원복.
        if busy:
            self.cancel_btn.config(state="normal", text="중단")
            self.force_btn.config(state="normal", text="강제 종료")
        else:
            self.cancel_btn.config(state="disabled", text="중단")
            self.force_btn.config(state="disabled", text="강제 종료")

    def _on_cancel(self):
        # 단순히 신호만 set — 진행 중인 파일은 끝까지 완료 후 다음 파일 직전에 탈출
        if self.cancel_event.is_set():
            return  # 이미 요청됨 — 중복 클릭 무시
        self.cancel_event.set()
        self._log("⏹ 중단 요청 — 현재 진행 중인 파일까지만 완료 후 정지합니다.")
        # 시각 피드백: 버튼을 비활성 + 텍스트 변경
        self.cancel_btn.config(state="disabled", text="중단 중…")

    # ---- 클립보드 모니터 ----
    # 표준 CF_* 정수 → 이름 매핑. EnumClipboardFormats 가 돌려주는 정수를
    # 사람이 읽을 수 있는 라벨로. 등록 포맷 (Biff12 등) 은 GetClipboardFormatName
    # 으로 따로 조회.
    _STANDARD_CF_NAMES = {
        1: "CF_TEXT",
        2: "CF_BITMAP",
        3: "CF_METAFILEPICT",
        4: "CF_SYLK",
        5: "CF_DIF",
        6: "CF_TIFF",
        7: "CF_OEMTEXT",
        8: "CF_DIB",
        9: "CF_PALETTE",
        10: "CF_PENDATA",
        11: "CF_RIFF",
        12: "CF_WAVE",
        13: "CF_UNICODETEXT",
        14: "CF_ENHMETAFILE",
        15: "CF_HDROP",
        16: "CF_LOCALE",
        17: "CF_DIBV5",
    }

    def _cf_name(self, fmt: int) -> str:
        if fmt in self._STANDARD_CF_NAMES:
            return self._STANDARD_CF_NAMES[fmt]
        try:
            return win32clipboard.GetClipboardFormatName(fmt)
        except Exception:
            return f"#{fmt}"

    # GetClipboardData 가 핸들 기반 (HBITMAP, HMETAFILE, HENHMETAFILE, HDROP)
    # 포맷에 대해 가끔 segfault — 폴러는 이런 포맷의 데이터는 절대 만지지 않음.
    # CF_DIB/CF_DIBV5 도 큰 비트맵일 때 pywin32 의 변환에서 죽는 케이스 보고됨.
    _CF_HANDLE_BASED = frozenset({
        2,   # CF_BITMAP (HBITMAP)
        3,   # CF_METAFILEPICT
        6,   # CF_TIFF (handle)
        8,   # CF_DIB — 대용량 DIB 변환에서 죽는 사례
        14,  # CF_ENHMETAFILE (HENHMETAFILE)
        15,  # CF_HDROP
        17,  # CF_DIBV5
    })

    def _poll_clipboard(self):
        """500ms 주기 폴. SequenceNumber 가 변했을 때만 enum/preview 갱신.

        워커가 클립보드/COM 을 잡고 있는 동안은 일시정지 (segfault 회피).
        핸들 기반 포맷 (CF_DIB, CF_ENHMETAFILE 등) 은 데이터 fetch 자체를 건너뜀.
        """
        # 워커가 작업 중이면 폴링 완전 중단 — 클립보드 동시 접근 = native crash 위험
        if self._worker_busy:
            self.clip_status.set("워커 작업 중 — 폴 일시정지")
            self.root.after(1000, self._poll_clipboard)
            return

        try:
            seq = win32clipboard.GetClipboardSequenceNumber()
        except Exception:
            seq = 0
        if seq == self._clip_last_seq:
            # 변화 없음 — 다음 폴만 예약하고 끝
            self.root.after(500, self._poll_clipboard)
            return
        self._clip_last_seq = seq

        # 짧은 재시도로 열기 — 워커가 잡고 있으면 다음 틱에 다시 시도
        opened = False
        for _ in range(3):
            try:
                win32clipboard.OpenClipboard()
                opened = True
                break
            except Exception:
                time.sleep(0.03)
        if not opened:
            self.clip_status.set(f"seq={seq} (열기 실패 — 워커 점유 중)")
            self.root.after(500, self._poll_clipboard)
            return

        formats = []  # [(fmt_int, name, size_bytes_or_None, data_or_None)]
        try:
            fmt = 0
            while True:
                try:
                    fmt = win32clipboard.EnumClipboardFormats(fmt)
                except Exception:
                    break
                if not fmt:
                    break
                name = self._cf_name(fmt)
                size = None
                data = None
                # 핸들 기반 포맷은 GetClipboardData 자체가 native crash 위험 — skip
                if fmt in self._CF_HANDLE_BASED:
                    size = -1  # "(handle)" 표시
                else:
                    try:
                        data = win32clipboard.GetClipboardData(fmt)
                        if data is None:
                            size = 0
                        elif isinstance(data, (bytes, bytearray)):
                            size = len(data)
                        elif isinstance(data, str):
                            size = len(data.encode("utf-8", errors="replace"))
                        else:
                            size = -1
                            data = None
                    except Exception:
                        size = None
                        data = None
                formats.append((fmt, name, size, data))

            # 텍스트 미리보기 — CF_UNICODETEXT 우선, 없으면 HTML Format, 없으면 첫 텍스트
            preview = ""
            preferred = (13, "HTML Format", "Rich Text Format", 1, 7)
            data_by_key = {}
            for fmt_i, name_i, size_i, data_i in formats:
                if isinstance(data_i, str):
                    data_by_key[fmt_i] = data_i
                    data_by_key[name_i] = data_i
            for key in preferred:
                if key in data_by_key:
                    preview = data_by_key[key]
                    break
        finally:
            try:
                win32clipboard.CloseClipboard()
            except Exception:
                pass

        # ---- UI 갱신 ----
        self.clip_status.set(f"seq={seq} · 포맷 {len(formats)}개")

        for iid in self.clip_tree.get_children():
            self.clip_tree.delete(iid)
        for fmt_i, name_i, size_i, _ in formats:
            if size_i is None:
                size_str = "?"
            elif size_i == -1:
                size_str = "(handle)"
            elif size_i < 1024:
                size_str = f"{size_i} B"
            elif size_i < 1024 * 1024:
                size_str = f"{size_i / 1024:.1f} KB"
            else:
                size_str = f"{size_i / 1024 / 1024:.2f} MB"
            self.clip_tree.insert("", "end", text=f"{name_i} ({fmt_i})", values=(size_str,))

        self.clip_preview.configure(state="normal")
        self.clip_preview.delete("1.0", "end")
        if preview:
            # 앞쪽 600자만 — 긴 HTML 잘림
            shown = preview[:600]
            if len(preview) > 600:
                shown += f"\n\n… (+{len(preview) - 600}자)"
            self.clip_preview.insert("1.0", shown)
        else:
            self.clip_preview.insert("1.0", "(텍스트 포맷 없음)")
        self.clip_preview.configure(state="disabled")

        self.root.after(500, self._poll_clipboard)

    def _on_force_stop(self):
        """Excel 프로세스를 taskkill 로 즉시 종료.

        - 진행 중인 COM 호출은 com_error 로 깨짐 → 현재 파일은 fail 처리됨
        - cancel_event + force_event 둘 다 set 됨
        - drm_clean_resave 가 force_event 를 보고 폴백 전략 (① openpyxl, ② zipfile)
          도 건너뛰고 즉시 break

        주의: 이 버튼은 *현재 워커가 띄운* Excel 만 죽임. 사용자가 별도로
        열어둔 Excel 창은 영향 없음 (PID 가 다름).
        """
        if self.force_event.is_set():
            return  # 이미 요청됨
        self.force_event.set()
        self.cancel_event.set()  # 다음 파일 진입도 막음
        self._log("⛔ 강제 종료 — Excel 프로세스 즉시 종료 중…")
        self.force_btn.config(state="disabled", text="종료 중…")
        self.cancel_btn.config(state="disabled", text="중단 중…")

        pid = self._excel_pid_holder[0]
        if not pid:
            self._log("  (워커가 띄운 Excel PID 가 아직 없음 — 신호만 set)")
            return
        try:
            # /F = force, /T = 자식 프로세스 포함, capture_output 으로 콘솔 안 보임
            result = subprocess.run(
                ["taskkill", "/F", "/T", "/PID", str(pid)],
                capture_output=True, timeout=5, text=True,
            )
            if result.returncode == 0:
                self._log(f"  Excel PID {pid} taskkill 성공")
            else:
                # 이미 죽었거나 권한 없음 등
                self._log(
                    f"  taskkill 반환코드 {result.returncode}: "
                    f"{(result.stderr or result.stdout or '').strip()}"
                )
        except Exception as exc:
            self._log(f"  taskkill 호출 실패: {exc}")

    @staticmethod
    def _short(text: str, maxlen: int = 70) -> str:
        """긴 파일명/라벨이 윈도우 폭을 밀어내지 않도록 가운데 생략."""
        if len(text) <= maxlen:
            return text
        keep = maxlen - 1  # '…' 한 글자
        head = keep // 2
        tail = keep - head
        return f"{text[:head]}…{text[-tail:]}"

    def _on_progress(self, current: int, total: int, src: str):
        def update():
            self.progress["maximum"] = total
            self.progress["value"] = current
            self.progress_text.set(
                f"전체 {current}/{total}: {self._short(Path(src).name)}"
            )
        self.root.after(0, update)

    def _on_file_progress(self, current: int, total: int, label: str):
        """파일 안에서의 sub-step 진행률 (도형 렌더링 / 시트 스캔 등)."""
        def update():
            self.file_progress["maximum"] = max(1, total)
            self.file_progress["value"] = min(current, total)
            self.file_progress_text.set(
                f"파일별 {current}/{total} — {self._short(label)}"
            )
        self.root.after(0, update)

    # ---- 스캔만 — 도형 카운트만 빠르게 돌고 임계치 초과 파일 분리 ----
    def _on_scan_only(self):
        if not HAS_PYWIN32:
            messagebox.showerror(
                "pywin32 필요",
                "Microsoft Excel 과 pywin32 가 필요합니다.\npip install pywin32",
            )
            return
        if not self.src_paths:
            messagebox.showerror("오류", "원본 엑셀 파일을 추가해 주세요.")
            return
        missing = [p for p in self.src_paths if not os.path.isfile(p)]
        if missing:
            messagebox.showerror(
                "오류",
                "다음 파일을 찾을 수 없습니다:\n\n" + "\n".join(missing[:10]),
            )
            return
        self.cancel_event.clear()
        self.force_event.clear()
        self._set_busy(True)
        self.progress["value"] = 0
        self.progress_text.set(
            f"스캔 시작 — {len(self.src_paths)}개, 임계치 {HEAVY_SHAPE_THRESHOLD} 도형…"
        )
        threading.Thread(
            target=self._scan_only_worker, args=(list(self.src_paths),),
            daemon=True,
        ).start()

    def _scan_only_worker(self, src_paths):
        """파일별로 도형 합산 → 임계치 초과면 즉시 pending/ 으로 이동.

        중단/충돌이 나도 이미 처리한 파일은 보존되도록 발견 즉시 이동.
        """
        self._worker_busy = True
        pythoncom.CoInitialize()
        excel = None
        moved = []   # [(src, target, shape_count)]
        kept = []    # [(src, shape_count)] — 임계치 이하
        errors = []  # [(src, errmsg)]

        def _start_excel():
            """Excel 인스턴스 새로 띄우고 PID 등록. 실패 시 None."""
            try:
                ex = win32com.client.DispatchEx("Excel.Application")
                ex.Visible = False
                ex.DisplayAlerts = False
                ex.AskToUpdateLinks = False
                try:
                    hwnd = int(ex.Hwnd)
                    if hwnd:
                        _, pid = win32process.GetWindowThreadProcessId(hwnd)
                        self._excel_pid_holder[0] = int(pid)
                except Exception:
                    pass
                return ex
            except Exception as exc:
                self._log(f"  Excel 시작 실패: {exc}")
                return None

        def _is_rpc_dead(exc):
            """COM 에러가 Excel 사망(RPC) 인가?"""
            try:
                code = getattr(exc, "hresult", None)
                if code is None and hasattr(exc, "args") and exc.args:
                    code = exc.args[0]
                return code in (-2147023174, -2147417848)
            except Exception:
                return False

        try:
            excel = _start_excel()
            if excel is None:
                self.root.after(
                    0, lambda: self._show_msgbox(
                        "showerror", "오류",
                        "Excel 시작 실패 — 스캔 불가",
                    ),
                )
                return
            total = len(src_paths)
            self._log(
                f"=== 스캔 + 즉시 이동 — {total}개, 임계치 {HEAVY_SHAPE_THRESHOLD} 도형 ==="
            )
            for i, src in enumerate(src_paths, start=1):
                if self.cancel_event.is_set() or self.force_event.is_set():
                    self._log(f"⏹ 스캔 중단 — {total - i + 1}개 건너뜀")
                    break
                src_p = Path(src)
                self._on_progress(i, total, src)
                self._on_file_progress(0, 1, f"열기: {src_p.name}")

                # Excel 죽었으면 재시작 후 진행
                if excel is None:
                    self._log("  ⚠ Excel 인스턴스 없음 — 재시작 시도")
                    excel = _start_excel()
                    if excel is None:
                        self._log(f"  ✗ 재시작 실패 — {src_p.name} 스킵")
                        errors.append((src, "Excel 시작 실패"))
                        continue

                try:
                    wb = _open_workbook_robust(excel, os.path.abspath(str(src_p)))
                except Exception as exc:
                    self._log(f"  스캔 열기 실패 ({src_p.name}): {exc}")
                    errors.append((src, str(exc)))
                    # RPC 사망이면 Excel 즉시 재시작 — 다음 파일들 위해
                    if _is_rpc_dead(exc):
                        self._log("  ⚠ Excel 프로세스 사망 (RPC) — 재시작")
                        try:
                            try:
                                excel.Quit()
                            except Exception:
                                pass
                        finally:
                            excel = _start_excel()
                    continue
                shape_total = 0
                rpc_died = False
                try:
                    sheet_count = int(wb.Sheets.Count)
                    for sh_idx in range(1, sheet_count + 1):
                        try:
                            shape_total += int(wb.Sheets(sh_idx).Shapes.Count)
                        except Exception as exc:
                            if _is_rpc_dead(exc):
                                rpc_died = True
                                break
                        self._on_file_progress(
                            sh_idx, sheet_count,
                            f"시트 {sh_idx}/{sheet_count} (누적 도형 {shape_total})",
                        )
                except Exception as exc:
                    if _is_rpc_dead(exc):
                        rpc_died = True
                    self._log(f"  스캔 중 오류 ({src_p.name}): {exc}")
                finally:
                    try:
                        wb.Close(SaveChanges=False)
                    except Exception:
                        pass

                if rpc_died:
                    self._log("  ⚠ Excel 프로세스 사망 (RPC) — 재시작 후 다음 파일")
                    errors.append((src, "Excel RPC 사망"))
                    try:
                        excel.Quit()
                    except Exception:
                        pass
                    excel = _start_excel()
                    continue

                if shape_total > HEAVY_SHAPE_THRESHOLD:
                    # 즉시 이동 — 원본 폴더 내 pending/
                    try:
                        pending_dir = src_p.parent / "pending"
                        pending_dir.mkdir(parents=True, exist_ok=True)
                        target = pending_dir / src_p.name
                        if target.exists():
                            stem, suf = target.stem, target.suffix
                            k = 2
                            while (pending_dir / f"{stem}_{k}{suf}").exists():
                                k += 1
                            target = pending_dir / f"{stem}_{k}{suf}"
                        shutil.move(str(src_p), str(target))
                        self._log(
                            f"  ↪ {src_p.name}: 도형 {shape_total}개 → pending 이동"
                        )
                        moved.append((src, str(target), shape_total))
                        self.root.after(
                            0, lambda p=src: self._remove_from_queue(p),
                        )
                    except Exception as exc:
                        self._log(f"  ! pending 이동 실패 ({src_p.name}): {exc}")
                        errors.append((src, f"이동 실패: {exc}"))
                else:
                    self._log(f"  · {src_p.name}: 도형 {shape_total}개")
                    kept.append((src, shape_total))
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
            self._excel_pid_holder[0] = 0
            self._worker_busy = False
            self.root.after(
                0, lambda: self._after_scan(moved, kept, errors),
            )

    def _after_scan(self, moved, kept, errors):
        """스캔 후 요약. 이미 이동은 워커에서 즉시 처리됨."""
        self._set_busy(False)
        self._on_file_progress(0, 1, "대기")
        total = len(moved) + len(kept)
        self._log(
            f"=== 스캔 완료 — {total}개 검사, 이동 {len(moved)}개, "
            f"유지 {len(kept)}개, 오류 {len(errors)}개 ==="
        )
        self.progress_text.set(
            f"스캔 완료: 검사 {total} / 이동 {len(moved)} / 오류 {len(errors)}"
        )
        msg = (
            f"검사 {total}개\n"
            f"이동 (도형 > {HEAVY_SHAPE_THRESHOLD}) {len(moved)}개\n"
            f"유지 {len(kept)}개\n"
            f"오류 {len(errors)}개"
        )
        if errors:
            preview = "\n".join(
                f"  - {Path(s).name}: {e[:80]}" for s, e in errors[:8]
            )
            msg += f"\n\n오류 일부:\n{preview}"
        messagebox.showinfo("스캔 완료", msg)

    # ---- DRM 우회 ----
    def _on_drm_clean(self):
        if not HAS_PYWIN32:
            messagebox.showerror(
                "pywin32 필요",
                "Microsoft Excel 과 pywin32 가 필요합니다.\npip install pywin32",
            )
            return
        if not self.src_paths:
            messagebox.showerror("오류", "원본 엑셀 파일을 추가해 주세요.")
            return
        missing = [p for p in self.src_paths if not os.path.isfile(p)]
        if missing:
            messagebox.showerror(
                "오류",
                "다음 파일을 찾을 수 없습니다:\n\n" + "\n".join(missing[:10]),
            )
            return
        out = self.output_dir.get().strip().strip('"')
        if not out:
            messagebox.showerror("오류", "출력 폴더를 선택해 주세요.")
            return
        # Clean/ : 결과 *_clean.xlsx, Org/ : 처리 끝난 원본 이동지.
        # 도형 많은 파일은 "스캔만" 버튼으로 별도 처리 — DRM 우회 자체는 사전
        # 스캔 없이 모든 큐 파일을 그대로 변환 시도.
        clean_dir = str(Path(out) / "Clean")
        org_dir = str(Path(out) / "Org")
        try:
            Path(clean_dir).mkdir(parents=True, exist_ok=True)
            Path(org_dir).mkdir(parents=True, exist_ok=True)
        except Exception as exc:
            messagebox.showerror("오류", f"출력 폴더 생성 실패: {exc}")
            return

        # 새 작업 시작 — 이전 취소/강제 신호 클리어
        self.cancel_event.clear()
        self.force_event.clear()
        self._set_busy(True)
        self.progress["value"] = 0
        self.progress_text.set("DRM 우회 사본 작성…")
        threading.Thread(
            target=self._drm_clean_worker,
            args=(
                list(self.src_paths), clean_dir, self.format_mode.get(), org_dir,
            ),
            daemon=True,
        ).start()

    def _on_done(self, src: str, status: str, dest):
        """워커 스레드에서 파일 단위로 호출.

        - ok / skip : 원본을 self._current_move_dir (Org) 로 이동 + Queue 제거
        - fail      : Queue 에 남겨 둠 (재시도 대상)
        """
        if status not in ("ok", "skip"):
            return
        target_dir = getattr(self, "_current_move_dir", "") or ""
        if target_dir:
            try:
                src_p = Path(src)
                if src_p.exists():
                    done_dir = Path(target_dir)
                    done_dir.mkdir(parents=True, exist_ok=True)
                    target = done_dir / src_p.name
                    if target.exists():
                        stem, suf = target.stem, target.suffix
                        n = 2
                        while (done_dir / f"{stem}_{n}{suf}").exists():
                            n += 1
                        target = done_dir / f"{stem}_{n}{suf}"
                    shutil.move(str(src_p), str(target))
                    self._log(f"  → 원본 이동: {target}")
            except Exception as exc:
                self._log(f"  ! 원본 이동 실패: {exc}")
        # Listbox 변경은 UI 스레드에서
        self.root.after(0, lambda p=src: self._remove_from_queue(p))

    def _drm_clean_worker(self, src_paths, output_dir, format_mode, move_dir):
        # _on_done 이 읽을 이동 폴더 스냅샷 (UI 변경 영향 없게 워커 시점에 고정).
        self._current_move_dir = move_dir
        # 이전 실행의 PID 가 남아 있으면 잘못된 프로세스를 죽일 위험 — 클리어
        self._excel_pid_holder[0] = 0
        # UI 클립보드 폴러 일시정지 — 워커와 동시 win32clipboard 호출 시 native
        # crash (특히 핸들 기반 포맷) 위험이 있어 작업 종료까지 완전 중단.
        self._worker_busy = True
        try:
            mode_label = {
                "fast": "빠름 (서식 미복사)",
                "balanced": "균등 분할 (서식 빠르게 복사)",
                "precise": "정밀 (셀별 — 느림)",
                "clipboard": "클립보드 XML (XMLSS — 모든 서식)",
            }.get(format_mode, format_mode)
            self._log(
                f"DRM 우회 사본 시작 — 파일 {len(src_paths)}개 / {mode_label}\n"
                f"  Clean → {output_dir}\n"
                f"  Org   → {move_dir}"
            )
            succeeded, failed = drm_clean_resave(
                src_paths,
                output_dir,
                log=self._log,
                on_progress=self._on_progress,
                on_done=self._on_done,
                format_mode=format_mode,
                should_cancel=self.cancel_event.is_set,
                should_force_stop=self.force_event.is_set,
                excel_pid_holder=self._excel_pid_holder,
                on_file_progress=self._on_file_progress,
            )
            forced = self.force_event.is_set()
            cancelled = self.cancel_event.is_set()
            if forced:
                self._log(
                    f"=== 강제 종료: 성공 {len(succeeded)}개, 실패 {len(failed)}개, "
                    f"남은 {len(src_paths) - len(succeeded) - len(failed)}개 건너뜀 ==="
                )
            elif cancelled:
                self._log(
                    f"=== 중단됨: 성공 {len(succeeded)}개, 실패 {len(failed)}개, "
                    f"남은 {len(src_paths) - len(succeeded) - len(failed)}개 건너뜀 ==="
                )
            else:
                self._log(
                    f"=== 완료: 성공 {len(succeeded)}개, 실패 {len(failed)}개 ==="
                )
            if self.open_output.get() and not cancelled and not forced:
                try:
                    # output_dir 는 Clean 서브폴더 — 사용자가 지정한 상위(Clean/Org 둘 다 보이는)를 연다
                    os.startfile(str(Path(output_dir).parent))
                except Exception:
                    pass
            if forced:
                self._show_msgbox(
                    "showwarning", "강제 종료",
                    f"Excel 프로세스를 즉시 종료했습니다.\n성공 {len(succeeded)}개, "
                    f"실패 {len(failed)}개, 미처리 "
                    f"{len(src_paths) - len(succeeded) - len(failed)}개\n\n"
                    f"강제 종료된 파일은 부분 출력본이 남아 있을 수 있으니 확인하세요.",
                )
            elif cancelled:
                self._show_msgbox(
                    "showinfo", "중단됨",
                    f"사용자 요청으로 중단했습니다.\n성공 {len(succeeded)}개, "
                    f"실패 {len(failed)}개, 미처리 "
                    f"{len(src_paths) - len(succeeded) - len(failed)}개",
                )
            elif failed:
                preview = "\n".join(
                    f"- {Path(p).name}: {err[:120]}" for p, err in failed[:10]
                )
                self._show_msgbox(
                    "showwarning", "일부 실패",
                    f"실패 {len(failed)}개:\n\n{preview}",
                )
            else:
                self._show_msgbox(
                    "showinfo", "완료",
                    f"{len(succeeded)}개 파일 변환 완료.\n"
                    f"  Clean → {output_dir}\n"
                    f"  Org   → {Path(output_dir).parent / 'Org'}",
                )
        except Exception as exc:
            tb = traceback.format_exc()
            self._log(f"!!! 워커 예외:\n{tb}")
            _crash_log(f"_drm_clean_worker 예외:\n{tb}")
            self._show_msgbox("showerror", "오류", f"DRM 우회 작업 실패: {exc}")
        finally:
            self._worker_busy = False  # UI 폴러 재개
            self.root.after(0, lambda: self._set_busy(False))


# ===========================================================================
# 진입점
# ===========================================================================
def main():
    # 크래시 훅 — Tk 띄우기 전에 설치해야 시작 자체에서 죽어도 흔적 남음
    _install_crash_hooks()
    _crash_log(f"=== 프로그램 시작 (PID {os.getpid()}) ===")

    # tkinterdnd2 는 자체 Tk 서브클래스 — 일반 tk.Tk() 위에서는 drop_target_register 가
    # 동작하지 않는다. 라이브러리 가용 시 그쪽을 root 로 사용.
    if HAS_TKDND:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()

    # Tk 콜백 (button command, after callback 등) 안에서 던진 예외는 기본적으로
    # stderr 에만 찍히고 GUI 는 계속 살아 있는데, 우리는 파일에도 남기고 싶음.
    def _tk_callback_exception(exc_type, exc_value, exc_tb):
        tb_str = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
        _crash_log(f"!!! Tk CALLBACK EXCEPTION:\n{tb_str}")
    root.report_callback_exception = _tk_callback_exception

    try:
        ttk.Style(root).theme_use("vista")
    except tk.TclError:
        pass
    App(root)
    try:
        root.mainloop()
    except Exception:
        _crash_log(
            f"!!! mainloop 종료 예외:\n{traceback.format_exc()}"
        )
        raise
    _crash_log("=== 정상 종료 ===")


if __name__ == "__main__":
    main()
