#!/usr/bin/env python3
"""Read-only OpenXML workbook structure scanner used by the WPF batch entry point.

This deliberately is not the COM universal-grid or AI-analysis pipeline.  It only
reads package metadata, merge layout, and a bounded set of header candidates and
writes a resumable, batch-scoped inventory below ``outputs/batches``.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import html
import io
import json
import os
import posixpath
import re
import sqlite3
import sys
import time
import zipfile
from collections import Counter, defaultdict
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree as ET

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - exercised by the command error path
    load_workbook = None


SCANNER_VERSION = "structure-scan-v1"
OPENXML_EXTENSIONS = {".xlsx", ".xlsm"}
BINARY_UNSUPPORTED_EXTENSIONS = {".xls", ".xlsb"}
INVENTORY_EXTENSIONS = OPENXML_EXTENSIONS | BINARY_UNSUPPORTED_EXTENSIONS | {".html", ".htm"}
MAX_ZIP_ENTRIES = 20_000
MAX_PACKAGE_UNCOMPRESSED_BYTES = 512 * 1024 * 1024
MAX_WORKSHEET_UNCOMPRESSED_BYTES = 96 * 1024 * 1024
MAX_COMPRESSION_RATIO = 250
MAX_DECLARED_CELLS = 2_000_000
MAX_SCANNED_CELLS = 250_000
MAX_MERGE_SAMPLE = 200
MAX_HEADER_ROWS = 80

SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


class QuarantinedPackageError(RuntimeError):
    """A package is intentionally not parsed further."""


def utc_now() -> str:
    return datetime.now(UTC).isoformat().replace("+00:00", "Z")


def json_text(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def atomic_write_text(path: Path, text: str, *, encoding: str = "utf-8") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.name + ".tmp")
    temporary.write_text(text, encoding=encoding)
    os.replace(temporary, path)


def atomic_write_json(path: Path, value: object) -> None:
    atomic_write_text(path, json.dumps(value, ensure_ascii=False, indent=2) + "\n")


def safe_batch_id(value: str) -> str:
    candidate = re.sub(r"[^A-Za-z0-9._-]+", "-", value.strip()).strip(".-")
    if not candidate or candidate != value or len(candidate) > 96:
        raise ValueError("batch ID must use only letters, digits, dot, underscore, or hyphen (max 96 characters).")
    return candidate


def source_fingerprint(path: Path) -> str:
    stat = path.stat()
    payload = f"{path.resolve()}|{stat.st_size}|{stat.st_mtime_ns}".encode("utf-8", errors="replace")
    return hashlib.sha256(payload).hexdigest()


def stable_source_id(relative_path: str) -> str:
    return hashlib.sha256(relative_path.casefold().encode("utf-8", errors="replace")).hexdigest()[:20]


def normalize_label(value: object) -> str:
    text = str(value or "").strip().casefold()
    return re.sub(r"[^0-9a-z가-힣]+", "", text)


def family_tag(relative_path: str) -> str:
    name = Path(relative_path).name.casefold()
    if "brs" in name:
        return "BRS"
    if "tiu" in name:
        return "TIU"
    if "msu" in name:
        return "MSU"
    return "Other"


def is_within(child: Path, parent: Path) -> bool:
    try:
        child.resolve().relative_to(parent.resolve())
        return True
    except ValueError:
        return False


def require_output_root(service_dir: Path) -> Path:
    root = (service_dir / "outputs" / "batches").resolve()
    if not is_within(root, service_dir / "outputs"):
        raise ValueError("Batch output root must remain below service outputs.")
    root.mkdir(parents=True, exist_ok=True)
    return root


def resolve_batch_directory(service_dir: Path, batch_id: str) -> Path:
    root = require_output_root(service_dir)
    target = (root / safe_batch_id(batch_id)).resolve()
    if not is_within(target, root):
        raise ValueError("Batch ID resolves outside outputs/batches.")
    return target


def open_state(path: Path) -> sqlite3.Connection:
    connection = sqlite3.connect(path)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA journal_mode=WAL")
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS items (
            relative_path TEXT PRIMARY KEY,
            source_path TEXT NOT NULL,
            extension TEXT NOT NULL,
            kind TEXT NOT NULL,
            size_bytes INTEGER NOT NULL,
            mtime_ns INTEGER NOT NULL,
            fingerprint TEXT NOT NULL,
            selected INTEGER NOT NULL,
            status TEXT NOT NULL,
            attempts INTEGER NOT NULL DEFAULT 0,
            result_path TEXT NOT NULL DEFAULT '',
            error_text TEXT NOT NULL DEFAULT '',
            started_at TEXT NOT NULL DEFAULT '',
            finished_at TEXT NOT NULL DEFAULT ''
        )
        """
    )
    connection.execute("CREATE INDEX IF NOT EXISTS idx_items_status ON items(status, relative_path)")
    connection.commit()
    return connection


def inventory_files(root: Path) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    for path in sorted((candidate for candidate in root.rglob("*") if candidate.is_file()), key=lambda item: str(item).casefold()):
        if path.name.startswith("~$"):
            continue
        extension = path.suffix.casefold()
        if extension not in INVENTORY_EXTENSIONS:
            continue
        relative = path.relative_to(root).as_posix()
        stat = path.stat()
        kind = "openxml" if extension in OPENXML_EXTENSIONS else "unsupported_binary" if extension in BINARY_UNSUPPORTED_EXTENSIONS else "non_workbook"
        records.append(
            {
                "relativePath": relative,
                "sourcePath": str(path.resolve()),
                "extension": extension,
                "kind": kind,
                "sizeBytes": stat.st_size,
                "mtimeNs": stat.st_mtime_ns,
                "fingerprint": source_fingerprint(path),
                "familyTag": family_tag(relative),
            }
        )
    return records


def pilot_selection(records: list[dict[str, object]], count: int) -> set[str]:
    eligible = [record for record in records if record["kind"] == "openxml"]
    if count <= 0 or count >= len(eligible):
        return {str(record["relativePath"]) for record in eligible}
    buckets: dict[str, list[dict[str, object]]] = defaultdict(list)
    for record in eligible:
        buckets[str(record["familyTag"])].append(record)
    for values in buckets.values():
        values.sort(key=lambda item: stable_source_id(str(item["relativePath"])))

    target = {
        "BRS": (count * 40 + 99) // 100,
        "TIU": (count * 25 + 99) // 100,
        "MSU": (count * 20 + 99) // 100,
    }
    target["Other"] = max(0, count - sum(target.values()))
    selected: list[dict[str, object]] = []
    for name in ("BRS", "TIU", "MSU", "Other"):
        selected.extend(buckets[name][: target[name]])
    if len(selected) < count:
        selected_ids = {str(record["relativePath"]) for record in selected}
        remaining = sorted(
            (record for record in eligible if str(record["relativePath"]) not in selected_ids),
            key=lambda item: stable_source_id(str(item["relativePath"])),
        )
        selected.extend(remaining[: count - len(selected)])
    return {str(record["relativePath"]) for record in selected[:count]}


def create_batch(service_dir: Path, root: Path, batch_id: str, pilot: int) -> Path:
    batch_dir = resolve_batch_directory(service_dir, batch_id)
    batch_file = batch_dir / "batch.json"
    if batch_dir.exists():
        if not batch_dir.is_dir():
            raise ValueError(f"Batch path is not a directory: {batch_id}")
        entries = list(batch_dir.iterdir())
        unexpected = [entry.name for entry in entries if entry.name != "logs"]
        logs_path = batch_dir / "logs"
        if unexpected or (logs_path.exists() and not logs_path.is_dir()):
            raise ValueError(f"Batch already has initialized artifacts: {batch_id}. Use --resume-batch {batch_id}.")
        # The WPF launcher creates its batch-scoped log before the child scanner
        # starts. A directory containing only that log is not an initialized batch.
    else:
        batch_dir.mkdir(parents=True, exist_ok=False)
    records = inventory_files(root)
    if not records:
        raise ValueError(f"No Excel or inventory files found below: {root}")
    selected = pilot_selection(records, pilot)
    batch = {
        "schemaVersion": "structure-scan-batch-v1",
        "scannerVersion": SCANNER_VERSION,
        "batchId": batch_id,
        "createdAt": utc_now(),
        "rootPath": str(root.resolve()),
        "options": {"pilot": pilot, "recursive": True, "readOnly": True, "usesCom": False},
        "discovered": len(records),
        "selectedInitially": len(selected),
    }
    atomic_write_json(batch_file, batch)
    connection = open_state(batch_dir / "state.sqlite")
    try:
        for record in records:
            relative = str(record["relativePath"])
            kind = str(record["kind"])
            if kind == "non_workbook":
                status = "NON_WORKBOOK"
            elif kind == "unsupported_binary":
                status = "UNSUPPORTED"
            elif pilot and relative not in selected:
                status = "DEFERRED"
            else:
                status = "PENDING"
            connection.execute(
                """
                INSERT INTO items(relative_path, source_path, extension, kind, size_bytes, mtime_ns, fingerprint, selected, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    relative,
                    record["sourcePath"],
                    record["extension"],
                    kind,
                    record["sizeBytes"],
                    record["mtimeNs"],
                    record["fingerprint"],
                    1 if relative in selected else 0,
                    status,
                ),
            )
        connection.commit()
    finally:
        connection.close()
    return batch_dir


def reject_unsafe_xml(raw: bytes, label: str) -> None:
    if b"<!DOCTYPE" in raw.upper() or b"<!ENTITY" in raw.upper():
        raise QuarantinedPackageError(f"Unsafe XML declaration in {label}.")


def xml_root(raw: bytes, label: str) -> ET.Element:
    reject_unsafe_xml(raw, label)
    try:
        return ET.fromstring(raw)
    except ET.ParseError as exc:
        raise QuarantinedPackageError(f"Invalid XML in {label}: {exc}") from exc


def relationship_target(target: str) -> str:
    target = target.replace("\\", "/")
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join("xl", target))


def zip_preflight(archive: zipfile.ZipFile) -> dict[str, object]:
    entries = archive.infolist()
    if len(entries) > MAX_ZIP_ENTRIES:
        raise QuarantinedPackageError(f"ZIP entry count exceeds limit ({len(entries)} > {MAX_ZIP_ENTRIES}).")
    uncompressed = sum(info.file_size for info in entries)
    compressed = sum(info.compress_size for info in entries)
    if uncompressed > MAX_PACKAGE_UNCOMPRESSED_BYTES:
        raise QuarantinedPackageError("ZIP package uncompressed size exceeds scan limit.")
    if compressed and uncompressed / compressed > MAX_COMPRESSION_RATIO:
        raise QuarantinedPackageError("ZIP compression ratio exceeds scan limit.")
    if any(info.flag_bits & 0x1 for info in entries):
        raise QuarantinedPackageError("Encrypted ZIP package is not scanned.")
    names = {info.filename for info in entries}
    required = {"xl/workbook.xml", "xl/_rels/workbook.xml.rels"}
    if not required.issubset(names):
        raise QuarantinedPackageError("OpenXML workbook metadata is missing.")
    return {"entryCount": len(entries), "compressedBytes": compressed, "uncompressedBytes": uncompressed}


def sheet_metadata_from_zip(path: Path) -> tuple[dict[str, object], list[dict[str, object]]]:
    with zipfile.ZipFile(path) as archive:
        package = zip_preflight(archive)
        workbook = xml_root(archive.read("xl/workbook.xml"), "xl/workbook.xml")
        rels = xml_root(archive.read("xl/_rels/workbook.xml.rels"), "xl/_rels/workbook.xml.rels")
        relation_map = {
            relation.attrib.get("Id", ""): relationship_target(relation.attrib.get("Target", ""))
            for relation in rels.findall(f"{{{PACKAGE_REL_NS}}}Relationship")
        }
        sheets: list[dict[str, object]] = []
        for order, sheet in enumerate(workbook.findall(f".//{{{SPREADSHEET_NS}}}sheet"), start=1):
            relationship_id = sheet.attrib.get(f"{{{OFFICE_REL_NS}}}id", "")
            part = relation_map.get(relationship_id, "")
            record: dict[str, object] = {
                "sheetIndex": order,
                "sheetName": sheet.attrib.get("name", f"Sheet{order}"),
                "sheetState": sheet.attrib.get("state", "visible"),
                "part": part,
                "declaredDimension": "",
                "mergeCount": 0,
                "mergeRangeSample": [],
                "drawingRelationshipCount": 0,
                "warnings": [],
            }
            info = archive.getinfo(part) if part in archive.namelist() else None
            if info is None:
                record["warnings"].append("Worksheet relationship target is missing from the package.")
                sheets.append(record)
                continue
            if info.file_size > MAX_WORKSHEET_UNCOMPRESSED_BYTES:
                record["warnings"].append("Worksheet XML exceeds scan limit; header scan will be truncated.")
                sheets.append(record)
                continue
            raw = archive.read(part)
            reject_unsafe_xml(raw, part)
            merges: list[str] = []
            dimension = ""
            drawings = 0
            try:
                for _, element in ET.iterparse(io.BytesIO(raw), events=("start",)):
                    name = element.tag.rsplit("}", 1)[-1]
                    if name == "dimension" and not dimension:
                        dimension = element.attrib.get("ref", "")
                    elif name == "mergeCell":
                        reference = element.attrib.get("ref", "")
                        if reference and len(merges) < MAX_MERGE_SAMPLE:
                            merges.append(reference)
                    elif name == "drawing":
                        drawings += 1
            except ET.ParseError as exc:
                raise QuarantinedPackageError(f"Invalid worksheet XML in {part}: {exc}") from exc
            record["declaredDimension"] = dimension
            record["mergeRangeSample"] = merges
            # Count all merge tags without retaining all raw coordinates.
            record["mergeCount"] = raw.count(b"<mergeCell ") + raw.count(b"<mergeCell>")
            record["drawingRelationshipCount"] = drawings
            sheets.append(record)
        content_types = archive.read("[Content_Types].xml") if "[Content_Types].xml" in archive.namelist() else b""
        package["hasVbaProject"] = b"vbaProject" in content_types
        return package, sheets


def declared_cells(dimension: str) -> int | None:
    if not dimension:
        return None
    matches = re.findall(r"([A-Z]+)(\d+)", dimension.upper())
    if not matches:
        return None

    def column_number(label: str) -> int:
        value = 0
        for letter in label:
            value = value * 26 + ord(letter) - ord("A") + 1
        return value

    values = [(column_number(column), int(row)) for column, row in matches]
    if len(values) == 1:
        return values[0][0] * values[0][1]
    return (abs(values[-1][0] - values[0][0]) + 1) * (abs(values[-1][1] - values[0][1]) + 1)


def header_token(value: object) -> str | None:
    normalized = normalize_label(value)
    if not normalized:
        return None
    if normalized == "input":
        return "INPUT"
    if normalized == "ok":
        return "OK"
    if normalized in {"totalng", "ngtotal", "totaldefect", "totaldefects"}:
        return "TOTAL_NG"
    if normalized in {"ngrate", "defectrate", "totalngrate"}:
        return "NG_RATE"
    if normalized in {"sample", "samples"}:
        return "SAMPLE"
    if normalized in {"average", "avg", "mean"}:
        return "AVERAGE"
    if normalized in {"max", "maximum"}:
        return "MAX"
    if normalized in {"min", "minimum"}:
        return "MIN"
    if normalized in {"normal", "baseline", "before"}:
        return "NORMAL_CUE"
    if normalized in {"test", "trial", "after"}:
        return "TEST_CUE"
    if normalized == "ng" or normalized.startswith("ng"):
        return "NG_CUE"
    return None


def column_label(column: int) -> str:
    letters = ""
    while column:
        column, remainder = divmod(column - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters


def header_range(cells: list[tuple[int, int, str]]) -> str:
    rows = [cell[0] for cell in cells]
    columns = [cell[1] for cell in cells]
    if not rows or not columns:
        return ""
    return f"{column_label(min(columns))}{min(rows)}:{column_label(max(columns))}{max(rows)}"


def classify_header_rows(sheet_name: str, state: str, rows: dict[int, list[tuple[int, int, str]]], merge_count: int) -> list[dict[str, object]]:
    sections: list[dict[str, object]] = []
    for row_number in sorted(rows):
        window: list[tuple[int, int, str]] = list(rows[row_number])
        window.extend(rows.get(row_number + 1, []))
        tokens = {token for _, _, token in window}
        range_address = header_range(window)
        base = {"sheetName": sheet_name, "sheetState": state, "headerRange": range_address, "mergeCount": merge_count}
        if {"INPUT", "OK", "TOTAL_NG", "NG_RATE"}.issubset(tokens):
            sections.append({**base, "type": "DEFECT_ACCOUNTING_LAYOUT_CANDIDATE", "confidence": "HIGH"})
        if {"SAMPLE", "AVERAGE", "MAX", "MIN"}.issubset(tokens):
            sections.append({**base, "type": "RAW_MEASUREMENT_SUMMARY_LAYOUT_CANDIDATE", "confidence": "HIGH"})
        if "NG_CUE" in tokens and "TOTAL_NG" in tokens:
            sections.append({**base, "type": "NG_BREAKDOWN_MATRIX_CANDIDATE", "confidence": "MEDIUM"})
        if {"NORMAL_CUE", "TEST_CUE"}.issubset(tokens):
            sections.append({**base, "type": "EXPLICIT_COHORT_COMPARISON_CANDIDATE", "confidence": "LOW"})
    unique: dict[tuple[str, str], dict[str, object]] = {}
    for section in sections:
        unique[(str(section["type"]), str(section["headerRange"]))] = section
    return list(unique.values())


def scan_headers(path: Path, sheet_metadata: list[dict[str, object]]) -> tuple[list[dict[str, object]], bool]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for .xlsx/.xlsm header scanning but is not installed.")
    meta_by_name = {str(record["sheetName"]): record for record in sheet_metadata}
    scanned: list[dict[str, object]] = []
    any_truncated = False
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        for worksheet in workbook.worksheets:
            metadata = meta_by_name.get(worksheet.title, {})
            record = dict(metadata)
            record.setdefault("sheetName", worksheet.title)
            record.setdefault("sheetState", worksheet.sheet_state)
            record["formulaCount"] = 0
            record["scannedCellCount"] = 0
            record["headerRows"] = []
            record["sections"] = []
            record.setdefault("warnings", [])
            declared = declared_cells(str(record.get("declaredDimension") or worksheet.calculate_dimension()))
            if declared is not None and declared > MAX_DECLARED_CELLS:
                record["warnings"].append("Declared worksheet dimension exceeds cell budget; header scan truncated.")
                record["scanStatus"] = "TRUNCATED"
                any_truncated = True
                scanned.append(record)
                continue
            token_rows: dict[int, list[tuple[int, int, str]]] = defaultdict(list)
            visible_rows: dict[int, list[str]] = defaultdict(list)
            truncated = False
            for row in worksheet.iter_rows():
                for cell in row:
                    record["scannedCellCount"] += 1
                    if int(record["scannedCellCount"]) > MAX_SCANNED_CELLS:
                        truncated = True
                        break
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        record["formulaCount"] += 1
                        continue
                    token = header_token(value)
                    if token is None:
                        continue
                    token_rows[cell.row].append((cell.row, cell.column, token))
                    if len(visible_rows[cell.row]) < 12:
                        visible_rows[cell.row].append(token)
                if truncated:
                    break
            record["headerRows"] = [
                {"row": row, "tokens": values}
                for row, values in sorted(visible_rows.items())[:MAX_HEADER_ROWS]
            ]
            record["sections"] = classify_header_rows(
                str(record["sheetName"]),
                str(record["sheetState"]),
                token_rows,
                int(record.get("mergeCount") or 0),
            )
            if truncated:
                record["warnings"].append("Header scan cell budget reached.")
                record["scanStatus"] = "TRUNCATED"
                any_truncated = True
            else:
                record["scanStatus"] = "OK"
            scanned.append(record)
    finally:
        workbook.close()
    return scanned, any_truncated


def primary_structure(sheets: Iterable[dict[str, object]], truncated: bool) -> str:
    section_count = sum(len(sheet.get("sections", [])) for sheet in sheets)
    if truncated:
        return "TRUNCATED_OR_UNREADABLE"
    if section_count == 0:
        return "NO_RECOGNIZED_TABLE"
    if section_count == 1:
        return "TABULAR_SINGLE"
    return "TABULAR_MULTI_SECTION"


def scan_openxml(path: Path, item: sqlite3.Row) -> dict[str, object]:
    package, metadata = sheet_metadata_from_zip(path)
    sheets, truncated = scan_headers(path, metadata)
    sections = [section for sheet in sheets for section in sheet.get("sections", [])]
    return {
        "schemaVersion": "structure-scan-result-v1",
        "scannerVersion": SCANNER_VERSION,
        "source": {
            "relativePath": item["relative_path"],
            "sourcePath": item["source_path"],
            "extension": item["extension"],
            "sizeBytes": item["size_bytes"],
            "mtimeNs": item["mtime_ns"],
            "fingerprint": item["fingerprint"],
            "familyTag": family_tag(item["relative_path"]),
        },
        "scanStatus": "TRUNCATED" if truncated else "OK",
        "readOnly": True,
        "usesCom": False,
        "package": package,
        "sheets": sheets,
        "primaryStructure": primary_structure(sheets, truncated),
        "structuralTypes": sorted({str(section["type"]) for section in sections}),
        "limitations": [
            "This is a header-and-merge layout scan only; it does not calculate formulas or assess data values.",
            "Candidate sections are not acceptance, quality, causality, or Test/Control conclusions.",
        ],
    }


def unsupported_result(item: sqlite3.Row) -> dict[str, object]:
    extension = str(item["extension"])
    status = "UNSUPPORTED_XLS" if extension == ".xls" else "UNSUPPORTED_XLSB"
    return {
        "schemaVersion": "structure-scan-result-v1",
        "scannerVersion": SCANNER_VERSION,
        "source": {
            "relativePath": item["relative_path"],
            "sourcePath": item["source_path"],
            "extension": extension,
            "sizeBytes": item["size_bytes"],
            "mtimeNs": item["mtime_ns"],
            "fingerprint": item["fingerprint"],
            "familyTag": family_tag(item["relative_path"]),
        },
        "scanStatus": "UNSUPPORTED",
        "primaryStructure": "TRUNCATED_OR_UNREADABLE",
        "structuralTypes": [],
        "limitations": [f"{status}: this read-only OpenXML scanner does not open {extension} files.", "No COM or automatic conversion fallback was used."],
    }


def append_event(batch_dir: Path, event: dict[str, object]) -> None:
    with (batch_dir / "events.jsonl").open("a", encoding="utf-8") as stream:
        stream.write(json.dumps(event, ensure_ascii=False, sort_keys=True) + "\n")
    with (batch_dir / "logs" / "batch.log").open("a", encoding="utf-8") as stream:
        stream.write(f"[{utc_now()}] {event['relativePath']} {event['status']}\n")


def current_fingerprint_matches(item: sqlite3.Row) -> bool:
    path = Path(str(item["source_path"]))
    return path.is_file() and source_fingerprint(path) == item["fingerprint"]


def write_result(batch_dir: Path, item: sqlite3.Row, result: dict[str, object]) -> str:
    relative = Path("results") / f"{stable_source_id(str(item['relative_path']))}.json"
    atomic_write_json(batch_dir / relative, result)
    return relative.as_posix()


def process_item(connection: sqlite3.Connection, batch_dir: Path, item: sqlite3.Row) -> None:
    if not current_fingerprint_matches(item):
        connection.execute(
            "UPDATE items SET status='CHANGED', error_text=?, finished_at=? WHERE relative_path=?",
            ("Source path, size, or mtime changed from the batch snapshot.", utc_now(), item["relative_path"]),
        )
        connection.commit()
        append_event(batch_dir, {"at": utc_now(), "relativePath": item["relative_path"], "status": "CHANGED"})
        return
    connection.execute(
        "UPDATE items SET status='SCANNING', attempts=attempts+1, started_at=?, error_text='' WHERE relative_path=?",
        (utc_now(), item["relative_path"]),
    )
    connection.commit()
    started = time.monotonic()
    try:
        if item["kind"] == "unsupported_binary":
            result = unsupported_result(item)
            status = "UNSUPPORTED"
        else:
            result = scan_openxml(Path(str(item["source_path"])), item)
            status = "TRUNCATED" if result["scanStatus"] == "TRUNCATED" else "SCANNED"
        result["elapsedSeconds"] = round(time.monotonic() - started, 3)
        if not current_fingerprint_matches(item):
            connection.execute(
                "UPDATE items SET status='CHANGED', error_text=?, finished_at=? WHERE relative_path=?",
                ("Source changed while it was being scanned.", utc_now(), item["relative_path"]),
            )
            connection.commit()
            append_event(batch_dir, {"at": utc_now(), "relativePath": item["relative_path"], "status": "CHANGED"})
            return
        result_path = write_result(batch_dir, item, result)
        connection.execute(
            "UPDATE items SET status=?, result_path=?, finished_at=? WHERE relative_path=?",
            (status, result_path, utc_now(), item["relative_path"]),
        )
        connection.commit()
        append_event(batch_dir, {"at": utc_now(), "relativePath": item["relative_path"], "status": status, "result": result_path})
    except (zipfile.BadZipFile, QuarantinedPackageError, KeyError) as exc:
        result = {
            "schemaVersion": "structure-scan-result-v1",
            "scannerVersion": SCANNER_VERSION,
            "source": {"relativePath": item["relative_path"], "sourcePath": item["source_path"], "extension": item["extension"], "fingerprint": item["fingerprint"]},
            "scanStatus": "QUARANTINED",
            "primaryStructure": "TRUNCATED_OR_UNREADABLE",
            "structuralTypes": [],
            "limitations": [str(exc)],
            "elapsedSeconds": round(time.monotonic() - started, 3),
        }
        result_path = write_result(batch_dir, item, result)
        connection.execute(
            "UPDATE items SET status='QUARANTINED', result_path=?, error_text=?, finished_at=? WHERE relative_path=?",
            (result_path, str(exc)[:2000], utc_now(), item["relative_path"]),
        )
        connection.commit()
        append_event(batch_dir, {"at": utc_now(), "relativePath": item["relative_path"], "status": "QUARANTINED", "error": str(exc)})
    except Exception as exc:  # Keep a per-file failure from aborting the corpus scan.
        connection.execute(
            "UPDATE items SET status='FAILED_RETRYABLE', error_text=?, finished_at=? WHERE relative_path=?",
            (f"{type(exc).__name__}: {exc}"[:2000], utc_now(), item["relative_path"]),
        )
        connection.commit()
        append_event(batch_dir, {"at": utc_now(), "relativePath": item["relative_path"], "status": "FAILED_RETRYABLE", "error": f"{type(exc).__name__}: {exc}"})


def item_rows(connection: sqlite3.Connection) -> list[sqlite3.Row]:
    return list(connection.execute("SELECT * FROM items ORDER BY relative_path"))


def result_for_row(batch_dir: Path, row: sqlite3.Row) -> dict[str, object]:
    path = str(row["result_path"] or "")
    if not path:
        return {}
    try:
        return json.loads((batch_dir / path).read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def build_outputs(batch_dir: Path, connection: sqlite3.Connection) -> dict[str, object]:
    rows = item_rows(connection)
    status_counts = Counter(str(row["status"]) for row in rows)
    structure_counts: Counter[str] = Counter()
    section_counts: Counter[str] = Counter()
    report_rows: list[dict[str, str]] = []
    failure_rows: list[dict[str, str]] = []
    for row in rows:
        result = result_for_row(batch_dir, row)
        structure = str(result.get("primaryStructure") or "")
        types = [str(value) for value in result.get("structuralTypes", [])]
        if structure:
            structure_counts[structure] += 1
        section_counts.update(types)
        sheet_count = len(result.get("sheets", [])) if isinstance(result.get("sheets"), list) else 0
        merge_count = sum(int(sheet.get("mergeCount") or 0) for sheet in result.get("sheets", []) if isinstance(sheet, dict))
        warnings = "; ".join(str(value) for value in result.get("limitations", []))
        report = {
            "relativePath": str(row["relative_path"]),
            "extension": str(row["extension"]),
            "familyTag": family_tag(str(row["relative_path"])),
            "status": str(row["status"]),
            "primaryStructure": structure,
            "structuralTypes": "; ".join(types),
            "sheetCount": str(sheet_count),
            "mergeCount": str(merge_count),
            "resultPath": str(row["result_path"]),
            "warningOrError": str(row["error_text"] or warnings),
        }
        report_rows.append(report)
        if row["status"] in {"FAILED_RETRYABLE", "QUARANTINED", "CHANGED"}:
            failure_rows.append(report)
    summary = {
        "schemaVersion": "structure-scan-summary-v1",
        "scannerVersion": SCANNER_VERSION,
        "generatedAt": utc_now(),
        "totalItems": len(rows),
        "statusCounts": dict(sorted(status_counts.items())),
        "primaryStructureCounts": dict(sorted(structure_counts.items())),
        "sectionCandidateCounts": dict(sorted(section_counts.items())),
        "classificationCsv": "classification.csv",
        "classificationHtml": "classification.html",
        "failuresCsv": "failures.csv",
    }
    atomic_write_json(batch_dir / "summary.json", summary)
    columns = list(report_rows[0].keys()) if report_rows else ["relativePath", "status"]
    for name, values in (("classification.csv", report_rows), ("failures.csv", failure_rows)):
        temporary = batch_dir / (name + ".tmp")
        with temporary.open("w", encoding="utf-8-sig", newline="") as stream:
            writer = csv.DictWriter(stream, fieldnames=columns)
            writer.writeheader()
            writer.writerows(values)
        os.replace(temporary, batch_dir / name)
    html_rows = "".join(
        "<tr>" + "".join(f"<td>{html.escape(row[column])}</td>" for column in columns) + "</tr>"
        for row in report_rows
    )
    count_rows = "".join(
        f"<li>{html.escape(name)}: {count}</li>" for name, count in sorted(status_counts.items())
    )
    document = f"""<!doctype html><html lang='ko'><head><meta charset='utf-8'><title>구조 사전검사</title>
<style>body{{font-family:Segoe UI,sans-serif;margin:24px;color:#172b4d}}table{{border-collapse:collapse;width:100%;font-size:12px}}th,td{{border:1px solid #d7dde7;padding:6px;text-align:left;vertical-align:top}}th{{background:#edf2f7}}code{{font-family:Consolas,monospace}}</style></head>
<body><h1>Excel 구조 사전검사</h1><p>이 결과는 header/merge layout 후보만 다루며 품질·비교·승인 결론을 만들지 않습니다. Excel COM, Office, universal-grid DB는 사용하지 않았습니다.</p>
<h2>상태</h2><ul>{count_rows}</ul><p><a href='classification.csv'>CSV</a> · <a href='summary.json'>JSON</a> · <a href='failures.csv'>실패 CSV</a></p>
<table><thead><tr>{''.join(f'<th>{html.escape(column)}</th>' for column in columns)}</tr></thead><tbody>{html_rows}</tbody></table></body></html>"""
    atomic_write_text(batch_dir / "classification.html", document)
    return summary


def pending_rows(connection: sqlite3.Connection, retry_failed: bool, limit: int, *, include_deferred: bool) -> list[sqlite3.Row]:
    connection.execute("UPDATE items SET status='INTERRUPTED' WHERE status='SCANNING'")
    connection.execute("UPDATE items SET status='PENDING' WHERE status='INTERRUPTED'")
    statuses = ["PENDING"]
    if include_deferred:
        statuses.append("DEFERRED")
    if retry_failed:
        statuses.append("FAILED_RETRYABLE")
    placeholders = ",".join("?" for _ in statuses)
    sql = f"SELECT * FROM items WHERE status IN ({placeholders}) ORDER BY relative_path"
    if limit > 0:
        sql += " LIMIT ?"
        return list(connection.execute(sql, (*statuses, limit)))
    return list(connection.execute(sql, statuses))


def run(args: argparse.Namespace) -> int:
    service_dir = Path(args.service_dir).resolve()
    if not service_dir.is_dir():
        raise ValueError(f"Service directory does not exist: {service_dir}")
    if bool(args.batch_folder) == bool(args.resume_batch):
        raise ValueError("Specify exactly one of --batch-folder or --resume-batch.")
    if args.pilot and args.limit:
        raise ValueError("--pilot and --limit cannot be used together.")
    is_new_batch = bool(args.batch_folder)
    if is_new_batch:
        root = Path(args.batch_folder)
        if not root.is_absolute() or not root.is_dir():
            raise ValueError("--batch-folder must be an existing absolute directory.")
        batch_id = args.batch_id or f"structure-scan-{datetime.now(UTC):%Y%m%dT%H%M%SZ}"
        batch_dir = create_batch(service_dir, root.resolve(), safe_batch_id(batch_id), args.pilot)
    else:
        if args.batch_id:
            raise ValueError("--batch-id cannot be used with --resume-batch.")
        batch_dir = resolve_batch_directory(service_dir, safe_batch_id(args.resume_batch))
        if not (batch_dir / "batch.json").is_file():
            raise ValueError(f"Batch does not exist: {args.resume_batch}")
    (batch_dir / "logs").mkdir(parents=True, exist_ok=True)
    connection = open_state(batch_dir / "state.sqlite")
    try:
        for item in pending_rows(connection, args.retry_failed, args.limit, include_deferred=not is_new_batch):
            process_item(connection, batch_dir, item)
        summary = build_outputs(batch_dir, connection)
    finally:
        connection.close()
    print(json.dumps({"status": "ok", "batchId": batch_dir.name, "batchDirectory": str(batch_dir), "summary": summary}, ensure_ascii=False))
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Read-only OpenXML structure scanner. It never starts Excel or COM.")
    parser.add_argument("--service-dir", required=True)
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument("--batch-folder")
    source.add_argument("--resume-batch")
    parser.add_argument("--batch-id")
    parser.add_argument("--pilot", type=int, default=0)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--retry-failed", action="store_true")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    if args.pilot < 0 or args.limit < 0:
        parser.error("--pilot and --limit must be non-negative.")
    try:
        return run(args)
    except (ValueError, RuntimeError) as exc:
        print(f"structure-scan error: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
