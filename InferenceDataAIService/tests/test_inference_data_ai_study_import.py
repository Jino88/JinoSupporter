from __future__ import annotations

import copy
import importlib.util
import json
import sqlite3
import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import Workbook

import inference_data_ai_query as query
import inference_data_ai_schema as schema
import inference_data_ai_source_ingest as source_ingest
import inference_data_ai_study_import as study_import


CLI_PATH = Path(__file__).parents[1] / "inference_data_ai_cli.py"
SPEC = importlib.util.spec_from_file_location("inference_data_ai_cli_for_study_import", CLI_PATH)
assert SPEC is not None and SPEC.loader is not None
cli = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(cli)


class CanonicalStudyImportTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory(dir=cli.SERVICE_DIR)
        self.root = Path(self.temp.name)
        self.source = self.root / "generic-review.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data"
        rows = [
            ["Condition", "Input", "NG", "Rate"],
            ["Changed 8 s", 10, 2, 0.2],
            ["Control 4 s", 10, 1, 0.1],
            ["Result", "Changed is higher", None, None],
        ]
        for row in rows:
            sheet.append(row)
        sheet["A5"] = "IV. Decision."
        sheet["B5"] = (
            "- Lot test frame new mold check SPK OK -> Continue move modul test"
        )
        sheet["B6"] = "=> Can use"
        sheet["B6"].data_type = "s"
        sheet["D2"].number_format = "0%"
        sheet["D3"].number_format = "0%"
        wide = workbook.create_sheet("WideData")
        wide.append(["Position", "R1", "R2", "R3", "R4"])
        wide.append([0, 10.1, 10.2, 10.3, "not numeric"])
        wide.append([5, 11.1, 11.2, 11.3, 12.4])
        wide.append(["extra row", None, None, None, None])
        header_axis = workbook.create_sheet("HeaderAxis")
        header_axis.append(["Specimen", "100.00Hz", "200.00Hz"])
        header_axis.append(["Sample-1", 1.1, 1.2])
        header_axis.append(["Sample-2", 2.1, 2.2])
        header_axis.append(["Sample-3", 3.1, 3.2])
        percent_data = workbook.create_sheet("PercentData")
        percent_data.append(["Specimen", "Rate"])
        percent_data.append(["Sample-1", 0.1])
        percent_data.append(["Sample-2", "12.5"])
        percent_data["B2"].number_format = "0%"
        percent_exact = workbook.create_sheet("PercentExact")
        percent_exact.append(["Rate"])
        percent_exact.append([0.220735789])
        percent_exact.append([0])
        percent_exact["A2"].number_format = "0.0%"
        percent_exact["A3"].number_format = "0.0%"
        average_data = workbook.create_sheet("AverageData")
        average_data.append(["Axis", "R1", "R2", "AVG"])
        average_data.append([100, 2, 4, 3])
        average_data.append([200, 10, 14, 12])
        formatted_identity = workbook.create_sheet("FormattedIdentity")
        formatted_identity.append(["Axis", 1, 1])
        formatted_identity.append([100, 2.5, 3.5])
        formatted_identity["B1"].number_format = '"18kPa #"?"_Before"'
        formatted_identity["C1"].number_format = '"18kPa #"?"_After"'
        count_ratio = workbook.create_sheet("CountRatio")
        count_ratio.append(
            [
                "Exact",
                "Spaced",
                "EA",
                "Narrative",
                "Range",
                "Bare ratio",
            ]
        )
        count_ratio.append(
            [
                "1/8 pcs",
                " 2 / 8 PCS ",
                "3/8 ea",
                "NG 4/8 pcs observed",
                "(0.4~0.5)mg",
                "5/8",
            ]
        )
        factor_data = workbook.create_sheet("FactorData")
        factor_data.append(["Arm", "Amount", "Other text"])
        factor_data.append(
            ["Normal (Line)", "1.56mg", "0.4~0.5mg"]
        )
        factor_data.append(
            ["Control explicit", "1.56 mg", "about 1.56mg"]
        )
        factor_data.append(
            ["Other", "1.56widget", "2026-07-18"]
        )
        factor_data.append(
            ["Other 2", "1/8 mg", "Model 123"]
        )
        unknown_quantity = workbook.create_sheet("UnknownQuantity")
        unknown_quantity.append(["Composite source", "Exact source"])
        unknown_quantity.append(["Lift target 5.6kg", "5.6kg"])
        unknown_quantity.append(["Drive at 8V", "8V"])
        unknown_quantity.append(["Wait 2 day before test", "2 day"])
        unknown_quantity.append(["Dry for 5min", "5min"])
        unknown_quantity.append(["Run condition (1st)", "1st"])
        unknown_quantity.append(["Total condition", "Total"])
        role_data = workbook.create_sheet("RoleData")
        role_data.append(["Arm label"])
        role_data.append(["ST"])
        role_data.append(["Standard condition"])
        grouped_reference = workbook.create_sheet("GroupedReference")
        grouped_reference.append(
            ["Ordered reference", "Mixed identity", "Out of order"]
        )
        for replicate in range(1, 11):
            grouped_reference.cell(
                row=replicate + 1,
                column=1,
                value=f"Normal #{replicate}",
            )
        grouped_reference["B2"] = "Test #1"
        grouped_reference["B3"] = "Normal #2"
        grouped_reference["C2"] = "Normal #2"
        grouped_reference["C3"] = "Normal #1"
        status_data = workbook.create_sheet("StatusData")
        status_data.append(["Replicate", "Status"])
        status_data.append(["R1", "PASSED"])
        status_data.append(["R2", "PASSED"])
        labeled_percent = workbook.create_sheet("LabeledPercent")
        labeled_percent.append(["Result"])
        labeled_percent.append(
            ["NG function 89.06% (Gauss  NG 71.88%)"]
        )
        comparison_series = workbook.create_sheet("ComparisonSeries")
        comparison_series.append(
            ["Axis", "Changed R1", "Control R7", "Reordered Axis", "Reordered R9"]
        )
        comparison_series.append([100, 1.1, 2.1, 200, 3.1])
        comparison_series.append([200, 1.2, 2.2, 100, 3.2])
        merged_identity = workbook.create_sheet("MergedIdentity")
        merged_identity.append(["Identity", "R1"])
        merged_identity.merge_cells("A2:A3")
        merged_identity["A2"] = "Merged batch"
        merged_identity["B2"] = 1.1
        merged_identity["B3"] = 1.2
        merged_identity["A4"].number_format = "@"
        merged_identity["B4"] = 1.3
        merged_identity["A5"] = "row-5"
        merged_identity["A6"] = "row-6"
        merged_identity.merge_cells("B5:B6")
        merged_identity["B5"] = 2.5
        merged_header = workbook.create_sheet("MergedHeader")
        merged_header.merge_cells("B3:G4")
        merged_header["B3"] = "RESULT CHECKING FO"
        merged_header["A5"] = "sample-1"
        merged_header["A6"] = "sample-2"
        merged_header["C5"] = 1.1
        merged_header["C6"] = 1.2
        merged_header["D5"] = 2.1
        merged_header["D6"] = 2.2
        merged_header["H5"] = 3.1
        merged_header["H6"] = 3.2
        workbook.save(self.source)
        workbook.close()
        self.db = self.root / "knowledge.sqlite"

    def tearDown(self) -> None:
        self.temp.cleanup()

    def prepare_source(self) -> tuple[sqlite3.Connection, dict]:
        connection = sqlite3.connect(self.db)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys=ON")
        cli.ensure_universal_schema(connection)
        payload = source_ingest.extract_workbook(self.source)
        capture = source_ingest.import_capture(connection, payload, captured_at="2026-07-17T00:00:00Z")
        bridge = source_ingest.bridge_capture_to_canonical_source(
            connection,
            dataset="Fixture",
            payload=payload,
            capture_result=capture,
            captured_at="2026-07-17T00:00:00Z",
        )
        connection.commit()
        return connection, bridge

    def manifest(self, revision_uid: str) -> dict:
        evidence = [{"sheet": "Data", "range": "A1:D4", "role": "SOURCE"}]
        return {
            "schemaVersion": "canonical-study-manifest-v1",
            "source": {
                "dataset": "Fixture",
                "sourcePath": str(self.source.resolve()),
                "revisionUid": revision_uid,
                "contentComplete": True,
            },
            "workbookAnalysis": {
                "key": "generic-cooling-review",
                "title": "Generic cooling review",
                "verificationStatus": "VERIFIED",
                "summary": "A changed condition and control were compared.",
                "evidence": evidence,
            },
            "studies": [
                {
                    "key": "cooling-time-vs-custom-ng",
                    "title": "Cooling duration versus custom NG",
                    "designType": "CONTROL_TEST",
                    "verificationStatus": "VERIFIED",
                    "comparabilityStatus": "VALID",
                    "confoundingStatus": "NONE",
                    "evidence": evidence,
                    "contexts": [
                        {
                            "key": "model",
                            "kind": "MODEL",
                            "originalValue": "Completely New Model",
                            "canonicalName": "Completely New Model",
                            "evidence": evidence,
                        }
                    ],
                    "factors": [
                        {
                            "key": "cooling-duration",
                            "originalLabel": "Unseen cooling duration",
                            "canonicalName": "Cooling duration",
                            "baselineCondition": "4 s",
                            "changedCondition": "8 s",
                            "isolationStatus": "ISOLATED",
                            "evidence": evidence,
                        }
                    ],
                    "arms": [
                        {
                            "key": "changed",
                            "role": "TEST",
                            "label": "Changed 8 s",
                            "sampleSize": 10,
                            "evidence": evidence,
                            "factorValues": [{"factor": "cooling-duration", "value": "8", "unit": "s"}],
                        },
                        {
                            "key": "control",
                            "role": "CONTROL",
                            "label": "Control 4 s",
                            "sampleSize": 10,
                            "evidence": evidence,
                            "factorValues": [
                                {"factor": "cooling-duration", "value": "4", "unit": "s", "isBaseline": True}
                            ],
                        },
                    ],
                    "outcomes": [
                        {
                            "key": "custom-ng",
                            "originalLabel": "Previously unseen custom NG",
                            "canonicalName": "Previously unseen custom NG",
                            "metricType": "defect_rate",
                            "unit": "%",
                            "denominatorBasis": "Input",
                            "favorableDirection": "LOWER",
                            "evidence": evidence,
                            "observations": [
                                {
                                    "key": "changed-value",
                                    "arm": "changed",
                                    "numerator": 2,
                                    "denominator": 10,
                                    "verificationStatus": "VERIFIED",
                                    "evidence": [{"sheet": "Data", "range": "A2:D2", "role": "OBSERVATION"}],
                                },
                                {
                                    "key": "control-value",
                                    "arm": "control",
                                    "numerator": 1,
                                    "denominator": 10,
                                    "verificationStatus": "VERIFIED",
                                    "evidence": [{"sheet": "Data", "range": "A3:D3", "role": "OBSERVATION"}],
                                },
                            ],
                        }
                    ],
                    "comparisons": [
                        {
                            "key": "changed-vs-control",
                            "comparedArm": "changed",
                            "controlArm": "control",
                            "designType": "CONTROL_TEST",
                            "matchingBasis": "same model, lot, line and date",
                            "validityStatus": "VALID",
                            "confoundingStatus": "NONE",
                            "verificationStatus": "VERIFIED",
                            "aggregationEligible": True,
                            "evidence": evidence,
                            "effects": [
                                {
                                    "outcome": "custom-ng",
                                    "effectType": "PERCENTAGE_POINT_CHANGE",
                                    "estimate": 10,
                                    "unit": "%p",
                                    "verificationStatus": "VERIFIED",
                                    "evidence": evidence,
                                }
                            ],
                        }
                    ],
                    "conclusions": [
                        {
                            "key": "observed-result",
                            "text": "Changed is higher",
                            "claimType": "SOURCE_CONCLUSION",
                            "causalStrength": "ASSOCIATION",
                            "verificationStatus": "VERIFIED",
                            "evidence": [
                                {
                                    "sheet": "Data",
                                    "range": "A4:B4",
                                    "role": "CONCLUSION",
                                    "sourceText": "Changed is higher",
                                }
                            ],
                        }
                    ],
                }
            ],
        }

    def unverified_manifest(self, revision_uid: str) -> dict:
        manifest = self.manifest(revision_uid)
        manifest["workbookAnalysis"]["verificationStatus"] = "NEEDS_REVIEW"
        study = manifest["studies"][0]
        study["verificationStatus"] = "NEEDS_REVIEW"
        study["comparabilityStatus"] = "UNASSESSED"
        study["confoundingStatus"] = "UNASSESSED"
        for observation in study["outcomes"][0]["observations"]:
            observation["verificationStatus"] = "NEEDS_REVIEW"
        comparison = study["comparisons"][0]
        comparison["validityStatus"] = "NEEDS_REVIEW"
        comparison["confoundingStatus"] = "UNASSESSED"
        comparison["verificationStatus"] = "NEEDS_REVIEW"
        comparison["aggregationEligible"] = False
        comparison["effects"] = []
        conclusion = study["conclusions"][0]
        conclusion["causalStrength"] = "DESCRIPTIVE"
        conclusion["verificationStatus"] = "NEEDS_REVIEW"
        return manifest

    def test_prevalidated_import_skips_only_repeated_source_claim_checks(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        manifest = self.unverified_manifest(bridge["revisionUid"])
        with (
            mock.patch.object(
                study_import,
                "validate_study_manifest",
                wraps=study_import.validate_study_manifest,
            ) as schema_validator,
            mock.patch.object(
                study_import,
                "make_database_evidence_checker",
                side_effect=AssertionError(
                    "Database evidence was already validated by DRAFT."
                ),
            ),
            mock.patch.object(
                study_import,
                "validate_numeric_observation_evidence",
                side_effect=AssertionError("Numeric evidence was revalidated."),
            ),
            mock.patch.object(
                study_import,
                "validate_factor_and_arm_evidence",
                side_effect=AssertionError("Factor evidence was revalidated."),
            ),
            mock.patch.object(
                study_import,
                "validate_comparison_representation_alignment",
                side_effect=AssertionError("Comparison evidence was revalidated."),
            ),
            mock.patch.object(
                study_import,
                "validate_conclusion_evidence",
                side_effect=AssertionError("Conclusion evidence was revalidated."),
            ),
        ):
            result = study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
                source_claims_prevalidated=True,
            )

        self.assertTrue(result["analysisUid"])
        schema_validator.assert_called_once_with(manifest)
        self.assertEqual(
            1,
            connection.execute(
                "SELECT COUNT(*) FROM workbook_analyses"
            ).fetchone()[0],
        )
        connection.close()

    def test_analysis_integrity_is_scoped_to_the_imported_analysis(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        result = study_import.import_study_manifest(
            connection,
            self.unverified_manifest(bridge["revisionUid"]),
            now_iso=cli.now_iso,
        )

        integrity = schema.validate_analysis_integrity(
            connection,
            workbook_analysis_id=result["workbookAnalysisId"],
        )

        self.assertTrue(integrity["ok"])
        self.assertEqual("WORKBOOK_ANALYSIS", integrity["scope"])
        self.assertEqual(1, integrity["counts"]["workbook_analyses"])
        self.assertEqual(1, integrity["counts"]["knowledge_studies"])
        self.assertEqual(2, integrity["counts"]["knowledge_observations"])
        connection.close()

    def test_cli_quarantine_excludes_then_same_key_reimport_restores(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        manifest = self.unverified_manifest(bridge["revisionUid"])
        imported = study_import.import_study_manifest(
            connection,
            manifest,
            now_iso=cli.now_iso,
        )
        public_analysis_id = connection.execute(
            """
            SELECT public_analysis_id
            FROM workbook_analyses
            WHERE analysis_uid=?
            """,
            (imported["analysisUid"],),
        ).fetchone()[0]
        public_data_id = connection.execute(
            "SELECT public_data_id FROM knowledge_studies"
        ).fetchone()[0]
        evidence_count = connection.execute(
            "SELECT COUNT(*) FROM evidence_items"
        ).fetchone()[0]
        connection.commit()
        connection.close()

        args = cli.build_parser().parse_args(
            [
                "analysis-quarantine",
                "--db",
                str(self.db),
                "--public-analysis-id",
                str(public_analysis_id),
                "--reason",
                "Numeric-only AI conclusion requires corrected reimport.",
            ]
        )
        with mock.patch.object(cli, "print_json") as printed:
            self.assertEqual(0, args.func(args))
        payload = printed.call_args.args[0]
        self.assertEqual("ok", payload["status"])
        self.assertEqual(
            public_analysis_id,
            payload["quarantine"]["publicAnalysisId"],
        )
        self.assertTrue(payload["quarantine"]["preservedEvidence"])

        connection = sqlite3.connect(self.db)
        connection.row_factory = sqlite3.Row
        try:
            hidden = query.build_evidence_pack(
                connection,
                str(public_data_id),
            )
            status = connection.execute(
                """
                SELECT analysis_status, verification_status
                FROM workbook_analyses
                WHERE public_analysis_id=?
                """,
                (public_analysis_id,),
            ).fetchone()
            child_statuses = {
                row[0]
                for row in connection.execute(
                    """
                    SELECT verification_status FROM knowledge_studies
                    UNION
                    SELECT verification_status FROM knowledge_observations
                    UNION
                    SELECT verification_status FROM knowledge_comparisons
                    UNION
                    SELECT verification_status FROM knowledge_claims
                    """
                )
            }
            self.assertEqual(("STALE", "STALE"), tuple(status))
            self.assertEqual({"STALE"}, child_statuses)
            self.assertEqual(
                evidence_count,
                connection.execute(
                    "SELECT COUNT(*) FROM evidence_items"
                ).fetchone()[0],
            )
            self.assertEqual(
                0,
                hidden["summary"]["relevantStudyCount"],
            )

            manifest["studies"][0]["title"] = (
                "Corrected cooling duration versus custom NG"
            )
            study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            connection.commit()
            restored = query.build_evidence_pack(
                connection,
                str(public_data_id),
            )
            restored_row = connection.execute(
                """
                SELECT wa.verification_status, s.verification_status, s.title
                FROM workbook_analyses wa
                JOIN knowledge_studies s
                  ON s.workbook_analysis_id=wa.workbook_analysis_id
                WHERE wa.public_analysis_id=?
                """,
                (public_analysis_id,),
            ).fetchone()
            quarantine_issue_status = connection.execute(
                """
                SELECT status
                FROM validation_issues
                WHERE entity_type='WORKBOOK_ANALYSIS'
                  AND entity_uid=?
                  AND validator_name='canonical-analysis-quarantine'
                """,
                (imported["analysisUid"],),
            ).fetchone()[0]
        finally:
            connection.close()

        self.assertEqual(1, restored["summary"]["relevantStudyCount"])
        self.assertEqual(
            (
                "NEEDS_REVIEW",
                "NEEDS_REVIEW",
                "Corrected cooling duration versus custom NG",
            ),
            tuple(restored_row),
        )
        self.assertEqual("RESOLVED", quarantine_issue_status)

    def test_quarantine_refuses_verified_analysis_or_child(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        manifest = self.manifest(bridge["revisionUid"])
        imported = study_import.import_study_manifest(
            connection,
            manifest,
            now_iso=cli.now_iso,
        )
        public_analysis_id = connection.execute(
            "SELECT public_analysis_id FROM workbook_analyses"
        ).fetchone()[0]
        with self.assertRaisesRegex(
            study_import.AnalysisQuarantineError,
            "VERIFIED analysis",
        ):
            study_import.quarantine_canonical_analysis(
                connection,
                public_analysis_id=public_analysis_id,
                reason="Must not quarantine approved data.",
                now_iso=cli.now_iso,
            )

        connection.execute(
            """
            UPDATE workbook_analyses
            SET verification_status='NEEDS_REVIEW'
            WHERE analysis_uid=?
            """,
            (imported["analysisUid"],),
        )
        with self.assertRaisesRegex(
            study_import.AnalysisQuarantineError,
            "VERIFIED child",
        ):
            study_import.quarantine_canonical_analysis(
                connection,
                public_analysis_id=public_analysis_id,
                reason="Still protected by a verified child.",
                now_iso=cli.now_iso,
            )
        status = connection.execute(
            """
            SELECT analysis_status, verification_status
            FROM workbook_analyses
            """
        ).fetchone()
        connection.close()
        self.assertEqual(
            ("VERIFIED", "NEEDS_REVIEW"),
            tuple(status),
        )

    def test_quarantine_requires_reason_and_refuses_review_decision(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        manifest = self.unverified_manifest(bridge["revisionUid"])
        imported = study_import.import_study_manifest(
            connection,
            manifest,
            now_iso=cli.now_iso,
        )
        public_analysis_id = connection.execute(
            "SELECT public_analysis_id FROM workbook_analyses"
        ).fetchone()[0]
        with self.assertRaisesRegex(
            study_import.AnalysisQuarantineError,
            "reason must not be empty",
        ):
            study_import.quarantine_canonical_analysis(
                connection,
                public_analysis_id=public_analysis_id,
                reason=" ",
                now_iso=cli.now_iso,
            )
        study_uid = connection.execute(
            "SELECT study_uid FROM knowledge_studies"
        ).fetchone()[0]
        connection.execute(
            """
            INSERT INTO review_decisions(
                decision_uid, entity_type, entity_uid, decision, reason,
                reviewer, decided_at
            ) VALUES (?, 'STUDY', ?, 'RETURN_TO_REVIEW', ?, 'fixture', ?)
            """,
            (
                "decision-quarantine-protection",
                study_uid,
                "Human review exists.",
                cli.now_iso(),
            ),
        )
        with self.assertRaisesRegex(
            study_import.AnalysisQuarantineError,
            "review decision",
        ):
            study_import.quarantine_canonical_analysis(
                connection,
                public_analysis_id=public_analysis_id,
                reason="Must not override human review.",
                now_iso=cli.now_iso,
            )
        status = connection.execute(
            """
            SELECT analysis_status, verification_status
            FROM workbook_analyses
            WHERE analysis_uid=?
            """,
            (imported["analysisUid"],),
        ).fetchone()
        connection.close()
        self.assertEqual(
            ("NEEDS_REVIEW", "NEEDS_REVIEW"),
            tuple(status),
        )

    def test_imports_unknown_domains_with_stable_ids_and_schema_candidates(self) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            first = study_import.import_study_manifest(connection, manifest, now_iso=cli.now_iso)
            first_data_id = connection.execute("SELECT public_data_id FROM knowledge_studies").fetchone()[0]
            first_effect_id = connection.execute("SELECT public_effect_id FROM knowledge_effects").fetchone()[0]
            second = study_import.import_study_manifest(connection, manifest, now_iso=cli.now_iso)
            integrity = cli.validate_knowledge_integrity(connection)
            counts = cli.knowledge_counts(connection)
            candidate_kinds = {
                row[0]
                for row in connection.execute(
                    "SELECT candidate_kind FROM knowledge_schema_candidates"
                )
            }
        finally:
            connection.close()

        self.assertEqual(first["analysisUid"], second["analysisUid"])
        self.assertRegex(first_data_id, r"^DATA-[0-9A-F]{12}$")
        self.assertRegex(first_effect_id, r"^EFF-[0-9A-F]{12}$")
        self.assertEqual(1, counts["knowledge_studies"])
        self.assertEqual(1, counts["knowledge_factors"])
        self.assertEqual(1, counts["knowledge_effects"])
        self.assertGreater(counts["evidence_items"], 0)
        self.assertIn("CONCEPT:CHANGED_FACTOR", candidate_kinds)
        self.assertIn("CONCEPT:OUTCOME", candidate_kinds)
        self.assertTrue(integrity["ok"], integrity)

    def test_true_source_decision_imports_source_conclusion(self) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            claim = connection.execute(
                """
                SELECT claim_type, claim_text, verification_status
                FROM knowledge_claims
                """
            ).fetchone()
            source_texts = [
                row[0]
                for row in connection.execute(
                    """
                    SELECT e.source_text
                    FROM evidence_items e
                    JOIN entity_evidence_links l
                      ON l.evidence_id=e.evidence_id
                    WHERE l.entity_type='CLAIM'
                    """
                )
            ]
        finally:
            connection.close()

        self.assertEqual(
            ("SOURCE_CONCLUSION", "Changed is higher", "VERIFIED"),
            tuple(claim),
        )
        self.assertIn("Changed is higher", source_texts)

    def test_ordered_multi_cell_source_conclusion_is_supported(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        first_line = (
            "- Lot test frame new mold check SPK OK -> Continue move modul test"
        )
        second_line = "=> Can use"
        try:
            manifest = self.manifest(bridge["revisionUid"])
            conclusion = manifest["studies"][0]["conclusions"][0]
            conclusion["text"] = f"{first_line} ; {second_line}"
            conclusion["evidence"] = [
                {
                    "sheet": "Data",
                    "range": "B5:B6",
                    "role": "CONCLUSION",
                    "sourceText": f"{first_line} ; {second_line}",
                }
            ]
            study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            stored = connection.execute(
                """
                SELECT claim_type, claim_text
                FROM knowledge_claims
                """
            ).fetchone()

            conclusion["text"] = f"{second_line} ; {first_line}"
            conclusion["evidence"][0]["sourceText"] = (
                f"{second_line} ; {first_line}"
            )
            with self.assertRaisesRegex(
                ValueError,
                "SOURCE_CONCLUSION requires directly cited",
            ):
                study_import.import_study_manifest(
                    connection,
                    manifest,
                    now_iso=cli.now_iso,
                )
        finally:
            connection.close()

        self.assertEqual(
            (
                "SOURCE_CONCLUSION",
                f"{first_line} ; {second_line}",
            ),
            tuple(stored),
        )

    def test_factor_whole_cell_quantity_requires_number_and_unit(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            study = manifest["studies"][0]
            factor = study["factors"][0]
            arm = study["arms"][0]
            factor["evidence"] = [
                {
                    "sheet": "FactorData",
                    "range": "B2",
                    "role": "FACTOR",
                }
            ]
            arm["evidence"] = [
                {
                    "sheet": "FactorData",
                    "range": "B2",
                    "role": "ARM",
                }
            ]
            factor_value = arm["factorValues"][0]
            factor_value.update(
                {
                    "value": "1.56mg",
                    "valueNumber": None,
                    "unit": "",
                }
            )
            with self.assertRaisesRegex(
                ValueError,
                r"valueNumber must equal 1\.56",
            ):
                study_import.import_study_manifest(
                    connection,
                    manifest,
                    now_iso=cli.now_iso,
                )

            factor_value["valueNumber"] = 1.56
            factor_value["unit"] = "mg"
            study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            stored = connection.execute(
                """
                SELECT afv.original_value, afv.value_number,
                       u.canonical_symbol
                FROM knowledge_arm_factor_values afv
                JOIN knowledge_units u ON u.unit_id=afv.unit_id
                JOIN knowledge_arms a ON a.arm_id=afv.arm_id
                WHERE a.arm_key='changed'
                """
            ).fetchone()
        finally:
            connection.close()

        self.assertEqual(("1.56mg", 1.56, "mg"), tuple(stored))

    def test_factor_quantity_gate_ignores_unsupported_text_forms(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            revision = study_import.resolve_manifest_revision(
                connection,
                {
                    "dataset": "Fixture",
                    "sourcePath": str(self.source.resolve()),
                    "revisionUid": bridge["revisionUid"],
                },
            )
            cases = (
                ("C2", "0.4~0.5mg"),
                ("C3", "about 1.56mg"),
                ("B4", "1.56widget"),
                ("C4", "2026-07-18"),
                ("B5", "1/8 mg"),
                ("C5", "Model 123"),
            )
            for cell_range, value_text in cases:
                with self.subTest(value=value_text):
                    manifest = self.manifest(bridge["revisionUid"])
                    study = manifest["studies"][0]
                    factor = study["factors"][0]
                    arm = study["arms"][0]
                    evidence = [
                        {
                            "sheet": "FactorData",
                            "range": cell_range,
                            "role": "SOURCE",
                        }
                    ]
                    factor["evidence"] = evidence
                    arm["evidence"] = evidence
                    arm["factorValues"][0].update(
                        {
                            "value": value_text,
                            "valueNumber": None,
                            "unit": "",
                        }
                    )
                    study_import.validate_factor_and_arm_evidence(
                        connection,
                        revision,
                        manifest,
                    )

            manifest = self.manifest(bridge["revisionUid"])
            study = manifest["studies"][0]
            unsupported = [
                {
                    "sheet": "FactorData",
                    "range": "A1",
                    "role": "SOURCE",
                    "sourceText": "1.56mg",
                }
            ]
            study["factors"][0]["evidence"] = unsupported
            study["arms"][0]["evidence"] = unsupported
            study["arms"][0]["factorValues"][0].update(
                {
                    "value": "1.56mg",
                    "valueNumber": 1.56,
                    "unit": "mg",
                }
            )
            with self.assertRaisesRegex(
                ValueError,
                "actual|whole-cell|exact",
            ):
                study_import.validate_factor_and_arm_evidence(
                    connection,
                    revision,
                    manifest,
                )
        finally:
            connection.close()

    def test_unknown_unit_quantity_tokens_require_exact_whole_cells(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            revision = study_import.resolve_manifest_revision(
                connection,
                {
                    "dataset": "Fixture",
                    "sourcePath": str(self.source.resolve()),
                    "revisionUid": bridge["revisionUid"],
                },
            )
            for row, token in (
                (2, "5.6kg"),
                (3, "8V"),
                (4, "2 day"),
                (5, "5min"),
            ):
                with self.subTest(token=token):
                    self.assertIsNone(
                        study_import.resolve_unit_id(connection, token.lstrip(
                            "0123456789. "
                        ))
                    )
                    manifest = self.manifest(bridge["revisionUid"])
                    study = manifest["studies"][0]
                    evidence = [
                        {
                            "sheet": "UnknownQuantity",
                            "range": f"A{row}",
                            "role": "SOURCE",
                        }
                    ]
                    study["factors"][0]["evidence"] = evidence
                    arm = study["arms"][0]
                    arm["evidence"] = evidence
                    arm["factorValues"][0].update(
                        {
                            "value": token,
                            "valueNumber": None,
                            "unit": "",
                        }
                    )
                    with self.assertRaisesRegex(
                        ValueError,
                        "exact whole-cell quantity",
                    ):
                        study_import.validate_factor_and_arm_evidence(
                            connection,
                            revision,
                            manifest,
                        )
        finally:
            connection.close()

    def test_compact_english_ordinals_are_not_quantity_tokens(self) -> None:
        for token in ("1st", "2nd", "3rd", "4th", "11th", "22ND"):
            with self.subTest(token=token):
                self.assertIsNone(
                    study_import._whole_cell_quantity_syntax(token)
                )
        for token in ("5.6kg", "8V", "2 day", "5min"):
            with self.subTest(token=token):
                self.assertIsNotNone(
                    study_import._whole_cell_quantity_syntax(token)
                )

    def test_ordinal_and_total_tokens_require_exact_whole_cell_evidence(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            revision = study_import.resolve_manifest_revision(
                connection,
                {
                    "dataset": "Fixture",
                    "sourcePath": str(self.source.resolve()),
                    "revisionUid": bridge["revisionUid"],
                },
            )
            for row, token in ((6, "1st"), (7, "Total")):
                with self.subTest(token=token):
                    manifest = self.manifest(bridge["revisionUid"])
                    study = manifest["studies"][0]
                    composite_evidence = [
                        {
                            "sheet": "UnknownQuantity",
                            "range": f"A{row}",
                            "role": "SOURCE",
                        }
                    ]
                    study["factors"][0]["evidence"] = composite_evidence
                    arm = study["arms"][0]
                    arm["evidence"] = composite_evidence
                    arm["factorValues"][0].update(
                        {
                            "value": token,
                            "valueNumber": None,
                            "unit": "",
                        }
                    )
                    with self.assertRaisesRegex(
                        ValueError,
                        "must not isolate",
                    ):
                        study_import.validate_factor_and_arm_evidence(
                            connection,
                            revision,
                            manifest,
                        )

                    exact_evidence = [
                        {
                            "sheet": "UnknownQuantity",
                            "range": f"B{row}",
                            "role": "SOURCE",
                        }
                    ]
                    study["factors"][0]["evidence"] = exact_evidence
                    arm["evidence"] = exact_evidence
                    study_import.validate_factor_and_arm_evidence(
                        connection,
                        revision,
                        manifest,
                    )
        finally:
            connection.close()

    def test_exact_unknown_unit_whole_cell_can_remain_unnormalized(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            revision = study_import.resolve_manifest_revision(
                connection,
                {
                    "dataset": "Fixture",
                    "sourcePath": str(self.source.resolve()),
                    "revisionUid": bridge["revisionUid"],
                },
            )
            for row, token in (
                (2, "5.6kg"),
                (3, "8V"),
                (4, "2 day"),
                (5, "5min"),
            ):
                with self.subTest(token=token):
                    manifest = self.manifest(bridge["revisionUid"])
                    study = manifest["studies"][0]
                    evidence = [
                        {
                            "sheet": "UnknownQuantity",
                            "range": f"B{row}",
                            "role": "SOURCE",
                        }
                    ]
                    study["factors"][0]["evidence"] = evidence
                    arm = study["arms"][0]
                    arm["evidence"] = evidence
                    arm["factorValues"][0].update(
                        {
                            "value": token,
                            "valueNumber": None,
                            "unit": "",
                        }
                    )
                    study_import.validate_factor_and_arm_evidence(
                        connection,
                        revision,
                        manifest,
                    )
        finally:
            connection.close()

    def test_normal_is_reference_and_can_be_comparison_control_arm(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            normal_arm = manifest["studies"][0]["arms"][1]
            normal_arm.update(
                {
                    "label": "Normal (Line)",
                    "condition": "Normal (Line)",
                    "role": "CONTROL",
                    "evidence": [
                        {
                            "sheet": "FactorData",
                            "range": "A2",
                            "role": "ARM",
                        }
                    ],
                }
            )
            with self.assertRaisesRegex(
                ValueError,
                "maps to REFERENCE",
            ):
                study_import.import_study_manifest(
                    connection,
                    manifest,
                    now_iso=cli.now_iso,
                )

            normal_arm["role"] = "REFERENCE"
            study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            stored = connection.execute(
                """
                SELECT reference_arm.arm_role, reference_arm.label,
                       c.control_arm_id=reference_arm.arm_id
                FROM knowledge_comparisons c
                JOIN knowledge_arms reference_arm
                  ON reference_arm.arm_id=c.control_arm_id
                """
            ).fetchone()

            manifest = self.manifest(bridge["revisionUid"])
            changed_arm = manifest["studies"][0]["arms"][0]
            changed_arm["role"] = "CONTROL"
            changed_arm["evidence"] = [
                {
                    "sheet": "Data",
                    "range": "A2",
                    "role": "ARM",
                }
            ]
            with self.assertRaisesRegex(
                ValueError,
                "explicit Control wording",
            ):
                study_import.import_study_manifest(
                    connection,
                    manifest,
                    now_iso=cli.now_iso,
                )
        finally:
            connection.close()

        self.assertEqual(
            ("REFERENCE", "Normal (Line)", 1),
            tuple(stored),
        )

    def test_bare_st_cannot_be_promoted_to_reference(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            revision = study_import.resolve_manifest_revision(
                connection,
                {
                    "dataset": "Fixture",
                    "sourcePath": str(self.source.resolve()),
                    "revisionUid": bridge["revisionUid"],
                },
            )
            manifest = self.manifest(bridge["revisionUid"])
            arm = manifest["studies"][0]["arms"][1]
            arm.update(
                {
                    "label": "ST",
                    "condition": "ST",
                    "role": "REFERENCE",
                    "evidence": [
                        {
                            "sheet": "RoleData",
                            "range": "A2",
                            "role": "ARM",
                        }
                    ],
                }
            )
            with self.assertRaisesRegex(
                ValueError,
                "bare abbreviation such as ST",
            ):
                study_import.validate_factor_and_arm_evidence(
                    connection,
                    revision,
                    manifest,
                )

            arm["role"] = "OTHER"
            study_import.validate_factor_and_arm_evidence(
                connection,
                revision,
                manifest,
            )

            arm.update(
                {
                    "label": "Standard condition",
                    "condition": "Standard condition",
                    "role": "REFERENCE",
                    "evidence": [
                        {
                            "sheet": "RoleData",
                            "range": "A3",
                            "role": "ARM",
                        }
                    ],
                }
            )
            study_import.validate_factor_and_arm_evidence(
                connection,
                revision,
                manifest,
            )
        finally:
            connection.close()

    def test_grouped_reference_requires_ordered_pure_replicate_cells(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            revision = study_import.resolve_manifest_revision(
                connection,
                {
                    "dataset": "Fixture",
                    "sourcePath": str(self.source.resolve()),
                    "revisionUid": bridge["revisionUid"],
                },
            )

            def grouped_manifest(
                evidence_range: str,
                *,
                label: str = "Normal #1 through Normal #10",
            ) -> dict:
                manifest = self.manifest(bridge["revisionUid"])
                arm = manifest["studies"][0]["arms"][1]
                arm.update(
                    {
                        "label": label,
                        "condition": label,
                        "role": "REFERENCE",
                        "evidence": [
                            {
                                "sheet": "GroupedReference",
                                "range": evidence_range,
                                "role": "ARM",
                            }
                        ],
                    }
                )
                return manifest

            study_import.validate_factor_and_arm_evidence(
                connection,
                revision,
                grouped_manifest("A2:A11"),
            )

            with self.assertRaisesRegex(
                ValueError,
                "mixed Test/Normal evidence",
            ):
                study_import.validate_factor_and_arm_evidence(
                    connection,
                    revision,
                    grouped_manifest("B2:B3"),
                )

            with self.assertRaisesRegex(
                ValueError,
                "ordered distinct",
            ):
                study_import.validate_factor_and_arm_evidence(
                    connection,
                    revision,
                    grouped_manifest("C2:C3"),
                )

            with self.assertRaisesRegex(
                ValueError,
                "Normal label must exactly match",
            ):
                study_import.validate_factor_and_arm_evidence(
                    connection,
                    revision,
                    grouped_manifest("A2:A11", label="Normal"),
                )
        finally:
            connection.close()

    def test_repeated_passed_cells_are_categorical_not_conclusions(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            study = manifest["studies"][0]
            study["comparisons"] = []
            conclusion = study["conclusions"][0]
            conclusion.update(
                {
                    "text": "PASSED",
                    "claimType": "SOURCE_CONCLUSION",
                    "causalStrength": "DESCRIPTIVE",
                    "evidence": [
                        {
                            "sheet": "StatusData",
                            "range": "B2",
                            "role": "CONCLUSION",
                            "sourceText": "PASSED",
                        }
                    ],
                }
            )
            with self.assertRaisesRegex(
                ValueError,
                "SOURCE_CONCLUSION requires directly cited",
            ):
                study_import.import_study_manifest(
                    connection,
                    manifest,
                    now_iso=cli.now_iso,
                )

            study["conclusions"] = []
            study["outcomes"] = [
                {
                    "key": "replicate-status",
                    "originalLabel": "Status",
                    "metricType": "categorical",
                    "unit": "",
                    "favorableDirection": "UNKNOWN",
                    "evidence": [
                        {
                            "sheet": "StatusData",
                            "range": "A1:B3",
                            "role": "OUTCOME",
                        }
                    ],
                    "observations": [
                        {
                            "key": "status-r1",
                            "arm": "changed",
                            "valueNumber": None,
                            "valueText": "PASSED",
                            "replicateKey": "R1",
                            "evidence": [
                                {
                                    "sheet": "StatusData",
                                    "range": "A2:B2",
                                    "role": "OBSERVATION",
                                }
                            ],
                        },
                        {
                            "key": "status-r2",
                            "arm": "changed",
                            "valueNumber": None,
                            "valueText": "PASSED",
                            "replicateKey": "R2",
                            "evidence": [
                                {
                                    "sheet": "StatusData",
                                    "range": "A3:B3",
                                    "role": "OBSERVATION",
                                }
                            ],
                        },
                    ],
                }
            ]
            study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            stored = [
                tuple(row)
                for row in connection.execute(
                    """
                    SELECT replicate_key, value_text
                    FROM knowledge_observations
                    ORDER BY replicate_key
                    """
                )
            ]
        finally:
            connection.close()

        self.assertEqual(
            [("R1", "PASSED"), ("R2", "PASSED")],
            stored,
        )

    def test_short_explicit_test_more_action_is_source_conclusion(
        self,
    ) -> None:
        for source_text in (
            "Test more",
            "Can use.",
            "- Follow standard:",
        ):
            with self.subTest(source_text=source_text):
                self.assertTrue(
                    study_import._source_conclusion_is_supported(
                        {"text": source_text},
                        [source_text],
                        source_text=source_text,
                    )
                )
        self.assertFalse(
            study_import._source_conclusion_is_supported(
                {"text": "PASSED"},
                ["PASSED"],
                source_text="PASSED",
            )
        )

    def test_numeric_only_synthesis_is_not_source_conclusion(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            conclusion = manifest["studies"][0]["conclusions"][0]
            conclusion["text"] = (
                "Changed 8 s has a higher NG count than Control 4 s"
            )
            conclusion["causalStrength"] = "DESCRIPTIVE"
            conclusion["evidence"] = [
                {
                    "sheet": "Data",
                    "range": "A2:D3",
                    "role": "CONCLUSION",
                    "sourceText": conclusion["text"],
                }
            ]
            with self.assertRaisesRegex(
                ValueError,
                "SOURCE_CONCLUSION requires directly cited captured narrative",
            ):
                study_import.import_study_manifest(
                    connection,
                    manifest,
                    now_iso=cli.now_iso,
                )

            conclusion["claimType"] = "AI_DERIVED_DESCRIPTIVE"
            study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            claim = connection.execute(
                """
                SELECT claim_type, verification_status
                FROM knowledge_claims
                """
            ).fetchone()
            observation_count = connection.execute(
                "SELECT COUNT(*) FROM knowledge_observations"
            ).fetchone()[0]
        finally:
            connection.close()

        self.assertEqual(
            ("AI_DERIVED_DESCRIPTIVE", "NEEDS_REVIEW"),
            tuple(claim),
        )
        self.assertEqual(2, observation_count)

    def test_out_of_bounds_evidence_rolls_back_import(self) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            manifest["studies"][0]["factors"][0]["evidence"][0]["range"] = "Z99:Z100"
            with self.assertRaisesRegex(ValueError, "outside Capture v2 bounds"):
                study_import.import_study_manifest(connection, manifest, now_iso=cli.now_iso)
            count = connection.execute("SELECT COUNT(*) FROM knowledge_studies").fetchone()[0]
        finally:
            connection.close()
        self.assertEqual(0, count)

    def test_numeric_observation_must_match_cited_capture_cells(self) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            manifest["studies"][0]["outcomes"][0]["observations"][0]["numerator"] = 7
            manifest["studies"][0]["comparisons"][0]["effects"][0]["estimate"] = 60
            with self.assertRaisesRegex(ValueError, "not present in its cited Capture v2 cells"):
                study_import.import_study_manifest(connection, manifest, now_iso=cli.now_iso)
            count = connection.execute("SELECT COUNT(*) FROM knowledge_studies").fetchone()[0]
        finally:
            connection.close()
        self.assertEqual(0, count)

    def test_explicit_count_ratio_text_supports_count_evidence(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            observation = (
                manifest["studies"][0]["outcomes"][0]["observations"][0]
            )
            revision = study_import.resolve_manifest_revision(
                connection,
                manifest["source"],
            )
            for cell, numerator in (("A2", 1), ("B2", 2), ("C2", 3)):
                with self.subTest(cell=cell):
                    observation["numerator"] = numerator
                    observation["denominator"] = 8
                    observation["evidence"] = [
                        {
                            "sheet": "CountRatio",
                            "range": cell,
                            "role": "OBSERVATION",
                        }
                    ]
                    study_import.validate_numeric_observation_evidence(
                        connection,
                        revision,
                        manifest,
                    )
        finally:
            connection.close()

    def test_count_ratio_parser_rejects_narrative_range_and_bare_ratio(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            observation = (
                manifest["studies"][0]["outcomes"][0]["observations"][0]
            )
            revision = study_import.resolve_manifest_revision(
                connection,
                manifest["source"],
            )
            rejected = (
                ("D2", "numerator", 4),
                ("E2", "valueNumber", 0.4),
                ("F2", "numerator", 5),
            )
            for cell, field, value in rejected:
                with self.subTest(cell=cell):
                    observation["valueNumber"] = None
                    observation["numerator"] = None
                    observation["denominator"] = None
                    observation[field] = value
                    observation["evidence"] = [
                        {
                            "sheet": "CountRatio",
                            "range": cell,
                            "role": "OBSERVATION",
                        }
                    ]
                    with self.assertRaisesRegex(
                        ValueError,
                        "not present in its cited Capture v2 cells",
                    ):
                        study_import.validate_numeric_observation_evidence(
                            connection,
                            revision,
                            manifest,
                        )
        finally:
            connection.close()

    def test_percentage_display_scale_is_accepted_from_number_format(self) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            observation = manifest["studies"][0]["outcomes"][0]["observations"][0]
            observation["valueNumber"] = 20
            observation["numerator"] = None
            observation["denominator"] = None
            observation["evidence"] = [
                {"sheet": "Data", "range": "D2", "role": "OBSERVATION"}
            ]
            study_import.validate_numeric_observation_evidence(
                connection,
                study_import.resolve_manifest_revision(connection, manifest["source"]),
                manifest,
            )
        finally:
            connection.close()

    def test_compound_text_percentage_requires_matching_outcome_label(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            outcome = manifest["studies"][0]["outcomes"][0]
            outcome["originalLabel"] = "NG function percentage"
            outcome["unit"] = "%"
            observation = outcome["observations"][0]
            observation["valueNumber"] = 89.06
            observation["numerator"] = None
            observation["denominator"] = None
            observation["evidence"] = [
                {
                    "sheet": "LabeledPercent",
                    "range": "A2",
                    "role": "OBSERVATION",
                }
            ]
            revision = study_import.resolve_manifest_revision(
                connection,
                manifest["source"],
            )
            study_import.validate_numeric_observation_evidence(
                connection,
                revision,
                manifest,
            )

            observation["valueNumber"] = 71.88
            with self.assertRaisesRegex(
                ValueError,
                "not present in its cited Capture v2 cells",
            ):
                study_import.validate_numeric_observation_evidence(
                    connection,
                    revision,
                    manifest,
                )

            outcome["originalLabel"] = "Gauss NG percentage"
            study_import.validate_numeric_observation_evidence(
                connection,
                revision,
                manifest,
            )
        finally:
            connection.close()

    def test_percent_formatted_scalar_observation_is_stored_in_display_scale(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            study = manifest["studies"][0]
            study["verificationStatus"] = "NEEDS_REVIEW"
            study["comparabilityStatus"] = "UNASSESSED"
            study["confoundingStatus"] = "UNASSESSED"
            study["comparisons"] = []
            study["conclusions"] = []
            observations = study["outcomes"][0]["observations"]
            observations[0] = {
                "key": "changed-value",
                "arm": "changed",
                "valueNumber": 20,
                "verificationStatus": "NEEDS_REVIEW",
                "evidence": [
                    {
                        "sheet": "Data",
                        "range": "D2",
                        "role": "OBSERVATION",
                    }
                ],
            }
            observations[1] = {
                "key": "control-value",
                "arm": "control",
                "valueNumber": 10,
                "verificationStatus": "NEEDS_REVIEW",
                "evidence": [
                    {
                        "sheet": "Data",
                        "range": "D3",
                        "role": "OBSERVATION",
                    }
                ],
            }

            study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            rows = list(
                connection.execute(
                    """
                    SELECT observation_key, value_number, details_json
                    FROM knowledge_observations
                    ORDER BY observation_key
                    """
                )
            )
        finally:
            connection.close()

        self.assertEqual(
            [("changed-value", 20.0), ("control-value", 10.0)],
            [(row["observation_key"], row["value_number"]) for row in rows],
        )
        self.assertEqual(
            {},
            json.loads(rows[0]["details_json"]),
        )
        self.assertEqual({}, json.loads(rows[1]["details_json"]))

    def test_percent_numeric_claim_requires_exact_human_scale(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            observation = (
                manifest["studies"][0]["outcomes"][0]["observations"][0]
            )
            observation["numerator"] = None
            observation["denominator"] = None
            observation["valueText"] = "22.1%"
            observation["evidence"] = [
                {
                    "sheet": "PercentExact",
                    "range": "A2",
                    "role": "OBSERVATION",
                }
            ]
            revision = study_import.resolve_manifest_revision(
                connection,
                manifest["source"],
            )

            for rejected_value in (0.220735789, 22.1):
                with self.subTest(rejected_value=rejected_value):
                    observation["valueNumber"] = rejected_value
                    with self.assertRaisesRegex(
                        ValueError,
                        "not present in its cited Capture v2 cells",
                    ):
                        study_import.validate_numeric_observation_evidence(
                            connection,
                            revision,
                            manifest,
                        )

            observation["valueNumber"] = 0.220735789 * 100.0
            study_import.validate_numeric_observation_evidence(
                connection,
                revision,
                manifest,
            )

            observation["valueNumber"] = 2
            observation["evidence"].append(
                {
                    "sheet": "Data",
                    "range": "C2",
                    "role": "OBSERVATION",
                }
            )
            with self.assertRaisesRegex(
                ValueError,
                "not present in its cited Capture v2 cells",
            ):
                study_import.validate_numeric_observation_evidence(
                    connection,
                    revision,
                    manifest,
                )

            observation["valueNumber"] = 0
            observation["evidence"] = [
                {
                    "sheet": "PercentExact",
                    "range": "A3",
                    "role": "OBSERVATION",
                }
            ]
            study_import.validate_numeric_observation_evidence(
                connection,
                revision,
                manifest,
            )
        finally:
            connection.close()

    @staticmethod
    def add_measurement_series(manifest: dict) -> None:
        study = manifest["studies"][0]
        study["comparisons"] = []
        study["measurementSeries"] = [
            {
                "key": "custom-ng-profile",
                "outcome": "custom-ng",
                "arm": "changed",
                "sheet": "WideData",
                "headerRange": "B1:D1",
                "valueRange": "B2:D3",
                "rowIdentityRange": "A2:A3",
                "axisSource": "ROW_IDENTITY",
                "axisLabel": "Position",
                "axisUnit": "mm",
                "valueUnit": "%",
                "stratumKey": "profile-a",
                "verificationStatus": "VERIFIED",
            }
        ]

    @staticmethod
    def add_aligned_comparison_series(manifest: dict) -> None:
        manifest["studies"][0]["measurementSeries"] = [
            {
                "key": "changed-raw",
                "seriesRole": "RAW",
                "outcome": "custom-ng",
                "arm": "changed",
                "sheet": "ComparisonSeries",
                "headerRange": "B1",
                "valueRange": "B2:B3",
                "rowIdentityRange": "A2:A3",
                "axisSource": "ROW_IDENTITY",
                "axisLabel": "Source axis",
                "axisUnit": "",
                "valueUnit": "mg",
                "stratumKey": "same-condition",
            },
            {
                "key": "control-raw",
                "seriesRole": "RAW",
                "outcome": "custom-ng",
                "arm": "control",
                "sheet": "ComparisonSeries",
                "headerRange": "C1",
                "valueRange": "C2:C3",
                "rowIdentityRange": "A2:A3",
                "axisSource": "ROW_IDENTITY",
                "axisLabel": "Source axis",
                "axisUnit": "",
                "valueUnit": "mg",
                "stratumKey": "same-condition",
            },
        ]

    def test_comparison_alignment_rejects_raw_versus_scalar(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            self.add_aligned_comparison_series(manifest)
            manifest["studies"][0]["measurementSeries"] = (
                manifest["studies"][0]["measurementSeries"][:1]
            )
            revision = study_import.resolve_manifest_revision(
                connection,
                manifest["source"],
            )
            with self.assertRaisesRegex(
                ValueError,
                "RAW measurementSeries versus scalar/summary",
            ):
                study_import.validate_comparison_representation_alignment(
                    connection,
                    revision,
                    manifest,
                )
        finally:
            connection.close()

    def test_comparison_alignment_accepts_independent_aligned_raw(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            self.add_aligned_comparison_series(manifest)
            revision = study_import.resolve_manifest_revision(
                connection,
                manifest["source"],
            )
            study_import.validate_comparison_representation_alignment(
                connection,
                revision,
                manifest,
            )
        finally:
            connection.close()

    def test_comparison_alignment_rejects_reordered_axis_or_unit(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            revision = study_import.resolve_manifest_revision(
                connection,
                self.manifest(bridge["revisionUid"])["source"],
            )
            for mismatch in ("axis", "unit"):
                with self.subTest(mismatch=mismatch):
                    manifest = self.manifest(bridge["revisionUid"])
                    self.add_aligned_comparison_series(manifest)
                    control_series = manifest["studies"][0][
                        "measurementSeries"
                    ][1]
                    if mismatch == "axis":
                        control_series.update(
                            {
                                "headerRange": "E1",
                                "valueRange": "E2:E3",
                                "rowIdentityRange": "D2:D3",
                            }
                        )
                    else:
                        control_series["valueUnit"] = "kPa"
                    with self.assertRaisesRegex(
                        ValueError,
                        "compatible value units and aligned ordered axis",
                    ):
                        study_import.validate_comparison_representation_alignment(
                            connection,
                            revision,
                            manifest,
                        )
        finally:
            connection.close()

    def test_comparison_alignment_accepts_shared_scalar_field(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            revision = study_import.resolve_manifest_revision(
                connection,
                manifest["source"],
            )
            study_import.validate_comparison_representation_alignment(
                connection,
                revision,
                manifest,
            )
        finally:
            connection.close()

    @staticmethod
    def add_standalone_average_series(manifest: dict) -> None:
        study = manifest["studies"][0]
        study["arms"].append(
            {
                "key": "combined-summary",
                "role": "OTHER",
                "label": "Combined summary",
                "evidence": [
                    {
                        "sheet": "AverageData",
                        "range": "A1:D3",
                        "role": "SOURCE",
                    }
                ],
                "factorValues": [],
            }
        )
        study["measurementSeries"] = [
            {
                "key": "raw-before-profile",
                "seriesRole": "RAW",
                "outcome": "custom-ng",
                "arm": "control",
                "sheet": "AverageData",
                "headerRange": "B1",
                "valueRange": "B2:B3",
                "rowIdentityRange": "A2:A3",
                "axisSource": "ROW_IDENTITY",
                "axisLabel": "Axis",
                "valueUnit": "%",
            },
            {
                "key": "raw-after-profile",
                "outcome": "custom-ng",
                "arm": "changed",
                "sheet": "AverageData",
                "headerRange": "C1",
                "valueRange": "C2:C3",
                "rowIdentityRange": "A2:A3",
                "axisSource": "ROW_IDENTITY",
                "axisLabel": "Axis",
                "valueUnit": "%",
            },
            {
                "key": "combined-average-profile",
                "seriesRole": "AGGREGATE",
                "aggregationFunction": "AVERAGE",
                "aggregateOfSeries": [
                    "raw-before-profile",
                    "raw-after-profile",
                ],
                "outcome": "custom-ng",
                "arm": "combined-summary",
                "sheet": "AverageData",
                "headerRange": "D1",
                "valueRange": "D2:D3",
                "rowIdentityRange": "A2:A3",
                "axisSource": "ROW_IDENTITY",
                "axisLabel": "Axis",
                "valueUnit": "%",
            },
        ]

    def test_measurement_series_expands_all_numeric_points_idempotently(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            self.add_measurement_series(manifest)
            first = study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            first_rows = list(
                connection.execute(
                    """
                    SELECT
                        ms.public_series_id, mp.public_point_id,
                        mp.row_ordinal, mp.column_ordinal, mp.axis_label,
                        mp.axis_value, mp.replicate_key, mp.stratum_key,
                        mp.value_number, mp.original_axis_unit,
                        mp.original_value_unit, mp.source_coordinate,
                        mp.axis_source_coordinate,
                        mp.replicate_source_coordinate
                    FROM knowledge_measurement_points mp
                    JOIN knowledge_measurement_series ms
                      ON ms.series_id=mp.series_id
                    ORDER BY mp.row_ordinal, mp.column_ordinal
                    """
                )
            )
            second = study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            second_ids = [
                row[0]
                for row in connection.execute(
                    """
                    SELECT public_point_id
                    FROM knowledge_measurement_points
                    ORDER BY row_ordinal, column_ordinal
                    """
                )
            ]
            series_evidence = connection.execute(
                """
                SELECT COUNT(*)
                FROM entity_evidence_links
                WHERE entity_type='MEASUREMENT_SERIES'
                """
            ).fetchone()[0]
        finally:
            connection.close()

        self.assertEqual(1, first["measurementSeries"])
        self.assertEqual(6, first["measurementPoints"])
        self.assertEqual(1, second["measurementSeries"])
        self.assertEqual(6, second["measurementPoints"])
        self.assertEqual(6, len(first_rows))
        self.assertRegex(first_rows[0]["public_series_id"], r"^SER-[0-9A-F]{12}$")
        self.assertRegex(first_rows[0]["public_point_id"], r"^MPT-[0-9A-F]{12}$")
        self.assertEqual(
            [row["public_point_id"] for row in first_rows],
            second_ids,
        )
        self.assertEqual(
            (1, 1, "0", 0.0, "R1", "profile-a", 10.1),
            (
                first_rows[0]["row_ordinal"],
                first_rows[0]["column_ordinal"],
                first_rows[0]["axis_label"],
                first_rows[0]["axis_value"],
                first_rows[0]["replicate_key"],
                first_rows[0]["stratum_key"],
                first_rows[0]["value_number"],
            ),
        )
        self.assertEqual("B2", first_rows[0]["source_coordinate"])
        self.assertEqual("A2", first_rows[0]["axis_source_coordinate"])
        self.assertEqual("B1", first_rows[0]["replicate_source_coordinate"])
        self.assertEqual("mm", first_rows[0]["original_axis_unit"])
        self.assertEqual("%", first_rows[0]["original_value_unit"])
        self.assertEqual(3, series_evidence)

    def test_numeric_preimport_gate_rejects_nonnumeric_series_cell(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            self.add_measurement_series(manifest)
            series = manifest["studies"][0]["measurementSeries"][0]
            series["headerRange"] = "B1:E1"
            series["valueRange"] = "B2:E3"
            revision = study_import.resolve_manifest_revision(
                connection,
                manifest["source"],
            )
            with self.assertRaisesRegex(
                ValueError,
                r"valueRange source cell E2 must be numeric",
            ):
                study_import.validate_numeric_observation_evidence(
                    connection,
                    revision,
                    manifest,
                )
        finally:
            connection.close()

    def test_measurement_series_header_axis_preserves_axis_and_replicate_roles(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            manifest["studies"][0]["comparisons"] = []
            manifest["studies"][0]["measurementSeries"] = [
                {
                    "key": "header-axis-profile",
                    "outcome": "custom-ng",
                    "arm": "changed",
                    "sheet": "HeaderAxis",
                    "headerRange": "B1:C1",
                    "valueRange": "B2:C3",
                    "rowIdentityRange": "A2:A3",
                    "axisSource": "HEADER",
                    "axisLabel": "Frequency",
                    "axisUnit": "Hz",
                    "valueUnit": "dB",
                }
            ]
            result = study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            series = connection.execute(
                """
                SELECT axis_source, verification_status
                FROM knowledge_measurement_series
                """
            ).fetchone()
            points = list(
                connection.execute(
                    """
                    SELECT
                        axis_label, axis_value, replicate_key,
                        axis_source_coordinate,
                        replicate_source_coordinate, source_coordinate
                    FROM knowledge_measurement_points
                    ORDER BY row_ordinal, column_ordinal
                    """
                )
            )
            series_review_issues = connection.execute(
                """
                SELECT COUNT(*)
                FROM validation_issues
                WHERE entity_type='MEASUREMENT_SERIES'
                  AND issue_code='NEEDS_REVIEW'
                  AND status='OPEN'
                """
            ).fetchone()[0]
        finally:
            connection.close()

        self.assertEqual(1, result["measurementSeries"])
        self.assertEqual(4, result["measurementPoints"])
        self.assertEqual(1, series_review_issues)
        self.assertEqual("HEADER", series["axis_source"])
        self.assertEqual("NEEDS_REVIEW", series["verification_status"])
        self.assertEqual(
            ("100.00Hz", 100.0, "Sample-1", "B1", "A2", "B2"),
            tuple(points[0]),
        )
        self.assertEqual(
            ("200.00Hz", 200.0, "Sample-2", "C1", "A3", "C3"),
            tuple(points[-1]),
        )

    def test_merged_covered_row_identity_uses_exact_anchor_provenance(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            study = manifest["studies"][0]
            study["comparisons"] = []
            study["measurementSeries"] = [
                {
                    "key": "merged-identity-point",
                    "seriesRole": "RAW",
                    "outcome": "custom-ng",
                    "arm": "changed",
                    "sheet": "MergedIdentity",
                    "headerRange": "B1",
                    "valueRange": "B3",
                    "rowIdentityRange": "A3",
                    "axisSource": "ROW_IDENTITY",
                    "axisLabel": "Batch",
                    "axisUnit": "",
                    "valueUnit": "",
                    "stratumKey": "",
                }
            ]
            study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            stored = connection.execute(
                """
                SELECT ms.row_identity_range, mp.axis_label,
                       mp.axis_source_coordinate, mp.source_coordinate
                FROM knowledge_measurement_series ms
                JOIN knowledge_measurement_points mp
                  ON mp.series_id=ms.series_id
                """
            ).fetchone()
        finally:
            connection.close()

        self.assertEqual(
            ("A3", "Merged batch", "A2", "B3"),
            tuple(stored),
        )

    def test_merged_covered_header_uses_exact_anchor_provenance(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            study = manifest["studies"][0]
            study["comparisons"] = []
            study["measurementSeries"] = [
                {
                    "key": "merged-header-profile",
                    "seriesRole": "RAW",
                    "outcome": "custom-ng",
                    "arm": "changed",
                    "sheet": "MergedHeader",
                    "headerRange": "C4",
                    "valueRange": "C5:C6",
                    "rowIdentityRange": "A5:A6",
                    "axisSource": "ROW_IDENTITY",
                    "axisLabel": "Sample",
                    "axisUnit": "",
                    "valueUnit": "",
                    "stratumKey": "",
                }
            ]
            study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            stored = connection.execute(
                """
                SELECT ms.header_range, mp.replicate_key,
                       mp.replicate_source_coordinate,
                       mp.source_coordinate
                FROM knowledge_measurement_series ms
                JOIN knowledge_measurement_points mp
                  ON mp.series_id=ms.series_id
                ORDER BY mp.row_ordinal
                LIMIT 1
                """
            ).fetchone()
        finally:
            connection.close()

        self.assertEqual(
            ("C4", "RESULT CHECKING FO", "B3", "C5"),
            tuple(stored),
        )

    def test_multi_column_headers_cannot_collapse_to_one_merge_anchor(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            manifest["studies"][0]["measurementSeries"] = [
                {
                    "key": "duplicate-merged-headers",
                    "seriesRole": "RAW",
                    "outcome": "custom-ng",
                    "arm": "changed",
                    "sheet": "MergedHeader",
                    "headerRange": "C4:D4",
                    "valueRange": "C5:D6",
                    "rowIdentityRange": "A5:A6",
                    "axisSource": "ROW_IDENTITY",
                    "axisLabel": "Sample",
                    "axisUnit": "",
                    "valueUnit": "",
                    "stratumKey": "",
                }
            ]
            with self.assertRaisesRegex(
                ValueError,
                "same merged anchor",
            ):
                study_import.validate_numeric_observation_evidence(
                    connection,
                    study_import.resolve_manifest_revision(
                        connection,
                        manifest["source"],
                    ),
                    manifest,
                )
        finally:
            connection.close()

    def test_unmerged_blank_header_remains_rejected(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            manifest["studies"][0]["measurementSeries"] = [
                {
                    "key": "blank-header-profile",
                    "seriesRole": "RAW",
                    "outcome": "custom-ng",
                    "arm": "changed",
                    "sheet": "MergedHeader",
                    "headerRange": "H4",
                    "valueRange": "H5:H6",
                    "rowIdentityRange": "A5:A6",
                    "axisSource": "ROW_IDENTITY",
                    "axisLabel": "Sample",
                    "axisUnit": "",
                    "valueUnit": "",
                    "stratumKey": "",
                }
            ]
            with self.assertRaisesRegex(
                ValueError,
                "missing captured source cell MergedHeader!H4|"
                "source cell H4 has no value",
            ):
                study_import.validate_numeric_observation_evidence(
                    connection,
                    study_import.resolve_manifest_revision(
                        connection,
                        manifest["source"],
                    ),
                    manifest,
                )
        finally:
            connection.close()

    def test_unmerged_blank_row_identity_remains_rejected(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            manifest["studies"][0]["measurementSeries"] = [
                {
                    "key": "blank-identity-point",
                    "seriesRole": "RAW",
                    "outcome": "custom-ng",
                    "arm": "changed",
                    "sheet": "MergedIdentity",
                    "headerRange": "B1",
                    "valueRange": "B4",
                    "rowIdentityRange": "A4",
                    "axisSource": "ROW_IDENTITY",
                    "axisLabel": "Batch",
                    "axisUnit": "",
                    "valueUnit": "",
                    "stratumKey": "",
                }
            ]
            with self.assertRaisesRegex(
                ValueError,
                "has no usable identity|has no value",
            ):
                study_import.validate_numeric_observation_evidence(
                    connection,
                    study_import.resolve_manifest_revision(
                        connection,
                        manifest["source"],
                    ),
                    manifest,
                )
        finally:
            connection.close()

    def test_merged_covered_numeric_value_remains_rejected(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            manifest["studies"][0]["measurementSeries"] = [
                {
                    "key": "covered-numeric-point",
                    "seriesRole": "RAW",
                    "outcome": "custom-ng",
                    "arm": "changed",
                    "sheet": "MergedIdentity",
                    "headerRange": "B1",
                    "valueRange": "B6",
                    "rowIdentityRange": "A6",
                    "axisSource": "ROW_IDENTITY",
                    "axisLabel": "Batch",
                    "axisUnit": "",
                    "valueUnit": "",
                    "stratumKey": "",
                }
            ]
            with self.assertRaisesRegex(
                ValueError,
                "source cell B6 has no value",
            ):
                study_import.validate_numeric_observation_evidence(
                    connection,
                    study_import.resolve_manifest_revision(
                        connection,
                        manifest["source"],
                    ),
                    manifest,
                )
        finally:
            connection.close()

    def test_measurement_series_preserves_aggregate_but_excludes_its_role(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            self.add_measurement_series(manifest)
            series = manifest["studies"][0]["measurementSeries"][0]
            series["aggregateReplicateRanges"] = ["D1"]
            study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            roles = [
                tuple(row)
                for row in connection.execute(
                    """
                    SELECT replicate_key, replicate_role, COUNT(*)
                    FROM knowledge_measurement_points
                    GROUP BY replicate_key, replicate_role
                    ORDER BY replicate_key
                    """
                )
            ]
            details = json.loads(
                connection.execute(
                    """
                    SELECT details_json
                    FROM knowledge_measurement_series
                    """
                ).fetchone()[0]
            )
        finally:
            connection.close()

        self.assertEqual(
            [
                ("R1", "RAW", 2),
                ("R2", "RAW", 2),
                ("R3", "AGGREGATE", 2),
            ],
            roles,
        )
        self.assertEqual(["D1"], details["aggregateReplicateRanges"])

    def test_average_aggregate_must_cover_all_raw_members(self) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            self.add_measurement_series(manifest)
            series = manifest["studies"][0]["measurementSeries"][0]
            series["sheet"] = "AverageData"
            series["headerRange"] = "B1:D1"
            series["valueRange"] = "B2:D3"
            series["rowIdentityRange"] = "A2:A3"
            series["aggregateReplicateRanges"] = ["D1"]
            study_import.validate_numeric_observation_evidence(
                connection,
                study_import.resolve_manifest_revision(
                    connection,
                    manifest["source"],
                ),
                manifest,
            )

            series["headerRange"] = "C1:D1"
            series["valueRange"] = "C2:D3"
            with self.assertRaisesRegex(
                ValueError,
                "does not equal the arithmetic mean",
            ):
                study_import.validate_numeric_observation_evidence(
                    connection,
                    study_import.resolve_manifest_revision(
                        connection,
                        manifest["source"],
                    ),
                    manifest,
                )
        finally:
            connection.close()

    def test_standalone_profile_average_references_two_raw_series(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            self.add_standalone_average_series(manifest)
            result = study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            roles = [
                tuple(row)
                for row in connection.execute(
                    """
                    SELECT ms.series_key, mp.replicate_role, COUNT(*)
                    FROM knowledge_measurement_points mp
                    JOIN knowledge_measurement_series ms
                      ON ms.series_id=mp.series_id
                    GROUP BY ms.series_key, mp.replicate_role
                    ORDER BY ms.series_key
                    """
                )
            ]
            aggregate_details = json.loads(
                connection.execute(
                    """
                    SELECT details_json
                    FROM knowledge_measurement_series
                    WHERE series_key='combined-average-profile'
                    """
                ).fetchone()[0]
            )
            aggregate_arm_role = connection.execute(
                """
                SELECT a.arm_role
                FROM knowledge_measurement_series ms
                JOIN knowledge_arms a ON a.arm_id=ms.arm_id
                WHERE ms.series_key='combined-average-profile'
                """
            ).fetchone()[0]
        finally:
            connection.close()

        self.assertEqual(3, result["measurementSeries"])
        self.assertEqual(6, result["measurementPoints"])
        self.assertEqual(
            [
                ("combined-average-profile", "AGGREGATE", 2),
                ("raw-after-profile", "RAW", 2),
                ("raw-before-profile", "RAW", 2),
            ],
            roles,
        )
        self.assertEqual("OTHER", aggregate_arm_role)
        self.assertEqual("AGGREGATE", aggregate_details["seriesRole"])
        self.assertEqual(
            "AVERAGE",
            aggregate_details["aggregationFunction"],
        )
        self.assertEqual(
            ["raw-before-profile", "raw-after-profile"],
            aggregate_details["aggregateOfSeries"],
        )
        self.assertEqual(
            [],
            aggregate_details["aggregateReplicateRanges"],
        )

    def test_standalone_single_axis_average_references_one_raw_series(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            self.add_standalone_average_series(manifest)
            study = manifest["studies"][0]
            raw = study["measurementSeries"][0]
            raw["headerRange"] = "B1:C1"
            raw["valueRange"] = "B2:C2"
            raw["rowIdentityRange"] = "A2"
            aggregate = study["measurementSeries"][2]
            aggregate["aggregateOfSeries"] = ["raw-before-profile"]
            aggregate["valueRange"] = "D2"
            aggregate["rowIdentityRange"] = "A2"
            study["measurementSeries"] = [raw, aggregate]
            study["comparisons"] = []

            result = study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            points = [
                tuple(row)
                for row in connection.execute(
                    """
                    SELECT ms.series_key, mp.replicate_role, mp.value_number
                    FROM knowledge_measurement_points mp
                    JOIN knowledge_measurement_series ms
                      ON ms.series_id=mp.series_id
                    ORDER BY ms.series_key, mp.column_ordinal
                    """
                )
            ]
        finally:
            connection.close()

        self.assertEqual(2, result["measurementSeries"])
        self.assertEqual(3, result["measurementPoints"])
        self.assertEqual(
            [
                ("combined-average-profile", "AGGREGATE", 3.0),
                ("raw-before-profile", "RAW", 2.0),
                ("raw-before-profile", "RAW", 4.0),
            ],
            points,
        )

    def test_standalone_average_rejects_mismatched_source_mean(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            self.add_standalone_average_series(manifest)
            aggregate = manifest["studies"][0]["measurementSeries"][2]
            aggregate["aggregateOfSeries"] = ["raw-before-profile"]
            with self.assertRaisesRegex(
                ValueError,
                "does not equal the arithmetic mean",
            ):
                study_import.import_study_manifest(
                    connection,
                    manifest,
                    now_iso=cli.now_iso,
                )
            counts = tuple(
                connection.execute(
                    """
                    SELECT
                        (SELECT COUNT(*)
                         FROM knowledge_measurement_series),
                        (SELECT COUNT(*)
                         FROM knowledge_measurement_points)
                    """
                ).fetchone()
            )
        finally:
            connection.close()

        self.assertEqual((0, 0), counts)

    def test_measurement_series_uses_source_formatted_header_identity(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            self.add_measurement_series(manifest)
            series = manifest["studies"][0]["measurementSeries"][0]
            series["sheet"] = "FormattedIdentity"
            series["headerRange"] = "B1:C1"
            series["valueRange"] = "B2:C2"
            series["rowIdentityRange"] = "A2:A2"
            study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            identities = [
                row[0]
                for row in connection.execute(
                    """
                    SELECT replicate_key
                    FROM knowledge_measurement_points
                    ORDER BY column_ordinal
                    """
                )
            ]
        finally:
            connection.close()

        self.assertEqual(
            ["18kPa #1_Before", "18kPa #1_After"],
            identities,
        )

    def test_measurement_series_rejects_non_numeric_or_missing_value_cell(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            self.add_measurement_series(manifest)
            series = manifest["studies"][0]["measurementSeries"][0]
            series["headerRange"] = "B1:E1"
            series["valueRange"] = "B2:E3"
            with self.assertRaisesRegex(ValueError, "must be numeric"):
                study_import.import_study_manifest(
                    connection,
                    manifest,
                    now_iso=cli.now_iso,
                )
            series["headerRange"] = "B1:D1"
            series["valueRange"] = "B2:D4"
            series["rowIdentityRange"] = "A2:A4"
            with self.assertRaisesRegex(
                ValueError,
                "missing captured source cell",
            ):
                study_import.import_study_manifest(
                    connection,
                    manifest,
                    now_iso=cli.now_iso,
                )
            counts = (
                connection.execute(
                    """
                    SELECT
                        (SELECT COUNT(*)
                         FROM knowledge_measurement_series),
                        (SELECT COUNT(*)
                         FROM knowledge_measurement_points)
                    """
                ).fetchone()
            )
        finally:
            connection.close()
        self.assertEqual((0, 0), tuple(counts))

    def test_measurement_series_reimport_removes_only_stale_points(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            self.add_measurement_series(manifest)
            initial_series = manifest["studies"][0]["measurementSeries"][0]
            initial_series["headerRange"] = "C1:D1"
            initial_series["valueRange"] = "C2:D3"
            study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            series_uid = connection.execute(
                "SELECT series_uid FROM knowledge_measurement_series"
            ).fetchone()[0]
            retained_point_id = connection.execute(
                """
                SELECT public_point_id
                FROM knowledge_measurement_points
                WHERE source_coordinate='C2'
                """
            ).fetchone()[0]
            reduced = copy.deepcopy(manifest)
            reduced_series = reduced["studies"][0]["measurementSeries"][0]
            reduced_series["headerRange"] = "B1:C1"
            reduced_series["valueRange"] = "B2:C3"
            study_import.import_study_manifest(
                connection,
                reduced,
                now_iso=cli.now_iso,
            )
            second_series_uid = connection.execute(
                "SELECT series_uid FROM knowledge_measurement_series"
            ).fetchone()[0]
            second_retained_point_id = connection.execute(
                """
                SELECT public_point_id
                FROM knowledge_measurement_points
                WHERE source_coordinate='C2'
                """
            ).fetchone()[0]
            coordinates = [
                row[0]
                for row in connection.execute(
                    """
                    SELECT source_coordinate
                    FROM knowledge_measurement_points
                    ORDER BY row_ordinal, column_ordinal
                    """
                )
            ]
            reduced["studies"][0]["measurementSeries"] = []
            study_import.import_study_manifest(
                connection,
                reduced,
                now_iso=cli.now_iso,
            )
            final_counts = tuple(
                connection.execute(
                    """
                    SELECT
                        (SELECT COUNT(*)
                         FROM knowledge_measurement_series),
                        (SELECT COUNT(*)
                         FROM knowledge_measurement_points)
                    """
                ).fetchone()
            )
        finally:
            connection.close()
        self.assertEqual(series_uid, second_series_uid)
        self.assertEqual(retained_point_id, second_retained_point_id)
        self.assertEqual(["B2", "C2", "B3", "C3"], coordinates)
        self.assertEqual((0, 0), final_counts)

    def test_measurement_series_reimport_allows_row_range_shift(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            manifest["studies"][0]["comparisons"] = []
            manifest["studies"][0]["measurementSeries"] = [
                {
                    "key": "shifted-row-profile",
                    "outcome": "custom-ng",
                    "arm": "changed",
                    "sheet": "HeaderAxis",
                    "headerRange": "B1:C1",
                    "valueRange": "B3:C4",
                    "rowIdentityRange": "A3:A4",
                    "axisSource": "HEADER",
                    "axisLabel": "Frequency",
                    "axisUnit": "Hz",
                    "valueUnit": "dB",
                    "verificationStatus": "VERIFIED",
                }
            ]
            study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            retained_point_id = connection.execute(
                """
                SELECT public_point_id
                FROM knowledge_measurement_points
                WHERE source_coordinate='C3'
                """
            ).fetchone()[0]
            shifted = copy.deepcopy(manifest)
            series = shifted["studies"][0]["measurementSeries"][0]
            series["valueRange"] = "B2:C3"
            series["rowIdentityRange"] = "A2:A3"
            study_import.import_study_manifest(
                connection,
                shifted,
                now_iso=cli.now_iso,
            )
            rows = list(
                connection.execute(
                    """
                    SELECT source_coordinate, public_point_id
                    FROM knowledge_measurement_points
                    ORDER BY row_ordinal, column_ordinal
                    """
                )
            )
        finally:
            connection.close()

        self.assertEqual(
            ["B2", "C2", "B3", "C3"],
            [row["source_coordinate"] for row in rows],
        )
        self.assertEqual(retained_point_id, rows[-1]["public_point_id"])

    def test_measurement_series_accepts_strict_numeric_text_and_percent_format(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            manifest = self.manifest(bridge["revisionUid"])
            manifest["studies"][0]["comparisons"] = []
            manifest["studies"][0]["measurementSeries"] = [
                {
                    "key": "percent-profile",
                    "outcome": "custom-ng",
                    "arm": "changed",
                    "sheet": "PercentData",
                    "headerRange": "B1",
                    "valueRange": "B2:B3",
                    "rowIdentityRange": "A2:A3",
                    "axisSource": "ROW_IDENTITY",
                    "axisLabel": "Specimen",
                    "axisUnit": "",
                    "valueUnit": "%",
                }
            ]
            study_import.import_study_manifest(
                connection,
                manifest,
                now_iso=cli.now_iso,
            )
            values = [
                row[0]
                for row in connection.execute(
                    """
                    SELECT value_number
                    FROM knowledge_measurement_points
                    ORDER BY row_ordinal
                    """
                )
            ]
        finally:
            connection.close()

        self.assertEqual([10.0, 12.5], values)

    def test_reimport_removes_only_stale_manifest_children_and_links(self) -> None:
        connection, bridge = self.prepare_source()
        try:
            full_manifest = self.manifest(bridge["revisionUid"])
            first = study_import.import_study_manifest(
                connection,
                full_manifest,
                now_iso=cli.now_iso,
            )
            workbook_analysis_id = first["workbookAnalysisId"]
            original_study_id = connection.execute(
                "SELECT study_id FROM knowledge_studies WHERE workbook_analysis_id=?",
                (workbook_analysis_id,),
            ).fetchone()[0]
            current_study_uid = connection.execute(
                "SELECT study_uid FROM knowledge_studies WHERE study_id=?",
                (original_study_id,),
            ).fetchone()[0]
            human_evidence_id = connection.execute(
                """
                INSERT INTO evidence_items(
                    evidence_uid, public_evidence_id, revision_id,
                    evidence_kind, sheet_name, start_row, start_col,
                    end_row, end_col, range_address, evidence_role,
                    source_text, note, verification_status, created_at
                ) VALUES (
                    'human-evidence-uid', 'EVD-HUMAN-1', ?, 'CELL_RANGE',
                    'Data', 1, 1, 1, 1, 'A1', 'HUMAN_NOTE',
                    'Human-added supporting note', '', 'NEEDS_REVIEW', ?
                )
                """,
                (bridge["revisionId"], cli.now_iso()),
            ).lastrowid
            connection.execute(
                """
                INSERT INTO entity_evidence_links(
                    entity_type, entity_uid, evidence_id,
                    evidence_role, claim_scope
                ) VALUES ('STUDY', ?, ?, 'HUMAN_NOTE', 'human review')
                """,
                (current_study_uid, human_evidence_id),
            )
            original_outcome_id = connection.execute(
                """
                SELECT o.outcome_id
                FROM knowledge_outcomes o
                JOIN knowledge_studies s ON s.study_id=o.study_id
                WHERE s.workbook_analysis_id=?
                """,
                (workbook_analysis_id,),
            ).fetchone()[0]
            stale_entities = list(
                connection.execute(
                    """
                    SELECT 'CONTEXT', c.context_uid
                    FROM knowledge_study_contexts c
                    JOIN knowledge_studies s ON s.study_id=c.study_id
                    WHERE s.workbook_analysis_id=?
                    UNION ALL
                    SELECT 'FACTOR', f.factor_uid
                    FROM knowledge_factors f
                    JOIN knowledge_studies s ON s.study_id=f.study_id
                    WHERE s.workbook_analysis_id=?
                    UNION ALL
                    SELECT 'COMPARISON', c.comparison_uid
                    FROM knowledge_comparisons c
                    JOIN knowledge_studies s ON s.study_id=c.study_id
                    WHERE s.workbook_analysis_id=?
                    UNION ALL
                    SELECT 'EFFECT', e.effect_uid
                    FROM knowledge_effects e
                    JOIN knowledge_comparisons c ON c.comparison_id=e.comparison_id
                    JOIN knowledge_studies s ON s.study_id=c.study_id
                    WHERE s.workbook_analysis_id=?
                    UNION ALL
                    SELECT 'CLAIM', c.claim_uid
                    FROM knowledge_claims c
                    WHERE c.workbook_analysis_id=?
                    """,
                    (workbook_analysis_id,) * 5,
                )
            )
            context_evidence_id = connection.execute(
                """
                SELECT evidence_id
                FROM entity_evidence_links
                WHERE entity_type='CONTEXT'
                LIMIT 1
                """
            ).fetchone()[0]
            connection.execute(
                """
                INSERT INTO entity_evidence_links(
                    entity_type, entity_uid, evidence_id, evidence_role, claim_scope
                ) VALUES ('EXTERNAL_NOTE', 'external-note-1', ?, 'SOURCE', '')
                """,
                (context_evidence_id,),
            )

            reduced = copy.deepcopy(full_manifest)
            study = reduced["studies"][0]
            study["contexts"] = []
            study["factors"] = []
            study["arms"] = [study["arms"][0]]
            study["arms"][0]["factorValues"] = []
            study["outcomes"][0]["observations"] = [
                study["outcomes"][0]["observations"][0]
            ]
            study["comparisons"] = []
            study["conclusions"] = []
            second = study_import.import_study_manifest(
                connection,
                reduced,
                now_iso=cli.now_iso,
            )
            third = study_import.import_study_manifest(
                connection,
                reduced,
                now_iso=cli.now_iso,
            )

            scoped_counts = connection.execute(
                """
                SELECT
                    (SELECT COUNT(*) FROM knowledge_studies s
                     WHERE s.workbook_analysis_id=?) AS studies,
                    (SELECT COUNT(*) FROM knowledge_study_contexts c
                     JOIN knowledge_studies s ON s.study_id=c.study_id
                     WHERE s.workbook_analysis_id=?) AS contexts,
                    (SELECT COUNT(*) FROM knowledge_factors f
                     JOIN knowledge_studies s ON s.study_id=f.study_id
                     WHERE s.workbook_analysis_id=?) AS factors,
                    (SELECT COUNT(*) FROM knowledge_arms a
                     JOIN knowledge_studies s ON s.study_id=a.study_id
                     WHERE s.workbook_analysis_id=?) AS arms,
                    (SELECT COUNT(*) FROM knowledge_outcomes o
                     JOIN knowledge_studies s ON s.study_id=o.study_id
                     WHERE s.workbook_analysis_id=?) AS outcomes,
                    (SELECT COUNT(*) FROM knowledge_observations v
                     JOIN knowledge_outcomes o ON o.outcome_id=v.outcome_id
                     JOIN knowledge_studies s ON s.study_id=o.study_id
                     WHERE s.workbook_analysis_id=?) AS observations,
                    (SELECT COUNT(*) FROM knowledge_comparisons c
                     JOIN knowledge_studies s ON s.study_id=c.study_id
                     WHERE s.workbook_analysis_id=?) AS comparisons,
                    (SELECT COUNT(*) FROM knowledge_effects e
                     JOIN knowledge_comparisons c ON c.comparison_id=e.comparison_id
                     JOIN knowledge_studies s ON s.study_id=c.study_id
                     WHERE s.workbook_analysis_id=?) AS effects,
                    (SELECT COUNT(*) FROM knowledge_claims c
                     WHERE c.workbook_analysis_id=?) AS claims,
                    (SELECT COUNT(*) FROM knowledge_arm_factor_values fv
                     JOIN knowledge_arms a ON a.arm_id=fv.arm_id
                     JOIN knowledge_studies s ON s.study_id=a.study_id
                     WHERE s.workbook_analysis_id=?) AS factor_values
                """,
                (workbook_analysis_id,) * 10,
            ).fetchone()
            current_study_id = connection.execute(
                "SELECT study_id FROM knowledge_studies WHERE workbook_analysis_id=?",
                (workbook_analysis_id,),
            ).fetchone()[0]
            current_outcome_id = connection.execute(
                """
                SELECT o.outcome_id
                FROM knowledge_outcomes o
                JOIN knowledge_studies s ON s.study_id=o.study_id
                WHERE s.workbook_analysis_id=?
                """,
                (workbook_analysis_id,),
            ).fetchone()[0]
            external_link_count = connection.execute(
                """
                SELECT COUNT(*)
                FROM entity_evidence_links
                WHERE entity_type='EXTERNAL_NOTE'
                  AND entity_uid='external-note-1'
                  AND evidence_id=?
                """,
                (context_evidence_id,),
            ).fetchone()[0]
            human_link_count = connection.execute(
                """
                SELECT COUNT(*)
                FROM entity_evidence_links
                WHERE entity_type='STUDY'
                  AND entity_uid=?
                  AND evidence_id=?
                  AND evidence_role='HUMAN_NOTE'
                """,
                (current_study_uid, human_evidence_id),
            ).fetchone()[0]
            stale_link_counts = [
                connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM entity_evidence_links
                    WHERE entity_type=? AND entity_uid=?
                    """,
                    (entity_type, entity_uid),
                ).fetchone()[0]
                for entity_type, entity_uid in stale_entities
            ]
        finally:
            connection.close()

        self.assertEqual(second["analysisUid"], third["analysisUid"])
        self.assertEqual(
            (1, 0, 0, 1, 1, 1, 0, 0, 0, 0),
            tuple(scoped_counts),
        )
        self.assertEqual(original_study_id, current_study_id)
        self.assertEqual(original_outcome_id, current_outcome_id)
        self.assertEqual(1, external_link_count)
        self.assertEqual(1, human_link_count)
        self.assertTrue(all(count == 0 for count in stale_link_counts))

    def test_changed_analysis_key_replaces_unreviewed_canonical_draft(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            first_manifest = self.manifest(bridge["revisionUid"])
            first_manifest["workbookAnalysis"][
                "verificationStatus"
            ] = "NEEDS_REVIEW"
            study = first_manifest["studies"][0]
            study["verificationStatus"] = "NEEDS_REVIEW"
            study["comparabilityStatus"] = "UNASSESSED"
            study["confoundingStatus"] = "UNASSESSED"
            comparison = study["comparisons"][0]
            comparison["verificationStatus"] = "NEEDS_REVIEW"
            comparison["validityStatus"] = "NEEDS_REVIEW"
            comparison["confoundingStatus"] = "UNASSESSED"
            comparison["aggregationEligible"] = False
            comparison["effects"] = []
            first = study_import.import_study_manifest(
                connection,
                first_manifest,
                now_iso=cli.now_iso,
            )
            old_analysis_uid = first["analysisUid"]
            second_manifest = copy.deepcopy(first_manifest)
            second_manifest["workbookAnalysis"]["key"] = (
                "generic-cooling-review-v2"
            )
            second = study_import.import_study_manifest(
                connection,
                second_manifest,
                now_iso=cli.now_iso,
            )
            analyses = list(
                connection.execute(
                    """
                    SELECT analysis_uid
                    FROM workbook_analyses
                    ORDER BY workbook_analysis_id
                    """
                )
            )
            studies = connection.execute(
                "SELECT COUNT(*) FROM knowledge_studies"
            ).fetchone()[0]
            old_links = connection.execute(
                """
                SELECT COUNT(*)
                FROM entity_evidence_links
                WHERE entity_type='WORKBOOK_ANALYSIS'
                  AND entity_uid=?
                """,
                (old_analysis_uid,),
            ).fetchone()[0]
        finally:
            connection.close()

        self.assertEqual(1, second["supersededAnalyses"])
        self.assertEqual(0, second["preservedStaleAnalyses"])
        self.assertEqual(
            [second["analysisUid"]],
            [row[0] for row in analyses],
        )
        self.assertEqual(1, studies)
        self.assertEqual(0, old_links)

    def test_changed_analysis_key_preserves_reviewed_draft_as_stale(
        self,
    ) -> None:
        connection, bridge = self.prepare_source()
        try:
            first_manifest = self.manifest(bridge["revisionUid"])
            first_manifest["workbookAnalysis"][
                "verificationStatus"
            ] = "NEEDS_REVIEW"
            study = first_manifest["studies"][0]
            study["verificationStatus"] = "NEEDS_REVIEW"
            study["comparabilityStatus"] = "UNASSESSED"
            study["confoundingStatus"] = "UNASSESSED"
            comparison = study["comparisons"][0]
            comparison["verificationStatus"] = "NEEDS_REVIEW"
            comparison["validityStatus"] = "NEEDS_REVIEW"
            comparison["confoundingStatus"] = "UNASSESSED"
            comparison["aggregationEligible"] = False
            comparison["effects"] = []
            first = study_import.import_study_manifest(
                connection,
                first_manifest,
                now_iso=cli.now_iso,
            )
            comparison_uid = connection.execute(
                "SELECT comparison_uid FROM knowledge_comparisons"
            ).fetchone()[0]
            connection.execute(
                """
                INSERT INTO review_decisions(
                    decision_uid, entity_type, entity_uid, decision,
                    reason, reviewer, decided_at
                ) VALUES (
                    'decision-1', 'COMPARISON', ?, 'RETURN_TO_REVIEW',
                    'Preserve review history', 'tester', ?
                )
                """,
                (comparison_uid, cli.now_iso()),
            )
            second_manifest = copy.deepcopy(first_manifest)
            second_manifest["workbookAnalysis"]["key"] = (
                "generic-cooling-review-v2"
            )
            second = study_import.import_study_manifest(
                connection,
                second_manifest,
                now_iso=cli.now_iso,
            )
            old_status = connection.execute(
                """
                SELECT verification_status
                FROM workbook_analyses
                WHERE analysis_uid=?
                """,
                (first["analysisUid"],),
            ).fetchone()[0]
            analysis_count = connection.execute(
                "SELECT COUNT(*) FROM workbook_analyses"
            ).fetchone()[0]
        finally:
            connection.close()

        self.assertEqual(0, second["supersededAnalyses"])
        self.assertEqual(1, second["preservedStaleAnalyses"])
        self.assertEqual("STALE", old_status)
        self.assertEqual(2, analysis_count)

    def test_import_validation_issues_are_synchronized_without_touching_human_issues(self) -> None:
        connection, bridge = self.prepare_source()
        try:
            needs_review = self.manifest(bridge["revisionUid"])
            needs_review["workbookAnalysis"]["verificationStatus"] = "NEEDS_REVIEW"
            study = needs_review["studies"][0]
            study["verificationStatus"] = "NEEDS_REVIEW"
            study["confoundingStatus"] = "UNASSESSED"
            study["comparisons"] = []
            result = study_import.import_study_manifest(
                connection,
                needs_review,
                now_iso=cli.now_iso,
            )
            workbook_analysis_id = result["workbookAnalysisId"]
            study_uid = connection.execute(
                "SELECT study_uid FROM knowledge_studies WHERE workbook_analysis_id=?",
                (workbook_analysis_id,),
            ).fetchone()[0]
            analysis_codes = {
                row[0]
                for row in connection.execute(
                    """
                    SELECT issue_code
                    FROM validation_issues
                    WHERE entity_type='WORKBOOK_ANALYSIS'
                      AND entity_uid=?
                      AND validator_name='canonical-study-import'
                    """,
                    (result["analysisUid"],),
                )
            }
            study_codes = {
                row[0]
                for row in connection.execute(
                    """
                    SELECT issue_code
                    FROM validation_issues
                    WHERE entity_type='STUDY'
                      AND entity_uid=?
                      AND validator_name='canonical-study-import'
                    """,
                    (study_uid,),
                )
            }
            connection.execute(
                """
                INSERT INTO validation_issues(
                    issue_uid, entity_type, entity_uid, issue_code, severity,
                    message, details_json, status, validator_name,
                    validator_version, created_at
                ) VALUES (
                    'human-issue-1', 'STUDY', ?, 'HUMAN_NOTE', 'INFO',
                    'Keep this review note.', '{}', 'ACCEPTED', 'human-review',
                    '1', '2026-07-17T00:00:00Z'
                )
                """,
                (study_uid,),
            )

            verified = self.manifest(bridge["revisionUid"])
            study_import.import_study_manifest(
                connection,
                verified,
                now_iso=cli.now_iso,
            )
            importer_issue_count = connection.execute(
                """
                SELECT COUNT(*)
                FROM validation_issues
                WHERE entity_uid IN (?, ?)
                  AND validator_name='canonical-study-import'
                """,
                (result["analysisUid"], study_uid),
            ).fetchone()[0]
            human_issue_count = connection.execute(
                """
                SELECT COUNT(*)
                FROM validation_issues
                WHERE issue_uid='human-issue-1'
                  AND status='ACCEPTED'
                  AND validator_name='human-review'
                """
            ).fetchone()[0]

            excluded = self.manifest(bridge["revisionUid"])
            excluded["workbookAnalysis"]["verificationStatus"] = "EXCLUDED"
            excluded["studies"] = []
            study_import.import_study_manifest(
                connection,
                excluded,
                now_iso=cli.now_iso,
            )
            excluded_codes = {
                row[0]
                for row in connection.execute(
                    """
                    SELECT issue_code
                    FROM validation_issues
                    WHERE entity_type='WORKBOOK_ANALYSIS'
                      AND entity_uid=?
                      AND validator_name='canonical-study-import'
                    """,
                    (result["analysisUid"],),
                )
            }
        finally:
            connection.close()

        self.assertEqual({"NEEDS_REVIEW"}, analysis_codes)
        self.assertTrue(
            {"NEEDS_REVIEW", "NO_COMPARISON", "CONFOUNDING_UNASSESSED"}
            <= study_codes
        )
        self.assertEqual(0, importer_issue_count)
        self.assertEqual(1, human_issue_count)
        self.assertEqual({"EXCLUDED"}, excluded_codes)


if __name__ == "__main__":
    unittest.main()
