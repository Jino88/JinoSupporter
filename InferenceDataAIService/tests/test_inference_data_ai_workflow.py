from __future__ import annotations

import json
import sqlite3
import tempfile
import threading
import unittest
from unittest import mock
from contextlib import closing, contextmanager
from pathlib import Path

from openpyxl import Workbook

import inference_data_ai_cli as cli
import inference_data_ai_content_coverage as content_coverage
import inference_data_ai_staged_draft_v2 as staged_v2
import inference_data_ai_workflow as workflow


class IncrementalWorkbookWorkflowTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory(dir=cli.SERVICE_DIR)
        self.root = Path(self.temp.name)
        self.database = self.root / "knowledge.sqlite"
        self.artifacts = self.root / "artifacts"
        cli.init_universal_db(self.database, "Fixture")

    def tearDown(self) -> None:
        self.temp.cleanup()

    def test_windows_system_error_marks_stale_workflow_owner(self) -> None:
        with mock.patch.object(
            workflow.os,
            "kill",
            side_effect=SystemError("invalid Windows PID"),
        ):
            self.assertFalse(workflow._pid_exists(424242))

    def test_retry_journal_snapshot_does_not_mix_attempt_stages(self) -> None:
        prior_result = {"publicAnalysisId": "ANALYSIS-OLD"}
        journal = {
            "attempt": 3,
            "semanticContracts": {"studyDraftPromptVersion": "v24"},
            "status": "NEEDS_REVIEW",
            "currentStage": "",
            "stages": {
                name: {
                    "status": "COMPLETED",
                    "finishedAt": f"old-{name}",
                }
                for name in workflow.STAGE_ORDER
            },
            "result": prior_result,
            "updatedAt": "old-updated",
            "finishedAt": "old-finished",
        }

        workflow._reset_journal_for_retry(
            journal,
            started_at="new-start",
        )

        self.assertEqual(4, journal["attempt"])
        self.assertEqual("RUNNING", journal["status"])
        self.assertIsNone(journal["result"])
        self.assertEqual("", journal["finishedAt"])
        self.assertTrue(
            all(
                item == {"status": "PENDING"}
                for item in journal["stages"].values()
            )
        )
        prior = journal["attemptHistory"][-1]
        self.assertEqual(3, prior["attempt"])
        self.assertEqual(prior_result, prior["result"])
        self.assertEqual(
            "COMPLETED",
            prior["stages"]["VERIFY"]["status"],
        )

    def test_starting_stage_invalidates_downstream_and_result(self) -> None:
        journal = {
            "stages": {
                name: {"status": "COMPLETED"}
                for name in workflow.STAGE_ORDER
            },
            "result": {"publicAnalysisId": "ANALYSIS-OLD"},
            "finishedAt": "old-finished",
        }

        workflow._invalidate_downstream_journal_stages(
            journal,
            "DRAFT",
        )

        self.assertEqual(
            {"status": "COMPLETED"},
            journal["stages"]["DRAFT"],
        )
        self.assertEqual(
            {"status": "PENDING"},
            journal["stages"]["IMPORT"],
        )
        self.assertEqual(
            {"status": "PENDING"},
            journal["stages"]["VERIFY"],
        )
        self.assertIsNone(journal["result"])
        self.assertEqual("", journal["finishedAt"])

    def test_invalidated_draft_quarantines_only_unverified_analysis(
        self,
    ) -> None:
        source_path = str(self.root / "source.xlsx")
        source = {
            "dataset": "Fixture",
            "sourcePath": source_path,
            "revisionUid": "revision-fixture",
            "contentSha256": "a" * 64,
        }
        with cli.connect_rw(self.database) as connection:
            document_id = int(
                connection.execute(
                    """
                    INSERT INTO source_documents(
                        document_uid, dataset, source_path,
                        original_file_name, created_at, updated_at
                    ) VALUES ('document-fixture', 'Fixture', ?,
                              'source.xlsx', 't0', 't0')
                    """,
                    (source_path,),
                ).lastrowid
            )
            revision_id = int(
                connection.execute(
                    """
                    INSERT INTO source_revisions(
                        revision_uid, document_id, source_fingerprint,
                        fingerprint_kind, content_sha256,
                        capture_contract, captured_at
                    ) VALUES ('revision-fixture', ?, ?, 'SHA256', ?,
                              'capture-v2-openxml-v1', 't0')
                    """,
                    (document_id, "a" * 64, "a" * 64),
                ).lastrowid
            )
            for analysis_key, public_id, status in (
                ("unverified", "ANALYSIS-UNVERIFIED", "NEEDS_REVIEW"),
                ("verified", "ANALYSIS-VERIFIED", "VERIFIED"),
            ):
                connection.execute(
                    """
                    INSERT INTO workbook_analyses(
                        analysis_uid, public_analysis_id, document_id,
                        revision_id, analysis_key, analysis_status,
                        verification_status, analyzer_name,
                        created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?,
                              'canonical-study-import', 't0', 't0')
                    """,
                    (
                        "analysis-" + analysis_key,
                        public_id,
                        document_id,
                        revision_id,
                        analysis_key,
                        status,
                        status,
                    ),
                )
            connection.commit()

        result = workflow._quarantine_invalidated_unverified_analyses(
            self.database,
            source=source,
            reason="Current draft contract rejected the old artifact.",
            now_iso=lambda: "t1",
        )

        self.assertEqual(
            ["ANALYSIS-UNVERIFIED"],
            result["quarantined"],
        )
        self.assertEqual(
            ["ANALYSIS-VERIFIED"],
            [
                item["publicAnalysisId"]
                for item in result["protected"]
            ],
        )
        with cli.connect_ro(self.database) as connection:
            statuses = {
                row["public_analysis_id"]: (
                    row["analysis_status"],
                    row["verification_status"],
                )
                for row in connection.execute(
                    """
                    SELECT public_analysis_id, analysis_status,
                           verification_status
                    FROM workbook_analyses
                    """
                )
            }
        self.assertEqual(
            ("STALE", "STALE"),
            statuses["ANALYSIS-UNVERIFIED"],
        )
        self.assertEqual(
            ("VERIFIED", "VERIFIED"),
            statuses["ANALYSIS-VERIFIED"],
        )

        second = workflow._quarantine_invalidated_unverified_analyses(
            self.database,
            source=source,
            reason="Repeated invalidation is idempotent.",
            now_iso=lambda: "t2",
        )
        self.assertEqual(
            ["ANALYSIS-UNVERIFIED"],
            second["alreadyStale"],
        )

    def test_locator_batch_cross_chunk_evidence_retries_each_chunk(
        self,
    ) -> None:
        source = {
            "revisionUid": "revision-1",
            "contentSha256": "a" * 64,
        }
        chunks = [
            {
                "chunkId": "chunk-a",
                "sheet": "Data",
                "range": "A1:A2",
                "contextCells": [],
            },
            {
                "chunkId": "chunk-b",
                "sheet": "Data",
                "range": "B1:B2",
                "contextCells": [],
            },
        ]
        jobs = [
            (chunk, self.root / f"{chunk['chunkId']}.json")
            for chunk in chunks
        ]
        submitted_sizes: list[int] = []

        def result_for(
            chunk: dict,
            evidence_range: str,
        ) -> dict:
            return {
                "schemaVersion": "semantic-locator-v1",
                "promptVersion": "semantic-locator-prompt-v1",
                "revisionUid": source["revisionUid"],
                "contentSha256": source["contentSha256"],
                "chunkId": chunk["chunkId"],
                "status": "CANDIDATES",
                "candidates": [
                    {
                        "key": f"candidate-{chunk['chunkId']}",
                        "title": "Candidate",
                        "summary": "Source-backed candidate.",
                        "designHint": "DESCRIPTIVE",
                        "contexts": [],
                        "changedFactors": [],
                        "outcomes": ["Outcome"],
                        "comparisonHints": [],
                        "evidence": [
                            {
                                "sheet": "Data",
                                "range": evidence_range,
                                "role": "CANDIDATE_REGION",
                            }
                        ],
                        "limitations": [],
                        "confidence": 0.8,
                    }
                ],
                "notes": [],
            }

        def validating_runner(
            batch: list[tuple[dict, Path]],
        ) -> list[dict]:
            submitted_sizes.append(len(batch))
            results = [
                result_for(
                    chunk,
                    (
                        "B1"
                        if len(batch) > 1 and chunk["chunkId"] == "chunk-a"
                        else str(chunk["range"]).split(":")[0]
                    ),
                )
                for chunk, _output in batch
            ]
            return [
                workflow.validate_locator_result(
                    result,
                    revision_uid=source["revisionUid"],
                    content_sha256=source["contentSha256"],
                    chunk=chunk,
                )
                for result, (chunk, _output) in zip(
                    results,
                    batch,
                    strict=True,
                )
            ]

        recovered, ai_calls = (
            workflow._run_locator_batch_with_singleton_retry(
                jobs,
                validating_runner,
            )
        )

        self.assertEqual([2, 1, 1], submitted_sizes)
        self.assertEqual(3, ai_calls)
        self.assertEqual(
            ["chunk-a", "chunk-b"],
            [result["chunkId"] for result in recovered],
        )
        self.assertEqual(
            ["A1", "B1"],
            [
                result["candidates"][0]["evidence"][0]["range"]
                for result in recovered
            ],
        )

    def tabular_source(self) -> Path:
        path = self.root / "novel-review.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Novel Data"
        sheet["B2"] = "Cryogenic dwell"
        sheet["C2"] = "Unseen response"
        sheet["B3"] = "9.5 s"
        sheet["C3"] = "Recorded"
        workbook.save(path)
        workbook.close()
        return path

    def staged_resume_source(self) -> Path:
        path = self.root / "staged-resume.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Resume Data"
        sheet["B2"] = "Narrative alpha"
        sheet["C2"] = "Narrative beta"
        sheet["B3"] = "Narrative gamma"
        sheet["C3"] = "Narrative delta"
        workbook.save(path)
        workbook.close()
        return path

    def empty_source(self) -> Path:
        path = self.root / "empty.xlsx"
        workbook = Workbook()
        workbook.save(path)
        workbook.close()
        return path

    def numeric_source(self) -> Path:
        path = self.root / "numeric-review.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Gauss Data"
        sheet["A1"] = "Gauss raw measurements"
        sheet["B1"] = "Test values"
        sheet["B2"] = 410
        sheet["B3"] = 412
        workbook.save(path)
        workbook.close()
        return path

    def formula_source(self) -> Path:
        path = self.root / "formula-review.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Formula Data"
        sheet["A1"] = "Condition"
        sheet["B1"] = "Raw response"
        sheet["C1"] = "Derived response"
        sheet["A2"] = "Test"
        sheet["B2"] = 4
        sheet["C2"] = "=B2/2"
        workbook.save(path)
        workbook.close()
        return path

    def multi_section_source(self) -> Path:
        path = self.root / "multi-section-review.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Review"
        sheet["A1"] = "Control arm"
        sheet["B1"] = "Observed stable"
        sheet["A10"] = "Test arm"
        sheet["B10"] = "Observed changed"
        workbook.save(path)
        workbook.close()
        return path

    def partially_missed_numeric_source(self) -> Path:
        path = self.root / "partially-missed-review.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Review"
        sheet["A1"] = "Candidate narrative"
        sheet["B1"] = "Candidate value"
        sheet["A2"] = "Condition"
        sheet["B2"] = "Recorded"
        sheet["A10"] = "Missed numeric result"
        sheet["B10"] = 19000
        sheet["A11"] = "Missed numeric result 2"
        sheet["B11"] = 20000
        workbook.save(path)
        workbook.close()
        return path

    def test_outcome_labels_without_results_are_semantically_terminal(self) -> None:
        manifest = {
            "studies": [
                {
                    "outcomes": [
                        {
                            "originalLabel": "Image-only comparison label",
                            "observations": [],
                        }
                    ],
                    "conclusions": [],
                }
            ]
        }
        self.assertTrue(
            workflow._draft_has_labels_but_no_reviewable_results(manifest)
        )
        manifest["studies"][0]["outcomes"][0]["observations"].append(
            {"valueText": "recorded"}
        )
        self.assertFalse(
            workflow._draft_has_labels_but_no_reviewable_results(manifest)
        )

    def test_tabular_workflow_is_review_gated_and_resumable(self) -> None:
        source = self.tabular_source()
        before = source.read_bytes()
        calls = {"locator": 0, "draft": 0}
        gate_entries: list[str] = []

        @contextmanager
        def pipeline_gate(name: str):
            gate_entries.append(name)
            yield

        def fake_locator_batch(**kwargs: object) -> list[dict]:
            calls["locator"] += 1
            source_identity = kwargs["source"]
            chunks = kwargs["chunks"]
            output_paths = kwargs["output_paths"]
            results = []
            for chunk in chunks:
                sheet_value = chunk["sheet"]
                sheet = (
                    sheet_value["title"]
                    if isinstance(sheet_value, dict)
                    else sheet_value
                )
                result = {
                    "schemaVersion": "semantic-locator-v1",
                    "promptVersion": "semantic-locator-prompt-v1",
                    "revisionUid": source_identity["revisionUid"],
                    "contentSha256": source_identity["contentSha256"],
                    "chunkId": chunk["chunkId"],
                    "status": "CANDIDATES",
                    "candidates": [
                        {
                            "key": f"candidate-{chunk['chunkId']}",
                            "title": "Novel review candidate",
                            "summary": "A source-backed review region.",
                            "designHint": "DESCRIPTIVE",
                            "contexts": [],
                            "changedFactors": ["Cryogenic dwell"],
                            "outcomes": ["Unseen response"],
                            "comparisonHints": [],
                            "evidence": [
                                {
                                    "sheet": sheet,
                                    "range": chunk["primaryRange"],
                                    "role": "CANDIDATE_REGION",
                                }
                            ],
                            "limitations": [
                                "No explicit control is present."
                            ],
                            "confidence": 0.8,
                        }
                    ],
                    "notes": [],
                }
                Path(output_paths[chunk["chunkId"]]).write_text(
                    json.dumps(result, ensure_ascii=False),
                    encoding="utf-8",
                )
                results.append(result)
            return results

        def fake_draft(**kwargs: object) -> dict:
            calls["draft"] += 1
            self.assertIs(True, kwargs["content_complete"])
            self.assertEqual(
                "canonical-study-manifest.json",
                Path(kwargs["output_path"]).name,
            )
            source_identity = kwargs["source"]
            chunk = kwargs["focused_chunks"][0]
            sheet_value = chunk["sheet"]
            sheet = (
                sheet_value["title"]
                if isinstance(sheet_value, dict)
                else sheet_value
            )
            evidence = [
                {
                    "sheet": sheet,
                    "range": chunk["primaryRange"],
                    "role": "SOURCE",
                    "sourceText": "Cryogenic dwell / Unseen response",
                    "note": "",
                }
            ]
            result = {
                "schemaVersion": "canonical-study-manifest-v1",
                "source": {
                    **source_identity,
                    "contentComplete": True,
                },
                "workbookAnalysis": {
                    "key": "novel-review",
                    "title": "Novel review",
                    "summary": "One descriptive Study requires review.",
                    "status": "NEEDS_REVIEW",
                    "verificationStatus": "NEEDS_REVIEW",
                    "limitations": [
                        "No source-backed control/comparison was identified."
                    ],
                    "evidence": evidence,
                },
                "studies": [
                    {
                        "key": "novel-study",
                        "title": "Cryogenic dwell descriptive review",
                        "purpose": "",
                        "hypothesis": "",
                        "objective": "",
                        "designType": "DESCRIPTIVE",
                        "comparisonBasis": "",
                        "verificationStatus": "NEEDS_REVIEW",
                        "comparabilityStatus": "UNASSESSED",
                        "confoundingStatus": "UNASSESSED",
                        "summary": "A condition and response are recorded.",
                        "limitations": [
                            "No valid comparison is present."
                        ],
                        "evidence": evidence,
                        "contexts": [],
                        "factors": [],
                        "arms": [
                            {
                                "key": "recorded",
                                "role": "OTHER",
                                "label": "Recorded",
                                "condition": "Recorded",
                                "sampleSize": None,
                                "sampleBasis": "",
                                "matchingBasis": "",
                                "factorValues": [],
                                "evidence": [
                                    {
                                        "sheet": sheet,
                                        "range": "C3",
                                        "role": "ARM",
                                        "sourceText": "Recorded",
                                        "note": "",
                                    }
                                ],
                            }
                        ],
                        "outcomes": [
                            {
                                "key": "unseen-response",
                                "originalLabel": "Unseen response",
                                "metricType": "CATEGORICAL",
                                "unit": "",
                                "favorableDirection": "UNKNOWN",
                                "evidence": [
                                    {
                                        "sheet": sheet,
                                        "range": "C2",
                                        "role": "OUTCOME_LABEL",
                                        "sourceText": "Unseen response",
                                        "note": "",
                                    }
                                ],
                                "observations": [
                                    {
                                        "key": "unseen-response-value",
                                        "arm": "recorded",
                                        "valueNumber": None,
                                        "valueText": "9.5 s",
                                        "numerator": None,
                                        "denominator": None,
                                        "ratePpm": None,
                                        "min": None,
                                        "max": None,
                                        "average": None,
                                        "sampleSize": None,
                                        "evidence": [
                                            {
                                                "sheet": sheet,
                                                "range": "B3",
                                                "role": "OBSERVATION",
                                                "sourceText": "9.5 s",
                                                "note": "",
                                            }
                                        ],
                                    }
                                ],
                            }
                        ],
                        "comparisons": [],
                        "conclusions": [],
                    }
                ],
            }
            kwargs["additional_validator"](result)
            Path(kwargs["output_path"]).write_text(
                json.dumps(result, ensure_ascii=False),
                encoding="utf-8",
            )
            return result

        with mock.patch.object(
            workflow,
            "import_study_manifest",
            wraps=workflow.import_study_manifest,
        ) as importer:
            first = workflow.ingest_workbook(
                database_path=self.database,
                source_path=source,
                artifact_root=self.artifacts,
                dataset="Fixture",
                locator_batch_runner=fake_locator_batch,
                draft_runner=fake_draft,
                locator_workers=1,
                pipeline_gate=pipeline_gate,
            )
        self.assertIs(
            True,
            importer.call_args.kwargs["source_claims_prevalidated"],
        )
        with mock.patch.object(
            workflow,
            "extract_workbook",
            side_effect=AssertionError(
                "An unchanged current Capture v2 revision must be reused."
            ),
        ):
            second = workflow.ingest_workbook(
                database_path=self.database,
                source_path=source,
                artifact_root=self.artifacts,
                dataset="Fixture",
                locator_batch_runner=fake_locator_batch,
                draft_runner=fake_draft,
                locator_workers=1,
            )

        self.assertEqual("NEEDS_REVIEW", first["status"])
        self.assertEqual(first["runId"], second["runId"])
        self.assertEqual(1, first["studies"])
        self.assertTrue(first["integrityOk"])
        self.assertFalse(first["imagesAnalyzed"])
        self.assertEqual({"locator": 1, "draft": 1}, calls)
        self.assertTrue(
            {"PACKET", "AI", "DB"} <= set(gate_entries)
        )
        self.assertNotIn("COM", gate_entries)
        self.assertEqual(before, source.read_bytes())
        self.assertEqual(
            "REUSED_CURRENT",
            json.loads(
                Path(second["journalPath"]).read_text(encoding="utf-8")
            )["stages"]["CAPTURE"]["result"]["action"],
        )

        journal = json.loads(
            Path(second["journalPath"]).read_text(encoding="utf-8")
        )
        self.assertEqual(2, journal["attempt"])
        self.assertEqual("NEEDS_REVIEW", journal["status"])
        self.assertEqual(
            "canonical-study-draft-prompt-v25",
            journal["semanticContracts"]["studyDraftPromptVersion"],
        )
        self.assertTrue(
            journal["stages"]["DRAFT"]["result"]["artifactReused"]
        )
        self.assertFalse(
            journal["stages"]["DRAFT"]["result"]["aiExecuted"]
        )
        self.assertTrue(
            all(
                journal["stages"][stage]["status"] == "COMPLETED"
                for stage in workflow.STAGE_ORDER
            )
        )
        with closing(sqlite3.connect(self.database)) as connection:
            issue_codes = {
                row[0]
                for row in connection.execute(
                    """
                    SELECT issue_code
                    FROM validation_issues
                    WHERE validator_name='canonical-study-import'
                    """
                )
            }
        self.assertTrue(
            {"NEEDS_REVIEW", "NO_COMPARISON", "CONFOUNDING_UNASSESSED"}
            <= issue_codes
        )
        provenance_path = (
            Path(second["artifactDirectory"])
            / "study-draft.provenance.json"
        )
        provenance = json.loads(
            provenance_path.read_text(encoding="utf-8")
        )
        self.assertEqual(
            "canonical-study-draft-prompt-v25",
            provenance["promptVersion"],
        )
        provenance["promptVersion"] = "superseded-prompt"
        provenance_path.write_text(
            json.dumps(provenance),
            encoding="utf-8",
        )
        second_journal_path = Path(second["journalPath"])
        stale_journal = json.loads(
            second_journal_path.read_text(encoding="utf-8")
        )
        stale_journal["semanticContracts"][
            "studyDraftPromptVersion"
        ] = "superseded-prompt"
        second_journal_path.write_text(
            json.dumps(stale_journal),
            encoding="utf-8",
        )
        with mock.patch.object(
            workflow,
            "extract_workbook",
            side_effect=AssertionError(
                "Prompt upgrades must not recapture an unchanged source."
            ),
        ):
            third = workflow.ingest_workbook(
                database_path=self.database,
                source_path=source,
                artifact_root=self.artifacts,
                dataset="Fixture",
                locator_batch_runner=fake_locator_batch,
                draft_runner=fake_draft,
                locator_workers=1,
            )
        third_journal = json.loads(
            Path(third["journalPath"]).read_text(encoding="utf-8")
        )
        self.assertEqual({"locator": 1, "draft": 2}, calls)
        self.assertEqual(3, third_journal["attempt"])
        self.assertTrue(
            third_journal["stages"]["DRAFT"]["result"]["aiExecuted"]
        )
        self.assertFalse(
            third_journal["stages"]["DRAFT"]["result"]["artifactReused"]
        )
        self.assertEqual(
            "canonical-study-draft-prompt-v25",
            third_journal["semanticContracts"][
                "studyDraftPromptVersion"
            ],
        )

    def test_formula_overlay_is_end_to_end_resumable_and_tamper_closed(
        self,
    ) -> None:
        source_path = self.formula_source()
        source_bytes = source_path.read_bytes()
        calls = {"locator": 0, "draft": 0}

        def evidence(
            sheet: str,
            address: str,
            role: str,
            source_text: str,
        ) -> dict:
            return {
                "sheet": sheet,
                "range": address,
                "role": role,
                "sourceText": source_text,
                "note": "",
            }

        def fake_locator_batch(**kwargs: object) -> list[dict]:
            calls["locator"] += 1
            source_identity = kwargs["source"]
            results = []
            for chunk in kwargs["chunks"]:
                formula_cells = [
                    cell
                    for cell in chunk["cells"]
                    if cell.get("formula")
                ]
                if formula_cells:
                    self.assertEqual(len(formula_cells), 1)
                    self.assertEqual(formula_cells[0]["formula"], "=B2/2")
                    self.assertEqual(formula_cells[0]["cachedValue"], 2)
                    self.assertEqual(
                        formula_cells[0]["valueSource"],
                        "DETERMINISTIC_FORMULA_DERIVED",
                    )
                sheet = chunk["sheet"]["title"]
                result = {
                    "schemaVersion": "semantic-locator-v1",
                    "promptVersion": "semantic-locator-prompt-v1",
                    "revisionUid": source_identity["revisionUid"],
                    "contentSha256": source_identity["contentSha256"],
                    "chunkId": chunk["chunkId"],
                    "status": "CANDIDATES",
                    "candidates": [
                        {
                            "key": "formula-candidate",
                            "title": "Formula candidate",
                            "summary": "Raw and derived responses.",
                            "designHint": "DESCRIPTIVE",
                            "contexts": [],
                            "changedFactors": ["Condition"],
                            "outcomes": [
                                "Raw response",
                                "Derived response",
                            ],
                            "comparisonHints": [],
                            "evidence": [
                                {
                                    "sheet": sheet,
                                    "range": "A1:C2",
                                    "role": "CANDIDATE_REGION",
                                }
                            ],
                            "limitations": [
                                "Single descriptive condition."
                            ],
                            "confidence": 0.9,
                        }
                    ],
                    "notes": [],
                }
                Path(kwargs["output_paths"][chunk["chunkId"]]).write_text(
                    json.dumps(result),
                    encoding="utf-8",
                )
                results.append(result)
            return results

        def fake_draft(**kwargs: object) -> dict:
            calls["draft"] += 1
            source_identity = kwargs["source"]
            chunk = kwargs["focused_chunks"][0]
            formula_cell = next(
                cell for cell in chunk["cells"] if cell.get("formula")
            )
            self.assertEqual(formula_cell["cachedValue"], 2)
            sheet = chunk["sheet"]["title"]
            source_evidence = [
                evidence(
                    sheet,
                    "A1:C2",
                    "SOURCE",
                    "Condition / Raw response / Derived response",
                ),
                evidence(
                    sheet,
                    "A1",
                    "FACTOR_LABEL",
                    "Condition",
                ),
                evidence(
                    sheet,
                    "A2",
                    "FACTOR_LEVEL",
                    "Test",
                ),
            ]
            result = {
                "schemaVersion": "canonical-study-manifest-v1",
                "source": {
                    **source_identity,
                    "contentComplete": True,
                },
                "workbookAnalysis": {
                    "key": "formula-review",
                    "title": "Formula review",
                    "summary": "Raw and deterministic formula values.",
                    "status": "NEEDS_REVIEW",
                    "verificationStatus": "NEEDS_REVIEW",
                    "limitations": [
                        "Single descriptive condition."
                    ],
                    "evidence": source_evidence,
                },
                "studies": [
                    {
                        "key": "formula-study",
                        "title": "Formula Study",
                        "purpose": "",
                        "hypothesis": "",
                        "objective": "",
                        "designType": "DESCRIPTIVE",
                        "comparisonBasis": "",
                        "verificationStatus": "NEEDS_REVIEW",
                        "comparabilityStatus": "UNASSESSED",
                        "confoundingStatus": "UNASSESSED",
                        "summary": "Two source-backed response values.",
                        "limitations": [
                            "No control comparison is present."
                        ],
                        "evidence": source_evidence,
                        "contexts": [],
                        "factors": [
                            {
                                "key": "condition",
                                "originalLabel": "Condition",
                                "canonicalName": "Condition",
                                "baselineCondition": "",
                                "changedCondition": "Test",
                                "isolationStatus": "UNASSESSED",
                                "evidence": [
                                    evidence(
                                        sheet,
                                        "A1",
                                        "FACTOR_LABEL",
                                        "Condition",
                                    )
                                ],
                            }
                        ],
                        "arms": [
                            {
                                "key": "test",
                                "role": "OTHER",
                                "label": "Test",
                                "condition": "Test",
                                "sampleSize": None,
                                "sampleBasis": "",
                                "matchingBasis": "",
                                "factorValues": [
                                    {
                                        "factor": "condition",
                                        "value": "Test",
                                        "unit": "",
                                    }
                                ],
                                "evidence": [
                                    evidence(
                                        sheet,
                                        "A2",
                                        "ARM",
                                        "Test",
                                    )
                                ],
                            }
                        ],
                        "outcomes": [
                            {
                                "key": "raw-response",
                                "originalLabel": "Raw response",
                                "metricType": "CONTINUOUS",
                                "unit": "",
                                "favorableDirection": "UNKNOWN",
                                "evidence": [
                                    evidence(
                                        sheet,
                                        "B1",
                                        "OUTCOME_LABEL",
                                        "Raw response",
                                    )
                                ],
                                "observations": [
                                    {
                                        "key": "raw-response-test",
                                        "arm": "test",
                                        "valueNumber": 4,
                                        "valueText": "4",
                                        "numerator": None,
                                        "denominator": None,
                                        "ratePpm": None,
                                        "min": None,
                                        "max": None,
                                        "average": None,
                                        "sampleSize": None,
                                        "evidence": [
                                            evidence(
                                                sheet,
                                                "B2",
                                                "OBSERVATION",
                                                "4",
                                            )
                                        ],
                                    }
                                ],
                            },
                            {
                                "key": "derived-response",
                                "originalLabel": "Derived response",
                                "metricType": "CONTINUOUS",
                                "unit": "",
                                "favorableDirection": "UNKNOWN",
                                "evidence": [
                                    evidence(
                                        sheet,
                                        "C1",
                                        "OUTCOME_LABEL",
                                        "Derived response",
                                    )
                                ],
                                "observations": [
                                    {
                                        "key": "derived-response-test",
                                        "arm": "test",
                                        "valueNumber": 2,
                                        "valueText": "2",
                                        "numerator": None,
                                        "denominator": None,
                                        "ratePpm": None,
                                        "min": None,
                                        "max": None,
                                        "average": None,
                                        "sampleSize": None,
                                        "evidence": [
                                            evidence(
                                                sheet,
                                                "C2",
                                                "OBSERVATION",
                                                "2",
                                            )
                                        ],
                                    }
                                ],
                            },
                        ],
                        "measurementSeries": [],
                        "comparisons": [],
                        "conclusions": [],
                    }
                ],
            }
            kwargs["additional_validator"](result)
            Path(kwargs["output_path"]).write_text(
                json.dumps(result),
                encoding="utf-8",
            )
            return result

        first = workflow.ingest_workbook(
            database_path=self.database,
            source_path=source_path,
            artifact_root=self.artifacts,
            dataset="Fixture",
            locator_batch_runner=fake_locator_batch,
            draft_runner=fake_draft,
            locator_workers=1,
            derive_formula_values=True,
        )
        overlay_path = (
            Path(first["artifactDirectory"])
            / "formula-derivation.overlay.json"
        )
        packet_path = (
            Path(first["artifactDirectory"])
            / "semantic-source-packet.json"
        )
        overlay = json.loads(overlay_path.read_text(encoding="utf-8"))
        packet = json.loads(packet_path.read_text(encoding="utf-8"))
        self.assertEqual(overlay["formulaCount"], 1)
        self.assertEqual(overlay["numericCount"], 1)
        self.assertEqual(overlay["errorCount"], 0)
        self.assertEqual(
            packet["inventory"]["formulaDerivation"]["overlaySha256"],
            overlay["overlaySha256"],
        )
        self.assertEqual(
            packet["inventory"]["formulaDerivation"]["captureMutated"],
            False,
        )
        self.assertEqual(
            packet["inventory"]["formulaDerivation"][
                "unresolvedFormulaCellCount"
            ],
            0,
        )
        projected_formula = next(
            cell
            for chunk in packet["chunks"]
            for cell in chunk["cells"]
            if cell.get("formula")
        )
        self.assertEqual(projected_formula["cachedValue"], 2)

        with mock.patch.object(
            workflow,
            "extract_workbook",
            side_effect=AssertionError(
                "Formula-overlay resume must reuse unchanged Capture v2."
            ),
        ):
            second = workflow.ingest_workbook(
                database_path=self.database,
                source_path=source_path,
                artifact_root=self.artifacts,
                dataset="Fixture",
                locator_batch_runner=fake_locator_batch,
                draft_runner=fake_draft,
                locator_workers=1,
                derive_formula_values=True,
            )
        self.assertEqual(first["runId"], second["runId"])
        self.assertEqual(calls, {"locator": 1, "draft": 1})
        self.assertEqual(source_bytes, source_path.read_bytes())
        journal = json.loads(
            Path(second["journalPath"]).read_text(encoding="utf-8")
        )
        formula_contract = journal["semanticContracts"][
            "formulaDerivation"
        ]
        self.assertEqual(
            formula_contract["overlaySha256"],
            overlay["overlaySha256"],
        )
        self.assertEqual(
            journal["stages"]["PACKET"]["result"][
                "formulaDerivation"
            ]["numericCount"],
            1,
        )

        with closing(sqlite3.connect(self.database)) as connection:
            capture_cache = connection.execute(
                """
                SELECT c.cached_value_json
                FROM capture_v2_cells c
                JOIN capture_v2_sheets s ON s.sheet_id=c.sheet_id
                JOIN capture_v2_revisions r ON r.revision_id=s.revision_id
                WHERE r.revision_uid=? AND c.coordinate='C2'
                """,
                (first["revisionUid"],),
            ).fetchone()[0]
        self.assertIsNone(capture_cache)

        tampered = dict(overlay)
        tampered["numericCount"] = 999
        overlay_path.write_text(json.dumps(tampered), encoding="utf-8")
        with self.assertRaises(ValueError):
            workflow.ingest_workbook(
                database_path=self.database,
                source_path=source_path,
                artifact_root=self.artifacts,
                dataset="Fixture",
                locator_batch_runner=fake_locator_batch,
                draft_runner=fake_draft,
                locator_workers=1,
                derive_formula_values=True,
            )
        self.assertEqual(calls, {"locator": 1, "draft": 1})

    def test_small_multi_section_workbook_uses_one_monolithic_call(
        self,
    ) -> None:
        source = self.multi_section_source()
        draft_calls = 0

        def fake_locator_batch(**kwargs: object) -> list[dict]:
            results = []
            source_identity = kwargs["source"]
            for chunk in kwargs["chunks"]:
                result = {
                    "schemaVersion": "semantic-locator-v1",
                    "promptVersion": "semantic-locator-prompt-v1",
                    "revisionUid": source_identity["revisionUid"],
                    "contentSha256": source_identity["contentSha256"],
                    "chunkId": chunk["chunkId"],
                    "status": "CANDIDATES",
                    "candidates": [
                        {
                            "key": f"candidate-{chunk['chunkId']}",
                            "title": "Cross-section comparison",
                            "summary": "Two explicit source arms.",
                            "designHint": "CONTROL_VS_TEST",
                            "contexts": [],
                            "changedFactors": [],
                            "outcomes": ["Observed result"],
                            "comparisonHints": [
                            "Control versus Test"
                            ],
                            "evidence": [
                                {
                                    "sheet": "Review",
                                    "range": chunk["primaryRange"],
                                    "role": "CANDIDATE_REGION",
                                }
                            ],
                            "limitations": [],
                            "confidence": 0.9,
                        }
                    ],
                    "notes": [],
                }
                Path(kwargs["output_paths"][chunk["chunkId"]]).write_text(
                    json.dumps(result),
                    encoding="utf-8",
                )
                results.append(result)
            return results

        def evidence(address: str) -> list[dict]:
            return [
                {
                    "sheet": "Review",
                    "range": address,
                    "role": "SOURCE",
                    "sourceText": address,
                    "note": "",
                }
            ]

        def fake_draft(**kwargs: object) -> dict:
            nonlocal draft_calls
            draft_calls += 1
            self.assertEqual(2, len(kwargs["focused_chunks"]))
            self.assertEqual(
                kwargs["expected_prompt_sha256"],
                __import__("hashlib").sha256(
                    kwargs["exact_prompt_text"].encode("utf-8")
                ).hexdigest(),
            )
            source_identity = kwargs["source"]
            combined_evidence = [
                *evidence("A1:B1"),
                *evidence("A10:B10"),
            ]

            def observation(
                key: str,
                arm: str,
                value: str,
                address: str,
            ) -> dict:
                return {
                    "key": key,
                    "arm": arm,
                    "valueNumber": None,
                    "valueText": value,
                    "numerator": None,
                    "denominator": None,
                    "ratePpm": None,
                    "min": None,
                    "max": None,
                    "average": None,
                    "sampleSize": None,
                    "evidence": evidence(address),
                }

            result = {
                "schemaVersion": "canonical-study-manifest-v1",
                "source": {
                    **source_identity,
                    "contentComplete": True,
                },
                "workbookAnalysis": {
                    "key": "multi-analysis",
                    "title": "Multi-section review",
                    "summary": "One source comparison spans two sections.",
                    "status": "NEEDS_REVIEW",
                    "verificationStatus": "NEEDS_REVIEW",
                    "limitations": [],
                    "evidence": combined_evidence,
                },
                "studies": [
                    {
                        "key": "cross-section-study",
                        "title": "Control versus Test",
                        "purpose": "",
                        "hypothesis": "",
                        "objective": "",
                        "designType": "CONTROL_VS_TEST",
                        "comparisonBasis": (
                            "Explicit Control and Test source arms."
                        ),
                        "verificationStatus": "NEEDS_REVIEW",
                        "comparabilityStatus": "UNASSESSED",
                        "confoundingStatus": "UNASSESSED",
                        "summary": "Qualitative observations are retained.",
                        "limitations": [],
                        "evidence": combined_evidence,
                        "contexts": [],
                        "factors": [],
                        "arms": [
                            {
                                "key": "baseline",
                                "role": "CONTROL",
                                "label": "Control arm",
                                "condition": "Control arm",
                                "sampleSize": None,
                                "sampleBasis": "",
                                "matchingBasis": "",
                                "factorValues": [],
                                "evidence": evidence("A1"),
                            },
                            {
                                "key": "test",
                                "role": "TEST",
                                "label": "Test arm",
                                "condition": "Test arm",
                                "sampleSize": None,
                                "sampleBasis": "",
                                "matchingBasis": "",
                                "factorValues": [],
                                "evidence": evidence("A10"),
                            },
                        ],
                        "outcomes": [
                            {
                                "key": "observed",
                                "originalLabel": "Observed result",
                                "metricType": "CATEGORICAL",
                                "unit": "",
                                "favorableDirection": "UNKNOWN",
                                "evidence": [
                                    *evidence("B1"),
                                    *evidence("B10"),
                                ],
                                "observations": [
                                    observation(
                                        "baseline-observation",
                                        "baseline",
                                        "Observed stable",
                                        "B1",
                                    ),
                                    observation(
                                        "test-observation",
                                        "test",
                                        "Observed changed",
                                        "B10",
                                    ),
                                ],
                            }
                        ],
                        "comparisons": [
                            {
                                "key": "baseline-v-test",
                                "comparedArm": "test",
                                "controlArm": "baseline",
                                "designType": "CONTROL_VS_TEST",
                                "matchingBasis": (
                                    "Explicit source arm labels."
                                ),
                                "validityStatus": "NEEDS_REVIEW",
                                "confoundingStatus": "UNASSESSED",
                                "verificationStatus": "NEEDS_REVIEW",
                                "aggregationEligible": False,
                                "evidence": combined_evidence,
                                "effects": [],
                            }
                        ],
                        "conclusions": [],
                    }
                ],
            }
            kwargs["additional_validator"](result)
            Path(kwargs["output_path"]).write_text(
                json.dumps(result),
                encoding="utf-8",
            )
            return result

        result = workflow.ingest_workbook(
            database_path=self.database,
            source_path=source,
            artifact_root=self.artifacts,
            dataset="Fixture",
            locator_batch_runner=fake_locator_batch,
            draft_runner=fake_draft,
            locator_workers=1,
            draft_monolithic_max_bytes=400_000,
        )

        self.assertEqual(1, draft_calls)
        manifest = json.loads(
            Path(result["manifestPath"]).read_text(encoding="utf-8")
        )
        self.assertEqual(
            1,
            len(manifest["studies"][0]["comparisons"]),
        )
        draft_result = json.loads(
            Path(result["journalPath"]).read_text(encoding="utf-8")
        )["stages"]["DRAFT"]["result"]
        self.assertEqual("MONOLITHIC", draft_result["mode"])
        self.assertFalse(draft_result["staged"])

    def test_unselected_numeric_section_fails_before_draft_and_import(
        self,
    ) -> None:
        source = self.partially_missed_numeric_source()

        def fake_locator_batch(**kwargs: object) -> list[dict]:
            source_identity = kwargs["source"]
            results = []
            for chunk in kwargs["chunks"]:
                candidate = int(chunk["sectionIndex"]) == 1
                result = {
                    "schemaVersion": "semantic-locator-v1",
                    "promptVersion": "semantic-locator-prompt-v1",
                    "revisionUid": source_identity["revisionUid"],
                    "contentSha256": source_identity["contentSha256"],
                    "chunkId": chunk["chunkId"],
                    "status": (
                        "CANDIDATES" if candidate else "NO_CANDIDATE"
                    ),
                    "candidates": (
                        [
                            {
                                "key": "narrative-candidate",
                                "title": "Narrative candidate",
                                "summary": "A source-backed region.",
                                "designHint": "DESCRIPTIVE",
                                "contexts": [],
                                "changedFactors": [],
                                "outcomes": [],
                                "comparisonHints": [],
                                "evidence": [
                                    {
                                        "sheet": "Review",
                                        "range": chunk["primaryRange"],
                                        "role": "CANDIDATE_REGION",
                                    }
                                ],
                                "limitations": [],
                                "confidence": 0.8,
                            }
                        ]
                        if candidate
                        else []
                    ),
                    "notes": [],
                }
                Path(kwargs["output_paths"][chunk["chunkId"]]).write_text(
                    json.dumps(result),
                    encoding="utf-8",
                )
                results.append(result)
            return results

        with self.assertRaisesRegex(
            workflow.IncrementalIngestError,
            "outside candidate-bearing sections",
        ):
            workflow.ingest_workbook(
                database_path=self.database,
                source_path=source,
                artifact_root=self.artifacts,
                dataset="Fixture",
                locator_batch_runner=fake_locator_batch,
                draft_runner=lambda **_kwargs: self.fail(
                    "unsafe selection must fail before drafting"
                ),
                locator_workers=1,
            )
        with closing(sqlite3.connect(self.database)) as connection:
            self.assertEqual(
                0,
                connection.execute(
                    "SELECT COUNT(*) FROM workbook_analyses"
                ).fetchone()[0],
            )

    def test_staged_draft_resumes_parts_and_blocks_partial_import(
        self,
    ) -> None:
        source = self.staged_resume_source()
        locator_calls = 0

        def fake_locator_batch(**kwargs: object) -> list[dict]:
            nonlocal locator_calls
            locator_calls += 1
            source_identity = kwargs["source"]
            chunks = kwargs["chunks"]
            output_paths = kwargs["output_paths"]
            results = []
            for chunk in chunks:
                sheet_value = chunk["sheet"]
                sheet = (
                    sheet_value["title"]
                    if isinstance(sheet_value, dict)
                    else sheet_value
                )
                result = {
                    "schemaVersion": "semantic-locator-v1",
                    "promptVersion": "semantic-locator-prompt-v1",
                    "revisionUid": source_identity["revisionUid"],
                    "contentSha256": source_identity["contentSha256"],
                    "chunkId": chunk["chunkId"],
                    "status": "CANDIDATES",
                    "candidates": [
                        {
                            "key": f"candidate-{chunk['chunkId']}",
                            "title": "Fragment candidate",
                            "summary": "One bounded source fragment.",
                            "designHint": "DESCRIPTIVE",
                            "contexts": [],
                            "changedFactors": [],
                            "outcomes": [],
                            "comparisonHints": [],
                            "evidence": [
                                {
                                    "sheet": sheet,
                                    "range": chunk["primaryRange"],
                                    "role": "CANDIDATE_REGION",
                                }
                            ],
                            "limitations": [],
                            "confidence": 0.8,
                        }
                    ],
                    "notes": [],
                }
                Path(output_paths[chunk["chunkId"]]).write_text(
                    json.dumps(result, ensure_ascii=False),
                    encoding="utf-8",
                )
                results.append(result)
            return results

        def fragment_value(envelope: dict) -> dict:
            records = []
            evidence = []
            for chunk in envelope["focusedChunks"]:
                sheet_value = chunk["sheet"]
                sheet = (
                    sheet_value["title"]
                    if isinstance(sheet_value, dict)
                    else sheet_value
                )
                evidence.append(
                    {
                        "sheet": sheet,
                        "range": chunk["primaryRange"],
                        "role": "SOURCE",
                    }
                )
            allowed = [
                *envelope["ownedSourceCellKeys"],
                *envelope["sharedAnchorCellKeys"],
            ]
            for study in envelope["registry"]["studies"]:
                identity_key = next(
                    (
                        key
                        for key in study["anchorEvidenceCellKeys"]
                        if key in allowed
                    ),
                    envelope["ownedSourceCellKeys"][0],
                )
                logical_id = study["logicalStudyId"]
                record_id = staged_v2.stable_record_id(
                    revision_uid=envelope["source"]["revisionUid"],
                    logical_study_id=logical_id,
                    record_type="STUDY_PATCH",
                    identity_cell_keys=[identity_key],
                    exact_source_label="Fragment Study",
                )
                records.append(
                    {
                        "recordType": "STUDY_PATCH",
                        "recordId": record_id,
                        "logicalStudyId": logical_id,
                        "identityCellKeys": [identity_key],
                        "exactSourceLabel": "Fragment Study",
                        "payload": {
                            "title": "Fragment Study",
                            "designType": "DESCRIPTIVE",
                            "summary": (
                                "Source-registered staged fragment."
                            ),
                        },
                        "evidence": evidence,
                    }
                )
            record_ids = [record["recordId"] for record in records]
            return {
                "schemaVersion": "study-draft-fragment-v2",
                "source": {
                    **envelope["source"],
                    "contentComplete": False,
                },
                "planId": envelope["planId"],
                "partId": envelope["partId"],
                "inputEnvelopeSha256": envelope[
                    "inputEnvelopeSha256"
                ],
                "records": records,
                "coverageDispositions": [
                    {
                        "sourceCellKey": key,
                        "disposition": "RECORD_EVIDENCE",
                        "recordIds": record_ids,
                        "reason": "Exact staged Study evidence.",
                    }
                    for key in envelope["ownedSourceCellKeys"]
                ],
            }

        first_attempt_part_ids: list[str] = []
        first_attempt_threads: set[int] = set()
        call_lock = threading.Lock()
        start_barrier = threading.Barrier(3)

        def failing_fragment(**kwargs: object) -> dict:
            envelope = kwargs["envelope"]
            kwargs["ai_call_observer"]()
            with call_lock:
                first_attempt_part_ids.append(envelope["partId"])
                first_attempt_threads.add(threading.get_ident())
                ordinal = len(first_attempt_part_ids)
            if ordinal <= 3:
                start_barrier.wait(timeout=5)
            if envelope["focusedChunks"][0]["primaryRange"] == "C2":
                raise RuntimeError("fixture fragment failure")
            return fragment_value(envelope)

        ingest_kwargs = {
            "database_path": self.database,
            "source_path": source,
            "artifact_root": self.artifacts,
            "dataset": "Fixture",
            "locator_batch_runner": fake_locator_batch,
            "locator_workers": 1,
            "max_cells": 1,
            "draft_monolithic_max_bytes": 1,
            "draft_fragment_max_chunks": 1,
            "draft_fragment_max_cells": 1,
            "draft_fragment_max_bytes": 100_000,
            "draft_fragment_workers": 3,
        }
        with self.assertRaisesRegex(
            RuntimeError,
            "fixture fragment failure",
        ):
            workflow.ingest_workbook(
                **ingest_kwargs,
                fragment_runner=failing_fragment,
            )

        run_directories = [
            path
            for path in self.artifacts.iterdir()
            if path.is_dir()
        ]
        self.assertEqual(1, len(run_directories))
        run_directory = run_directories[0]
        plan = json.loads(
            (run_directory / "study-draft-plan.json").read_text(
                encoding="utf-8"
            )
        )
        self.assertGreater(len(plan["parts"]), 1)
        self.assertTrue(
            all(
                part["promptBytes"]
                <= plan["limits"]["maxPromptBytes"]
                for part in plan["parts"]
            )
        )
        self.assertEqual(
            min(
                ingest_kwargs["draft_fragment_workers"],
                len(plan["parts"]),
            ),
            len(first_attempt_part_ids),
        )
        self.assertLess(
            len(first_attempt_part_ids),
            len(plan["parts"]),
        )
        self.assertGreaterEqual(len(first_attempt_threads), 3)
        self.assertFalse(
            (run_directory / "canonical-study-manifest.json").exists()
        )
        completed_part_count = len(
            list(
                (run_directory / "draft-parts-v2").glob(
                    "*.provenance.json"
                )
            )
        )
        self.assertEqual(
            len(first_attempt_part_ids) - 1,
            completed_part_count,
        )
        with closing(sqlite3.connect(self.database)) as connection:
            self.assertEqual(
                0,
                connection.execute(
                    "SELECT COUNT(*) FROM workbook_analyses"
                ).fetchone()[0],
            )

        resumed_part_ids: list[str] = []

        def resumed_fragment(**kwargs: object) -> dict:
            envelope = kwargs["envelope"]
            self.assertTrue(callable(kwargs["ai_call_observer"]))
            resumed_part_ids.append(envelope["partId"])
            return fragment_value(envelope)

        result = workflow.ingest_workbook(
            **ingest_kwargs,
            fragment_runner=resumed_fragment,
        )

        self.assertEqual(
            len(plan["parts"]) - completed_part_count,
            len(resumed_part_ids),
        )
        self.assertEqual(1, locator_calls)
        registry = json.loads(
            (run_directory / "study-registry-v2.json").read_text(
                encoding="utf-8"
            )
        )
        self.assertEqual(len(registry["studies"]), result["studies"])
        manifest = json.loads(
            Path(result["manifestPath"]).read_text(encoding="utf-8")
        )
        self.assertIs(True, manifest["source"]["contentComplete"])
        self.assertEqual(
            len(registry["studies"]),
            len({study["key"] for study in manifest["studies"]}),
        )
        self.assertTrue(
            all(
                not study["comparisons"]
                for study in manifest["studies"]
            )
        )
        journal = json.loads(
            Path(result["journalPath"]).read_text(encoding="utf-8")
        )
        draft_result = journal["stages"]["DRAFT"]["result"]
        self.assertTrue(draft_result["staged"])
        self.assertEqual(len(plan["parts"]), draft_result["partCount"])
        self.assertEqual(
            completed_part_count,
            draft_result["partsReused"],
        )
        self.assertEqual(
            len(plan["parts"]) - completed_part_count,
            draft_result["partsExecuted"],
        )
        self.assertFalse(draft_result["aiExecuted"])
        third = workflow.ingest_workbook(
            **ingest_kwargs,
            fragment_runner=lambda **_kwargs: self.fail(
                "verified final staged artifacts must resume without AI"
            ),
        )
        third_draft = json.loads(
            Path(third["journalPath"]).read_text(encoding="utf-8")
        )["stages"]["DRAFT"]["result"]
        self.assertTrue(third_draft["artifactReused"])
        self.assertFalse(third_draft["aiExecuted"])

    def test_incomplete_quantitative_draft_fails_before_analysis_import(
        self,
    ) -> None:
        source = self.numeric_source()

        def fake_locator_batch(**kwargs: object) -> list[dict]:
            source_identity = kwargs["source"]
            output_paths = kwargs["output_paths"]
            results = []
            for chunk in kwargs["chunks"]:
                result = {
                    "schemaVersion": "semantic-locator-v1",
                    "promptVersion": "semantic-locator-prompt-v1",
                    "revisionUid": source_identity["revisionUid"],
                    "contentSha256": source_identity["contentSha256"],
                    "chunkId": chunk["chunkId"],
                    "status": "CANDIDATES",
                    "candidates": [
                        {
                            "key": "gauss-panel",
                            "title": "Gauss panel",
                            "summary": "A numeric source panel.",
                            "designHint": "DESCRIPTIVE",
                            "contexts": [],
                            "changedFactors": [],
                            "outcomes": ["Gauss"],
                            "comparisonHints": [],
                            "evidence": [
                                {
                                    "sheet": "Gauss Data",
                                    "range": chunk["primaryRange"],
                                    "role": "CANDIDATE_REGION",
                                }
                            ],
                            "limitations": [],
                            "confidence": 0.8,
                        }
                    ],
                    "notes": [],
                }
                Path(output_paths[chunk["chunkId"]]).write_text(
                    json.dumps(result),
                    encoding="utf-8",
                )
                results.append(result)
            return results

        def incomplete_draft(**kwargs: object) -> dict:
            source_identity = kwargs["source"]
            result = {
                "schemaVersion": "canonical-study-manifest-v1",
                "source": {
                    **source_identity,
                    "contentComplete": True,
                },
                "workbookAnalysis": {
                    "key": "incomplete-gauss-analysis",
                    "title": "Incomplete Gauss analysis",
                    "summary": "The numeric panel was not represented.",
                    "status": "NEEDS_REVIEW",
                    "verificationStatus": "NEEDS_REVIEW",
                    "limitations": [],
                    "evidence": [
                        {
                            "sheet": "Gauss Data",
                            "range": "A1:B3",
                            "role": "SOURCE",
                            "sourceText": "Gauss raw measurements",
                            "note": "",
                        }
                    ],
                },
                "studies": [
                    {
                        "key": "empty-gauss-study",
                        "title": "Empty Gauss Study",
                        "designType": "DESCRIPTIVE",
                        "verificationStatus": "NEEDS_REVIEW",
                        "comparabilityStatus": "UNASSESSED",
                        "confoundingStatus": "UNASSESSED",
                        "evidence": [],
                        "contexts": [],
                        "factors": [],
                        "arms": [],
                        "outcomes": [],
                        "measurementSeries": [],
                        "comparisons": [],
                        "conclusions": [],
                        "limitations": [],
                    }
                ],
            }
            kwargs["additional_validator"](result)
            Path(kwargs["output_path"]).write_text(
                json.dumps(result),
                encoding="utf-8",
            )
            return result

        with self.assertRaisesRegex(
            content_coverage.ContentCoverageError,
            "Gauss Data!B2",
        ):
            workflow.ingest_workbook(
                database_path=self.database,
                source_path=source,
                artifact_root=self.artifacts,
                dataset="Fixture",
                locator_batch_runner=fake_locator_batch,
                draft_runner=incomplete_draft,
                locator_workers=1,
            )

        with closing(sqlite3.connect(self.database)) as connection:
            self.assertEqual(
                0,
                connection.execute(
                    "SELECT COUNT(*) FROM workbook_analyses"
                ).fetchone()[0],
            )

    def test_empty_workbook_completes_without_ai_and_preserves_terminal_status(self) -> None:
        source = self.empty_source()

        def unexpected_ai(**_: object):
            raise AssertionError("AI must not run for a terminal workbook.")

        result = workflow.ingest_workbook(
            database_path=self.database,
            source_path=source,
            artifact_root=self.artifacts,
            dataset="Fixture",
            locator_batch_runner=unexpected_ai,
            draft_runner=unexpected_ai,
        )

        self.assertEqual("EXCLUDED", result["status"])
        self.assertEqual("EMPTY_WORKBOOK", result["workbookStatus"])
        self.assertEqual(0, result["studies"])
        with closing(sqlite3.connect(self.database)) as connection:
            row = connection.execute(
                """
                SELECT capture_status, source_content_status, is_current
                FROM source_revisions
                WHERE revision_uid=?
                """,
                (result["revisionUid"],),
            ).fetchone()
        self.assertEqual(("CAPTURED", "EMPTY_WORKBOOK", 1), row)
        journal = json.loads(
            Path(result["journalPath"]).read_text(encoding="utf-8")
        )
        self.assertEqual(0, journal["stages"]["LOCATOR"]["result"]["aiCalls"])
        self.assertFalse(journal["source"]["imagesAnalyzed"])

        with closing(sqlite3.connect(self.database)) as connection:
            connection.execute(
                """
                UPDATE source_revisions
                SET source_content_status='CAPTURED'
                WHERE revision_uid=?
                """,
                (result["revisionUid"],),
            )
            workflow.ensure_knowledge_schema(connection, workflow.utc_now_iso)
            restored = connection.execute(
                """
                SELECT capture_status, source_content_status
                FROM source_revisions
                WHERE revision_uid=?
                """,
                (result["revisionUid"],),
            ).fetchone()
        self.assertEqual(("CAPTURED", "EMPTY_WORKBOOK"), restored)


if __name__ == "__main__":
    unittest.main()
