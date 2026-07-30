from __future__ import annotations

import argparse
import hashlib
import importlib.util
import json
import sqlite3
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


SCANNER_PATH = Path(__file__).parents[1] / "inference_data_ai_structure_scan.py"
SPEC = importlib.util.spec_from_file_location("inference_data_ai_structure_scan", SCANNER_PATH)
assert SPEC is not None and SPEC.loader is not None
scanner = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = scanner
SPEC.loader.exec_module(scanner)


def checksum(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


class StructureScanTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.service = self.root / "service"
        self.input = self.root / "input"
        self.service.mkdir()
        self.input.mkdir()

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def create_workbook(self, name: str, headers: list[str], *, hidden_summary: bool = False) -> Path:
        path = self.input / name
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Report"
        worksheet.merge_cells("A1:D1")
        worksheet["A1"] = "Source report title"
        for column, header in enumerate(headers, start=1):
            worksheet.cell(2, column, header)
        worksheet.append([67, 67, 0, 0])
        if hidden_summary:
            hidden = workbook.create_sheet("Hidden summary")
            hidden.sheet_state = "hidden"
            hidden.append(["Sample", "Average", "Max", "Min"])
        workbook.save(path)
        workbook.close()
        return path

    def run_batch(self, **overrides: object) -> int:
        options: dict[str, object] = {
            "service_dir": str(self.service),
            "batch_folder": str(self.input),
            "resume_batch": None,
            "batch_id": "fixture-batch",
            "pilot": 0,
            "limit": 0,
            "retry_failed": False,
        }
        options.update(overrides)
        return scanner.run(argparse.Namespace(**options))

    def state_rows(self) -> dict[str, str]:
        path = self.service / "outputs" / "batches" / "fixture-batch" / "state.sqlite"
        connection = sqlite3.connect(path)
        try:
            return {path: status for path, status in connection.execute("SELECT relative_path, status FROM items")}
        finally:
            connection.close()

    def test_scans_merged_defect_layout_without_modifying_source_or_main_db(self) -> None:
        source = self.create_workbook("BRS_defect.xlsx", ["Input", "OK", "Total NG", "NG rate"], hidden_summary=True)
        before = checksum(source)

        self.assertEqual(0, self.run_batch())

        batch = self.service / "outputs" / "batches" / "fixture-batch"
        self.assertEqual(before, checksum(source))
        self.assertFalse((self.service / "outputs" / "universal-grid").exists())
        self.assertTrue((batch / "state.sqlite").is_file())
        self.assertTrue((batch / "classification.csv").is_file())
        self.assertTrue((batch / "classification.html").is_file())
        self.assertTrue((batch / "summary.json").is_file())
        self.assertEqual({"BRS_defect.xlsx": "SCANNED"}, self.state_rows())

        result_file = next((batch / "results").glob("*.json"))
        result = json.loads(result_file.read_text(encoding="utf-8"))
        self.assertTrue(result["usesCom"] is False)
        self.assertIn("DEFECT_ACCOUNTING_LAYOUT_CANDIDATE", result["structuralTypes"])
        report = next(sheet for sheet in result["sheets"] if sheet["sheetName"] == "Report")
        self.assertEqual(1, report["mergeCount"])
        hidden = next(sheet for sheet in result["sheets"] if sheet["sheetName"] == "Hidden summary")
        self.assertEqual("hidden", hidden["sheetState"])
        self.assertIn("RAW_MEASUREMENT_SUMMARY_LAYOUT_CANDIDATE", result["structuralTypes"])

    def test_pilot_defers_remaining_file_then_resume_scans_it(self) -> None:
        self.create_workbook("BRS_one.xlsx", ["Input", "OK", "Total NG", "NG rate"])
        self.create_workbook("TIU_two.xlsx", ["Sample", "Average", "Max", "Min"])

        self.assertEqual(0, self.run_batch(pilot=1))
        first_state = self.state_rows()
        self.assertEqual(1, list(first_state.values()).count("SCANNED"))
        self.assertEqual(1, list(first_state.values()).count("DEFERRED"))

        self.assertEqual(
            0,
            self.run_batch(
                batch_folder=None,
                resume_batch="fixture-batch",
                batch_id=None,
                pilot=0,
            ),
        )
        self.assertEqual({"BRS_one.xlsx": "SCANNED", "TIU_two.xlsx": "SCANNED"}, self.state_rows())

    def test_quarantines_invalid_xlsx_and_records_unsupported_binary_and_html(self) -> None:
        (self.input / "bad.xlsx").write_bytes(b"not a zip package")
        (self.input / "legacy.xls").write_bytes(b"old binary")
        (self.input / "review.html").write_text("<html>untrusted</html>", encoding="utf-8")

        self.assertEqual(0, self.run_batch())

        self.assertEqual(
            {"bad.xlsx": "QUARANTINED", "legacy.xls": "UNSUPPORTED", "review.html": "NON_WORKBOOK"},
            self.state_rows(),
        )
        batch = self.service / "outputs" / "batches" / "fixture-batch"
        failures = (batch / "failures.csv").read_text(encoding="utf-8-sig")
        self.assertIn("bad.xlsx", failures)
        self.assertNotIn("review.html", failures)

    def test_allows_wpf_precreated_logs_only_directory_but_rejects_initialized_artifacts(self) -> None:
        self.create_workbook("BRS_fixture.xlsx", ["Input", "OK", "Total NG", "NG rate"])
        batch = self.service / "outputs" / "batches" / "fixture-batch"
        logs = batch / "logs"
        logs.mkdir(parents=True)
        (logs / "wpf.log").write_text("launcher log", encoding="utf-8")

        self.assertEqual(0, self.run_batch())
        self.assertTrue((batch / "batch.json").is_file())
        self.assertTrue((logs / "wpf.log").is_file())

        collision = self.service / "outputs" / "batches" / "collision"
        collision.mkdir(parents=True)
        (collision / "state.sqlite").write_bytes(b"not a state database")
        with self.assertRaisesRegex(ValueError, "initialized artifacts"):
            self.run_batch(batch_id="collision")


if __name__ == "__main__":
    unittest.main()
