from __future__ import annotations

import argparse
import hashlib
import importlib.util
import sqlite3
import sys
import tempfile
import unittest
from contextlib import contextmanager
from datetime import date
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).parents[1]
SCANNER_PATH = ROOT / "inference_data_ai_structure_scan.py"
CAPTURE_PATH = ROOT / "inference_data_ai_numeric_capture.py"
REVIEW_PATH = ROOT / "inference_data_ai_numeric_review.py"
RENDERER_PATH = ROOT / "inference_data_ai_numeric_renderer.py"


def load_module(name: str, path: Path):
    specification = importlib.util.spec_from_file_location(name, path)
    assert specification is not None and specification.loader is not None
    module = importlib.util.module_from_spec(specification)
    sys.modules[name] = module
    specification.loader.exec_module(module)
    return module


scanner = load_module("inference_data_ai_structure_scan", SCANNER_PATH)
capture = load_module("inference_data_ai_numeric_capture", CAPTURE_PATH)
review = load_module("inference_data_ai_numeric_review", REVIEW_PATH)
renderer = load_module("inference_data_ai_numeric_renderer", RENDERER_PATH)


def checksum(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


class NumericCaptureTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.service = self.root / "service"
        self.input = self.root / "input"
        self.service.mkdir()
        self.input.mkdir()
        self.source = self.create_workbook()
        self.create_structure_batch()

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def create_workbook(self) -> Path:
        path = self.input / "BRS_rate.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Rate"
        sheet.merge_cells("A1:F1")
        sheet["A1"] = "This is a narrative report title and is not a table label."
        for column, value in enumerate(["Date", "Type", "Input", "OK", "Total NG", "NG rate"], start=1):
            sheet.cell(3, column, value)
        sheet.append([date(2025, 12, 17), "Normal", 110, 107, 3, 3 / 110])
        sheet.append([date(2025, 12, 17), "Test", 62, 61, 1, 1 / 62])
        sheet.append([date(2025, 12, 18), "Test", 100, 98, 2, 2 / 100])
        sheet["G5"] = "=E5/C5"
        hidden = workbook.create_sheet("Hidden")
        hidden.sheet_state = "hidden"
        hidden.append(["Sample", "Average", "Max", "Min"])
        hidden.append([50, 435.62, 444, 420])
        workbook.save(path)
        workbook.close()
        return path

    def create_structure_batch(self) -> None:
        result = scanner.run(
            argparse.Namespace(
                service_dir=str(self.service),
                batch_folder=str(self.input),
                resume_batch=None,
                batch_id="fixture-batch",
                pilot=0,
                limit=0,
                retry_failed=False,
            )
        )
        self.assertEqual(0, result)

    def run_capture(self, **overrides: object) -> int:
        options: dict[str, object] = {
            "service_dir": str(self.service),
            "structure_batch": "fixture-batch",
            "limit": 0,
            "progress_every": 0,
            "force": False,
        }
        options.update(overrides)
        return capture.run(argparse.Namespace(**options))

    @contextmanager
    def connection(self):
        connection = sqlite3.connect(self.service / "outputs" / "batches" / "fixture-batch" / "numeric-capture.sqlite")
        connection.row_factory = sqlite3.Row
        try:
            yield connection
        finally:
            connection.close()

    def test_captures_numeric_facts_dates_formula_and_structural_labels_without_com(self) -> None:
        before = checksum(self.source)

        self.assertEqual(0, self.run_capture())

        self.assertEqual(before, checksum(self.source))
        self.assertFalse((self.service / "outputs" / "universal-grid").exists())
        batch = self.service / "outputs" / "batches" / "fixture-batch"
        self.assertTrue((batch / "numeric-capture.sqlite").is_file())
        self.assertTrue((batch / "numeric-capture-summary.json").is_file())
        self.assertTrue((batch / "numeric-capture.html").is_file())
        with self.connection() as connection:
            workbook = connection.execute("SELECT * FROM capture_workbooks").fetchone()
            self.assertEqual("CAPTURED", workbook["capture_status"])
            self.assertEqual(2, workbook["sheet_count_captured"])
            self.assertGreaterEqual(workbook["numeric_cell_count"], 12)
            self.assertEqual(3, workbook["date_cell_count"])
            self.assertEqual(1, workbook["formula_count"])
            self.assertEqual(1, connection.execute("SELECT COUNT(*) FROM captured_merge_ranges").fetchone()[0])
            self.assertEqual(1, connection.execute("SELECT COUNT(*) FROM formula_cells").fetchone()[0])
            table = connection.execute(
                "SELECT * FROM numeric_table_candidates WHERE candidate_type='DEFECT_RATE_NUMERIC_TABLE'"
            ).fetchone()
            self.assertIsNotNone(table)
            labels = {
                row[0]
                for row in connection.execute(
                    "SELECT label_text FROM numeric_table_labels WHERE table_id=?", (table["table_id"],)
                )
            }
            self.assertTrue({"Input", "OK", "Total NG", "NG rate", "Normal", "Test"}.issubset(labels))
            self.assertNotIn("This is a narrative report title and is not a table label.", labels)
            dates = {row[0] for row in connection.execute("SELECT date_value FROM date_cells")}
            self.assertEqual({"2025-12-17", "2025-12-18"}, dates)

    def test_force_recapture_is_idempotent_and_numeric_text_percent_is_normalized(self) -> None:
        self.assertEqual(("0.42", 0.42, "TEXT_PERCENT"), capture.numeric_value("42%"))
        self.assertEqual(("-1234.5", -1234.5, "TEXT_NUMBER"), capture.numeric_value("(1,234.5)"))
        self.assertEqual(0, self.run_capture())
        with self.connection() as connection:
            initial = connection.execute("SELECT COUNT(*) FROM numeric_cells").fetchone()[0]
        self.assertEqual(0, self.run_capture(force=True))
        with self.connection() as connection:
            self.assertEqual(initial, connection.execute("SELECT COUNT(*) FROM numeric_cells").fetchone()[0])
            self.assertEqual(2, connection.execute("SELECT attempts FROM capture_workbooks").fetchone()[0])

    def test_separates_side_by_side_numeric_tables_before_header_mapping(self) -> None:
        regions = capture.numeric_table_regions(
            [(row, column) for row in (1, 2, 3) for column in (1, 2, 3, 9, 10, 11)]
        )
        self.assertEqual([(1, 3, 1, 3, 9), (1, 3, 9, 11, 9)], regions)

    def test_changed_source_is_not_captured_against_old_snapshot(self) -> None:
        self.source.write_bytes(self.source.read_bytes() + b"changed")
        self.assertEqual(0, self.run_capture())
        with self.connection() as connection:
            row = connection.execute("SELECT capture_status, error_text FROM capture_workbooks").fetchone()
            self.assertEqual("CHANGED", row["capture_status"])
            self.assertIn("differs", row["error_text"])
            self.assertEqual(0, connection.execute("SELECT COUNT(*) FROM numeric_cells").fetchone()[0])

    def test_extracts_only_same_date_explicit_test_normal_defect_comparison(self) -> None:
        self.assertEqual(0, self.run_capture())
        self.assertEqual(
            0,
            review.run(argparse.Namespace(service_dir=str(self.service), structure_batch="fixture-batch")),
        )
        with self.connection() as connection:
            facts = list(
                connection.execute(
                    """
                    SELECT condition_role, measurement_date, computed_ng_rate, fact_status
                    FROM numeric_review_facts ORDER BY measurement_date, condition_role
                    """
                )
            )
            self.assertEqual(
                [("NORMAL", "2025-12-17", "OBSERVED"), ("TEST", "2025-12-17", "OBSERVED"), ("TEST", "2025-12-18", "OBSERVED")],
                [(row["condition_role"], row["measurement_date"], row["fact_status"]) for row in facts],
            )
            comparisons = list(
                connection.execute(
                    """
                    SELECT measurement_date, test_ng_rate, normal_ng_rate, absolute_delta, comparison_status
                    FROM test_normal_comparisons ORDER BY measurement_date
                    """
                )
            )
            self.assertEqual(["VALID", "NO_SAME_DAY_NORMAL"], [row["comparison_status"] for row in comparisons])
            self.assertAlmostEqual(1 / 62, comparisons[0]["test_ng_rate"])
            self.assertAlmostEqual(3 / 110, comparisons[0]["normal_ng_rate"])
            self.assertAlmostEqual(1 / 62 - 3 / 110, comparisons[0]["absolute_delta"])
            self.assertIsNone(comparisons[1]["normal_ng_rate"])

    def test_renders_one_numeric_only_html_per_captured_workbook(self) -> None:
        self.assertEqual(0, self.run_capture())
        self.assertEqual(0, review.run(argparse.Namespace(service_dir=str(self.service), structure_batch="fixture-batch")))
        self.assertEqual(0, renderer.run(argparse.Namespace(service_dir=str(self.service), structure_batch="fixture-batch")))
        batch = self.service / "outputs" / "batches" / "fixture-batch"
        self.assertTrue((batch / "numeric-report-index.html").is_file())
        reports = list((batch / "numeric-reports").glob("*.html"))
        self.assertEqual(1, len(reports))
        page = reports[0].read_text(encoding="utf-8")
        self.assertIn("Test", page)
        self.assertIn("Normal", page)
        self.assertIn("불량률", page)
        self.assertNotIn("A1:", page)
        self.assertNotIn("numeric-capture.sqlite", page)
        with self.connection() as connection:
            self.assertEqual(1, connection.execute("SELECT COUNT(*) FROM numeric_html_reports").fetchone()[0])


if __name__ == "__main__":
    unittest.main()
