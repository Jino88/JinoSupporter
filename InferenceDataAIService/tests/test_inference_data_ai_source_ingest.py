from __future__ import annotations

import importlib.util
import json
import sqlite3
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side


MODULE_PATH = Path(__file__).parents[1] / "inference_data_ai_source_ingest.py"
SPEC = importlib.util.spec_from_file_location("inference_data_ai_source_ingest", MODULE_PATH)
assert SPEC is not None and SPEC.loader is not None
ingest = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(ingest)


class SourceIngestExtractionTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)

    def tearDown(self) -> None:
        self.temp.cleanup()

    def fixture_path(self, name: str = "representative.xlsx") -> Path:
        path = self.root / name
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data"
        sheet["A1"] = 0
        sheet["A1"].font = Font(name="Arial", bold=True, color="FF112233")
        sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
        sheet["A1"].border = Border(left=Side(style="thin", color="FF000000"))
        sheet["A1"].alignment = Alignment(horizontal="center", wrap_text=True)
        sheet["A1"].protection = Protection(locked=False, hidden=True)
        sheet["B1"] = 0.26
        sheet["B1"].number_format = "0.0%"
        sheet["C1"] = "Result"
        sheet["A2"] = "Merged heading"
        sheet.merge_cells("A2:B2")
        sheet["A3"] = "Sample"
        sheet["B3"] = 4
        sheet["C3"] = "=A1+B3"
        sheet["C4"].border = Border(bottom=Side(style="thin", color="FF000000"))
        sheet.row_dimensions[2].height = 31.5
        sheet.row_dimensions[2].hidden = True
        sheet.column_dimensions["B"].width = 22.25
        sheet.column_dimensions["B"].hidden = True
        sheet.freeze_panes = "B2"
        sheet.auto_filter.ref = "A1:C3"

        hidden = workbook.create_sheet("Hidden")
        hidden.sheet_state = "hidden"
        hidden["A1"] = "not a table"
        workbook.save(path)
        workbook.close()
        return path

    def test_sparse_capture_preserves_source_fidelity_and_ignores_images(self) -> None:
        path = self.fixture_path()
        before = path.stat()
        payload = ingest.extract_workbook(path)
        after = path.stat()

        self.assertEqual(ingest.CAPTURE_SCHEMA_VERSION, payload["schemaVersion"])
        self.assertEqual(ingest.CAPTURE_CONTRACT, payload["captureContract"])
        self.assertEqual(ingest.sha256_file(path), payload["source"]["contentSha256"])
        self.assertEqual((before.st_size, before.st_mtime_ns), (after.st_size, after.st_mtime_ns))
        self.assertEqual("IGNORED", payload["extractor"]["imageHandling"])
        self.assertFalse(
            any(
                key.casefold() == "images"
                for sheet in payload["workbook"]["sheets"]
                for key in sheet
            )
        )

        workbook = payload["workbook"]
        self.assertEqual("CAPTURED", workbook["status"])
        self.assertFalse(workbook["isTrulyEmpty"])
        self.assertEqual(2, workbook["sheetCount"])
        self.assertEqual(1, workbook["tabularSheetCount"])

        data = workbook["sheets"][0]
        hidden = workbook["sheets"][1]
        self.assertEqual("visible", data["sheetState"])
        self.assertEqual("hidden", hidden["sheetState"])
        self.assertEqual("NO_TABULAR_EVIDENCE", hidden["status"])
        self.assertEqual("B2", data["freezePanes"])
        self.assertEqual("A1:C3", data["autoFilter"])
        self.assertEqual("A1:C4", data["usedBounds"]["address"])
        self.assertEqual("A1:C4", data["contentBounds"]["address"])

        cells = {cell["coordinate"]: cell for cell in data["cells"]}
        self.assertEqual(0, cells["A1"]["rawValue"])
        self.assertEqual(0, cells["A1"]["displayValue"])
        self.assertTrue(cells["A1"]["style"]["font"]["bold"])
        self.assertEqual("solid", cells["A1"]["style"]["fill"]["type"])
        self.assertEqual("thin", cells["A1"]["style"]["border"]["left"]["style"])
        self.assertEqual("center", cells["A1"]["style"]["alignment"]["horizontal"])
        self.assertFalse(cells["A1"]["style"]["protection"]["locked"])
        self.assertEqual(0.26, cells["B1"]["rawValue"])
        self.assertEqual("0.0%", cells["B1"]["numberFormat"])
        self.assertEqual("A2:B2", cells["A2"]["mergeRange"])
        self.assertEqual("anchor", cells["A2"]["mergeRole"])
        self.assertIn("B2", cells)
        self.assertIsNone(cells["B2"]["rawValue"])
        self.assertEqual("covered", cells["B2"]["mergeRole"])

        formula = cells["C3"]
        self.assertEqual("=A1+B3", formula["formula"])
        self.assertIsNone(formula["rawValue"])
        self.assertIsNone(formula["cachedValue"])
        self.assertIsNone(formula["displayValue"])
        self.assertEqual("f", formula["dataType"])
        self.assertIn("C4", cells)
        self.assertIsNone(cells["C4"]["rawValue"])
        self.assertEqual("thin", cells["C4"]["style"]["border"]["bottom"]["style"])
        self.assertEqual(1, data["structuralCellCount"])

        self.assertEqual(
            [{"row": 2, "height": 31.5, "hidden": True}],
            data["rowDimensions"],
        )
        self.assertEqual(22.25, data["columnDimensions"][0]["width"])
        self.assertTrue(data["columnDimensions"][0]["hidden"])
        self.assertEqual("A2:B2", data["mergedRanges"][0]["address"])

    def test_empty_workbook_and_non_tabular_workbook_are_distinct(self) -> None:
        empty_path = self.root / "empty.xlsx"
        workbook = Workbook()
        workbook.save(empty_path)
        workbook.close()

        empty = ingest.extract_workbook(empty_path)
        self.assertEqual("EMPTY_WORKBOOK", empty["workbook"]["status"])
        self.assertTrue(empty["workbook"]["isTrulyEmpty"])
        self.assertEqual("EMPTY", empty["workbook"]["sheets"][0]["status"])
        self.assertIsNone(empty["workbook"]["sheets"][0]["usedBounds"])
        self.assertEqual([], empty["workbook"]["sheets"][0]["cells"])

        narrative_path = self.root / "narrative.xlsx"
        workbook = Workbook()
        workbook.active["D7"] = "Narrative only"
        workbook.save(narrative_path)
        workbook.close()

        narrative = ingest.extract_workbook(narrative_path)
        self.assertEqual("NO_TABULAR_EVIDENCE", narrative["workbook"]["status"])
        sheet = narrative["workbook"]["sheets"][0]
        self.assertFalse(sheet["isTrulyEmpty"])
        self.assertFalse(sheet["hasTabularEvidence"])
        self.assertEqual("D7:D7", sheet["usedBounds"]["address"])
        self.assertEqual("D7:D7", sheet["contentBounds"]["address"])

    def test_json_serialization_is_deterministic_and_round_trips(self) -> None:
        payload = ingest.extract_workbook(self.fixture_path())
        first = ingest.capture_json_bytes(payload)
        second = ingest.capture_json_bytes(payload)
        self.assertEqual(first, second)

        output = ingest.write_capture_json(payload, self.root / "capture.json")
        self.assertEqual(first, output.read_bytes())
        self.assertEqual(payload, ingest.read_capture_json(output))

    def test_rejects_non_xlsx_sources(self) -> None:
        source = self.root / "not-xlsx.xlsm"
        source.write_bytes(b"not a workbook")
        with self.assertRaises(ingest.UnsupportedSourceError):
            ingest.extract_workbook(source)


class SourceIngestDatabaseTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.source = self.root / "revisioned.xlsx"
        self.connection = sqlite3.connect(":memory:")
        self.connection.row_factory = sqlite3.Row

    def tearDown(self) -> None:
        self.connection.close()
        self.temp.cleanup()

    def save_source(self, value: int) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Group"
        sheet["B1"] = "Value"
        sheet["A2"] = "A"
        sheet["B2"] = value
        workbook.save(self.source)
        workbook.close()

    def test_import_skips_current_sha_and_marks_changed_revision_stale(self) -> None:
        self.save_source(1)
        first_payload = ingest.extract_workbook(self.source)
        first = ingest.import_capture(
            self.connection,
            first_payload,
            captured_at="2026-07-17T00:00:00Z",
        )
        self.assertEqual("IMPORTED", first["action"])
        first_bytes = self.source.read_bytes()

        skipped = ingest.import_capture(
            self.connection,
            first_payload,
            captured_at="2026-07-17T00:01:00Z",
        )
        self.assertEqual("SKIPPED", skipped["action"])
        self.assertEqual(first["revisionId"], skipped["revisionId"])

        self.save_source(2)
        second_payload = ingest.extract_workbook(self.source)
        second = ingest.import_capture(
            self.connection,
            second_payload,
            captured_at="2026-07-17T00:02:00Z",
        )
        self.assertEqual("IMPORTED", second["action"])
        self.assertNotEqual(first["contentSha256"], second["contentSha256"])

        revisions = self.connection.execute(
            """
            SELECT revision_id, capture_status, is_current, stale_at
            FROM capture_v2_revisions
            ORDER BY revision_id
            """
        ).fetchall()
        self.assertEqual(2, len(revisions))
        self.assertEqual("STALE", revisions[0]["capture_status"])
        self.assertEqual(0, revisions[0]["is_current"])
        self.assertEqual("2026-07-17T00:02:00Z", revisions[0]["stale_at"])
        self.assertEqual("CAPTURED", revisions[1]["capture_status"])
        self.assertEqual(1, revisions[1]["is_current"])

        cell = self.connection.execute(
            """
            SELECT c.raw_value_json, c.number_format, c.style_json
            FROM capture_v2_cells c
            JOIN capture_v2_sheets s ON s.sheet_id=c.sheet_id
            WHERE s.revision_id=? AND c.coordinate='B2'
            """,
            (second["revisionId"],),
        ).fetchone()
        self.assertEqual(2, json.loads(cell["raw_value_json"]))
        self.assertEqual("General", cell["number_format"])
        self.assertIsInstance(json.loads(cell["style_json"]), dict)

        # Reverting bytes reactivates the complete known source revision.
        self.source.write_bytes(first_bytes)
        reactivated_payload = ingest.extract_workbook(self.source)
        reactivated = ingest.import_capture(
            self.connection,
            reactivated_payload,
            captured_at="2026-07-17T00:03:00Z",
        )
        self.assertEqual("REACTIVATED", reactivated["action"])
        self.assertEqual(first["revisionId"], reactivated["revisionId"])
        current = self.connection.execute(
            "SELECT revision_id FROM capture_v2_revisions WHERE is_current=1"
        ).fetchone()
        self.assertEqual(first["revisionId"], current["revision_id"])

    def test_failed_import_rolls_back_without_staling_current_revision(self) -> None:
        self.save_source(1)
        payload = ingest.extract_workbook(self.source)
        imported = ingest.import_capture(self.connection, payload)

        invalid = json.loads(json.dumps(payload))
        invalid["source"]["contentSha256"] = "f" * 64
        invalid["workbook"]["sheets"][0]["cells"].append(
            {
                "row": 1,
                "column": 1,
                "coordinate": "A1",
                "rawValue": "duplicate",
                "formula": None,
                "cachedValue": None,
                "displayValue": "duplicate",
                "dataType": "s",
                "cachedDataType": None,
                "numberFormat": "General",
                "styleId": 0,
                "style": {},
                "mergeRange": None,
                "mergeRole": "none",
            }
        )
        with self.assertRaises(sqlite3.IntegrityError):
            ingest.import_capture(self.connection, invalid)

        current = self.connection.execute(
            """
            SELECT revision_id, capture_status, is_current
            FROM capture_v2_revisions
            WHERE is_current=1
            """
        ).fetchone()
        self.assertEqual(imported["revisionId"], current["revision_id"])
        self.assertEqual("CAPTURED", current["capture_status"])
        self.assertEqual(1, current["is_current"])
        self.assertEqual(
            1,
            self.connection.execute(
                "SELECT COUNT(*) FROM capture_v2_revisions"
            ).fetchone()[0],
        )

    def test_preinstalled_schema_preserves_caller_savepoint(self) -> None:
        self.save_source(1)
        payload = ingest.extract_workbook(self.source)
        ingest.ensure_capture_v2_schema(self.connection)
        self.connection.execute("SAVEPOINT caller_item")
        imported = ingest.import_capture(self.connection, payload)
        self.connection.execute("RELEASE SAVEPOINT caller_item")
        self.assertEqual("IMPORTED", imported["action"])

    def test_capture_revision_verifier_checks_counts_and_source_sha(self) -> None:
        self.save_source(1)
        payload = ingest.extract_workbook(self.source)
        imported = ingest.import_capture(self.connection, payload)
        result = ingest.verify_capture_revision(
            self.connection,
            imported["revisionId"],
            verify_source_sha256=True,
        )
        self.assertTrue(result["ok"], result)
        self.connection.execute(
            """
            UPDATE capture_v2_sheets
            SET captured_cell_count=captured_cell_count+1
            WHERE revision_id=?
            """,
            (imported["revisionId"],),
        )
        invalid = ingest.verify_capture_revision(
            self.connection,
            imported["revisionId"],
        )
        self.assertFalse(invalid["ok"])
        self.assertTrue(any("cell count mismatch" in error for error in invalid["errors"]))

    def test_sheet_count_metadata_can_be_reconciled_from_cells(
        self,
    ) -> None:
        self.save_source(1)
        payload = ingest.extract_workbook(self.source)
        imported = ingest.import_capture(self.connection, payload)
        self.connection.execute(
            """
            UPDATE capture_v2_sheets
            SET formula_cell_count=formula_cell_count+4,
                nonempty_cell_count=nonempty_cell_count+4,
                structural_cell_count=structural_cell_count-4
            WHERE revision_id=?
            """,
            (imported["revisionId"],),
        )
        invalid = ingest.verify_capture_revision(
            self.connection,
            imported["revisionId"],
        )
        self.assertEqual(
            1,
            sum(
                "formula count mismatch" in error
                for error in invalid["errors"]
            ),
        )

        repaired = ingest.reconcile_capture_sheet_counts(
            self.connection,
            imported["revisionId"],
        )
        verified = ingest.verify_capture_revision(
            self.connection,
            imported["revisionId"],
        )
        counts = self.connection.execute(
            """
            SELECT nonempty_cell_count, structural_cell_count,
                   captured_cell_count, formula_cell_count
            FROM capture_v2_sheets
            WHERE revision_id=? AND sheet_index=1
            """,
            (imported["revisionId"],),
        ).fetchone()

        self.assertEqual(1, repaired["repairedSheetCount"])
        self.assertTrue(verified["ok"], verified)
        self.assertEqual(
            int(counts["captured_cell_count"]),
            int(counts["nonempty_cell_count"])
            + int(counts["structural_cell_count"]),
        )


if __name__ == "__main__":
    unittest.main()
