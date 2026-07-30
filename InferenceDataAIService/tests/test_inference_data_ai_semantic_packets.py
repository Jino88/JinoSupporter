from __future__ import annotations

import importlib.util
import json
import sqlite3
import tempfile
import unittest
from collections import Counter
from pathlib import Path

from openpyxl import Workbook


SERVICE_DIR = Path(__file__).parents[1]


def load_module(name: str, file_name: str):
    specification = importlib.util.spec_from_file_location(name, SERVICE_DIR / file_name)
    assert specification is not None and specification.loader is not None
    module = importlib.util.module_from_spec(specification)
    specification.loader.exec_module(module)
    return module


capture = load_module("semantic_packet_capture_fixture", "inference_data_ai_source_ingest.py")
packets = load_module("inference_data_ai_semantic_packets", "inference_data_ai_semantic_packets.py")


def recursive_keys(value):
    if isinstance(value, dict):
        for key, child in value.items():
            yield str(key)
            yield from recursive_keys(child)
    elif isinstance(value, list):
        for child in value:
            yield from recursive_keys(child)


class SemanticSourcePacketTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.database = self.root / "capture-v2.sqlite"
        self.connection = sqlite3.connect(self.database)

    def tearDown(self) -> None:
        self.connection.close()
        self.temp.cleanup()

    def import_wide_fixture(self) -> dict:
        source = self.root / "wide-generic.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Unmapped Ω source"
        for column in range(1, 8):
            sheet.cell(1, column, f"UNKNOWN_Ω_TERM_{column}")
        sheet["A2"] = "=1+2"
        sheet["B2"] = 0
        sheet.row_dimensions[2].hidden = True
        sheet.column_dimensions["B"].hidden = True
        sheet.column_dimensions["B"].width = 19.5
        sheet["A5"] = "UNLISTED_MERGED_CONCEPT"
        sheet.merge_cells("A5:B5")
        sheet.freeze_panes = "B2"
        sheet.auto_filter.ref = "A1:G2"
        sheet.sheet_state = "hidden"
        workbook.create_sheet("Visible empty")
        workbook.save(source)
        workbook.close()
        return capture.capture_and_import(
            self.connection,
            source,
            captured_at="2026-07-17T01:00:00Z",
        )

    def import_empty_fixture(self) -> dict:
        source = self.root / "empty.xlsx"
        workbook = Workbook()
        workbook.save(source)
        workbook.close()
        return capture.capture_and_import(
            self.connection,
            source,
            captured_at="2026-07-17T02:00:00Z",
        )

    def import_no_tabular_fixture(self) -> dict:
        source = self.root / "no-tabular.xlsx"
        workbook = Workbook()
        workbook.active["D7"] = "NOVEL_UNKNOWN_NARRATIVE_Ω"
        workbook.save(source)
        workbook.close()
        return capture.capture_and_import(
            self.connection,
            source,
            captured_at="2026-07-17T03:00:00Z",
        )

    def test_complete_unique_coverage_and_domain_neutral_source_fidelity(self) -> None:
        imported = self.import_wide_fixture()
        self.connection.execute("PRAGMA query_only=ON")

        result = packets.build_semantic_source_packets(
            self.connection,
            revision_id=imported["revisionId"],
            max_cells=3,
            max_rows=2,
            empty_row_gap=1,
        )

        self.assertEqual("semantic-source-packet-v1", result["schemaVersion"])
        inventory = result["inventory"]
        self.assertEqual("COMPLETE", inventory["coverage"]["status"])
        self.assertEqual(10, inventory["coverage"]["expectedCellCount"])
        self.assertEqual(10, inventory["coverage"]["packetCellCount"])
        self.assertEqual([], inventory["coverage"]["duplicateCellKeys"])
        self.assertEqual([], inventory["coverage"]["missingCellKeys"])

        all_cells = [
            cell
            for chunk in result["chunks"]
            for cell in chunk["cells"]
        ]
        source_keys = [cell["sourceCellKey"] for cell in all_cells]
        self.assertEqual(10, len(source_keys))
        self.assertEqual(10, len(set(source_keys)))
        self.assertTrue(all(count == 1 for count in Counter(source_keys).values()))
        self.assertEqual(
            {
                "A1",
                "B1",
                "C1",
                "D1",
                "E1",
                "F1",
                "G1",
                "A2",
                "B2",
                "A5",
            },
            {cell["coordinate"] for cell in all_cells},
        )

        raw_values = {cell["coordinate"]: cell["rawValue"] for cell in all_cells}
        self.assertEqual("UNKNOWN_Ω_TERM_1", raw_values["A1"])
        self.assertEqual("UNLISTED_MERGED_CONCEPT", raw_values["A5"])
        self.assertEqual(0, raw_values["B2"])

        formula = next(cell for cell in all_cells if cell["coordinate"] == "A2")
        self.assertEqual("=1+2", formula["formula"])
        self.assertIsNone(formula["rawValue"])
        self.assertIsNone(formula["cachedValue"])
        self.assertIsNone(formula["displayValue"])
        self.assertTrue(formula["hidden"]["sheet"])
        self.assertTrue(formula["hidden"]["row"])
        self.assertEqual("FORMULA_NO_CACHE", formula["valueSource"])
        self.assertTrue(formula["primary"])
        self.assertFalse(formula["contextOnly"])

        hidden_column = next(
            cell for cell in all_cells if cell["coordinate"] == "B2"
        )
        self.assertTrue(hidden_column["hidden"]["column"])
        hidden_column_chunk = next(
            chunk
            for chunk in result["chunks"]
            if any(cell["coordinate"] == "B2" for cell in chunk["cells"])
        )
        self.assertEqual(
            19.5,
            next(
                dimension
                for dimension in hidden_column_chunk["columnDimensions"]
                if dimension["key"] == "B"
            )["width"],
        )

        merged = next(cell for cell in all_cells if cell["coordinate"] == "A5")
        self.assertEqual("A5:B5", merged["mergeRange"])
        self.assertEqual("anchor", merged["mergeRole"])
        row_five_chunk = next(
            chunk
            for chunk in result["chunks"]
            if any(cell["coordinate"] == "A5" for cell in chunk["cells"])
        )
        self.assertEqual(2, row_five_chunk["sectionIndex"])
        self.assertEqual("A5:B5", row_five_chunk["mergedRanges"][0]["address"])

        for chunk in result["chunks"]:
            self.assertLessEqual(len(chunk["cells"]), 3)
            self.assertLessEqual(chunk["bounds"]["rowCount"], 2)
            self.assertEqual(
                imported["revisionId"],
                chunk["sourceRevision"]["revisionId"],
            )
            self.assertEqual(
                imported["contentSha256"],
                chunk["sourceRevision"]["contentSha256"],
            )
            self.assertEqual(chunk["packetId"], chunk["chunkId"])
            self.assertTrue(chunk["primaryRange"])
            self.assertFalse(chunk["truncated"])
            self.assertTrue(chunk["styleDictionary"])

        self.assertFalse(
            any(key.casefold().startswith("image") for key in recursive_keys(result))
        )
        self.assertEqual(
            packets.packet_json_bytes(result),
            packets.packet_json_bytes(
                packets.build_semantic_source_packets(
                    self.connection,
                    revision_id=imported["revisionId"],
                    max_cells=3,
                    max_rows=2,
                    empty_row_gap=1,
                )
            ),
        )

    def test_single_wide_row_is_split_into_ordered_column_segments(self) -> None:
        imported = self.import_wide_fixture()
        result = packets.build_semantic_source_packets(
            self.connection,
            revision_id=imported["revisionId"],
            max_cells=3,
            max_rows=20,
        )

        wide = [
            chunk
            for chunk in result["chunks"]
            if chunk["splitReason"] == "WIDE_ROW"
        ]
        self.assertEqual(3, len(wide))
        self.assertEqual(
            [(1, 3), (4, 6), (7, 7)],
            [
                (
                    chunk["rowSegment"]["minColumn"],
                    chunk["rowSegment"]["maxColumn"],
                )
                for chunk in wide
            ],
        )
        self.assertEqual(
            [1, 2, 3],
            [chunk["rowSegment"]["segmentIndex"] for chunk in wide],
        )
        self.assertTrue(
            all(chunk["rowSegment"]["segmentCount"] == 3 for chunk in wide)
        )
        self.assertEqual([], wide[0]["contextCells"])
        self.assertEqual(
            ["A1", "B1"],
            [cell["coordinate"] for cell in wide[1]["contextCells"]],
        )
        self.assertTrue(
            all(cell["contextOnly"] and not cell["primary"] for cell in wide[1]["contextCells"])
        )

    def test_empty_workbook_emits_terminal_packet_without_chunks(self) -> None:
        imported = self.import_empty_fixture()
        result = packets.build_semantic_source_packets(
            self.connection,
            revision_id=imported["revisionId"],
        )

        self.assertEqual("EMPTY_WORKBOOK", result["inventory"]["workbook"]["status"])
        self.assertEqual([], result["chunks"])
        self.assertEqual(1, len(result["terminalPackets"]))
        terminal = result["terminalPackets"][0]
        self.assertEqual("TERMINAL", terminal["packetType"])
        self.assertEqual("EMPTY_WORKBOOK", terminal["terminalStatus"])
        self.assertEqual({"type": "WORKBOOK"}, terminal["scope"])
        self.assertEqual([], terminal["cells"])
        self.assertEqual("COMPLETE", result["inventory"]["coverage"]["status"])
        self.assertEqual(0, result["inventory"]["coverage"]["expectedCellCount"])
        self.assertFalse(result["inventory"]["contentCompleteForManifest"])

    def test_no_tabular_terminal_keeps_unknown_source_cell_in_coverage(self) -> None:
        imported = self.import_no_tabular_fixture()
        result = packets.build_semantic_source_packets(
            self.connection,
            revision_id=imported["revisionId"],
        )

        self.assertEqual(
            "NO_TABULAR_EVIDENCE",
            result["inventory"]["workbook"]["status"],
        )
        self.assertEqual(1, len(result["terminalPackets"]))
        self.assertEqual(
            "NO_TABULAR_EVIDENCE",
            result["terminalPackets"][0]["terminalStatus"],
        )
        self.assertEqual(1, len(result["chunks"]))
        self.assertEqual(
            "NOVEL_UNKNOWN_NARRATIVE_Ω",
            result["chunks"][0]["cells"][0]["rawValue"],
        )
        self.assertEqual("COMPLETE", result["inventory"]["coverage"]["status"])
        self.assertEqual(1, result["inventory"]["coverage"]["packetCellCount"])

    def test_database_path_entry_point_is_read_only(self) -> None:
        imported = self.import_wide_fixture()
        self.connection.commit()

        result = packets.build_semantic_source_packets_from_db(
            self.database,
            revision_id=imported["revisionId"],
            max_cells=3,
        )
        self.assertEqual("COMPLETE", result["inventory"]["coverage"]["status"])
        selected_by_source = packets.build_semantic_source_packets_from_db(
            self.database,
            source_path=result["inventory"]["sourceRevision"]["sourcePath"],
            max_cells=3,
        )
        self.assertEqual(
            imported["revisionId"],
            selected_by_source["inventory"]["sourceRevision"]["revisionId"],
        )
        with packets.connect_capture_v2_readonly(self.database) as connection:
            with self.assertRaises(sqlite3.OperationalError):
                connection.execute(
                    "UPDATE capture_v2_documents SET file_name='changed'"
                )

    def test_coverage_validator_detects_duplicates(self) -> None:
        chunk = {
            "cells": [
                {"sourceCellKey": "r:1:A1"},
                {"sourceCellKey": "r:1:A1"},
            ]
        }
        validation = packets.validate_packet_coverage(["r:1:A1"], [chunk])
        self.assertEqual("INVALID", validation["status"])
        self.assertEqual(["r:1:A1"], validation["duplicateCellKeys"])


if __name__ == "__main__":
    unittest.main()
