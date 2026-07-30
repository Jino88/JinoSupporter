#!/usr/bin/env python3
"""Read-only numeric workbook capture for a completed structure-scan batch.

The scanner intentionally stores only structural candidates.  This program is
the next stage: it captures *numeric-table facts* into a batch-scoped SQLite
database without starting Excel, recalculating formulas, or touching the
existing universal-grid database.

Free-form narrative is not persisted as review data.  Short text labels are
kept only when they are adjacent to numeric table regions and are therefore
needed to identify columns, dates, Test, and Normal conditions later.
"""
from __future__ import annotations

import argparse
import csv
import html
import json
import math
import os
import re
import sqlite3
import sys
import time
import zipfile
from collections import Counter, defaultdict
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterator
from xml.etree import ElementTree as ET

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - command error path
    load_workbook = None

import inference_data_ai_structure_scan as structure


CAPTURE_VERSION = "numeric-capture-v1"
MAX_CAPTURED_CELLS_PER_SHEET = 1_000_000
MAX_CAPTURED_MERGES_PER_SHEET = 100_000
MAX_HEADER_ROWS_PER_REGION = 2
MAX_HEADER_LABEL_LENGTH = 500
MAX_TABLE_COLUMN_GAP = 3


class CaptureLimitError(RuntimeError):
    """A workbook is valid but cannot be completely captured under a safety limit."""


def utc_now() -> str:
    return datetime.now(UTC).isoformat().replace("+00:00", "Z")


def atomic_write_text(path: Path, value: str, *, encoding: str = "utf-8") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.name + ".tmp")
    temporary.write_text(value, encoding=encoding)
    os.replace(temporary, path)


def atomic_write_json(path: Path, value: object) -> None:
    atomic_write_text(path, json.dumps(value, ensure_ascii=False, indent=2) + "\n")


def resolve_batch_directory(service_dir: Path, batch_id: str) -> Path:
    return structure.resolve_batch_directory(service_dir, structure.safe_batch_id(batch_id))


def source_fingerprint(path: Path) -> str:
    return structure.source_fingerprint(path)


def open_capture_db(path: Path) -> sqlite3.Connection:
    connection = sqlite3.connect(path)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys=ON")
    connection.execute("PRAGMA journal_mode=WAL")
    connection.executescript(
        """
        CREATE TABLE IF NOT EXISTS capture_metadata (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS capture_workbooks (
            workbook_id INTEGER PRIMARY KEY,
            relative_path TEXT NOT NULL UNIQUE,
            source_path TEXT NOT NULL,
            extension TEXT NOT NULL,
            snapshot_fingerprint TEXT NOT NULL,
            current_fingerprint TEXT NOT NULL DEFAULT '',
            structure_scan_status TEXT NOT NULL,
            capture_status TEXT NOT NULL,
            attempts INTEGER NOT NULL DEFAULT 0,
            sheet_count_expected INTEGER NOT NULL DEFAULT 0,
            sheet_count_captured INTEGER NOT NULL DEFAULT 0,
            numeric_cell_count INTEGER NOT NULL DEFAULT 0,
            formula_count INTEGER NOT NULL DEFAULT 0,
            date_cell_count INTEGER NOT NULL DEFAULT 0,
            table_candidate_count INTEGER NOT NULL DEFAULT 0,
            merge_count INTEGER NOT NULL DEFAULT 0,
            error_text TEXT NOT NULL DEFAULT '',
            started_at TEXT NOT NULL DEFAULT '',
            finished_at TEXT NOT NULL DEFAULT ''
        );
        CREATE INDEX IF NOT EXISTS idx_capture_workbooks_status
            ON capture_workbooks(capture_status, relative_path);

        CREATE TABLE IF NOT EXISTS captured_sheets (
            sheet_id INTEGER PRIMARY KEY,
            workbook_id INTEGER NOT NULL REFERENCES capture_workbooks(workbook_id) ON DELETE CASCADE,
            sheet_index INTEGER NOT NULL,
            sheet_name TEXT NOT NULL,
            sheet_state TEXT NOT NULL,
            declared_dimension TEXT NOT NULL DEFAULT '',
            max_row INTEGER NOT NULL,
            max_column INTEGER NOT NULL,
            merge_count INTEGER NOT NULL DEFAULT 0,
            numeric_cell_count INTEGER NOT NULL DEFAULT 0,
            formula_count INTEGER NOT NULL DEFAULT 0,
            date_cell_count INTEGER NOT NULL DEFAULT 0,
            capture_status TEXT NOT NULL,
            warning_text TEXT NOT NULL DEFAULT '',
            UNIQUE(workbook_id, sheet_index)
        );
        CREATE INDEX IF NOT EXISTS idx_captured_sheets_workbook ON captured_sheets(workbook_id, sheet_index);

        CREATE TABLE IF NOT EXISTS captured_merge_ranges (
            sheet_id INTEGER NOT NULL REFERENCES captured_sheets(sheet_id) ON DELETE CASCADE,
            range_ref TEXT NOT NULL,
            PRIMARY KEY(sheet_id, range_ref)
        );

        CREATE TABLE IF NOT EXISTS numeric_cells (
            sheet_id INTEGER NOT NULL REFERENCES captured_sheets(sheet_id) ON DELETE CASCADE,
            row_index INTEGER NOT NULL,
            column_index INTEGER NOT NULL,
            source_kind TEXT NOT NULL,
            value_text TEXT NOT NULL,
            numeric_value REAL NOT NULL,
            formula_text TEXT NOT NULL DEFAULT '',
            PRIMARY KEY(sheet_id, row_index, column_index)
        );
        CREATE INDEX IF NOT EXISTS idx_numeric_cells_sheet_position
            ON numeric_cells(sheet_id, row_index, column_index);

        CREATE TABLE IF NOT EXISTS formula_cells (
            sheet_id INTEGER NOT NULL REFERENCES captured_sheets(sheet_id) ON DELETE CASCADE,
            row_index INTEGER NOT NULL,
            column_index INTEGER NOT NULL,
            formula_text TEXT NOT NULL,
            cached_value_text TEXT NOT NULL DEFAULT '',
            cached_numeric_value REAL,
            cached_is_date INTEGER NOT NULL DEFAULT 0,
            PRIMARY KEY(sheet_id, row_index, column_index)
        );

        CREATE TABLE IF NOT EXISTS date_cells (
            sheet_id INTEGER NOT NULL REFERENCES captured_sheets(sheet_id) ON DELETE CASCADE,
            row_index INTEGER NOT NULL,
            column_index INTEGER NOT NULL,
            date_value TEXT NOT NULL,
            source_kind TEXT NOT NULL,
            PRIMARY KEY(sheet_id, row_index, column_index)
        );
        CREATE INDEX IF NOT EXISTS idx_date_cells_sheet_date
            ON date_cells(sheet_id, date_value);

        CREATE TABLE IF NOT EXISTS numeric_table_candidates (
            table_id INTEGER PRIMARY KEY,
            sheet_id INTEGER NOT NULL REFERENCES captured_sheets(sheet_id) ON DELETE CASCADE,
            start_row INTEGER NOT NULL,
            end_row INTEGER NOT NULL,
            start_column INTEGER NOT NULL,
            end_column INTEGER NOT NULL,
            header_start_row INTEGER NOT NULL,
            header_end_row INTEGER NOT NULL,
            numeric_cell_count INTEGER NOT NULL,
            candidate_type TEXT NOT NULL,
            confidence TEXT NOT NULL,
            UNIQUE(sheet_id, start_row, end_row, start_column, end_column)
        );
        CREATE INDEX IF NOT EXISTS idx_numeric_table_candidates_sheet ON numeric_table_candidates(sheet_id, start_row);

        CREATE TABLE IF NOT EXISTS numeric_table_labels (
            table_id INTEGER NOT NULL REFERENCES numeric_table_candidates(table_id) ON DELETE CASCADE,
            row_index INTEGER NOT NULL,
            column_index INTEGER NOT NULL,
            label_text TEXT NOT NULL,
            label_role TEXT NOT NULL,
            PRIMARY KEY(table_id, row_index, column_index)
        );
        """
    )
    connection.execute(
        "INSERT OR REPLACE INTO capture_metadata(key, value) VALUES ('schemaVersion', ?)",
        ("numeric-capture-db-v1",),
    )
    connection.execute(
        "INSERT OR REPLACE INTO capture_metadata(key, value) VALUES ('captureVersion', ?)", (CAPTURE_VERSION,))
    connection.commit()
    return connection


def source_rows(batch_dir: Path) -> list[sqlite3.Row]:
    state_path = batch_dir / "state.sqlite"
    if not state_path.is_file():
        raise ValueError(f"Structure scan state database is missing: {state_path}")
    source = sqlite3.connect(state_path)
    source.row_factory = sqlite3.Row
    try:
        return list(
            source.execute(
                """
                SELECT relative_path, source_path, extension, kind, fingerprint, status
                FROM items
                WHERE kind='openxml'
                ORDER BY relative_path
                """
            )
        )
    finally:
        source.close()


def numeric_value(value: object) -> tuple[str, float, str] | None:
    """Return exact text, a SQLite-compatible number, and source kind.

    Excel frequently stores numeric-looking table values as text.  Those are
    captured as ``TEXT_NUMBER`` rather than silently discarded.  Boolean and
    date values are deliberately excluded from metric numbers.
    """
    if isinstance(value, bool) or value is None or isinstance(value, (date, datetime)):
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            decimal_value = Decimal(str(value))
        except InvalidOperation:
            return None
        if not decimal_value.is_finite():
            return None
        as_float = float(decimal_value)
        if not math.isfinite(as_float):
            return None
        return format(decimal_value, "f"), as_float, "NUMBER"
    if not isinstance(value, str):
        return None
    text = value.strip()
    if not text or len(text) > 128:
        return None
    compact = text.replace(",", "").replace(" ", "")
    sign = ""
    if compact.startswith("(") and compact.endswith(")"):
        sign, compact = "-", compact[1:-1]
    is_percent = compact.endswith("%")
    if is_percent:
        compact = compact[:-1]
    if not re.fullmatch(r"[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?", sign + compact):
        return None
    try:
        decimal_value = Decimal(sign + compact)
        if is_percent:
            decimal_value /= Decimal("100")
    except InvalidOperation:
        return None
    if not decimal_value.is_finite():
        return None
    as_float = float(decimal_value)
    if not math.isfinite(as_float):
        return None
    return format(decimal_value, "f"), as_float, "TEXT_PERCENT" if is_percent else "TEXT_NUMBER"


def normalized_date(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def display_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def all_merge_ranges(path: Path, sheet_parts: dict[str, str]) -> dict[str, list[str]]:
    """Read all merge references without loading worksheets into editable mode."""
    result: dict[str, list[str]] = {name: [] for name in sheet_parts}
    with zipfile.ZipFile(path) as archive:
        # ``sheet_metadata_from_zip`` already performs OpenXML preflight and
        # rejects unsafe worksheet XML before this streaming pass.
        for sheet_name, part in sheet_parts.items():
            if not part or part not in archive.namelist():
                continue
            values: list[str] = []
            try:
                with archive.open(part) as stream:
                    for _, element in ET.iterparse(stream, events=("start",)):
                        if element.tag.rsplit("}", 1)[-1] != "mergeCell":
                            continue
                        reference = element.attrib.get("ref", "")
                        if reference:
                            values.append(reference)
                        if len(values) > MAX_CAPTURED_MERGES_PER_SHEET:
                            raise CaptureLimitError(
                                f"Merge range count exceeds capture limit ({MAX_CAPTURED_MERGES_PER_SHEET}) on {sheet_name}."
                            )
            except ET.ParseError as exc:
                raise structure.QuarantinedPackageError(f"Invalid merge XML in {part}: {exc}") from exc
            result[sheet_name] = values
    return result


def numeric_table_regions(numeric_positions: list[tuple[int, int]]) -> list[tuple[int, int, int, int, int]]:
    """Form numeric table candidates from contiguous row *and* column bands.

    A worksheet can place two reports side by side.  Grouping only by rows
    would combine their duplicated ``Input``/``Total NG`` headers and make the
    source columns ambiguous.  A gap of four or more entirely non-numeric
    columns is treated as a new table band; smaller gaps remain part of the
    same table to tolerate text-only condition columns inside a table.
    """
    by_row: dict[int, list[int]] = defaultdict(list)
    for row, column in numeric_positions:
        by_row[row].append(column)
    if not by_row:
        return []
    row_bands: list[list[int]] = []
    current_rows: list[int] = []
    for row in sorted(by_row):
        if current_rows and row - current_rows[-1] > 3:
            row_bands.append(current_rows)
            current_rows = []
        current_rows.append(row)
    if current_rows:
        row_bands.append(current_rows)
    regions: list[tuple[int, int, int, int, int]] = []
    for rows in row_bands:
        columns = sorted({column for row in rows for column in by_row[row]})
        if not columns:
            continue
        column_bands: list[list[int]] = []
        current_columns: list[int] = []
        for column in columns:
            if current_columns and column - current_columns[-1] > MAX_TABLE_COLUMN_GAP:
                column_bands.append(current_columns)
                current_columns = []
            current_columns.append(column)
        if current_columns:
            column_bands.append(current_columns)
        for band in column_bands:
            start_column, end_column = min(band), max(band)
            cells = sum(1 for row in rows for column in by_row[row] if start_column <= column <= end_column)
            regions.append((rows[0], rows[-1], start_column, end_column, cells))
    return regions


def candidate_type(header_texts: list[str]) -> tuple[str, str]:
    tokens = {structure.header_token(value) for value in header_texts}
    tokens.discard(None)
    if {"INPUT", "OK", "TOTAL_NG", "NG_RATE"}.issubset(tokens):
        return "DEFECT_RATE_NUMERIC_TABLE", "HIGH"
    if {"SAMPLE", "AVERAGE", "MAX", "MIN"}.issubset(tokens):
        return "MEASUREMENT_SUMMARY_NUMERIC_TABLE", "HIGH"
    if {"NORMAL_CUE", "TEST_CUE"}.issubset(tokens):
        return "TEST_NORMAL_NUMERIC_TABLE", "MEDIUM"
    return "NUMERIC_TABLE_UNCLASSIFIED", "LOW"


def capture_sheet(
    connection: sqlite3.Connection,
    workbook_id: int,
    sheet_index: int,
    formula_sheet: Any,
    cached_sheet: Any,
    metadata: dict[str, object],
    merge_ranges: list[str],
) -> dict[str, int]:
    max_row = int(formula_sheet.max_row or 0)
    max_column = int(formula_sheet.max_column or 0)
    declared_cells = max_row * max_column
    if declared_cells > MAX_CAPTURED_CELLS_PER_SHEET:
        raise CaptureLimitError(
            f"Declared cell count exceeds capture limit ({declared_cells} > {MAX_CAPTURED_CELLS_PER_SHEET}) on {formula_sheet.title}."
        )
    cursor = connection.execute(
        """
        INSERT INTO captured_sheets(
            workbook_id, sheet_index, sheet_name, sheet_state, declared_dimension,
            max_row, max_column, merge_count, capture_status
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'CAPTURING')
        """,
        (
            workbook_id,
            sheet_index,
            formula_sheet.title,
            formula_sheet.sheet_state,
            str(metadata.get("declaredDimension") or formula_sheet.calculate_dimension()),
            max_row,
            max_column,
            len(merge_ranges),
        ),
    )
    sheet_id = int(cursor.lastrowid)
    connection.executemany(
        "INSERT INTO captured_merge_ranges(sheet_id, range_ref) VALUES (?, ?)",
        ((sheet_id, item) for item in merge_ranges),
    )

    numeric_rows: list[tuple[int, int]] = []
    text_by_row: dict[int, list[tuple[int, str]]] = defaultdict(list)
    numeric_count = 0
    formula_count = 0
    date_count = 0
    visits = 0
    numeric_records: list[tuple[object, ...]] = []
    formula_records: list[tuple[object, ...]] = []
    date_records: list[tuple[object, ...]] = []

    formula_rows = formula_sheet.iter_rows()
    cached_rows = cached_sheet.iter_rows()
    for row_index, (formula_row, cached_row) in enumerate(zip(formula_rows, cached_rows), start=1):
        for column_index, (formula_cell, cached_cell) in enumerate(zip(formula_row, cached_row), start=1):
            visits += 1
            if visits > MAX_CAPTURED_CELLS_PER_SHEET:
                raise CaptureLimitError(
                    f"Cell iteration exceeds capture limit ({MAX_CAPTURED_CELLS_PER_SHEET}) on {formula_sheet.title}."
                )
            # Read-only worksheets use EmptyCell instances for leading/trailing
            # blanks.  They deliberately have no coordinate attributes, so the
            # iterator position is the authoritative coordinate here.
            row, column = row_index, column_index
            value = getattr(formula_cell, "value", None)
            cached_value = getattr(cached_cell, "value", None)
            formula_text = ""
            if getattr(formula_cell, "data_type", None) == "f" or (isinstance(value, str) and value.startswith("=")):
                formula_count += 1
                formula_text = str(value)
                cached_number = numeric_value(cached_value)
                cached_date = normalized_date(cached_value)
                formula_records.append(
                    (
                        sheet_id,
                        row,
                        column,
                        formula_text,
                        display_value(cached_value),
                        None if cached_number is None else cached_number[1],
                        1 if cached_date else 0,
                    )
                )
                if cached_date:
                    date_count += 1
                    date_records.append((sheet_id, row, column, cached_date, "FORMULA_CACHED"))
                elif cached_number is not None:
                    numeric_count += 1
                    numeric_rows.append((row, column))
                    numeric_records.append(
                        (sheet_id, row, column, "FORMULA_CACHED", cached_number[0], cached_number[1], formula_text)
                    )
                continue

            actual_date = normalized_date(value)
            if actual_date:
                date_count += 1
                date_records.append((sheet_id, row, column, actual_date, "DATE"))
                continue
            number = numeric_value(value)
            if number is not None:
                numeric_count += 1
                numeric_rows.append((row, column))
                numeric_records.append((sheet_id, row, column, number[2], number[0], number[1], ""))
                continue
            if isinstance(value, str):
                label = value.strip()
                if label and len(label) <= MAX_HEADER_LABEL_LENGTH:
                    text_by_row[row].append((column, label))

    connection.executemany(
        """
        INSERT INTO numeric_cells(sheet_id, row_index, column_index, source_kind, value_text, numeric_value, formula_text)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        numeric_records,
    )
    connection.executemany(
        """
        INSERT INTO formula_cells(sheet_id, row_index, column_index, formula_text, cached_value_text, cached_numeric_value, cached_is_date)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        formula_records,
    )
    connection.executemany(
        "INSERT INTO date_cells(sheet_id, row_index, column_index, date_value, source_kind) VALUES (?, ?, ?, ?, ?)",
        date_records,
    )

    candidate_count = 0
    for start_row, end_row, start_column, end_column, cell_count in numeric_table_regions(numeric_rows):
        header_start = max(1, start_row - MAX_HEADER_ROWS_PER_REGION)
        header_end = max(header_start, start_row - 1)
        header_entries = [
            (row, column, label)
            for row in range(header_start, header_end + 1)
            for column, label in text_by_row.get(row, [])
            if start_column <= column <= end_column
        ]
        row_label_entries = [
            (row, column, label)
            for row in range(start_row, end_row + 1)
            for column, label in text_by_row.get(row, [])
            if start_column - 3 <= column <= end_column + 3
        ]
        kind, confidence = candidate_type([item[2] for item in header_entries])
        cursor = connection.execute(
            """
            INSERT INTO numeric_table_candidates(
                sheet_id, start_row, end_row, start_column, end_column,
                header_start_row, header_end_row, numeric_cell_count, candidate_type, confidence
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (sheet_id, start_row, end_row, start_column, end_column, header_start, header_end, cell_count, kind, confidence),
        )
        table_id = int(cursor.lastrowid)
        connection.executemany(
            """
            INSERT INTO numeric_table_labels(table_id, row_index, column_index, label_text, label_role)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                (table_id, row, column, label, "HEADER") for row, column, label in header_entries
            ),
        )
        connection.executemany(
            """
            INSERT OR IGNORE INTO numeric_table_labels(table_id, row_index, column_index, label_text, label_role)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                (table_id, row, column, label, "ROW_LABEL") for row, column, label in row_label_entries
            ),
        )
        candidate_count += 1

    connection.execute(
        """
        UPDATE captured_sheets
        SET numeric_cell_count=?, formula_count=?, date_cell_count=?, capture_status='CAPTURED'
        WHERE sheet_id=?
        """,
        (numeric_count, formula_count, date_count, sheet_id),
    )
    return {
        "numeric": numeric_count,
        "formulas": formula_count,
        "dates": date_count,
        "tables": candidate_count,
        "merges": len(merge_ranges),
    }


def delete_existing_capture(connection: sqlite3.Connection, workbook_id: int) -> None:
    connection.execute("DELETE FROM captured_sheets WHERE workbook_id=?", (workbook_id,))
    connection.execute(
        """
        UPDATE capture_workbooks
        SET sheet_count_captured=0, numeric_cell_count=0, formula_count=0,
            date_cell_count=0, table_candidate_count=0, merge_count=0, error_text=''
        WHERE workbook_id=?
        """,
        (workbook_id,),
    )


def ensure_workbook(connection: sqlite3.Connection, item: sqlite3.Row) -> sqlite3.Row:
    connection.execute(
        """
        INSERT INTO capture_workbooks(
            relative_path, source_path, extension, snapshot_fingerprint, structure_scan_status, capture_status
        ) VALUES (?, ?, ?, ?, ?, 'PENDING')
        ON CONFLICT(relative_path) DO UPDATE SET
            source_path=excluded.source_path,
            extension=excluded.extension,
            snapshot_fingerprint=excluded.snapshot_fingerprint,
            structure_scan_status=excluded.structure_scan_status
        """,
        (
            item["relative_path"],
            item["source_path"],
            item["extension"],
            item["fingerprint"],
            item["status"],
        ),
    )
    return connection.execute("SELECT * FROM capture_workbooks WHERE relative_path=?", (item["relative_path"],)).fetchone()


def capture_workbook(connection: sqlite3.Connection, item: sqlite3.Row, *, force: bool) -> str:
    record = ensure_workbook(connection, item)
    assert record is not None
    path = Path(str(item["source_path"]))
    if not path.is_file() or source_fingerprint(path) != str(item["fingerprint"]):
        connection.execute(
            "UPDATE capture_workbooks SET capture_status='CHANGED', error_text=?, finished_at=? WHERE workbook_id=?",
            ("Source path, size, or modified time differs from the structure-scan snapshot.", utc_now(), record["workbook_id"]),
        )
        connection.commit()
        return "CHANGED"
    if not force and record["capture_status"] == "CAPTURED" and record["current_fingerprint"] == item["fingerprint"]:
        connection.commit()
        return "SKIPPED"
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for numeric capture but is not installed.")

    workbook_id = int(record["workbook_id"])
    delete_existing_capture(connection, workbook_id)
    connection.execute(
        """
        UPDATE capture_workbooks
        SET capture_status='CAPTURING', attempts=attempts+1, current_fingerprint=?, started_at=?, finished_at='', error_text=''
        WHERE workbook_id=?
        """,
        (item["fingerprint"], utc_now(), workbook_id),
    )
    connection.commit()
    started = time.monotonic()
    formula_workbook = None
    cached_workbook = None
    try:
        _, package_sheets = structure.sheet_metadata_from_zip(path)
        part_by_name = {str(value["sheetName"]): str(value.get("part") or "") for value in package_sheets}
        metadata_by_name = {str(value["sheetName"]): value for value in package_sheets}
        merge_by_name = all_merge_ranges(path, part_by_name)
        formula_workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
        cached_workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        if formula_workbook.sheetnames != cached_workbook.sheetnames:
            raise RuntimeError("Formula and cached-value workbook sheet lists differ.")
        totals: Counter[str] = Counter()
        for sheet_index, (formula_sheet, cached_sheet) in enumerate(zip(formula_workbook.worksheets, cached_workbook.worksheets), start=1):
            if formula_sheet.title != cached_sheet.title:
                raise RuntimeError("Formula and cached-value worksheet order differs.")
            totals.update(
                capture_sheet(
                    connection,
                    workbook_id,
                    sheet_index,
                    formula_sheet,
                    cached_sheet,
                    metadata_by_name.get(formula_sheet.title, {}),
                    merge_by_name.get(formula_sheet.title, []),
                )
            )
        if source_fingerprint(path) != str(item["fingerprint"]):
            raise RuntimeError("Source path, size, or modified time changed while numeric capture was running.")
        connection.execute(
            """
            UPDATE capture_workbooks
            SET capture_status='CAPTURED', sheet_count_expected=?, sheet_count_captured=?,
                numeric_cell_count=?, formula_count=?, date_cell_count=?, table_candidate_count=?,
                merge_count=?, finished_at=?, error_text=''
            WHERE workbook_id=?
            """,
            (
                len(formula_workbook.worksheets),
                len(formula_workbook.worksheets),
                totals["numeric"],
                totals["formulas"],
                totals["dates"],
                totals["tables"],
                totals["merges"],
                utc_now(),
                workbook_id,
            ),
        )
        connection.commit()
        return "CAPTURED"
    except CaptureLimitError as exc:
        connection.rollback()
        connection.execute("DELETE FROM captured_sheets WHERE workbook_id=?", (workbook_id,))
        connection.execute(
            "UPDATE capture_workbooks SET capture_status='TRUNCATED', error_text=?, finished_at=? WHERE workbook_id=?",
            (str(exc), utc_now(), workbook_id),
        )
        connection.commit()
        return "TRUNCATED"
    except (zipfile.BadZipFile, structure.QuarantinedPackageError, KeyError) as exc:
        connection.rollback()
        connection.execute("DELETE FROM captured_sheets WHERE workbook_id=?", (workbook_id,))
        connection.execute(
            "UPDATE capture_workbooks SET capture_status='QUARANTINED', error_text=?, finished_at=? WHERE workbook_id=?",
            (f"{type(exc).__name__}: {exc}"[:2000], utc_now(), workbook_id),
        )
        connection.commit()
        return "QUARANTINED"
    except Exception as exc:
        connection.rollback()
        connection.execute("DELETE FROM captured_sheets WHERE workbook_id=?", (workbook_id,))
        connection.execute(
            "UPDATE capture_workbooks SET capture_status='FAILED_RETRYABLE', error_text=?, finished_at=? WHERE workbook_id=?",
            (f"{type(exc).__name__}: {exc}"[:2000], utc_now(), workbook_id),
        )
        connection.commit()
        return "FAILED_RETRYABLE"
    finally:
        if formula_workbook is not None:
            formula_workbook.close()
        if cached_workbook is not None:
            cached_workbook.close()
        _ = started  # Kept for debugger timing without adding per-cell event noise.


def write_outputs(batch_dir: Path, connection: sqlite3.Connection) -> dict[str, object]:
    rows = list(connection.execute("SELECT * FROM capture_workbooks ORDER BY relative_path"))
    statuses = Counter(str(row["capture_status"]) for row in rows)
    source_statuses = Counter(str(row["structure_scan_status"]) for row in rows)
    totals = connection.execute(
        """
        SELECT COALESCE(SUM(numeric_cell_count), 0), COALESCE(SUM(formula_count), 0),
               COALESCE(SUM(date_cell_count), 0), COALESCE(SUM(table_candidate_count), 0)
        FROM capture_workbooks WHERE capture_status='CAPTURED'
        """
    ).fetchone()
    summary = {
        "schemaVersion": "numeric-capture-summary-v1",
        "captureVersion": CAPTURE_VERSION,
        "generatedAt": utc_now(),
        "usesCom": False,
        "sourceWorkbookCount": len(rows),
        "captureStatusCounts": dict(sorted(statuses.items())),
        "structureScanStatusCounts": dict(sorted(source_statuses.items())),
        "capturedNumericCellCount": int(totals[0]),
        "capturedFormulaCount": int(totals[1]),
        "capturedDateCellCount": int(totals[2]),
        "numericTableCandidateCount": int(totals[3]),
        "database": "numeric-capture.sqlite",
        "classificationCsv": "numeric-capture.csv",
        "limitations": [
            "Only numeric-table facts, dates, formulas, merge ranges, and short adjacent header labels are captured.",
            "Formulas are never recalculated; cached values are retained only when present in the workbook.",
            "No Test–Normal comparison or quality conclusion is produced by this capture stage.",
        ],
    }
    atomic_write_json(batch_dir / "numeric-capture-summary.json", summary)
    report_rows = [
        {
            "relativePath": str(row["relative_path"]),
            "status": str(row["capture_status"]),
            "sourceScanStatus": str(row["structure_scan_status"]),
            "sheetCount": str(row["sheet_count_captured"]),
            "numericCells": str(row["numeric_cell_count"]),
            "formulas": str(row["formula_count"]),
            "dates": str(row["date_cell_count"]),
            "numericTableCandidates": str(row["table_candidate_count"]),
            "warningOrError": str(row["error_text"]),
        }
        for row in rows
    ]
    columns = list(report_rows[0]) if report_rows else ["relativePath", "status"]
    temporary = batch_dir / "numeric-capture.csv.tmp"
    with temporary.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=columns)
        writer.writeheader()
        writer.writerows(report_rows)
    os.replace(temporary, batch_dir / "numeric-capture.csv")
    status_items = "".join(f"<li>{html.escape(name)}: {count}</li>" for name, count in sorted(statuses.items()))
    body_rows = "".join(
        "<tr>" + "".join(f"<td>{html.escape(row[column])}</td>" for column in columns) + "</tr>"
        for row in report_rows
    )
    document = f"""<!doctype html><html lang='ko'><head><meta charset='utf-8'><title>숫자 표 원본 적재</title>
<style>body{{font-family:Segoe UI,sans-serif;margin:24px;color:#172b4d}}table{{border-collapse:collapse;width:100%;font-size:12px}}th,td{{border:1px solid #d7dde7;padding:6px;text-align:left;vertical-align:top}}th{{background:#edf2f7}}</style></head>
<body><h1>숫자 표 원본 적재</h1><p>Excel/COM을 사용하지 않았습니다. 이 결과는 숫자 사실 적재 현황이며, 비교·품질 판정을 포함하지 않습니다.</p>
<h2>상태</h2><ul>{status_items}</ul><p><a href='numeric-capture.csv'>CSV</a> · <a href='numeric-capture-summary.json'>JSON</a></p>
<table><thead><tr>{''.join(f'<th>{html.escape(column)}</th>' for column in columns)}</tr></thead><tbody>{body_rows}</tbody></table></body></html>"""
    atomic_write_text(batch_dir / "numeric-capture.html", document)
    return summary


def run(args: argparse.Namespace) -> int:
    service_dir = Path(args.service_dir).resolve()
    if not service_dir.is_dir():
        raise ValueError(f"Service directory does not exist: {service_dir}")
    batch_dir = resolve_batch_directory(service_dir, args.structure_batch)
    if not (batch_dir / "batch.json").is_file():
        raise ValueError(f"Structure batch does not exist: {args.structure_batch}")
    records = source_rows(batch_dir)
    if not records:
        raise ValueError("The selected structure batch has no .xlsx/.xlsm source files.")
    if args.limit < 0 or args.progress_every < 0:
        raise ValueError("--limit must be non-negative.")
    connection = open_capture_db(batch_dir / "numeric-capture.sqlite")
    try:
        # Register the entire frozen source snapshot before a pilot run.  This
        # makes pending, changed, and captured workbook counts reconcilable and
        # lets the renderer create one status page per source workbook.
        for item in records:
            ensure_workbook(connection, item)
        connection.commit()
        selected = records[: args.limit] if args.limit else records
        run_counts: Counter[str] = Counter()
        for index, item in enumerate(selected, start=1):
            status = capture_workbook(connection, item, force=args.force)
            run_counts[status] += 1
            if status not in {"CAPTURED", "SKIPPED"} or (
                args.progress_every and (index == 1 or index == len(selected) or index % args.progress_every == 0)
            ):
                print(
                    json.dumps(
                        {"progress": f"{index}/{len(selected)}", "relativePath": item["relative_path"], "status": status},
                        ensure_ascii=False,
                    )
                )
        summary = write_outputs(batch_dir, connection)
    finally:
        connection.close()
    print(json.dumps({"status": "ok", "batchDirectory": str(batch_dir), "runStatusCounts": dict(sorted(run_counts.items())), "summary": summary}, ensure_ascii=False))
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Capture read-only numeric Excel facts into a batch-scoped SQLite database. Never uses Excel or COM.")
    parser.add_argument("--service-dir", required=True)
    parser.add_argument("--structure-batch", required=True)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--progress-every", type=int, default=25)
    parser.add_argument("--force", action="store_true", help="Recapture matching source files in the batch-scoped numeric DB.")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    try:
        return run(args)
    except (ValueError, RuntimeError) as exc:
        print(f"numeric-capture error: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
