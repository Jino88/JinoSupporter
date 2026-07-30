"""Deterministic, read-only Capture v2 ingestion for DRM-free ``.xlsx`` files.

This module captures source fidelity only.  It does not infer study meaning,
classify tables, extract images, or evaluate formulas.  The sparse cell walk
uses openpyxl's instantiated cell map and adds coordinates covered by merged
ranges; it deliberately does not iterate the full ``max_row * max_column``
rectangle.
"""

from __future__ import annotations

import hashlib
import json
import math
import sqlite3
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils.cell import (
        get_column_letter,
        range_boundaries,
    )
except ImportError:  # pragma: no cover - exercised through the explicit error
    load_workbook = None
    MergedCell = object  # type: ignore[assignment,misc]


CAPTURE_SCHEMA_VERSION = "input-data-openxml-capture-v2"
CAPTURE_CONTRACT = "openpyxl-sparse-source-capture-v2.0"
COM_CAPTURE_CONTRACT = "excel-com-fixed-grid-capture-v2.1"
SUPPORTED_CAPTURE_CONTRACTS = frozenset(
    {CAPTURE_CONTRACT, COM_CAPTURE_CONTRACT}
)
EXTRACTOR_NAME = "inference_data_ai_source_ingest"
EXTRACTOR_VERSION = "2.0"


class CaptureError(RuntimeError):
    """Base class for source-capture failures."""


class UnsupportedSourceError(CaptureError):
    """Raised when Capture v2 is asked to read a non-XLSX source."""


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def sha256_file(path: str | Path, chunk_size: int = 1024 * 1024) -> str:
    """Return a content-only SHA-256 fingerprint without modifying the file."""

    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        while chunk := stream.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest()


def _json_text(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _portable_value(value: Any) -> Any:
    """Convert an openpyxl value to deterministic, loss-aware JSON data."""

    if value is None or isinstance(value, (str, bool, int)):
        return value
    if isinstance(value, float):
        if math.isnan(value):
            return {"type": "float", "value": "NaN"}
        if math.isinf(value):
            return {"type": "float", "value": "Infinity" if value > 0 else "-Infinity"}
        return value
    if isinstance(value, datetime):
        return {"type": "datetime", "value": value.isoformat()}
    if isinstance(value, date):
        return {"type": "date", "value": value.isoformat()}
    if isinstance(value, time):
        return {"type": "time", "value": value.isoformat()}
    if isinstance(value, timedelta):
        return {"type": "timedelta", "seconds": value.total_seconds()}
    if isinstance(value, Decimal):
        return {"type": "decimal", "value": str(value)}
    if isinstance(value, bytes):
        return {"type": "bytes", "hex": value.hex()}
    return {"type": type(value).__name__, "value": str(value)}


def _compact(mapping: dict[str, Any]) -> dict[str, Any]:
    return {
        key: value
        for key, value in mapping.items()
        if value is not None and value != "" and value != [] and value != {}
    }


def _color_payload(color: Any) -> dict[str, Any]:
    if color is None:
        return {}
    color_type = getattr(color, "type", None)
    return _compact(
        {
            "type": color_type,
            "rgb": getattr(color, "rgb", None) if color_type == "rgb" else None,
            "indexed": getattr(color, "indexed", None) if color_type == "indexed" else None,
            "theme": getattr(color, "theme", None) if color_type == "theme" else None,
            "tint": getattr(color, "tint", None),
            "auto": getattr(color, "auto", None) if color_type == "auto" else None,
        }
    )


def _side_payload(side: Any) -> dict[str, Any]:
    if side is None:
        return {}
    return _compact(
        {
            "style": getattr(side, "style", None),
            "color": _color_payload(getattr(side, "color", None)),
        }
    )


def _style_payload(cell: Any) -> dict[str, Any]:
    font = getattr(cell, "font", None)
    fill = getattr(cell, "fill", None)
    border = getattr(cell, "border", None)
    alignment = getattr(cell, "alignment", None)
    protection = getattr(cell, "protection", None)
    return _compact(
        {
            "font": _compact(
                {
                    "name": getattr(font, "name", None),
                    "size": getattr(font, "sz", None),
                    "bold": getattr(font, "b", None),
                    "italic": getattr(font, "i", None),
                    "underline": getattr(font, "u", None),
                    "strike": getattr(font, "strike", None),
                    "vertAlign": getattr(font, "vertAlign", None),
                    "color": _color_payload(getattr(font, "color", None)),
                }
            ),
            "fill": _compact(
                {
                    "type": getattr(fill, "fill_type", None),
                    "fgColor": _color_payload(getattr(fill, "fgColor", None)),
                    "bgColor": _color_payload(getattr(fill, "bgColor", None)),
                }
            ),
            "border": _compact(
                {
                    side_name: _side_payload(getattr(border, side_name, None))
                    for side_name in (
                        "left",
                        "right",
                        "top",
                        "bottom",
                        "diagonal",
                        "vertical",
                        "horizontal",
                    )
                }
                | {
                    "outline": getattr(border, "outline", None),
                    "diagonalUp": getattr(border, "diagonalUp", None),
                    "diagonalDown": getattr(border, "diagonalDown", None),
                }
            ),
            "alignment": _compact(
                {
                    "horizontal": getattr(alignment, "horizontal", None),
                    "vertical": getattr(alignment, "vertical", None),
                    "textRotation": getattr(alignment, "textRotation", None),
                    "wrapText": getattr(alignment, "wrapText", None),
                    "shrinkToFit": getattr(alignment, "shrinkToFit", None),
                    "indent": getattr(alignment, "indent", None),
                    "relativeIndent": getattr(alignment, "relativeIndent", None),
                    "justifyLastLine": getattr(alignment, "justifyLastLine", None),
                    "readingOrder": getattr(alignment, "readingOrder", None),
                }
            ),
            "protection": _compact(
                {
                    "locked": getattr(protection, "locked", None),
                    "hidden": getattr(protection, "hidden", None),
                }
            ),
        }
    )


def _bounds_payload(min_row: int, min_column: int, max_row: int, max_column: int) -> dict[str, Any]:
    return {
        "minRow": min_row,
        "minColumn": min_column,
        "maxRow": max_row,
        "maxColumn": max_column,
        "rowCount": max_row - min_row + 1,
        "columnCount": max_column - min_column + 1,
        "address": (
            f"{get_column_letter(min_column)}{min_row}:"
            f"{get_column_letter(max_column)}{max_row}"
        ),
    }


def _bounds_from_coordinates(coordinates: Iterable[tuple[int, int]]) -> dict[str, Any] | None:
    iterator = iter(coordinates)
    try:
        first_row, first_column = next(iterator)
    except StopIteration:
        return None
    min_row = max_row = first_row
    min_column = max_column = first_column
    for row, column in iterator:
        min_row = min(min_row, row)
        min_column = min(min_column, column)
        max_row = max(max_row, row)
        max_column = max(max_column, column)
    return _bounds_payload(min_row, min_column, max_row, max_column)


def _reported_bounds(worksheet: Any) -> dict[str, Any] | None:
    instantiated = getattr(worksheet, "_cells", {})
    if not instantiated:
        return None
    dimension = worksheet.calculate_dimension()
    min_column, min_row, max_column, max_row = range_boundaries(dimension)
    return _bounds_payload(min_row, min_column, max_row, max_column)


def _merge_payloads(worksheet: Any) -> tuple[list[dict[str, Any]], dict[tuple[int, int], tuple[str, str]]]:
    ranges: list[tuple[int, int, int, int, str]] = []
    for merged_range in worksheet.merged_cells.ranges:
        min_column, min_row, max_column, max_row = range_boundaries(str(merged_range))
        ranges.append((min_row, min_column, max_row, max_column, str(merged_range)))
    ranges.sort()

    payloads: list[dict[str, Any]] = []
    membership: dict[tuple[int, int], tuple[str, str]] = {}
    for min_row, min_column, max_row, max_column, address in ranges:
        payloads.append(
            {
                **_bounds_payload(min_row, min_column, max_row, max_column),
                "anchor": f"{get_column_letter(min_column)}{min_row}",
            }
        )
        for row in range(min_row, max_row + 1):
            for column in range(min_column, max_column + 1):
                role = "anchor" if (row, column) == (min_row, min_column) else "covered"
                membership[(row, column)] = (address, role)
    return payloads, membership


def _cell_payload(
    cell: Any | None,
    cached_cell: Any | None,
    row: int,
    column: int,
    merge_info: tuple[str, str] | None,
) -> dict[str, Any]:
    is_placeholder = cell is None or isinstance(cell, MergedCell)
    value = None if cell is None else getattr(cell, "value", None)
    data_type = "n" if cell is None else str(getattr(cell, "data_type", "n") or "n")
    is_formula = not is_placeholder and data_type == "f"
    cached_value = getattr(cached_cell, "value", None) if is_formula and cached_cell is not None else None
    cached_data_type = (
        str(getattr(cached_cell, "data_type", "") or "")
        if is_formula and cached_cell is not None and cached_value is not None
        else None
    )

    # Formula text is kept only in formula; it is never presented as a display
    # value.  openpyxl does not calculate formulas, so a missing cached value is
    # represented explicitly as null.
    raw_value = None if is_formula else _portable_value(value)
    formula_value = getattr(value, "text", value)
    formula = str(formula_value) if is_formula and formula_value is not None else None
    display_value = _portable_value(cached_value) if is_formula else raw_value

    style_id = int(getattr(cell, "style_id", 0) or 0) if cell is not None else 0
    number_format = str(getattr(cell, "number_format", "General") or "General") if cell is not None else "General"
    payload = {
        "row": row,
        "column": column,
        "coordinate": f"{get_column_letter(column)}{row}",
        "rawValue": raw_value,
        "formula": formula,
        "cachedValue": _portable_value(cached_value) if is_formula else None,
        "displayValue": display_value,
        "dataType": data_type,
        "cachedDataType": cached_data_type,
        "numberFormat": number_format,
        "styleId": style_id,
        "style": _style_payload(cell) if cell is not None else {},
        "mergeRange": merge_info[0] if merge_info else None,
        "mergeRole": merge_info[1] if merge_info else "none",
    }
    return payload


def _row_dimensions(worksheet: Any) -> list[dict[str, Any]]:
    dimensions: list[dict[str, Any]] = []
    for index, dimension in sorted(worksheet.row_dimensions.items()):
        if dimension.height is None and not bool(dimension.hidden):
            continue
        dimensions.append(
            {
                "row": int(index),
                "height": dimension.height,
                "hidden": bool(dimension.hidden),
            }
        )
    return dimensions


def _column_dimensions(worksheet: Any) -> list[dict[str, Any]]:
    dimensions: list[dict[str, Any]] = []
    ordered = sorted(
        worksheet.column_dimensions.items(),
        key=lambda item: (int(getattr(item[1], "min", 0) or 0), item[0]),
    )
    for key, dimension in ordered:
        if dimension.width is None and not bool(dimension.hidden):
            continue
        min_column = int(getattr(dimension, "min", 0) or 0)
        max_column = int(getattr(dimension, "max", 0) or min_column)
        dimensions.append(
            {
                "key": str(key),
                "minColumn": min_column,
                "maxColumn": max_column,
                "width": dimension.width,
                "hidden": bool(dimension.hidden),
            }
        )
    return dimensions


def _freeze_panes(worksheet: Any) -> str | None:
    pane = worksheet.freeze_panes
    if pane is None:
        return None
    return str(getattr(pane, "coordinate", pane))


def _defined_names(workbook: Any) -> list[dict[str, Any]]:
    names: list[dict[str, Any]] = []
    for defined_name in workbook.defined_names.values():
        names.append(
            _compact(
                {
                    "name": getattr(defined_name, "name", None),
                    "value": getattr(defined_name, "attr_text", None),
                    "localSheetId": getattr(defined_name, "localSheetId", None),
                    "hidden": getattr(defined_name, "hidden", None),
                    "function": getattr(defined_name, "function", None),
                    "vbProcedure": getattr(defined_name, "vbProcedure", None),
                }
            )
        )
    names.sort(key=lambda item: (str(item.get("name", "")), int(item.get("localSheetId", -1))))
    return names


def _workbook_metadata(workbook: Any) -> dict[str, Any]:
    properties = workbook.properties
    calculation = getattr(workbook, "calculation", None)
    return {
        "properties": _compact(
            {
                "title": properties.title,
                "subject": properties.subject,
                "creator": properties.creator,
                "keywords": properties.keywords,
                "description": properties.description,
                "lastModifiedBy": properties.lastModifiedBy,
                "category": properties.category,
                "contentStatus": properties.contentStatus,
                "identifier": properties.identifier,
                "language": properties.language,
                "created": _portable_value(properties.created),
                "modified": _portable_value(properties.modified),
                "version": properties.version,
                "revision": properties.revision,
            }
        ),
        "dateEpoch": str(workbook.epoch),
        "activeSheetIndex": int(workbook.index(workbook.active)),
        "calculation": _compact(
            {
                "mode": getattr(calculation, "calcMode", None),
                "fullCalcOnLoad": getattr(calculation, "fullCalcOnLoad", None),
                "forceFullCalc": getattr(calculation, "forceFullCalc", None),
            }
        ),
        "definedNames": _defined_names(workbook),
    }


def _sheet_payload(
    worksheet: Any,
    cached_worksheet: Any | None,
    sheet_index: int,
) -> dict[str, Any]:
    merges, merge_membership = _merge_payloads(worksheet)
    instantiated = getattr(worksheet, "_cells", {})

    nonempty_coordinates: set[tuple[int, int]] = set()
    formula_coordinates: set[tuple[int, int]] = set()
    structural_coordinates: set[tuple[int, int]] = set()
    for coordinate, cell in instantiated.items():
        if isinstance(cell, MergedCell):
            continue
        value = getattr(cell, "value", None)
        if getattr(cell, "data_type", None) == "f":
            nonempty_coordinates.add(coordinate)
            formula_coordinates.add(coordinate)
        elif value is not None:
            nonempty_coordinates.add(coordinate)
        if (
            bool(getattr(cell, "has_style", False))
            or getattr(cell, "comment", None) is not None
            or getattr(cell, "hyperlink", None) is not None
        ):
            structural_coordinates.add(coordinate)

    captured_coordinates = nonempty_coordinates | structural_coordinates | set(merge_membership)
    cached_cells = getattr(cached_worksheet, "_cells", {}) if cached_worksheet is not None else {}
    cells = [
        _cell_payload(
            instantiated.get((row, column)),
            cached_cells.get((row, column)),
            row,
            column,
            merge_membership.get((row, column)),
        )
        for row, column in sorted(captured_coordinates)
    ]

    nonempty_rows = {row for row, _ in nonempty_coordinates}
    nonempty_columns = {column for _, column in nonempty_coordinates}
    has_tabular_evidence = (
        len(nonempty_coordinates) >= 4
        and len(nonempty_rows) >= 2
        and len(nonempty_columns) >= 2
    )
    is_truly_empty = not nonempty_coordinates and not merges
    status = (
        "EMPTY"
        if is_truly_empty
        else "CAPTURED"
        if has_tabular_evidence
        else "NO_TABULAR_EVIDENCE"
    )

    auto_filter = str(worksheet.auto_filter.ref) if worksheet.auto_filter.ref else None
    return {
        "sheetIndex": sheet_index,
        "title": worksheet.title,
        "sheetState": worksheet.sheet_state,
        "status": status,
        "isTrulyEmpty": is_truly_empty,
        "hasTabularEvidence": has_tabular_evidence,
        "usedBounds": _reported_bounds(worksheet),
        "contentBounds": _bounds_from_coordinates(sorted(captured_coordinates)),
        "nonEmptyCellCount": len(nonempty_coordinates),
        "structuralCellCount": len(structural_coordinates - nonempty_coordinates),
        "capturedCellCount": len(captured_coordinates),
        "formulaCellCount": len(formula_coordinates),
        "mergeCount": len(merges),
        "freezePanes": _freeze_panes(worksheet),
        "autoFilter": auto_filter,
        "sheetMetadata": {
            "selected": bool(getattr(worksheet.sheet_view, "tabSelected", False)),
            "showGridLines": getattr(worksheet.sheet_view, "showGridLines", None),
            "zoomScale": getattr(worksheet.sheet_view, "zoomScale", None),
        },
        "rowDimensions": _row_dimensions(worksheet),
        "columnDimensions": _column_dimensions(worksheet),
        "mergedRanges": merges,
        "cells": cells,
    }


def _apply_cached_formula_values(sheet: dict[str, Any], cached_worksheet: Any) -> None:
    cached_cells = getattr(cached_worksheet, "_cells", {})
    for cell in sheet["cells"]:
        if not cell.get("formula"):
            continue
        coordinate = (int(cell["row"]), int(cell["column"]))
        cached_cell = cached_cells.get(coordinate)
        cached_value = getattr(cached_cell, "value", None) if cached_cell is not None else None
        cell["cachedValue"] = _portable_value(cached_value)
        cell["displayValue"] = _portable_value(cached_value)
        cell["cachedDataType"] = (
            str(getattr(cached_cell, "data_type", "") or "")
            if cached_cell is not None and cached_value is not None
            else None
        )


def extract_workbook(source_path: str | Path) -> dict[str, Any]:
    """Extract an XLSX source into the deterministic Capture v2 JSON contract."""

    source = Path(source_path).expanduser().resolve()
    if source.suffix.casefold() != ".xlsx":
        raise UnsupportedSourceError("Capture v2 accepts DRM-free .xlsx files only.")
    if not source.is_file():
        raise FileNotFoundError(source)
    if load_workbook is None:
        raise CaptureError("openpyxl is required for Capture v2.")

    before = source.stat()
    content_sha256 = sha256_file(source)
    formula_workbook = None
    cached_workbook = None
    try:
        # Normal (non-read-only) mode is intentional: it exposes sparse
        # instantiated cells, merged ranges, dimensions, and style metadata.
        formula_workbook = load_workbook(
            source,
            read_only=False,
            data_only=False,
            keep_links=False,
        )
        sheet_names = list(formula_workbook.sheetnames)
        metadata = _workbook_metadata(formula_workbook)
        sheets = [
            _sheet_payload(
                formula_workbook[title],
                None,
                index,
            )
            for index, title in enumerate(sheet_names, start=1)
        ]
        # Do not retain two complete workbook object graphs simultaneously.
        formula_workbook.close()
        formula_workbook = None

        cached_workbook = load_workbook(
            source,
            read_only=False,
            data_only=True,
            keep_links=False,
        )
        if sheet_names != cached_workbook.sheetnames:
            raise CaptureError("Formula and cached workbook views have different sheet orders.")
        for sheet, title in zip(sheets, sheet_names):
            _apply_cached_formula_values(sheet, cached_workbook[title])

        tabular_sheet_count = sum(bool(sheet["hasTabularEvidence"]) for sheet in sheets)
        nonempty_sheet_count = sum(not bool(sheet["isTrulyEmpty"]) for sheet in sheets)
        if not nonempty_sheet_count:
            status = "EMPTY_WORKBOOK"
        elif not tabular_sheet_count:
            status = "NO_TABULAR_EVIDENCE"
        else:
            status = "CAPTURED"

        payload = {
            "schemaVersion": CAPTURE_SCHEMA_VERSION,
            "captureContract": CAPTURE_CONTRACT,
            "extractor": {
                "name": EXTRACTOR_NAME,
                "version": EXTRACTOR_VERSION,
                "formulaEvaluation": False,
                "imageHandling": "IGNORED",
                "cellTraversal": "SPARSE_INSTANTIATED_PLUS_MERGE_COORDINATES",
                "tabularEvidenceRule": "AT_LEAST_4_VALUES_ACROSS_2_ROWS_AND_2_COLUMNS",
            },
            "source": {
                "sourcePath": str(source),
                "fileName": source.name,
                "extension": source.suffix.lower(),
                "sizeBytes": before.st_size,
                "mtimeNs": before.st_mtime_ns,
                "contentSha256": content_sha256,
            },
            "workbook": {
                "status": status,
                "isTrulyEmpty": nonempty_sheet_count == 0,
                "sheetCount": len(sheets),
                "nonEmptySheetCount": nonempty_sheet_count,
                "tabularSheetCount": tabular_sheet_count,
                "metadata": metadata,
                "sheets": sheets,
            },
        }
    finally:
        if formula_workbook is not None:
            formula_workbook.close()
        if cached_workbook is not None:
            cached_workbook.close()

    after = source.stat()
    if (
        before.st_size != after.st_size
        or before.st_mtime_ns != after.st_mtime_ns
        or content_sha256 != sha256_file(source)
    ):
        raise CaptureError("Source changed while Capture v2 was reading it.")
    return payload


def capture_json_bytes(payload: dict[str, Any]) -> bytes:
    """Serialize a capture deterministically for storage or hashing."""

    return (_json_text(payload) + "\n").encode("utf-8")


def write_capture_json(payload: dict[str, Any], destination: str | Path) -> Path:
    """Write a deterministic JSON artifact using an atomic same-directory replace."""

    target = Path(destination).expanduser().resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_name(f".{target.name}.tmp")
    temporary.write_bytes(capture_json_bytes(payload))
    temporary.replace(target)
    return target


def read_capture_json(path: str | Path) -> dict[str, Any]:
    with Path(path).open("r", encoding="utf-8") as stream:
        payload = json.load(stream)
    _validate_payload(payload)
    return payload


def _validate_payload(payload: dict[str, Any]) -> None:
    if payload.get("schemaVersion") != CAPTURE_SCHEMA_VERSION:
        raise CaptureError("Unsupported Capture v2 schemaVersion.")
    if payload.get("captureContract") not in SUPPORTED_CAPTURE_CONTRACTS:
        raise CaptureError("Unsupported Capture v2 captureContract.")
    source = payload.get("source")
    workbook = payload.get("workbook")
    if not isinstance(source, dict) or not isinstance(workbook, dict):
        raise CaptureError("Capture payload requires source and workbook objects.")
    fingerprint = source.get("contentSha256")
    if not isinstance(fingerprint, str) or len(fingerprint) != 64:
        raise CaptureError("Capture payload requires a SHA-256 content fingerprint.")
    if not isinstance(workbook.get("sheets"), list):
        raise CaptureError("Capture payload requires a workbook.sheets list.")


def capture_v2_ddl() -> str:
    """Return the additive SQLite Capture v2 migration DDL."""

    return """
    PRAGMA foreign_keys = ON;

    CREATE TABLE IF NOT EXISTS capture_v2_documents (
        document_id INTEGER PRIMARY KEY AUTOINCREMENT,
        source_path TEXT NOT NULL UNIQUE,
        file_name TEXT NOT NULL,
        source_kind TEXT NOT NULL DEFAULT 'XLSX',
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS capture_v2_runs (
        run_id INTEGER PRIMARY KEY AUTOINCREMENT,
        dataset TEXT NOT NULL,
        input_path TEXT NOT NULL,
        started_at TEXT NOT NULL,
        finished_at TEXT NOT NULL DEFAULT '',
        total_files INTEGER NOT NULL DEFAULT 0,
        succeeded INTEGER NOT NULL DEFAULT 0,
        failed INTEGER NOT NULL DEFAULT 0,
        skipped INTEGER NOT NULL DEFAULT 0,
        reactivated INTEGER NOT NULL DEFAULT 0,
        options_json TEXT NOT NULL DEFAULT '{}'
    );

    CREATE TABLE IF NOT EXISTS capture_v2_ingest_items (
        ingest_item_id INTEGER PRIMARY KEY AUTOINCREMENT,
        run_id INTEGER NOT NULL,
        source_path TEXT NOT NULL,
        content_sha256 TEXT NOT NULL DEFAULT '',
        action TEXT NOT NULL,
        capture_revision_id INTEGER,
        canonical_revision_id INTEGER,
        message TEXT NOT NULL DEFAULT '',
        started_at TEXT NOT NULL,
        finished_at TEXT NOT NULL DEFAULT '',
        FOREIGN KEY(run_id) REFERENCES capture_v2_runs(run_id) ON DELETE CASCADE,
        UNIQUE(run_id, source_path)
    );

    CREATE INDEX IF NOT EXISTS idx_capture_v2_items_action
        ON capture_v2_ingest_items(run_id, action);

    CREATE TABLE IF NOT EXISTS capture_v2_revisions (
        revision_id INTEGER PRIMARY KEY AUTOINCREMENT,
        revision_uid TEXT NOT NULL UNIQUE,
        document_id INTEGER NOT NULL,
        content_sha256 TEXT NOT NULL,
        capture_contract TEXT NOT NULL,
        extractor_name TEXT NOT NULL,
        extractor_version TEXT NOT NULL,
        size_bytes INTEGER NOT NULL,
        mtime_ns INTEGER NOT NULL,
        capture_status TEXT NOT NULL
            CHECK(capture_status IN ('CAPTURED','NO_TABULAR_EVIDENCE','EMPTY_WORKBOOK','STALE')),
        is_current INTEGER NOT NULL DEFAULT 1 CHECK(is_current IN (0,1)),
        capture_json_sha256 TEXT NOT NULL,
        captured_at TEXT NOT NULL,
        stale_at TEXT,
        FOREIGN KEY(document_id) REFERENCES capture_v2_documents(document_id) ON DELETE CASCADE,
        UNIQUE(document_id, content_sha256, capture_contract)
    );

    CREATE UNIQUE INDEX IF NOT EXISTS idx_capture_v2_one_current_revision
        ON capture_v2_revisions(document_id) WHERE is_current=1;
    CREATE INDEX IF NOT EXISTS idx_capture_v2_revision_sha
        ON capture_v2_revisions(content_sha256, capture_contract);

    CREATE TABLE IF NOT EXISTS capture_v2_workbooks (
        revision_id INTEGER PRIMARY KEY,
        workbook_status TEXT NOT NULL,
        is_truly_empty INTEGER NOT NULL CHECK(is_truly_empty IN (0,1)),
        sheet_count INTEGER NOT NULL,
        nonempty_sheet_count INTEGER NOT NULL,
        tabular_sheet_count INTEGER NOT NULL,
        metadata_json TEXT NOT NULL,
        FOREIGN KEY(revision_id) REFERENCES capture_v2_revisions(revision_id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS capture_v2_sheets (
        sheet_id INTEGER PRIMARY KEY AUTOINCREMENT,
        revision_id INTEGER NOT NULL,
        sheet_index INTEGER NOT NULL,
        title TEXT NOT NULL,
        sheet_state TEXT NOT NULL,
        capture_status TEXT NOT NULL,
        is_truly_empty INTEGER NOT NULL CHECK(is_truly_empty IN (0,1)),
        has_tabular_evidence INTEGER NOT NULL CHECK(has_tabular_evidence IN (0,1)),
        nonempty_cell_count INTEGER NOT NULL,
        structural_cell_count INTEGER NOT NULL DEFAULT 0,
        captured_cell_count INTEGER NOT NULL,
        formula_cell_count INTEGER NOT NULL,
        merge_count INTEGER NOT NULL,
        used_bounds_json TEXT,
        content_bounds_json TEXT,
        freeze_panes TEXT,
        auto_filter TEXT,
        metadata_json TEXT NOT NULL,
        FOREIGN KEY(revision_id) REFERENCES capture_v2_revisions(revision_id) ON DELETE CASCADE,
        UNIQUE(revision_id, sheet_index)
    );

    CREATE INDEX IF NOT EXISTS idx_capture_v2_sheet_revision
        ON capture_v2_sheets(revision_id, sheet_index);

    CREATE TABLE IF NOT EXISTS capture_v2_cells (
        sheet_id INTEGER NOT NULL,
        row_index INTEGER NOT NULL,
        column_index INTEGER NOT NULL,
        coordinate TEXT NOT NULL,
        raw_value_json TEXT,
        formula_text TEXT,
        cached_value_json TEXT,
        display_value_json TEXT,
        data_type TEXT NOT NULL,
        cached_data_type TEXT,
        number_format TEXT NOT NULL,
        style_id INTEGER NOT NULL,
        style_json TEXT NOT NULL,
        merge_range TEXT,
        merge_role TEXT NOT NULL,
        PRIMARY KEY(sheet_id, row_index, column_index),
        FOREIGN KEY(sheet_id) REFERENCES capture_v2_sheets(sheet_id) ON DELETE CASCADE
    ) WITHOUT ROWID;

    CREATE INDEX IF NOT EXISTS idx_capture_v2_cell_coordinate
        ON capture_v2_cells(sheet_id, coordinate);

    CREATE TABLE IF NOT EXISTS capture_v2_merged_ranges (
        sheet_id INTEGER NOT NULL,
        address TEXT NOT NULL,
        min_row INTEGER NOT NULL,
        min_column INTEGER NOT NULL,
        max_row INTEGER NOT NULL,
        max_column INTEGER NOT NULL,
        anchor_coordinate TEXT NOT NULL,
        PRIMARY KEY(sheet_id, address),
        FOREIGN KEY(sheet_id) REFERENCES capture_v2_sheets(sheet_id) ON DELETE CASCADE
    ) WITHOUT ROWID;

    CREATE TABLE IF NOT EXISTS capture_v2_row_dimensions (
        sheet_id INTEGER NOT NULL,
        row_index INTEGER NOT NULL,
        height REAL,
        hidden INTEGER NOT NULL CHECK(hidden IN (0,1)),
        PRIMARY KEY(sheet_id, row_index),
        FOREIGN KEY(sheet_id) REFERENCES capture_v2_sheets(sheet_id) ON DELETE CASCADE
    ) WITHOUT ROWID;

    CREATE TABLE IF NOT EXISTS capture_v2_column_dimensions (
        sheet_id INTEGER NOT NULL,
        dimension_key TEXT NOT NULL,
        min_column INTEGER NOT NULL,
        max_column INTEGER NOT NULL,
        width REAL,
        hidden INTEGER NOT NULL CHECK(hidden IN (0,1)),
        PRIMARY KEY(sheet_id, dimension_key),
        FOREIGN KEY(sheet_id) REFERENCES capture_v2_sheets(sheet_id) ON DELETE CASCADE
    ) WITHOUT ROWID;
    """


def ensure_capture_v2_schema(connection: sqlite3.Connection) -> None:
    already_installed = (
        connection.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='capture_v2_revisions'"
        ).fetchone()
        is not None
    )
    if not already_installed:
        connection.executescript(capture_v2_ddl())
    columns = {
        str(row[1])
        for row in connection.execute("PRAGMA table_info(capture_v2_sheets)")
    }
    if "structural_cell_count" not in columns:
        connection.execute(
            "ALTER TABLE capture_v2_sheets ADD COLUMN structural_cell_count INTEGER NOT NULL DEFAULT 0"
        )


def _revision_uid(
    source_path: str,
    content_sha256: str,
    capture_contract: str = CAPTURE_CONTRACT,
) -> str:
    digest = hashlib.sha256(
        f"{source_path}\x1f{content_sha256}\x1f{capture_contract}".encode("utf-8")
    ).hexdigest()
    return f"capture_revision_{digest[:24]}"


def _json_column(value: Any) -> str | None:
    return None if value is None else _json_text(value)


def _insert_cells(
    connection: sqlite3.Connection,
    sheet_id: int,
    cells: Iterable[dict[str, Any]],
) -> None:
    connection.executemany(
        """
        INSERT INTO capture_v2_cells(
            sheet_id, row_index, column_index, coordinate,
            raw_value_json, formula_text, cached_value_json, display_value_json,
            data_type, cached_data_type, number_format, style_id, style_json,
            merge_range, merge_role
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            (
                sheet_id,
                int(cell["row"]),
                int(cell["column"]),
                str(cell["coordinate"]),
                _json_column(cell.get("rawValue")),
                cell.get("formula"),
                _json_column(cell.get("cachedValue")),
                _json_column(cell.get("displayValue")),
                str(cell.get("dataType") or ""),
                cell.get("cachedDataType"),
                str(cell.get("numberFormat") or "General"),
                int(cell.get("styleId") or 0),
                _json_text(cell.get("style") or {}),
                cell.get("mergeRange"),
                str(cell.get("mergeRole") or "none"),
            )
            for cell in cells
        ),
    )


def _insert_sheet(
    connection: sqlite3.Connection,
    revision_id: int,
    sheet: dict[str, Any],
) -> int:
    cursor = connection.execute(
        """
        INSERT INTO capture_v2_sheets(
            revision_id, sheet_index, title, sheet_state, capture_status,
            is_truly_empty, has_tabular_evidence, nonempty_cell_count,
            structural_cell_count, captured_cell_count, formula_cell_count, merge_count,
            used_bounds_json, content_bounds_json, freeze_panes, auto_filter,
            metadata_json
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            revision_id,
            int(sheet["sheetIndex"]),
            str(sheet["title"]),
            str(sheet["sheetState"]),
            str(sheet["status"]),
            int(bool(sheet["isTrulyEmpty"])),
            int(bool(sheet["hasTabularEvidence"])),
            int(sheet["nonEmptyCellCount"]),
            int(sheet.get("structuralCellCount") or 0),
            int(sheet["capturedCellCount"]),
            int(sheet["formulaCellCount"]),
            int(sheet["mergeCount"]),
            _json_column(sheet.get("usedBounds")),
            _json_column(sheet.get("contentBounds")),
            sheet.get("freezePanes"),
            sheet.get("autoFilter"),
            _json_text(sheet.get("sheetMetadata") or {}),
        ),
    )
    sheet_id = int(cursor.lastrowid)
    _insert_cells(connection, sheet_id, sheet.get("cells") or [])

    connection.executemany(
        """
        INSERT INTO capture_v2_merged_ranges(
            sheet_id, address, min_row, min_column, max_row, max_column,
            anchor_coordinate
        ) VALUES (?,?,?,?,?,?,?)
        """,
        (
            (
                sheet_id,
                str(merged["address"]),
                int(merged["minRow"]),
                int(merged["minColumn"]),
                int(merged["maxRow"]),
                int(merged["maxColumn"]),
                str(merged["anchor"]),
            )
            for merged in sheet.get("mergedRanges") or []
        ),
    )
    connection.executemany(
        """
        INSERT INTO capture_v2_row_dimensions(sheet_id, row_index, height, hidden)
        VALUES (?,?,?,?)
        """,
        (
            (
                sheet_id,
                int(dimension["row"]),
                dimension.get("height"),
                int(bool(dimension.get("hidden"))),
            )
            for dimension in sheet.get("rowDimensions") or []
        ),
    )
    connection.executemany(
        """
        INSERT INTO capture_v2_column_dimensions(
            sheet_id, dimension_key, min_column, max_column, width, hidden
        ) VALUES (?,?,?,?,?,?)
        """,
        (
            (
                sheet_id,
                str(dimension["key"]),
                int(dimension["minColumn"]),
                int(dimension["maxColumn"]),
                dimension.get("width"),
                int(bool(dimension.get("hidden"))),
            )
            for dimension in sheet.get("columnDimensions") or []
        ),
    )
    return sheet_id


def import_capture(
    connection: sqlite3.Connection,
    payload: dict[str, Any],
    *,
    captured_at: str | None = None,
) -> dict[str, Any]:
    """Atomically import one workbook capture.

    The same current ``contentSha256 + captureContract`` is skipped.  A changed
    source makes the previous current revision stale.  If source bytes later
    revert to a known revision, that complete revision is reactivated instead
    of duplicating its cell rows.
    """

    _validate_payload(payload)
    ensure_capture_v2_schema(connection)
    timestamp = captured_at or utc_now_iso()
    source = payload["source"]
    workbook = payload["workbook"]
    capture_contract = str(payload["captureContract"])
    extractor = payload.get("extractor") or {}
    extractor_name = str(extractor.get("name") or EXTRACTOR_NAME)
    extractor_version = str(
        extractor.get("version") or EXTRACTOR_VERSION
    )
    source_path = str(Path(source["sourcePath"]).expanduser().resolve())
    content_sha256 = str(source["contentSha256"])
    source_kind = str(
        source.get("extension") or Path(source_path).suffix or "EXCEL"
    ).lstrip(".").upper()
    capture_json_sha256 = hashlib.sha256(capture_json_bytes(payload)).hexdigest()

    connection.execute("SAVEPOINT capture_v2_workbook")
    try:
        connection.execute(
            """
            INSERT INTO capture_v2_documents(
                source_path, file_name, source_kind, created_at, updated_at
            ) VALUES (?,?,?,?,?)
            ON CONFLICT(source_path) DO UPDATE SET
                file_name=excluded.file_name,
                source_kind=excluded.source_kind,
                updated_at=excluded.updated_at
            """,
            (
                source_path,
                str(source["fileName"]),
                source_kind,
                timestamp,
                timestamp,
            ),
        )
        document_id = int(
            connection.execute(
                "SELECT document_id FROM capture_v2_documents WHERE source_path=?",
                (source_path,),
            ).fetchone()[0]
        )

        current = connection.execute(
            """
            SELECT revision_id, content_sha256, capture_contract
            FROM capture_v2_revisions
            WHERE document_id=? AND is_current=1
            """,
            (document_id,),
        ).fetchone()
        if (
            current is not None
            and str(current[1]) == content_sha256
            and str(current[2]) == capture_contract
        ):
            connection.execute("RELEASE SAVEPOINT capture_v2_workbook")
            return {
                "action": "SKIPPED",
                "documentId": document_id,
                "revisionId": int(current[0]),
                "contentSha256": content_sha256,
                "captureContract": capture_contract,
            }

        connection.execute(
            """
            UPDATE capture_v2_revisions
            SET is_current=0, capture_status='STALE', stale_at=?
            WHERE document_id=? AND is_current=1
            """,
            (timestamp, document_id),
        )

        known = connection.execute(
            """
            SELECT revision_id
            FROM capture_v2_revisions
            WHERE document_id=? AND content_sha256=? AND capture_contract=?
            """,
            (document_id, content_sha256, capture_contract),
        ).fetchone()
        if known is not None:
            revision_id = int(known[0])
            connection.execute(
                """
                UPDATE capture_v2_revisions
                SET is_current=1, capture_status=?, size_bytes=?, mtime_ns=?,
                    capture_json_sha256=?, captured_at=?, stale_at=NULL
                WHERE revision_id=?
                """,
                (
                    str(workbook["status"]),
                    int(source["sizeBytes"]),
                    int(source["mtimeNs"]),
                    capture_json_sha256,
                    timestamp,
                    revision_id,
                ),
            )
            action = "REACTIVATED"
        else:
            cursor = connection.execute(
                """
                INSERT INTO capture_v2_revisions(
                    revision_uid, document_id, content_sha256, capture_contract,
                    extractor_name, extractor_version, size_bytes, mtime_ns,
                    capture_status, is_current, capture_json_sha256, captured_at
                ) VALUES (?,?,?,?,?,?,?,?,?,1,?,?)
                """,
                (
                    _revision_uid(
                        source_path,
                        content_sha256,
                        capture_contract,
                    ),
                    document_id,
                    content_sha256,
                    capture_contract,
                    extractor_name,
                    extractor_version,
                    int(source["sizeBytes"]),
                    int(source["mtimeNs"]),
                    str(workbook["status"]),
                    capture_json_sha256,
                    timestamp,
                ),
            )
            revision_id = int(cursor.lastrowid)
            connection.execute(
                """
                INSERT INTO capture_v2_workbooks(
                    revision_id, workbook_status, is_truly_empty, sheet_count,
                    nonempty_sheet_count, tabular_sheet_count, metadata_json
                ) VALUES (?,?,?,?,?,?,?)
                """,
                (
                    revision_id,
                    str(workbook["status"]),
                    int(bool(workbook["isTrulyEmpty"])),
                    int(workbook["sheetCount"]),
                    int(workbook["nonEmptySheetCount"]),
                    int(workbook["tabularSheetCount"]),
                    _json_text(workbook.get("metadata") or {}),
                ),
            )
            for sheet in workbook["sheets"]:
                _insert_sheet(connection, revision_id, sheet)
            action = "IMPORTED"

        connection.execute("RELEASE SAVEPOINT capture_v2_workbook")
        return {
            "action": action,
            "documentId": document_id,
            "revisionId": revision_id,
            "contentSha256": content_sha256,
            "captureContract": capture_contract,
        }
    except Exception:
        connection.execute("ROLLBACK TO SAVEPOINT capture_v2_workbook")
        connection.execute("RELEASE SAVEPOINT capture_v2_workbook")
        raise


def capture_and_import(
    connection: sqlite3.Connection,
    source_path: str | Path,
    *,
    captured_at: str | None = None,
) -> dict[str, Any]:
    payload = extract_workbook(source_path)
    return import_capture(connection, payload, captured_at=captured_at)


def bridge_capture_to_canonical_source(
    connection: sqlite3.Connection,
    *,
    dataset: str,
    payload: dict[str, Any],
    capture_result: dict[str, Any],
    captured_at: str | None = None,
) -> dict[str, Any]:
    """Link an imported Capture v2 revision to the canonical source layer.

    ``ensure_knowledge_schema`` must already have been applied by the caller.
    The bridge is intentionally separate from raw capture so this module stays
    usable with a standalone Capture v2 database.
    """

    from inference_data_ai_schema import stable_uid

    timestamp = captured_at or utc_now_iso()
    source = payload["source"]
    source_path = str(Path(source["sourcePath"]).expanduser().resolve())
    source_kind = str(
        source.get("extension") or Path(source_path).suffix or "EXCEL"
    ).lstrip(".").upper()
    capture_revision_id = int(capture_result["revisionId"])
    capture_revision = connection.execute(
        """
        SELECT revision_uid, content_sha256, capture_contract, extractor_name,
               extractor_version, size_bytes, mtime_ns, capture_status, is_current
        FROM capture_v2_revisions
        WHERE revision_id=?
        """,
        (capture_revision_id,),
    ).fetchone()
    if not capture_revision:
        raise CaptureError(f"Capture v2 revision is missing: {capture_revision_id}")
    source_revision_columns = {
        str(row[1])
        for row in connection.execute("PRAGMA table_info(source_revisions)")
    }
    if "capture_v2_revision_id" not in source_revision_columns:
        raise CaptureError("Canonical source schema is missing capture_v2_revision_id.")

    document_uid = stable_uid("document", dataset, source_path)
    connection.execute(
        """
        INSERT INTO source_documents(
            document_uid, dataset, source_path, original_file_name,
            source_kind, lifecycle_status, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, 'ACTIVE', ?, ?)
        ON CONFLICT(dataset, source_path) DO UPDATE SET
            original_file_name=excluded.original_file_name,
            lifecycle_status='ACTIVE',
            updated_at=excluded.updated_at
        """,
        (
            document_uid,
            dataset,
            source_path,
            str(source["fileName"]),
            source_kind,
            timestamp,
            timestamp,
        ),
    )
    document_id = int(
        connection.execute(
            "SELECT document_id FROM source_documents WHERE dataset=? AND source_path=?",
            (dataset, source_path),
        ).fetchone()[0]
    )
    if int(capture_revision["is_current"]):
        connection.execute(
            """
            UPDATE source_revisions
            SET is_current=0, capture_status='STALE'
            WHERE document_id=?
            """,
            (document_id,),
        )
    connection.execute(
        """
        INSERT INTO source_revisions(
            revision_uid, document_id, legacy_workbook_id,
            source_fingerprint, fingerprint_kind, content_sha256,
            size_bytes, mtime_ns, extractor_name, extractor_version,
            capture_contract, capture_status, source_content_status,
            is_current, captured_at, capture_v2_revision_id
        ) VALUES (
            ?, ?, NULL, ?, 'SHA256', ?, ?, ?, ?, ?, ?,
            'CAPTURED', ?, ?, ?, ?
        )
        ON CONFLICT(revision_uid) DO UPDATE SET
            document_id=excluded.document_id,
            source_fingerprint=excluded.source_fingerprint,
            fingerprint_kind='SHA256',
            content_sha256=excluded.content_sha256,
            size_bytes=excluded.size_bytes,
            mtime_ns=excluded.mtime_ns,
            extractor_name=excluded.extractor_name,
            extractor_version=excluded.extractor_version,
            capture_contract=excluded.capture_contract,
            capture_status=CASE
                WHEN excluded.is_current=1 THEN 'CAPTURED'
                ELSE 'STALE'
            END,
            source_content_status=excluded.source_content_status,
            is_current=excluded.is_current,
            captured_at=excluded.captured_at,
            capture_v2_revision_id=excluded.capture_v2_revision_id
        """,
        (
            str(capture_revision["revision_uid"]),
            document_id,
            str(capture_revision["content_sha256"]),
            str(capture_revision["content_sha256"]),
            int(capture_revision["size_bytes"]),
            int(capture_revision["mtime_ns"]),
            str(capture_revision["extractor_name"]),
            str(capture_revision["extractor_version"]),
            str(capture_revision["capture_contract"]),
            str(capture_revision["capture_status"]),
            int(capture_revision["is_current"]),
            timestamp,
            capture_revision_id,
        ),
    )
    canonical_revision_id = int(
        connection.execute(
            "SELECT revision_id FROM source_revisions WHERE revision_uid=?",
            (str(capture_revision["revision_uid"]),),
        ).fetchone()[0]
    )
    return {
        "documentId": document_id,
        "documentUid": document_uid,
        "revisionId": canonical_revision_id,
        "revisionUid": str(capture_revision["revision_uid"]),
        "captureV2RevisionId": capture_revision_id,
        "workbookStatus": str(payload["workbook"]["status"]),
    }


def verify_capture_revision(
    connection: sqlite3.Connection,
    revision_id: int,
    *,
    verify_source_sha256: bool = False,
) -> dict[str, Any]:
    revision = connection.execute(
        """
        SELECT r.*, d.source_path
        FROM capture_v2_revisions r
        JOIN capture_v2_documents d ON d.document_id=r.document_id
        WHERE r.revision_id=?
        """,
        (revision_id,),
    ).fetchone()
    if not revision:
        raise CaptureError(f"Capture v2 revision is missing: {revision_id}")
    errors: list[str] = []
    workbook = connection.execute(
        "SELECT * FROM capture_v2_workbooks WHERE revision_id=?",
        (revision_id,),
    ).fetchone()
    if not workbook:
        errors.append("workbook metadata row is missing")
        sheets: list[Any] = []
    else:
        sheets = connection.execute(
            "SELECT * FROM capture_v2_sheets WHERE revision_id=? ORDER BY sheet_index",
            (revision_id,),
        ).fetchall()
        if len(sheets) != int(workbook["sheet_count"]):
            errors.append(f"sheet count mismatch: {len(sheets)} != {workbook['sheet_count']}")
    for sheet in sheets:
        sheet_id = int(sheet["sheet_id"])
        actual_cells = int(
            connection.execute(
                "SELECT COUNT(*) FROM capture_v2_cells WHERE sheet_id=?",
                (sheet_id,),
            ).fetchone()[0]
        )
        actual_formulas = int(
            connection.execute(
                """
                SELECT COUNT(*) FROM capture_v2_cells
                WHERE sheet_id=? AND formula_text IS NOT NULL AND formula_text<>''
                """,
                (sheet_id,),
            ).fetchone()[0]
        )
        actual_merges = int(
            connection.execute(
                "SELECT COUNT(*) FROM capture_v2_merged_ranges WHERE sheet_id=?",
                (sheet_id,),
            ).fetchone()[0]
        )
        if actual_cells != int(sheet["captured_cell_count"]):
            errors.append(
                f"{sheet['title']}: cell count mismatch {actual_cells} != {sheet['captured_cell_count']}"
            )
        if actual_formulas != int(sheet["formula_cell_count"]):
            errors.append(
                f"{sheet['title']}: formula count mismatch {actual_formulas} != {sheet['formula_cell_count']}"
            )
        if actual_merges != int(sheet["merge_count"]):
            errors.append(
                f"{sheet['title']}: merge count mismatch {actual_merges} != {sheet['merge_count']}"
            )
    if verify_source_sha256:
        source_path = Path(str(revision["source_path"]))
        if not source_path.is_file():
            errors.append("source file is missing")
        else:
            actual_sha256 = sha256_file(source_path)
            if actual_sha256 != str(revision["content_sha256"]):
                errors.append("source SHA-256 no longer matches the captured revision")
    canonical_bridge_ok: bool | None = None
    source_revision_columns = {
        str(row[1])
        for row in connection.execute("PRAGMA table_info(source_revisions)")
    }
    if "capture_v2_revision_id" in source_revision_columns:
        canonical_bridge_ok = (
            connection.execute(
                """
                SELECT 1 FROM source_revisions
                WHERE capture_v2_revision_id=? AND revision_uid=?
                LIMIT 1
                """,
                (revision_id, str(revision["revision_uid"])),
            ).fetchone()
            is not None
        )
        if not canonical_bridge_ok:
            errors.append("canonical source revision bridge is missing")
    return {
        "revisionId": revision_id,
        "revisionUid": str(revision["revision_uid"]),
        "sourcePath": str(revision["source_path"]),
        "isCurrent": bool(revision["is_current"]),
        "ok": not errors,
        "errors": errors,
        "canonicalBridgeOk": canonical_bridge_ok,
        "sheetCount": len(sheets),
    }


def reconcile_capture_sheet_counts(
    connection: sqlite3.Connection,
    revision_id: int,
) -> dict[str, Any]:
    """Repair sheet aggregates from the capture cells actually persisted."""

    sheets = connection.execute(
        """
        SELECT sheet_id, title, nonempty_cell_count,
               structural_cell_count, captured_cell_count,
               formula_cell_count
        FROM capture_v2_sheets
        WHERE revision_id=?
        ORDER BY sheet_index
        """,
        (revision_id,),
    ).fetchall()
    repaired: list[dict[str, Any]] = []
    for sheet in sheets:
        sheet_id = int(sheet["sheet_id"])
        counts = connection.execute(
            """
            SELECT
                COUNT(*) AS captured,
                SUM(
                    CASE WHEN
                        raw_value_json IS NOT NULL
                        OR formula_text IS NOT NULL
                        OR cached_value_json IS NOT NULL
                        OR display_value_json IS NOT NULL
                    THEN 1 ELSE 0 END
                ) AS nonempty,
                SUM(
                    CASE WHEN
                        formula_text IS NOT NULL
                        AND formula_text<>''
                    THEN 1 ELSE 0 END
                ) AS formulas
            FROM capture_v2_cells
            WHERE sheet_id=?
            """,
            (sheet_id,),
        ).fetchone()
        captured_count = int(counts["captured"])
        nonempty_count = int(counts["nonempty"] or 0)
        formula_count = int(counts["formulas"] or 0)
        structural_count = captured_count - nonempty_count
        before = {
            "nonempty": int(sheet["nonempty_cell_count"]),
            "structural": int(sheet["structural_cell_count"]),
            "captured": int(sheet["captured_cell_count"]),
            "formulas": int(sheet["formula_cell_count"]),
        }
        after = {
            "nonempty": nonempty_count,
            "structural": structural_count,
            "captured": captured_count,
            "formulas": formula_count,
        }
        if before == after:
            continue
        connection.execute(
            """
            UPDATE capture_v2_sheets
            SET nonempty_cell_count=?,
                structural_cell_count=?,
                captured_cell_count=?,
                formula_cell_count=?
            WHERE sheet_id=?
            """,
            (
                nonempty_count,
                structural_count,
                captured_count,
                formula_count,
                sheet_id,
            ),
        )
        repaired.append(
            {
                "sheetId": sheet_id,
                "title": str(sheet["title"]),
                "before": before,
                "after": after,
            }
        )
    return {
        "revisionId": revision_id,
        "repairedSheetCount": len(repaired),
        "sheets": repaired,
    }


def verify_capture_database(
    connection: sqlite3.Connection,
    *,
    current_only: bool = True,
    verify_source_sha256: bool = False,
) -> dict[str, Any]:
    where = "WHERE is_current=1" if current_only else ""
    revision_ids = [
        int(row[0])
        for row in connection.execute(
            f"SELECT revision_id FROM capture_v2_revisions {where} ORDER BY revision_id"
        )
    ]
    revisions = [
        verify_capture_revision(
            connection,
            revision_id,
            verify_source_sha256=verify_source_sha256,
        )
        for revision_id in revision_ids
    ]
    unfinished_runs = int(
        connection.execute(
            "SELECT COUNT(*) FROM capture_v2_runs WHERE finished_at=''"
        ).fetchone()[0]
    )
    failed_items = int(
        connection.execute(
            "SELECT COUNT(*) FROM capture_v2_ingest_items WHERE action='FAILED'"
        ).fetchone()[0]
    )
    invalid = sum(not item["ok"] for item in revisions)
    return {
        "ok": invalid == 0 and unfinished_runs == 0,
        "checked": len(revisions),
        "valid": len(revisions) - invalid,
        "invalid": invalid,
        "unfinishedRuns": unfinished_runs,
        "failedItems": failed_items,
        "sourceSha256Verified": verify_source_sha256,
        "revisions": revisions,
    }


__all__ = [
    "CAPTURE_CONTRACT",
    "CAPTURE_SCHEMA_VERSION",
    "COM_CAPTURE_CONTRACT",
    "SUPPORTED_CAPTURE_CONTRACTS",
    "CaptureError",
    "UnsupportedSourceError",
    "capture_and_import",
    "capture_json_bytes",
    "capture_v2_ddl",
    "bridge_capture_to_canonical_source",
    "ensure_capture_v2_schema",
    "extract_workbook",
    "import_capture",
    "read_capture_json",
    "reconcile_capture_sheet_counts",
    "sha256_file",
    "verify_capture_database",
    "verify_capture_revision",
    "write_capture_json",
]
