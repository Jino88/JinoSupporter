from __future__ import annotations

import ctypes
import json
import os
import re
import shutil
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory

import win32com.client
import pythoncom
from PIL import Image, ImageGrab

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

MAX_PROGRAM_CELLS_PER_SHEET = int(os.environ.get("JINO_MAX_PROGRAM_CELLS_PER_SHEET", "100000"))
MAX_CELL_TEXT_CHARS = 500
MAX_RENDER_REGIONS_PER_SHEET = 12
FAST_CONFIDENCE_THRESHOLD = 70
_PROGRESS_PATH: Path | None = None

TABLE_HEADER_KEYWORDS = {
    "date",
    "line",
    "type",
    "input",
    "ok",
    "ng",
    "rate",
    "total",
    "no",
    "item",
    "defect",
    "result",
    "sigma",
    "hearing",
    "vision",
    "spl",
    "thd",
    "noise",
    "touch",
}


def progress(message: str) -> None:
    text = f"PROGRESS\t{message}"
    try:
        print(text, flush=True)
    except Exception:
        pass
    if _PROGRESS_PATH is None:
        return
    try:
        _PROGRESS_PATH.parent.mkdir(parents=True, exist_ok=True)
        with _PROGRESS_PATH.open("a", encoding="utf-8") as fh:
            fh.write(f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
    except Exception:
        pass


def safe_name(value: str) -> str:
    stem = Path(value).stem
    stem = re.sub(r"[^\w._-]+", "_", stem, flags=re.UNICODE).strip("._")
    return (stem or "workbook")[:140]


def column_name(column: int) -> str:
    name = ""
    while column > 0:
        column, rem = divmod(column - 1, 26)
        name = chr(65 + rem) + name
    return name or "A"


def cell_address(row: int, column: int) -> str:
    return f"{column_name(column)}{row}"


def range_address(row1: int, col1: int, row2: int, col2: int) -> str:
    start = cell_address(row1, col1)
    end = cell_address(row2, col2)
    return start if start == end else f"{start}:{end}"


def row_col_range_address(region: dict) -> str:
    return range_address(
        int(region["rowStart"]),
        int(region["columnStart"]),
        int(region["rowEnd"]),
        int(region["columnEnd"]),
    )


def clipped_text(value, limit: int = MAX_CELL_TEXT_CHARS) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if len(text) <= limit:
        return text
    return text[:limit] + "...[truncated]"


def excel_color_to_hex(value) -> str:
    try:
        color = int(value)
    except Exception:
        return ""
    if color < 0 or color in {0, 16777215}:
        return ""
    red = color & 0xFF
    green = (color >> 8) & 0xFF
    blue = (color >> 16) & 0xFF
    return f"#{red:02X}{green:02X}{blue:02X}"


def alignment_name(value) -> str:
    names = {
        -4108: "center",
        -4131: "left",
        -4152: "right",
        -4130: "justify",
        -4117: "distributed",
        1: "general",
    }
    try:
        return names.get(int(value), str(value))
    except Exception:
        return ""


def cell_display_text(cell) -> str:
    formula_text = ""
    try:
        formula_text = clipped_text(cell.Formula)
    except Exception:
        formula_text = ""

    # Excel shows narrow error cells as "#####"; keep the underlying error text.
    if formula_text.startswith("#") and formula_text.endswith("!"):
        return formula_text

    for attr in ("Text", "Value2", "Value"):
        try:
            text = clipped_text(getattr(cell, attr))
            if text:
                return text
        except Exception:
            continue
    return ""


def is_numeric_like(text: str) -> bool:
    value = (text or "").strip().replace(",", "")
    if not value:
        return False
    if value.endswith("%"):
        value = value[:-1]
    if value.startswith("(") and value.endswith(")"):
        value = "-" + value[1:-1]
    try:
        float(value)
        return True
    except Exception:
        return False


def has_table_keyword(text: str) -> bool:
    normalized = re.sub(r"[^A-Za-z0-9+/%]+", " ", text or "").lower()
    words = set(normalized.split())
    if words & TABLE_HEADER_KEYWORDS:
        return True
    return any(keyword in normalized for keyword in ("ng rate", "total ng", "no sound"))


def regions_overlap_or_touch(a: dict, b: dict, row_gap: int = 0, col_gap: int = 0) -> bool:
    return not (
        int(a["rowEnd"]) + row_gap < int(b["rowStart"])
        or int(b["rowEnd"]) + row_gap < int(a["rowStart"])
        or int(a["columnEnd"]) + col_gap < int(b["columnStart"])
        or int(b["columnEnd"]) + col_gap < int(a["columnStart"])
    )


def merge_region(a: dict, b: dict) -> dict:
    merged = dict(a)
    merged["rowStart"] = min(int(a["rowStart"]), int(b["rowStart"]))
    merged["rowEnd"] = max(int(a["rowEnd"]), int(b["rowEnd"]))
    merged["columnStart"] = min(int(a["columnStart"]), int(b["columnStart"]))
    merged["columnEnd"] = max(int(a["columnEnd"]), int(b["columnEnd"]))
    reasons = []
    for item in (a.get("reason", ""), b.get("reason", "")):
        if item and item not in reasons:
            reasons.append(item)
    if reasons:
        merged["reason"] = "; ".join(reasons)
    return merged


def column_overlap_ratio(a: dict, b: dict) -> float:
    left = max(int(a["columnStart"]), int(b["columnStart"]))
    right = min(int(a["columnEnd"]), int(b["columnEnd"]))
    overlap = max(0, right - left + 1)
    width = min(
        int(a["columnEnd"]) - int(a["columnStart"]) + 1,
        int(b["columnEnd"]) - int(b["columnStart"]) + 1,
    )
    return (overlap / width) if width > 0 else 0.0


def should_merge_table_regions(a: dict, b: dict) -> bool:
    if a.get("kind") != "table" or b.get("kind") != "table":
        return False
    row_gap = max(int(b["rowStart"]) - int(a["rowEnd"]) - 1, int(a["rowStart"]) - int(b["rowEnd"]) - 1, 0)
    return row_gap <= 1 and column_overlap_ratio(a, b) >= 0.65


def merge_lookup(merges: list[dict]) -> dict[tuple[int, int], dict]:
    lookup: dict[tuple[int, int], dict] = {}
    for merge in merges:
        try:
            for row in range(int(merge["rowStart"]), int(merge["rowEnd"]) + 1):
                for col in range(int(merge["columnStart"]), int(merge["columnEnd"]) + 1):
                    lookup[(row, col)] = merge
        except Exception:
            continue
    return lookup


def clear_clipboard() -> None:
    user32 = ctypes.windll.user32
    if user32.OpenClipboard(None):
        try:
            user32.EmptyClipboard()
        finally:
            user32.CloseClipboard()


def copy_range_png(ws, row1: int, col1: int, row2: int, col2: int, out_path: Path) -> tuple[int, int]:
    rng = ws.Range(ws.Cells(row1, col1), ws.Cells(row2, col2))
    last_error: Exception | None = None
    for _ in range(6):
        try:
            clear_clipboard()
            rng.CopyPicture(Appearance=1, Format=2)
            time.sleep(0.45)
            img = ImageGrab.grabclipboard()
            if img is None:
                raise RuntimeError("clipboard did not contain an image")
            out_path.parent.mkdir(parents=True, exist_ok=True)
            img.save(out_path)
            clear_clipboard()
            return img.size
        except Exception as exc:
            last_error = exc
            time.sleep(1.2)
    raise RuntimeError(f"failed to render range image: {last_error}")


def used_bounds(ws) -> tuple[int, int, int, int]:
    used = ws.UsedRange
    row1 = int(used.Row)
    col1 = int(used.Column)
    row2 = row1 + int(used.Rows.Count) - 1
    col2 = col1 + int(used.Columns.Count) - 1
    return row1, col1, row2, col2


def shape_ranges(ws, row1: int, col1: int, row2: int, col2: int) -> list[dict]:
    results: list[dict] = []
    try:
        count = int(ws.Shapes.Count)
    except Exception:
        return results

    for index in range(1, count + 1):
        try:
            shape = ws.Shapes(index)
            top_left = shape.TopLeftCell
            bottom_right = shape.BottomRightCell
            sr1 = max(row1, int(top_left.Row))
            sc1 = max(col1, int(top_left.Column))
            sr2 = min(row2, int(bottom_right.Row))
            sc2 = min(col2, int(bottom_right.Column))
            if sr2 < sr1 or sc2 < sc1:
                continue
            name = clipped_text(getattr(shape, "Name", ""), 160)
            shape_type = ""
            try:
                shape_type = str(int(shape.Type))
            except Exception:
                shape_type = ""
            results.append(
                {
                    "shapeIndex": index,
                    "name": name,
                    "type": shape_type,
                    "address": range_address(sr1, sc1, sr2, sc2),
                    "rowStart": sr1,
                    "rowEnd": sr2,
                    "columnStart": sc1,
                    "columnEnd": sc2,
                }
            )
        except Exception:
            continue

    results.sort(key=lambda item: (item["rowStart"], item["columnStart"], item["rowEnd"], item["columnEnd"]))
    return results


def normalize_merge_record(
    row_start: int,
    column_start: int,
    row_end: int,
    column_end: int,
    text: str,
    source: str,
) -> dict:
    return {
        "address": range_address(row_start, column_start, row_end, column_end),
        "rowStart": row_start,
        "rowEnd": row_end,
        "columnStart": column_start,
        "columnEnd": column_end,
        "text": text,
        "source": source,
    }


def merged_cells_from_com(ws, row1: int, col1: int, row2: int, col2: int) -> list[dict]:
    results: list[dict] = []
    seen: set[str] = set()

    for row in range(row1, row2 + 1):
        for col in range(col1, col2 + 1):
            try:
                cell = ws.Cells(row, col)
                if not bool(cell.MergeCells):
                    continue
                area = cell.MergeArea
                area_row1 = int(area.Row)
                area_col1 = int(area.Column)
                area_row2 = area_row1 + int(area.Rows.Count) - 1
                area_col2 = area_col1 + int(area.Columns.Count) - 1
                address = range_address(area_row1, area_col1, area_row2, area_col2)
                if address in seen:
                    continue
                seen.add(address)
                try:
                    text = cell_display_text(area.Cells(1, 1))
                except Exception:
                    text = ""
                results.append(normalize_merge_record(area_row1, area_col1, area_row2, area_col2, text, "excel_com"))
            except Exception:
                continue

    results.sort(key=lambda item: (item["rowStart"], item["columnStart"], item["rowEnd"], item["columnEnd"]))
    return results


def merged_cells_from_openpyxl(workbook_path: Path, sheet_name: str, row1: int, col1: int, row2: int, col2: int) -> list[dict]:
    if load_workbook is None or workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return []

    try:
        wb = load_workbook(workbook_path, data_only=True, read_only=False)
    except Exception:
        return []

    try:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        results: list[dict] = []
        for merge_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merge_range.bounds
            if max_row < row1 or min_row > row2 or max_col < col1 or min_col > col2:
                continue
            value = ws.cell(min_row, min_col).value
            results.append(
                normalize_merge_record(
                    min_row,
                    min_col,
                    max_row,
                    max_col,
                    clipped_text(value),
                    "openpyxl",
                )
            )
        results.sort(key=lambda item: (item["rowStart"], item["columnStart"], item["rowEnd"], item["columnEnd"]))
        return results
    finally:
        try:
            wb.close()
        except Exception:
            pass


def merge_cell_sources(com_merges: list[dict], fallback_merges: list[dict]) -> list[dict]:
    merged: dict[str, dict] = {}
    for item in fallback_merges:
        merged[item["address"]] = item
    for item in com_merges:
        existing = merged.get(item["address"])
        if existing is not None and existing.get("source") != item.get("source"):
            item = dict(item)
            item["fallbackSource"] = existing.get("source", "")
        merged[item["address"]] = item

    results = list(merged.values())
    results.sort(key=lambda item: (item["rowStart"], item["columnStart"], item["rowEnd"], item["columnEnd"]))
    return results


def merged_cells(ws, workbook_path: Path, sheet_name: str, row1: int, col1: int, row2: int, col2: int) -> list[dict]:
    com_merges = merged_cells_from_com(ws, row1, col1, row2, col2)
    fallback_merges = merged_cells_from_openpyxl(workbook_path, sheet_name, row1, col1, row2, col2)
    return merge_cell_sources(com_merges, fallback_merges)


def cell_style(cell) -> dict:
    style: dict = {}
    try:
        if bool(cell.Font.Bold):
            style["bold"] = True
    except Exception:
        pass
    try:
        fill = excel_color_to_hex(cell.Interior.Color)
        if fill:
            style["fillColor"] = fill
    except Exception:
        pass
    try:
        font_color = excel_color_to_hex(cell.Font.Color)
        if font_color:
            style["fontColor"] = font_color
    except Exception:
        pass
    try:
        h_align = alignment_name(cell.HorizontalAlignment)
        if h_align and h_align != "general":
            style["horizontalAlignment"] = h_align
    except Exception:
        pass
    try:
        v_align = alignment_name(cell.VerticalAlignment)
        if v_align:
            style["verticalAlignment"] = v_align
    except Exception:
        pass
    try:
        number_format = clipped_text(cell.NumberFormat, 120)
        if number_format and number_format.lower() != "general":
            style["numberFormat"] = number_format
    except Exception:
        pass
    return style


def normalized_display_text(text: str, raw_value, number_format: str) -> str:
    if not number_format or "%" not in number_format or not isinstance(raw_value, (int, float)):
        return text
    if "%" in text:
        return text
    decimals = 0
    match = re.search(r"0\.(0+)%", number_format)
    if match:
        decimals = len(match.group(1))
    return f"{raw_value * 100:.{decimals}f}%"


def program_extract(ws, row1: int, col1: int, row2: int, col2: int, merges: list[dict]) -> dict:
    cells: list[dict] = []
    merge_by_coord = merge_lookup(merges)
    truncated = False

    for row in range(row1, row2 + 1):
        for col in range(col1, col2 + 1):
            if len(cells) >= MAX_PROGRAM_CELLS_PER_SHEET:
                truncated = True
                break
            try:
                cell = ws.Cells(row, col)
                raw_value = None
                try:
                    raw_value = cell.Value2
                except Exception:
                    raw_value = None
                raw_text = clipped_text(raw_value)
                formula = ""
                try:
                    if bool(cell.HasFormula):
                        formula = clipped_text(cell.Formula, MAX_CELL_TEXT_CHARS)
                except Exception:
                    formula = ""

                merge = merge_by_coord.get((row, col))
                is_merge_anchor = bool(
                    merge
                    and int(merge["rowStart"]) == row
                    and int(merge["columnStart"]) == col
                )

                text = cell_display_text(cell)
                style = cell_style(cell)
                number_format = style.get("numberFormat", "")
                normalized_text = normalized_display_text(text, raw_value, number_format)
                merge_text = str(merge.get("text", "")) if merge else ""
                inherited_merge_text = bool(merge and not is_merge_anchor and merge_text and not text)
                if inherited_merge_text:
                    text = merge_text
                    normalized_text = merge_text

                if not raw_text and not text and not formula and not is_merge_anchor and not inherited_merge_text:
                    continue

                entry = {
                    "row": row,
                    "column": col,
                    "address": cell_address(row, col),
                    "text": text,
                }
                if raw_text and raw_text != text:
                    entry["rawValue"] = raw_text
                if normalized_text and normalized_text != text:
                    entry["normalizedText"] = normalized_text
                if number_format:
                    entry["numberFormat"] = number_format
                if formula:
                    entry["formula"] = formula
                if merge:
                    entry["isMerged"] = True
                    entry["isMergeAnchor"] = is_merge_anchor
                    entry["mergeAddress"] = merge.get("address", "")
                    entry["mergeText"] = merge_text
                    if inherited_merge_text:
                        entry["inheritedMergeText"] = True
                if style:
                    entry["style"] = style
                cells.append(entry)
            except Exception:
                continue
        if truncated:
            break

    row_heights: list[dict] = []
    for row in range(row1, row2 + 1):
        try:
            row_heights.append({"row": row, "height": round(float(ws.Rows(row).RowHeight), 2)})
        except Exception:
            continue

    column_widths: list[dict] = []
    for col in range(col1, col2 + 1):
        try:
            column_widths.append(
                {
                    "column": col,
                    "name": column_name(col),
                    "width": round(float(ws.Columns(col).ColumnWidth), 2),
                }
            )
        except Exception:
            continue

    return {
        "source": "Excel COM cell read; merged cell text is expanded to every cell in the merged range, with inheritedMergeText=true on non-anchor cells.",
        "cellCount": len(cells),
        "truncated": truncated,
        "maxCells": MAX_PROGRAM_CELLS_PER_SHEET,
        "cells": cells,
        "rowHeights": row_heights,
        "columnWidths": column_widths,
    }


def effective_bounds(
    used: tuple[int, int, int, int],
    program: dict,
    merges: list[dict],
    shapes: list[dict],
) -> dict:
    row1, col1, row2, col2 = used
    coords: list[tuple[int, int]] = []
    for cell in program.get("cells", []):
        try:
            coords.append((int(cell["row"]), int(cell["column"])))
        except Exception:
            continue
    for merge in merges:
        if not str(merge.get("text", "")).strip():
            continue
        try:
            coords.append((int(merge["rowStart"]), int(merge["columnStart"])))
            coords.append((int(merge["rowEnd"]), int(merge["columnEnd"])))
        except Exception:
            continue
    for shape in shapes:
        try:
            coords.append((int(shape["rowStart"]), int(shape["columnStart"])))
            coords.append((int(shape["rowEnd"]), int(shape["columnEnd"])))
        except Exception:
            continue

    if not coords:
        return {
            "rowStart": row1,
            "rowEnd": row2,
            "columnStart": col1,
            "columnEnd": col2,
            "source": "usedRange",
        }

    return {
        "rowStart": max(row1, min(row for row, _ in coords)),
        "rowEnd": min(row2, max(row for row, _ in coords)),
        "columnStart": max(col1, min(col for _, col in coords)),
        "columnEnd": min(col2, max(col for _, col in coords)),
        "source": "programExtract+mergedCells+shapes",
    }


def row_profiles(program: dict) -> dict[int, dict]:
    profiles: dict[int, dict] = {}
    for cell in program.get("cells", []):
        try:
            row = int(cell["row"])
            col = int(cell["column"])
        except Exception:
            continue
        text = str(cell.get("text", ""))
        profile = profiles.setdefault(
            row,
            {
                "row": row,
                "columns": set(),
                "numeric": 0,
                "keywords": 0,
                "texts": [],
            },
        )
        profile["columns"].add(col)
        profile["texts"].append(text)
        if is_numeric_like(text) or is_numeric_like(str(cell.get("normalizedText", ""))):
            profile["numeric"] += 1
        if has_table_keyword(text):
            profile["keywords"] += 1

    for profile in profiles.values():
        columns = sorted(profile["columns"])
        profile["count"] = len(columns)
        profile["columnStart"] = columns[0] if columns else 0
        profile["columnEnd"] = columns[-1] if columns else 0
        profile["span"] = (columns[-1] - columns[0] + 1) if columns else 0
    return profiles


def is_table_like_profile(profile: dict) -> bool:
    count = int(profile.get("count", 0))
    numeric = int(profile.get("numeric", 0))
    keywords = int(profile.get("keywords", 0))
    span = int(profile.get("span", 0))
    if count >= 5:
        return True
    if count >= 3 and (numeric >= 2 or keywords > 0 or span >= 4):
        return True
    return False


def group_consecutive_rows(rows: list[int], max_gap: int = 1) -> list[list[int]]:
    if not rows:
        return []
    groups: list[list[int]] = [[rows[0]]]
    for row in rows[1:]:
        if row - groups[-1][-1] <= max_gap + 1:
            groups[-1].append(row)
        else:
            groups.append([row])
    return groups


def region_from_rows(
    rows: list[int],
    profiles: dict[int, dict],
    merges: list[dict],
    effective: dict,
    row_padding_top: int = 2,
    row_padding_bottom: int = 1,
    column_padding: int = 1,
) -> dict:
    row_start = min(rows)
    row_end = max(rows)
    cols: list[int] = []

    padded_row_start = max(int(effective["rowStart"]), row_start - row_padding_top)
    padded_row_end = min(int(effective["rowEnd"]), row_end + row_padding_bottom)

    for row in range(padded_row_start, padded_row_end + 1):
        profile = profiles.get(row)
        if not profile:
            continue
        cols.extend(profile["columns"])

    for merge in merges:
        try:
            mr1 = int(merge["rowStart"])
            mr2 = int(merge["rowEnd"])
            mc1 = int(merge["columnStart"])
            mc2 = int(merge["columnEnd"])
        except Exception:
            continue
        if mr2 < padded_row_start or mr1 > padded_row_end:
            continue
        cols.extend([mc1, mc2])

    if not cols:
        cols = [int(effective["columnStart"]), int(effective["columnEnd"])]

    column_start = max(int(effective["columnStart"]), min(cols) - column_padding)
    column_end = min(int(effective["columnEnd"]), max(cols) + column_padding)
    return {
        "kind": "table",
        "rowStart": padded_row_start,
        "rowEnd": padded_row_end,
        "columnStart": column_start,
        "columnEnd": column_end,
        "sourceRows": f"{row_start}-{row_end}",
        "reason": "multiple aligned cells/numeric values/table header keywords",
    }


def shape_render_regions(shapes: list[dict], effective: dict) -> list[dict]:
    regions: list[dict] = []
    for shape in shapes:
        try:
            row_start = max(int(effective["rowStart"]), int(shape["rowStart"]) - 1)
            row_end = min(int(effective["rowEnd"]), int(shape["rowEnd"]) + 1)
            col_start = max(int(effective["columnStart"]), int(shape["columnStart"]) - 1)
            col_end = min(int(effective["columnEnd"]), int(shape["columnEnd"]) + 1)
        except Exception:
            continue
        regions.append(
            {
                "kind": "image_or_shape",
                "rowStart": row_start,
                "rowEnd": row_end,
                "columnStart": col_start,
                "columnEnd": col_end,
                "sourceShape": shape.get("address", ""),
                "reason": "Excel shape/picture area",
            }
        )
    return regions


def detect_render_regions(
    program: dict,
    merges: list[dict],
    shapes: list[dict],
    effective: dict,
) -> tuple[list[dict], list[dict], list[str]]:
    profiles = row_profiles(program)
    table_rows = sorted(row for row, profile in profiles.items() if is_table_like_profile(profile))
    warnings: list[str] = []
    regions: list[dict] = []

    for rows in group_consecutive_rows(table_rows):
        numeric_total = sum(int(profiles[row].get("numeric", 0)) for row in rows if row in profiles)
        cell_total = sum(int(profiles[row].get("count", 0)) for row in rows if row in profiles)
        keyword_total = sum(int(profiles[row].get("keywords", 0)) for row in rows if row in profiles)
        if len(rows) < 2 and numeric_total < 2 and keyword_total == 0:
            continue
        if cell_total < 6:
            continue
        regions.append(region_from_rows(rows, profiles, merges, effective))

    regions.extend(shape_render_regions(shapes, effective))

    compacted: list[dict] = []
    for region in sorted(regions, key=lambda item: (item["rowStart"], item["columnStart"], item["rowEnd"], item["columnEnd"])):
        if (
            compacted
            and (
                should_merge_table_regions(compacted[-1], region)
                or (
                    region.get("kind") != "table"
                    and region.get("kind") == compacted[-1].get("kind")
                    and regions_overlap_or_touch(compacted[-1], region, row_gap=0, col_gap=0)
                )
            )
        ):
            compacted[-1] = merge_region(compacted[-1], region)
        else:
            compacted.append(region)

    if len(compacted) > MAX_RENDER_REGIONS_PER_SHEET:
        warnings.append(
            f"Detected {len(compacted)} render regions; only the first {MAX_RENDER_REGIONS_PER_SHEET} were rendered."
        )
        compacted = compacted[:MAX_RENDER_REGIONS_PER_SHEET]

    covered_rows: set[int] = set()
    for region in compacted:
        for row in range(int(region["rowStart"]), int(region["rowEnd"]) + 1):
            covered_rows.add(row)

    text_rows = sorted(row for row, profile in profiles.items() if row not in covered_rows and int(profile.get("count", 0)) > 0)
    text_only_ranges: list[dict] = []
    for rows in group_consecutive_rows(text_rows, max_gap=0):
        cols: list[int] = []
        for row in rows:
            cols.extend(profiles[row]["columns"])
        if not cols:
            continue
        text_only_ranges.append(
            {
                "rowStart": min(rows),
                "rowEnd": max(rows),
                "columnStart": min(cols),
                "columnEnd": max(cols),
                "rangeAddress": range_address(min(rows), min(cols), max(rows), max(cols)),
                "reason": "text-only; exact text is in programExtract and no PNG was rendered",
            }
        )

    for index, region in enumerate(compacted, start=1):
        region["renderRegionIndex"] = index
        region["rangeAddress"] = row_col_range_address(region)

    return compacted, text_only_ranges, warnings


def merge_semantics_for_sheet(merges: list[dict], render_regions: list[dict]) -> list[dict]:
    semantics: list[dict] = []
    for merge in merges:
        try:
            row_start = int(merge["rowStart"])
            row_end = int(merge["rowEnd"])
            col_start = int(merge["columnStart"])
            col_end = int(merge["columnEnd"])
        except Exception:
            continue

        row_span = row_end - row_start + 1
        col_span = col_end - col_start + 1
        text = str(merge.get("text", "")).strip()
        role = "merged_cell"
        meaning = "Excel merged range preserved from workbook"
        if col_span >= 3 and row_span <= 2:
            if any(regions_overlap_or_touch(merge, region) and region.get("kind") == "table" for region in render_regions):
                role = "column_group_or_section_header"
                meaning = "horizontal merged label; may define a table section or parent column group"
            else:
                role = "section_title"
                meaning = "wide horizontal merged title/section outside detected table image regions"
        elif row_span >= 2 and col_span == 1:
            if has_table_keyword(text):
                role = "vertical_column_header"
                meaning = "vertical merged table header"
            else:
                role = "body_inherited_value"
                meaning = "merged value inherited by rows in this range"
        elif row_span >= 2 and col_span >= 2:
            role = "block_label"
            meaning = "multi-row/multi-column merged block"

        semantics.append(
            {
                "text": text,
                "range": merge.get("address", ""),
                "source": "excel",
                "roleHint": role,
                "meaningHint": meaning,
                "rowStart": row_start,
                "rowEnd": row_end,
                "columnStart": col_start,
                "columnEnd": col_end,
                "appliesTo": [f"rows {row_start}-{row_end}", f"columns {column_name(col_start)}-{column_name(col_end)}"],
            }
        )
    return semantics


def cell_text(cell: dict) -> str:
    return str(cell.get("normalizedText") or cell.get("text") or "").strip()


def cells_by_row(program: dict) -> dict[int, dict[int, dict]]:
    rows: dict[int, dict[int, dict]] = {}
    for cell in program.get("cells", []):
        try:
            row = int(cell["row"])
            col = int(cell["column"])
        except Exception:
            continue
        rows.setdefault(row, {})[col] = cell
    return rows


def row_matrix(
    row_numbers: list[int],
    by_row: dict[int, dict[int, dict]],
    col_start: int,
    col_end: int,
) -> list[list[str]]:
    matrix: list[list[str]] = []
    for row in row_numbers:
        values: list[str] = []
        for col in range(col_start, col_end + 1):
            values.append(cell_text(by_row.get(row, {}).get(col, {})))
        if any(value.strip() for value in values):
            matrix.append(values)
    return matrix


def row_cell_objects(
    row: int,
    by_row: dict[int, dict[int, dict]],
    col_start: int,
    col_end: int,
    header_labels: dict[int, str],
) -> list[dict]:
    values: list[dict] = []
    for col in range(col_start, col_end + 1):
        cell = by_row.get(row, {}).get(col)
        if not cell:
            continue
        text = cell_text(cell)
        if not text:
            continue
        item = {
            "address": cell.get("address", cell_address(row, col)),
            "column": column_name(col),
            "header": header_labels.get(col, ""),
            "text": text,
        }
        raw_value = str(cell.get("rawValue", "")).strip()
        if raw_value and raw_value != text:
            item["rawValue"] = raw_value
        merge_address = str(cell.get("mergeAddress", "")).strip()
        if merge_address:
            item["mergeAddress"] = merge_address
        values.append(item)
    return values


def combine_header_labels(header_rows: list[int], by_row: dict[int, dict[int, dict]], col_start: int, col_end: int) -> dict[int, str]:
    labels: dict[int, str] = {}
    for col in range(col_start, col_end + 1):
        parts: list[str] = []
        for row in header_rows:
            text = cell_text(by_row.get(row, {}).get(col, {}))
            if text and text not in parts:
                parts.append(text)
        labels[col] = " / ".join(parts)
    return labels


def first_data_row(rows: list[int], profiles: dict[int, dict]) -> int | None:
    for row in rows:
        profile = profiles.get(row, {})
        if int(profile.get("numeric", 0)) >= 2 and int(profile.get("count", 0)) >= 3:
            return row
    for row in rows:
        profile = profiles.get(row, {})
        if int(profile.get("numeric", 0)) >= 1 and int(profile.get("count", 0)) >= 3:
            return row
    return rows[1] if len(rows) > 1 else (rows[0] if rows else None)


def merge_overlaps_region(merge: dict, region: dict) -> bool:
    return not (
        int(merge["rowEnd"]) < int(region["rowStart"])
        or int(merge["rowStart"]) > int(region["rowEnd"])
        or int(merge["columnEnd"]) < int(region["columnStart"])
        or int(merge["columnStart"]) > int(region["columnEnd"])
    )


def table_merged_cells(merges: list[dict], semantics: list[dict], region: dict) -> list[dict]:
    semantics_by_range = {item.get("range", ""): item for item in semantics}
    results: list[dict] = []
    for merge in merges:
        if not merge_overlaps_region(merge, region):
            continue
        address = str(merge.get("address", ""))
        semantic = semantics_by_range.get(address, {})
        results.append(
            {
                "text": str(merge.get("text", "")).strip(),
                "range": address,
                "source": "excel",
                "roleHint": semantic.get("roleHint", ""),
                "meaningHint": semantic.get("meaningHint", ""),
                "appliesTo": semantic.get("appliesTo", []),
            }
        )
    return results


def confidence_for_table(
    region: dict,
    rows: list[int],
    header_rows: list[int],
    body_rows: list[int],
    profiles: dict[int, dict],
    merged_cells: list[dict],
) -> tuple[int, list[str], list[str]]:
    score = 20
    reasons: list[str] = ["deterministic table candidate from aligned Excel cells"]
    issues: list[str] = []
    keyword_total = sum(int(profiles.get(row, {}).get("keywords", 0)) for row in header_rows + body_rows)
    numeric_body_rows = [row for row in body_rows if int(profiles.get(row, {}).get("numeric", 0)) >= 2]
    horizontal_merges = [item for item in merged_cells if "column_group" in str(item.get("roleHint", ""))]
    vertical_merges = [item for item in merged_cells if str(item.get("roleHint", "")) in {"vertical_column_header", "body_inherited_value"}]

    if keyword_total >= 3:
        score += 20
        reasons.append(f"header/table keywords matched {keyword_total} time(s)")
    elif keyword_total > 0:
        score += 10
        reasons.append("some table keywords matched")
    else:
        issues.append("few/no table header keywords")

    if len(numeric_body_rows) >= 2:
        score += 20
        reasons.append(f"{len(numeric_body_rows)} numeric data row(s) detected")
    elif len(numeric_body_rows) == 1:
        score += 8
        issues.append("only one numeric data row")
    else:
        issues.append("no repeated numeric body rows")

    if horizontal_merges:
        score += 12
        reasons.append("horizontal merged header/group cells detected")
    if vertical_merges:
        score += 10
        reasons.append("vertical merged headers or inherited body values detected")

    body_patterns = [set(profiles.get(row, {}).get("columns", set())) for row in body_rows if profiles.get(row)]
    if len(body_patterns) >= 2:
        first = body_patterns[0]
        similarities: list[float] = []
        for pattern in body_patterns[1:]:
            union = len(first | pattern)
            similarities.append((len(first & pattern) / union) if union else 0)
        if similarities and sum(similarities) / len(similarities) >= 0.65:
            score += 12
            reasons.append("body row column pattern is stable")
        else:
            score -= 10
            issues.append("body row column pattern varies")

    if not header_rows:
        score -= 15
        issues.append("header rows were not separated")
    if len(rows) < 2:
        score -= 20
        issues.append("table candidate has fewer than two populated rows")

    if str(region.get("kind", "")) != "table":
        score -= 20
        issues.append("candidate was not classified as table")

    return max(0, min(100, score)), reasons, issues


def title_for_table(region: dict, rows: list[int], by_row: dict[int, dict[int, dict]]) -> str:
    row_start = min(rows) if rows else int(region["rowStart"])
    for row in range(row_start - 1, max(0, row_start - 4), -1):
        texts = [cell_text(cell) for _, cell in sorted(by_row.get(row, {}).items())]
        texts = [text for text in texts if text]
        if texts:
            return " ".join(texts[:3])
    return f"Table {region.get('rangeAddress', row_col_range_address(region))}"


def build_table_candidate(sheet: dict, region: dict) -> dict | None:
    program = sheet.get("programExtract", {})
    by_row = cells_by_row(program)
    profiles = row_profiles(program)
    col_start = int(region["columnStart"])
    col_end = int(region["columnEnd"])
    rows = [
        row for row in range(int(region["rowStart"]), int(region["rowEnd"]) + 1)
        if any(col_start <= col <= col_end and cell_text(cell) for col, cell in by_row.get(row, {}).items())
    ]
    if not rows:
        return None

    data_row = first_data_row(rows, profiles)
    if data_row is None:
        return None

    header_rows = [row for row in rows if row < data_row]
    if not header_rows and len(rows) > 1:
        header_rows = [rows[0]]
        body_rows = rows[1:]
    else:
        body_rows = [row for row in rows if row >= data_row]

    header_labels = combine_header_labels(header_rows, by_row, col_start, col_end)
    merged_cells = table_merged_cells(sheet.get("mergedCells", []), sheet.get("mergeSemantics", []), region)
    confidence, reasons, issues = confidence_for_table(region, rows, header_rows, body_rows, profiles, merged_cells)

    normalized_rows: list[dict] = []
    for row in body_rows:
        cells = row_cell_objects(row, by_row, col_start, col_end, header_labels)
        if not cells:
            continue
        numeric_count = sum(1 for item in cells if is_numeric_like(item.get("text", "")))
        normalized_rows.append(
            {
                "sourceDisplayRows": f"row {row}",
                "cells": cells,
                "numericCellCount": numeric_count,
                "inheritedFields": [],
            }
        )

    return {
        "title": title_for_table(region, rows, by_row),
        "sourceRange": region.get("rangeAddress", row_col_range_address(region)),
        "confidence": confidence,
        "confidenceReasons": reasons,
        "confidenceIssues": issues,
        "needsAiRefine": confidence < 80,
        "layout": {
            "headerRows": row_matrix(header_rows, by_row, col_start, col_end),
            "bodyRows": row_matrix(body_rows, by_row, col_start, col_end),
        },
        "mergedCells": merged_cells,
        "normalizedRows": normalized_rows,
        "headers": [header_labels.get(col, "") or column_name(col) for col in range(col_start, col_end + 1)],
        "rows": row_matrix(body_rows, by_row, col_start, col_end),
        "notes": ["Deterministic table candidate generated before AI refinement."],
    }


def visible_text_items(sheet: dict, table_regions: list[dict]) -> list[dict]:
    items: list[dict] = []
    order = 1
    for cell in sorted(sheet.get("programExtract", {}).get("cells", []), key=lambda item: (item.get("row", 0), item.get("column", 0))):
        text = cell_text(cell)
        if not text:
            continue
        cell_region = {
            "rowStart": int(cell.get("row", 0)),
            "rowEnd": int(cell.get("row", 0)),
            "columnStart": int(cell.get("column", 0)),
            "columnEnd": int(cell.get("column", 0)),
        }
        if any(regions_overlap_or_touch(cell_region, region) for region in table_regions):
            continue
        role = "paragraph" if len(text) >= 20 else "table_text"
        if has_table_keyword(text):
            role = "label"
        if re.match(r"^[IVX]+\.", text, flags=re.IGNORECASE):
            role = "section_header"
        items.append({"order": order, "role": role, "text": text})
        order += 1
        if len(items) >= 300:
            break
    return items


def build_deterministic_structure(request: dict) -> dict:
    output_sheets: list[dict] = []
    sheet_confidences: list[int] = []
    needs_ai = False
    notes: list[str] = [
        "Fast deterministic structure generated from Excel COM/openpyxl before Codex refinement.",
        "All source merged cells are preserved; Codex refinement is skipped only when confidence is high.",
    ]

    for sheet in request.get("sheets", []):
        table_regions = [chunk for chunk in sheet.get("chunks", []) if chunk.get("kind") == "table"]
        tables: list[dict] = []
        for region in table_regions:
            table = build_table_candidate(sheet, region)
            if table is not None:
                tables.append(table)

        if tables:
            sheet_confidence = min(table.get("confidence", 0) for table in tables)
        elif sheet.get("programExtract", {}).get("cellCount", 0) > 0:
            sheet_confidence = 55
        else:
            sheet_confidence = 20

        if sheet_confidence < 80:
            needs_ai = True

        sheet_confidences.append(sheet_confidence)
        output_sheets.append(
            {
                "sheetIndex": sheet.get("sheetIndex", 0),
                "sheetName": sheet.get("sheetName", ""),
                "confidence": sheet_confidence,
                "needsAiRefine": sheet_confidence < 80,
                "chunksRead": [chunk.get("chunkIndex", 0) for chunk in sheet.get("chunks", [])],
                "visibleText": visible_text_items(sheet, table_regions),
                "sections": [],
                "tables": tables,
                "unreadableAreas": [] if tables else ["No confident table candidate was generated before AI refinement."],
            }
        )

    overall_confidence = min(sheet_confidences) if sheet_confidences else 0
    return {
        "fileName": request.get("fileName", ""),
        "extractionMode": "fast_deterministic",
        "confidence": overall_confidence,
        "needsAiRefine": needs_ai or overall_confidence < 80,
        "sheets": output_sheets,
        "extractionNotes": notes,
    }


def open_workbook_with_retries(excel, workbook_path: Path):
    last_error: Exception | None = None
    path = str(workbook_path)
    missing = pythoncom.Missing
    for attempt in range(1, 7):
        try:
            excel.Workbooks.Open(
                path,
                0,      # UpdateLinks
                True,   # ReadOnly
                missing, # Format
                missing, # Password
                missing, # WriteResPassword
                True,   # IgnoreReadOnlyRecommended
                missing, # Origin
                missing, # Delimiter
                False,  # Editable
                False,  # Notify
                missing, # Converter
                False,  # AddToMru
                True,   # Local
                0,      # CorruptLoad
            )
            return excel.Workbooks(int(excel.Workbooks.Count))
        except Exception as exc:
            last_error = exc
            time.sleep(min(1.5 * attempt, 8))
    raise RuntimeError(f"Excel failed to open staged workbook after retries: {last_error}")


def windows_long_path(path: Path) -> str:
    text = str(path.absolute())
    if not sys.platform.startswith("win") or text.startswith("\\\\?\\"):
        return text
    if text.startswith("\\\\"):
        return "\\\\?\\UNC\\" + text[2:]
    return "\\\\?\\" + text


def stage_workbook_for_excel(workbook_path: Path) -> tuple[TemporaryDirectory[str], Path]:
    temp_dir = TemporaryDirectory(prefix="jino_excel_render_")
    suffix = workbook_path.suffix if workbook_path.suffix.lower() in {".xlsx", ".xlsm", ".xls"} else ".xlsx"
    staged_path = Path(temp_dir.name) / f"workbook{suffix}"
    shutil.copy2(windows_long_path(workbook_path), staged_path)
    return temp_dir, staged_path


def render_workbook(source_path: Path, staged_path: Path, image_dir: Path, chunk_rows: int, overlap_rows: int, render_images: bool = True) -> dict:
    progress(f"Excel COM starting: {source_path.name}")
    image_dir.mkdir(parents=True, exist_ok=True)
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AskToUpdateLinks = False
    try:
        excel.ScreenUpdating = False
    except Exception:
        pass
    try:
        excel.EnableEvents = False
    except Exception:
        pass
    try:
        excel.Calculation = -4135  # xlCalculationManual
    except Exception:
        pass
    try:
        excel.AutomationSecurity = 3
    except Exception:
        pass

    try:
        progress("Opening workbook through Excel COM...")
        wb = open_workbook_with_retries(excel, staged_path)
        sheet_count = int(wb.Worksheets.Count)
        progress(f"Workbook opened. Sheets: {sheet_count}")
        request = {
            "fileName": source_path.name,
            "workbookPath": str(source_path),
            "renderSourcePath": str(staged_path),
            "renderedAt": datetime.now().isoformat(timespec="seconds"),
            "readMode": "Use programExtract first for exact displayed cells. Merged cell text is expanded to every cell in the merged range; inheritedMergeText marks non-anchor inherited values. Use mergedCells for actual Excel merged ranges (Excel COM, supplemented by openpyxl). Use rendered images only for visual verification. Structure extraction only.",
            "sheets": [],
        }

        slug = safe_name(source_path.name)
        workbook_image_dir = image_dir / slug
        workbook_image_dir.mkdir(parents=True, exist_ok=True)

        try:
            for sheet_index in range(1, sheet_count + 1):
                ws = wb.Worksheets(sheet_index)
                ws.Activate()
                row1, col1, row2, col2 = used_bounds(ws)
                sheet_name = str(ws.Name)
                progress(f"Sheet {sheet_index}/{sheet_count} '{sheet_name}': reading cells {range_address(row1, col1, row2, col2)}")
                merges = merged_cells(ws, staged_path, sheet_name, row1, col1, row2, col2)
                shapes = shape_ranges(ws, row1, col1, row2, col2)
                program = program_extract(ws, row1, col1, row2, col2, merges)
                effective = effective_bounds((row1, col1, row2, col2), program, merges, shapes)
                render_regions, text_only_ranges, render_warnings = detect_render_regions(program, merges, shapes, effective)
                progress(
                    f"Sheet {sheet_index}/{sheet_count} '{sheet_name}': "
                    f"{program.get('cellCount', 0)} cells, {len(merges)} merged ranges"
                )
                if not render_images:
                    render_regions = []
                    render_warnings = [*render_warnings, "Image rendering skipped for text-only extraction."]
                sheet = {
                    "sheetIndex": sheet_index,
                    "sheetName": sheet_name,
                    "usedRange": {
                        "rowStart": row1,
                        "rowEnd": row2,
                        "columnStart": col1,
                        "columnEnd": col2,
                    },
                    "effectiveRange": {
                        **effective,
                        "rangeAddress": row_col_range_address(effective),
                    },
                    "mergedCells": merges,
                    "mergeSemantics": merge_semantics_for_sheet(merges, render_regions),
                    "shapes": shapes,
                    "programExtract": program,
                    "renderPolicy": {
                        "mode": "table_and_visual_regions_only",
                        "description": "Text-only areas are kept in programExtract but are not rendered as PNG. PNG is created only for detected table/shape regions that benefit from visual verification.",
                        "renderedRegionCount": len(render_regions),
                        "textOnlyRangeCount": len(text_only_ranges),
                        "warnings": render_warnings,
                    },
                    "textOnlyRanges": text_only_ranges,
                    "chunks": [],
                }

                chunk_index = 1
                if render_regions:
                    progress(f"Sheet {sheet_index}/{sheet_count} '{sheet_name}': rendering {len(render_regions)} visual region(s)")
                for region in render_regions:
                    start = int(region["rowStart"])
                    end = int(region["rowEnd"])
                    region_col1 = int(region["columnStart"])
                    region_col2 = int(region["columnEnd"])
                    region_kind = safe_name(region.get("kind", "region"))
                    image_path = workbook_image_dir / (
                        f"sheet_{sheet_index:02d}_chunk_{chunk_index:03d}_{region_kind}_"
                        f"r{start}-{end}_c{region_col1}-{region_col2}.png"
                    )
                    progress(
                        f"Sheet {sheet_index}/{sheet_count} '{sheet_name}': "
                        f"rendering region {chunk_index}/{len(render_regions)} {range_address(start, region_col1, end, region_col2)}"
                    )
                    width, height = copy_range_png(ws, start, region_col1, end, region_col2, image_path)
                    sheet["chunks"].append(
                        {
                            "chunkIndex": chunk_index,
                            "kind": region.get("kind", "region"),
                            "rangeAddress": region.get("rangeAddress", range_address(start, region_col1, end, region_col2)),
                            "rowStart": start,
                            "rowEnd": end,
                            "columnStart": region_col1,
                            "columnEnd": region_col2,
                            "reason": region.get("reason", ""),
                            "sourceRows": region.get("sourceRows", ""),
                            "sourceShape": region.get("sourceShape", ""),
                            "imagePath": str(image_path),
                            "imageWidth": width,
                            "imageHeight": height,
                        }
                    )
                    chunk_index += 1

                request["sheets"].append(sheet)
        finally:
            progress("Closing workbook...")
            wb.Close(False)

        progress("Building deterministic structure from extracted cells...")
        deterministic = build_deterministic_structure(request)
        request["deterministicStructure"] = deterministic
        request["deterministicSummary"] = {
            "confidence": deterministic.get("confidence", 0),
            "needsAiRefine": deterministic.get("needsAiRefine", True),
            "mode": deterministic.get("extractionMode", "fast_deterministic"),
            "sheetCount": len(deterministic.get("sheets", [])),
            "tableCount": sum(len(sheet.get("tables", [])) for sheet in deterministic.get("sheets", [])),
        }
        progress("Excel COM extraction finished.")
        return request
    finally:
        excel.Quit()


def main() -> int:
    if len(sys.argv) < 4:
        print("usage: input_data_test_render_excel.py workbookPath imageDir renderedRequestPath [chunkRows] [overlapRows]", file=sys.stderr)
        return 2

    workbook_path = Path(sys.argv[1])
    image_dir = Path(sys.argv[2])
    request_path = Path(sys.argv[3])
    chunk_rows = int(sys.argv[4]) if len(sys.argv) > 4 else 36
    overlap_rows = int(sys.argv[5]) if len(sys.argv) > 5 else 3
    render_images = "--text-only" not in sys.argv[6:] and "--no-images" not in sys.argv[6:]
    global _PROGRESS_PATH
    if "--progress" in sys.argv:
        idx = sys.argv.index("--progress")
        if idx + 1 < len(sys.argv):
            _PROGRESS_PATH = Path(sys.argv[idx + 1])

    temp_dir: TemporaryDirectory[str] | None = None
    staged_path: Path | None = None
    try:
        progress("Staging workbook for Excel COM...")
        temp_dir, staged_path = stage_workbook_for_excel(workbook_path)
        result = render_workbook(workbook_path, staged_path, image_dir, chunk_rows, overlap_rows, render_images)
        request_path.parent.mkdir(parents=True, exist_ok=True)
        progress("Writing rendered request JSON...")
        request_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        print(request_path)
        return 0
    except Exception as exc:
        request_path.parent.mkdir(parents=True, exist_ok=True)
        failure = {
            "fileName": workbook_path.name,
            "workbookPath": str(workbook_path),
            "workbookPathLength": len(str(workbook_path)),
            "stagedPath": str(staged_path or ""),
            "error": str(exc),
            "traceback": traceback.format_exc(),
        }
        request_path.write_text(json.dumps(failure, ensure_ascii=False, indent=2), encoding="utf-8")
        print(str(exc), file=sys.stderr)
        return 1
    finally:
        if temp_dir is not None:
            temp_dir.cleanup()


if __name__ == "__main__":
    raise SystemExit(main())
